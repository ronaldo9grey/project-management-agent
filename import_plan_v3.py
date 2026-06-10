#!/usr/bin/env python3
"""
项目计划版本导入脚本 - 隆林铝厂空压机集中控制项目研究 V3
"""
import openpyxl
import psycopg2
import psycopg2.extras
from datetime import datetime, date, timedelta
from typing import Dict, List, Optional, Tuple
import os
import sys

# 配置
DB_CONFIG = {
    'host': 'localhost',
    'database': 'project_cost_tracking',
    'user': 'yjydb',
    'password': 'qv52A03xcxAQCoDglUJelm4Sb'
}

PROJECT_ID = '20'
NEW_VERSION = 'V3'
EXCEL_FILE = '/home/ubuntu/info/隆林铝厂空压站集中控制运行系统进度表.xlsx'

def excel_serial_to_date(serial) -> Optional[date]:
    """Excel序列号转日期"""
    if serial is None or serial == '':
        return None
    try:
        serial = int(serial)
        base = date(1899, 12, 30)
        return base + timedelta(days=serial)
    except:
        return None

def parse_excel_file() -> List[Dict]:
    """解析Excel文件，返回任务列表"""
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws = wb['Sheet1']
    
    tasks = []
    current_phase = None
    current_phase_num = 0
    current_node = None
    row_idx = 3
    
    while row_idx <= ws.max_row:
        phase = ws.cell(row=row_idx, column=1).value
        node = ws.cell(row=row_idx, column=2).value
        task_name = ws.cell(row=row_idx, column=3).value
        category = ws.cell(row=row_idx, column=4).value
        start = ws.cell(row=row_idx, column=5).value
        end = ws.cell(row=row_idx, column=6).value
        
        if phase:
            current_phase = phase
            current_phase_num = int(phase.split('.')[0])
        
        if task_name and category == '计划':
            plan_start = excel_serial_to_date(start)
            plan_end = excel_serial_to_date(end)
            actual_start = None
            actual_end = None
            
            # 检查下一行是否有实际时间
            if row_idx + 1 <= ws.max_row:
                next_cat = ws.cell(row=row_idx + 1, column=4).value
                if next_cat == '实际':
                    actual_start = excel_serial_to_date(ws.cell(row=row_idx + 1, column=5).value)
                    actual_end = excel_serial_to_date(ws.cell(row=row_idx + 1, column=6).value)
                    row_idx += 1
            
            # 解析任务编号
            task_num = task_name.split()[0]  # e.g., "1.1"
            parts = task_num.split('.')
            phase_task = parts[0]
            node_task = parts[1] if len(parts) > 1 else '1'
            
            # 生成task_id
            task_id = f"{PROJECT_ID}_{phase_task}_{node_task}_{NEW_VERSION}"
            
            tasks.append({
                'task_id': task_id,
                'task_name': task_name,
                'phase': current_phase,
                'phase_num': current_phase_num,
                'node': current_node or node,
                'plan_start': plan_start,
                'plan_end': plan_end,
                'actual_start': actual_start,
                'actual_end': actual_end,
                'progress': 100.0 if actual_end else 0.0,
                'status': '已完成' if actual_end else '未开始'
            })
        
        row_idx += 1
    
    return tasks

def import_v3_tasks(tasks: List[Dict], dry_run: bool = False):
    """导入V3任务"""
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    
    try:
        # 1. 将现有V2标记为非最新
        print(f"\n[步骤1] 将项目{PROJECT_ID}的V2任务标记为 is_latest=false")
        if not dry_run:
            cur.execute("""
                UPDATE project_tasks 
                SET is_latest = false, update_time = NOW()
                WHERE project_id = %s AND is_latest = true AND is_deleted = false
            """, (PROJECT_ID,))
            print(f"  已更新 {cur.rowcount} 条记录")
        else:
            print("  [DRY RUN] 跳过更新")
        
        # 2. 插入V3任务
        print(f"\n[步骤2] 插入V3任务（共{len(tasks)}个）")
        inserted = 0
        skipped = 0
        
        for task in tasks:
            # 检查是否已存在
            cur.execute("""
                SELECT task_id FROM project_tasks 
                WHERE task_id = %s AND is_deleted = false
            """, (task['task_id'],))
            
            if cur.fetchone():
                skipped += 1
                continue
            
            if not dry_run:
                cur.execute("""
                    INSERT INTO project_tasks (
                        task_id, project_id, task_name, 
                        start_date, end_date, 
                        progress, status, actual_end_date,
                        task_level, parent_task_id,
                        is_latest, is_deleted, create_time, update_time
                    ) VALUES (
                        %s, %s, %s,
                        %s, %s,
                        %s, %s, %s,
                        %s, %s,
                        true, false, NOW(), NOW()
                    )
                """, (
                    task['task_id'],
                    PROJECT_ID,
                    task['task_name'],
                    task['plan_start'],
                    task['plan_end'],
                    task['progress'],
                    task['status'],
                    task['actual_end'],
                    1,  # task_level
                    None,  # parent_task_id
                ))
                inserted += 1
                print(f"  ✓ {task['task_id']} | {task['task_name']} | {task['status']} | {task['plan_start']}~{task['plan_end']}")
            else:
                inserted += 1
                print(f"  [DRY RUN] {task['task_id']} | {task['task_name']} | {task['status']}")
        
        if not dry_run:
            conn.commit()
            print(f"\n✅ 导入完成：插入 {inserted} 条，跳过 {skipped} 条")
        else:
            print(f"\n[DRY RUN] 将插入 {inserted} 条，跳过 {skipped} 条")
        
        return inserted, skipped
        
    except Exception as e:
        conn.rollback()
        print(f"❌ 导入失败: {e}")
        raise
    finally:
        cur.close()
        conn.close()

def verify_import():
    """验证导入结果"""
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    
    print("\n=== 验证导入结果 ===")
    
    # V3任务统计
    cur.execute("""
        SELECT COUNT(*) as cnt,
               COUNT(CASE WHEN status = '已完成' THEN 1 END) as completed,
               COUNT(CASE WHEN status = '未开始' THEN 1 END) as pending
        FROM project_tasks 
        WHERE project_id = %s AND task_id LIKE %s AND is_deleted = false
    """, (PROJECT_ID, f'%{NEW_VERSION}'))
    r = cur.fetchone()
    print(f"V3任务: {r['cnt']} 个")
    print(f"  已完成: {r['completed']} 个")
    print(f"  未开始: {r['pending']} 个")
    
    # 检查is_latest
    cur.execute("""
        SELECT is_latest, COUNT(*) as cnt
        FROM project_tasks 
        WHERE project_id = %s AND is_deleted = false
        GROUP BY is_latest
    """, (PROJECT_ID,))
    print("\nis_latest分布:")
    for r in cur.fetchall():
        print(f"  latest={r['is_latest']}: {r['cnt']} 条")
    
    # 版本分布
    cur.execute("""
        SELECT SUBSTRING(task_id FROM 'V([0-9]+)') as ver, 
               is_latest, COUNT(*) as cnt
        FROM project_tasks 
        WHERE project_id = %s AND is_deleted = false
        GROUP BY ver, is_latest
        ORDER BY ver
    """, (PROJECT_ID,))
    print("\n版本分布:")
    for r in cur.fetchall():
        print(f"  V{r['ver']}: latest={r['is_latest']}, {r['cnt']} 条")
    
    cur.close()
    conn.close()

if __name__ == '__main__':
    import argparse
    
    parser = argparse.ArgumentParser(description='导入项目计划V3')
    parser.add_argument('--dry-run', action='store_true', help='仅预览，不实际执行')
    parser.add_argument('--verify', action='store_true', help='仅验证不导入')
    args = parser.parse_args()
    
    if args.verify:
        verify_import()
    else:
        print("=" * 60)
        print("项目计划V3导入脚本")
        print("=" * 60)
        
        print(f"\n项目: {PROJECT_ID}")
        print(f"Excel: {EXCEL_FILE}")
        print(f"新版本: {NEW_VERSION}")
        
        print("\n解析Excel文件...")
        tasks = parse_excel_file()
        print(f"解析到 {len(tasks)} 个任务")
        
        # 显示已完成任务
        completed = [t for t in tasks if t['status'] == '已完成']
        print(f"\n已完成任务 ({len(completed)} 个):")
        for t in completed[:10]:
            print(f"  {t['task_id']} | {t['task_name']} | 实际:{t['actual_end']}")
        if len(completed) > 10:
            print(f"  ... 还有 {len(completed) - 10} 个")
        
        import_v3_tasks(tasks, dry_run=args.dry_run)
        
        if not args.dry_run:
            verify_import()
