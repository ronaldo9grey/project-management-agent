#!/usr/bin/env python3
"""
生成测试用的项目计划Excel文件
"""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime, timedelta

def create_plan_excel(filename, version=1):
    """创建项目计划Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "项目计划"
    
    # 表头
    headers = ["任务名称", "负责人", "开始日期", "结束日期", "计划工时", "状态", "备注"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    
    # 设置列宽
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 30
    
    # 任务数据 - 版本1
    tasks_v1 = [
        ["需求调研", "admin", datetime(2026, 4, 1), datetime(2026, 4, 5), 40, "未开始", "收集用户需求"],
        ["系统设计", "admin", datetime(2026, 4, 6), datetime(2026, 4, 15), 80, "未开始", "架构设计和UI设计"],
        ["数据库设计", "admin", datetime(2026, 4, 10), datetime(2026, 4, 14), 32, "未开始", "数据库表结构设计"],
        ["后端开发", "admin", datetime(2026, 4, 16), datetime(2026, 5, 15), 200, "未开始", "API开发"],
        ["前端开发", "admin", datetime(2026, 4, 16), datetime(2026, 5, 15), 200, "未开始", "界面开发"],
        ["接口联调", "admin", datetime(2026, 5, 16), datetime(2026, 5, 25), 60, "未开始", "前后端联调"],
        ["测试", "admin", datetime(2026, 5, 26), datetime(2026, 6, 5), 80, "未开始", "功能测试和性能测试"],
        ["部署上线", "admin", datetime(2026, 6, 6), datetime(2026, 6, 10), 24, "未开始", "生产环境部署"],
    ]
    
    # 任务数据 - 版本2（有修改）
    tasks_v2 = [
        ["需求调研", "admin", datetime(2026, 4, 1), datetime(2026, 4, 3), 24, "已完成", "需求已确认"],  # 工时减少
        ["系统设计", "admin", datetime(2026, 4, 4), datetime(2026, 4, 12), 72, "进行中", "架构设计完成"],  # 日期提前，工时减少
        ["数据库设计", "admin", datetime(2026, 4, 8), datetime(2026, 4, 11), 28, "已完成", "表结构设计完成"],  # 工时减少
        ["后端开发", "admin", datetime(2026, 4, 13), datetime(2026, 5, 12), 220, "进行中", "核心API开发中"],  # 工时增加
        ["前端开发", "admin", datetime(2026, 4, 13), datetime(2026, 5, 12), 180, "进行中", "界面开发中"],  # 工时减少
        ["新增-安全模块", "admin", datetime(2026, 5, 13), datetime(2026, 5, 18), 40, "未开始", "安全认证模块"],  # 新增任务
        ["接口联调", "admin", datetime(2026, 5, 13), datetime(2026, 5, 20), 50, "未开始", "联调测试"],  # 工时减少
        ["测试", "admin", datetime(2026, 5, 21), datetime(2026, 5, 30), 80, "未开始", "功能测试和性能测试"],
        # 部署上线任务被删除
    ]
    
    tasks = tasks_v1 if version == 1 else tasks_v2
    
    for row_idx, task in enumerate(tasks, 2):
        for col_idx, value in enumerate(task, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if isinstance(value, datetime):
                cell.number_format = 'YYYY-MM-DD'
    
    wb.save(filename)
    print(f"已创建: {filename} ({len(tasks)}个任务)")
    return len(tasks)

if __name__ == "__main__":
    # 创建两个版本的计划
    count1 = create_plan_excel("/home/ubuntu/.openclaw/workspace/plan_v1.xlsx", version=1)
    count2 = create_plan_excel("/home/ubuntu/.openclaw/workspace/plan_v2.xlsx", version=2)
    print(f"\n版本1: {count1}个任务")
    print(f"版本2: {count2}个任务")
