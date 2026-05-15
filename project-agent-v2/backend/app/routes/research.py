"""
研发项目工时归集模块 API路由
"""

from fastapi import APIRouter, HTTPException, Depends
from pydantic import BaseModel
from typing import List, Optional
from datetime import datetime, date, timedelta
import calendar
import chinese_calendar as cc  # 使用 chinese-calendar 库（包含调休补班日）
from sqlalchemy import text
from app.database import get_connection
from app.logger import get_logger

logger = get_logger(__name__)
router = APIRouter(prefix="/agent/api/agent/research", tags=["research"])


# ============== 工作日判断 ==============

def is_workday(d: date) -> bool:
    """判断是否为工作日（使用 chinese-calendar，包含调休补班日）"""
    return cc.is_workday(d)


def get_workdays_in_month(year: int, month: int) -> List[int]:
    """获取指定月份的工作日列表"""
    days_in_month = calendar.monthrange(year, month)[1]
    workdays = []
    for day in range(1, days_in_month + 1):
        d = date(year, month, day)
        if is_workday(d):
            workdays.append(day)
    return workdays


# ============== Pydantic Models ==============

class ProjectCreate(BaseModel):
    name: str
    project_type: Optional[str] = None
    project_source: Optional[str] = None
    collaboration_form: Optional[str] = None
    expected_output: Optional[str] = None
    technical_goal: Optional[str] = None
    start_date: date
    end_date: date
    current_stage: Optional[str] = None
    budget_2026: Optional[float] = None
    total_budget: Optional[float] = None
    description: Optional[str] = None


class MemberCreate(BaseModel):
    member_name: str
    member_type: str = "contract_member"
    is_manager: bool = False
    employment_end_date: Optional[date] = None


# ============== 全局工时归集 ==============

@router.get("/allocations/monthly")
def get_monthly_allocations(year: int = 2026, month: int = None):
    """获取月度工时大表（人员×日期）"""
    if month is None:
        month = datetime.now().month
    
    days_in_month = calendar.monthrange(year, month)[1]
    
    with get_connection() as conn:
        result = conn.execute(text("""
            SELECT wha.member_name, wha.allocation_date::text as allocation_date, wha.hours,
                   rp.name as project_name, rp.id as project_id
            FROM work_hour_allocation wha
            JOIN research_projects rp ON wha.project_id = rp.id
            WHERE EXTRACT(YEAR FROM wha.allocation_date) = :year
              AND EXTRACT(MONTH FROM wha.allocation_date) = :month
            ORDER BY wha.member_name, wha.allocation_date
        """), {"year": year, "month": month})
        
        data = {}
        for row in result:
            member = row.member_name
            day = int(row.allocation_date.split('-')[2])
            
            if member not in data:
                data[member] = {d: [] for d in range(1, days_in_month + 1)}
            
            data[member][day].append({
                "project_name": row.project_name,
                "hours": row.hours,
                "project_id": row.project_id
            })
        
        summary = {}
        for member, days in data.items():
            total = sum(sum(p["hours"] for p in day_data) if day_data else 0 for day_data in days.values())
            summary[member] = {"total_hours": total, "days_worked": sum(1 for d in days.values() if d)}
        
        return {"year": year, "month": month, "days_in_month": days_in_month, "data": data, "summary": summary}


@router.post("/allocate-all")
def allocate_all_projects():
    """一键归集所有项目的工时（排除周末和节假日，包含补班日）"""
    with get_connection() as conn:
        # 清空旧数据
        conn.execute(text("DELETE FROM work_hour_allocation"))
        
        # 获取所有项目
        projects_result = conn.execute(text("""
            SELECT id, name, start_date, end_date, budget_2026
            FROM research_projects WHERE status = 'ongoing'
        """))
        projects = {p.id: p for p in projects_result}
        
        if not projects:
            return {"message": "没有项目", "total_records": 0}
        
        # 收集每个人员参与的项目信息
        member_info = {}
        
        for proj_id, proj in projects.items():
            members = conn.execute(text("""
                SELECT member_name, is_manager, employment_end_date::text as employment_end_date
                FROM research_project_members WHERE project_id = :id
            """), {"id": proj_id})
            
            for m in members:
                name = m.member_name
                if name not in member_info:
                    member_info[name] = {
                        "projects": [],
                        "is_manager": False,
                        "employment_end": m.employment_end_date
                    }
                
                if m.is_manager:
                    member_info[name]["is_manager"] = True
                
                member_info[name]["projects"].append({
                    "id": proj_id,
                    "name": proj.name,
                    "budget": float(proj.budget_2026 or 100),
                    "start": proj.start_date,
                    "end": proj.end_date
                })
        
        year_start = date(2026, 1, 1)
        year_end = date(2026, 12, 31)
        insert_count = 0
        
        # 对每个人员分配工时
        for member_name, info in member_info.items():
            is_manager = info["is_manager"]
            employment_end = info["employment_end"]
            member_projects = info["projects"]
            
            # 计算有效结束日期
            if employment_end:
                member_end = min(date.fromisoformat(employment_end), year_end)
            else:
                member_end = year_end
            
            # 收集该人员在各项目中的有效工作日
            proj_workdays = {}
            for proj in member_projects:
                proj_id = proj["id"]
                proj_start = max(proj["start"], year_start)
                proj_end = min(proj["end"], member_end)
                
                workdays = []
                current = proj_start
                while current <= proj_end:
                    if is_workday(current):  # 使用 chinese-calendar 判断
                        workdays.append(current)
                    current += timedelta(days=1)
                
                proj_workdays[proj_id] = workdays
            
            # 合并所有项目的工作日
            all_workdays = set()
            for days in proj_workdays.values():
                all_workdays.update(days)
            
            if not all_workdays:
                continue
            
            # 按项目预算比例分配天数
            total_budget = sum(p["budget"] for p in member_projects)
            proj_days_target = {}
            for proj in member_projects:
                ratio = proj["budget"] / total_budget if total_budget > 0 else 1 / len(member_projects)
                proj_days_target[proj["id"]] = int(len(all_workdays) * ratio)
            
            # 补齐差额
            total_assigned = sum(proj_days_target.values())
            diff = len(all_workdays) - total_assigned
            if diff > 0:
                max_proj = max(member_projects, key=lambda p: p["budget"])["id"]
                proj_days_target[max_proj] += diff
            
            # 随机分配日期给各项目
            import random
            date_pool = sorted(list(all_workdays))
            random.shuffle(date_pool)
            
            hours_per_day = 2 if is_manager else 8
            
            for proj in member_projects:
                proj_id = proj["id"]
                days_needed = proj_days_target[proj_id]
                
                proj_valid_dates = [d for d in date_pool if d in proj_workdays[proj_id]]
                assigned_dates = proj_valid_dates[:days_needed]
                
                for d in assigned_dates:
                    conn.execute(text("""
                        INSERT INTO work_hour_allocation (project_id, member_name, allocation_date, hours)
                        VALUES (:pid, :name, :date, :hours)
                    """), {"pid": proj_id, "name": member_name, "date": d, "hours": hours_per_day})
                    insert_count += 1
                
                for d in assigned_dates:
                    if d in date_pool:
                        date_pool.remove(d)
        
        conn.commit()
        logger.info(f"工时归集完成: {insert_count}条")
        
        return {"message": "归集完成", "total_records": insert_count}


@router.get("/allocations/project-summary")
def get_project_summary(year: int = 2026, month: int = None):
    """获取项目维度工时汇总"""
    with get_connection() as conn:
        projects_info = conn.execute(text("""
            SELECT id, name, start_date, end_date, budget_2026
            FROM research_projects WHERE status = 'ongoing'
            ORDER BY id
        """))
        
        projects_base = {p.id: {
            "id": p.id,
            "name": p.name,
            "start_date": str(p.start_date),
            "end_date": str(p.end_date),
            "budget_2026": float(p.budget_2026 or 0),
            "members": {},
            "monthly_total": {},
            "daily_total": {},
            "annual_total": 0
        } for p in projects_info}
        
        if not projects_base:
            return {"year": year, "month": month, "projects": []}
        
        if month is None:
            result = conn.execute(text("""
                SELECT 
                    wha.project_id,
                    wha.member_name,
                    rpm.is_manager,
                    rpm.employment_end_date::text as employment_end,
                    EXTRACT(MONTH FROM wha.allocation_date) as month,
                    COUNT(*) as days,
                    SUM(wha.hours) as total_hours
                FROM work_hour_allocation wha
                LEFT JOIN research_project_members rpm 
                    ON rpm.project_id = wha.project_id AND rpm.member_name = wha.member_name
                WHERE EXTRACT(YEAR FROM wha.allocation_date) = :year
                GROUP BY wha.project_id, wha.member_name, rpm.is_manager, rpm.employment_end_date, EXTRACT(MONTH FROM wha.allocation_date)
                ORDER BY wha.project_id, wha.member_name, month
            """), {"year": year})
            
            for row in result:
                proj_id = row.project_id
                if proj_id not in projects_base:
                    continue
                    
                member_name = row.member_name
                month_num = int(row.month)
                
                if member_name not in projects_base[proj_id]["members"]:
                    projects_base[proj_id]["members"][member_name] = {
                        "is_manager": row.is_manager or False,
                        "employment_end": row.employment_end,
                        "monthly_hours": {},
                        "daily_hours": {},
                        "total_hours": 0
                    }
                
                projects_base[proj_id]["members"][member_name]["monthly_hours"][month_num] = int(row.total_hours)
                projects_base[proj_id]["members"][member_name]["total_hours"] += int(row.total_hours)
                
                if month_num not in projects_base[proj_id]["monthly_total"]:
                    projects_base[proj_id]["monthly_total"][month_num] = 0
                projects_base[proj_id]["monthly_total"][month_num] += int(row.total_hours)
            
        else:
            days_in_month = calendar.monthrange(year, month)[1]
            workdays = get_workdays_in_month(year, month)
            
            result = conn.execute(text("""
                SELECT 
                    wha.project_id,
                    wha.member_name,
                    rpm.is_manager,
                    rpm.employment_end_date::text as employment_end,
                    EXTRACT(DAY FROM wha.allocation_date) as day,
                    SUM(wha.hours) as hours
                FROM work_hour_allocation wha
                LEFT JOIN research_project_members rpm 
                    ON rpm.project_id = wha.project_id AND rpm.member_name = wha.member_name
                WHERE EXTRACT(YEAR FROM wha.allocation_date) = :year
                  AND EXTRACT(MONTH FROM wha.allocation_date) = :month
                GROUP BY wha.project_id, wha.member_name, rpm.is_manager, rpm.employment_end_date, EXTRACT(DAY FROM wha.allocation_date)
                ORDER BY wha.project_id, wha.member_name, day
            """), {"year": year, "month": month})
            
            for row in result:
                proj_id = row.project_id
                if proj_id not in projects_base:
                    continue
                    
                member_name = row.member_name
                day_num = int(row.day)
                
                if member_name not in projects_base[proj_id]["members"]:
                    projects_base[proj_id]["members"][member_name] = {
                        "is_manager": row.is_manager or False,
                        "employment_end": row.employment_end,
                        "monthly_hours": {},
                        "daily_hours": {},
                        "total_hours": 0
                    }
                
                projects_base[proj_id]["members"][member_name]["daily_hours"][day_num] = int(row.hours)
                projects_base[proj_id]["members"][member_name]["total_hours"] += int(row.hours)
                
                if day_num not in projects_base[proj_id]["daily_total"]:
                    projects_base[proj_id]["daily_total"][day_num] = 0
                projects_base[proj_id]["daily_total"][day_num] += int(row.hours)
            
            for proj in projects_base.values():
                proj["workdays"] = workdays
                proj["days_in_month"] = days_in_month
        
        for proj in projects_base.values():
            if month is None:
                proj["annual_total"] = sum(proj["monthly_total"].values())
            else:
                proj["monthly_total_sum"] = sum(proj["daily_total"].values())
        
        return {
            "year": year, 
            "month": month, 
            "projects": list(projects_base.values()), 
            "workdays": get_workdays_in_month(year, month) if month else None
        }


@router.get("/allocations/export")
def export_allocations(year: int = 2026, month: int = None):
    """导出工时表为Excel（多sheet，带样式）"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    import io
    import base64
    
    # 获取项目汇总数据
    summary = get_project_summary(year, month)
    wb = Workbook()
    
    # 样式定义
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    total_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    month_names = ['一月', '二月', '三月', '四月', '五月', '六月', 
                   '七月', '八月', '九月', '十月', '十一月', '十二月']
    
    # ============ Sheet 1: 合计 ============
    ws = wb.active
    ws.title = '合计'
    ws['A1'] = f'{year}年研发项目工时汇总'
    ws['A1'].font = Font(bold=True, size=14)
    
    # 表头
    if month:
        headers = ['项目名称', '人员数', f'{month}月工时(h)']
    else:
        headers = ['项目名称', '人员数'] + month_names + ['年度合计(h)']
    
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_align
    
    # 数据行
    total_all = 0
    for row_idx, proj in enumerate(summary['projects'], 4):
        ws.cell(row=row_idx, column=1, value=proj['name']).border = border
        ws.cell(row=row_idx, column=2, value=len(proj['members'])).border = border
        
        if month:
            h = proj.get('monthly_total_sum', 0)
            c = ws.cell(row=row_idx, column=3, value=h)
            c.border = border
            c.alignment = center_align
            total_all += h
        else:
            for col_idx, m in enumerate(range(1, 13), 3):
                c = ws.cell(row=row_idx, column=col_idx, value=proj['monthly_total'].get(m, 0) or '')
                c.border = border
                c.alignment = center_align
            tc = ws.cell(row=row_idx, column=15, value=proj['annual_total'])
            tc.border = border
            tc.alignment = center_align
            tc.fill = total_fill
            total_all += proj['annual_total']
    
    # 合计行
    total_row = len(summary['projects']) + 4
    ws.cell(row=total_row, column=1, value='合计').font = Font(bold=True)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=total_row, column=c)
        cell.border = border
        cell.fill = total_fill
        cell.font = Font(bold=True)
    ws.cell(row=total_row, column=len(headers), value=total_all)
    
    # 调整列宽
    ws.column_dimensions['A'].width = 40
    for c in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 12
    
    # ============ 各项目Sheet ============
    for proj in summary['projects']:
        if not proj['members']:
            continue
        
        pws = wb.create_sheet(title=proj['name'][:31])
        pws['A1'] = proj['name']
        pws['A1'].font = Font(bold=True, size=12)
        pws['A2'] = f"工期: {proj['start_date']} ~ {proj['end_date']}"
        
        # 表头
        if month:
            workdays = proj.get('workdays', [])
            ph = ['人员', '管理'] + [f'{d}日' for d in workdays] + ['合计(h)']
        else:
            ph = ['人员', '管理'] + month_names + ['合计(h)']
        
        for ci, h in enumerate(ph, 1):
            c = pws.cell(row=4, column=ci, value=h)
            c.font = header_font
            c.fill = header_fill
            c.border = border
            c.alignment = center_align
        
        # 数据行
        for row_idx, (member_name, member_data) in enumerate(proj['members'].items(), 5):
            pws.cell(row=row_idx, column=1, value=member_name).border = border
            pws.cell(row=row_idx, column=2, value='是' if member_data['is_manager'] else '').border = border
            pws.cell(row=row_idx, column=2).alignment = center_align
            
            if month:
                for ci, d in enumerate(workdays, 3):
                    c = pws.cell(row=row_idx, column=ci, value=member_data['daily_hours'].get(d, 0) or '')
                    c.border = border
                    c.alignment = center_align
                tc = pws.cell(row=row_idx, column=len(ph), value=member_data['total_hours'])
            else:
                for ci, m in enumerate(range(1, 13), 3):
                    c = pws.cell(row=row_idx, column=ci, value=member_data['monthly_hours'].get(m, 0) or '')
                    c.border = border
                    c.alignment = center_align
                tc = pws.cell(row=row_idx, column=len(ph), value=member_data['total_hours'])
            
            tc.border = border
            tc.fill = total_fill
            tc.alignment = center_align
        
        # 合计行
        sr = len(proj['members']) + 5
        pws.cell(row=sr, column=1, value='合计').font = Font(bold=True)
        pws.cell(row=sr, column=1).border = border
        
        if month:
            for ci, d in enumerate(workdays, 3):
                c = pws.cell(row=sr, column=ci, value=proj['daily_total'].get(d, 0) or '')
                c.border = border
                c.fill = total_fill
                c.font = Font(bold=True)
            pws.cell(row=sr, column=len(ph), value=proj.get('monthly_total_sum', 0))
        else:
            for ci, m in enumerate(range(1, 13), 3):
                c = pws.cell(row=sr, column=ci, value=proj['monthly_total'].get(m, 0) or '')
                c.border = border
                c.fill = total_fill
                c.font = Font(bold=True)
            pws.cell(row=sr, column=len(ph), value=proj['annual_total'])
        
        pws.cell(row=sr, column=len(ph)).border = border
        pws.cell(row=sr, column=len(ph)).fill = total_fill
        pws.cell(row=sr, column=len(ph)).font = Font(bold=True)
        
        # 调整列宽
        pws.column_dimensions['A'].width = 15
        pws.column_dimensions['B'].width = 8
        for c in range(3, len(ph) + 1):
            pws.column_dimensions[get_column_letter(c)].width = 10
    
    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # 转换为base64
    excel_base64 = base64.b64encode(output.read()).decode('utf-8')
    
    filename = f"研发工时_{year}年"
    if month:
        filename += f"{month}月"
    filename += ".xlsx"
    
    return {
        "excel_data": excel_base64,
        "filename": filename,
        "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    }


# ============== 项目 CRUD
# ============== 项目 CRUD ==============

@router.get("/projects")
def list_projects():
    with get_connection() as conn:
        result = conn.execute(text("""
            SELECT rp.*, (SELECT COUNT(*) FROM research_project_members WHERE project_id = rp.id) as member_count
            FROM research_projects rp ORDER BY rp.id
        """))
        projects = []
        for row in result:
            projects.append({
                "id": row.id, "name": row.name, "project_type": row.project_type,
                "project_source": row.project_source, "collaboration_form": row.collaboration_form,
                "expected_output": row.expected_output, "technical_goal": row.technical_goal,
                "start_date": str(row.start_date), "end_date": str(row.end_date),
                "current_stage": row.current_stage, "budget_2026": float(row.budget_2026) if row.budget_2026 else None,
                "total_budget": float(row.total_budget) if row.total_budget else None,
                "description": row.description, "status": row.status,
                "member_count": row.member_count, "created_at": str(row.created_at)
            })
        return {"projects": projects}


@router.get("/projects/{project_id}")
def get_project(project_id: int):
    with get_connection() as conn:
        result = conn.execute(text("SELECT * FROM research_projects WHERE id = :id"), {"id": project_id})
        row = result.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="项目不存在")
        
        project = {
            "id": row.id, "name": row.name, "project_type": row.project_type,
            "project_source": row.project_source, "collaboration_form": row.collaboration_form,
            "expected_output": row.expected_output, "technical_goal": row.technical_goal,
            "start_date": str(row.start_date), "end_date": str(row.end_date),
            "current_stage": row.current_stage, "budget_2026": float(row.budget_2026) if row.budget_2026 else None,
            "total_budget": float(row.total_budget) if row.total_budget else None,
            "description": row.description, "status": row.status
        }
        
        member_result = conn.execute(text("""
            SELECT id, member_name, member_type, is_manager, employment_end_date::text as employment_end_date
            FROM research_project_members WHERE project_id = :id ORDER BY member_type, member_name
        """), {"id": project_id})
        
        project["members"] = [{"id": m.id, "member_name": m.member_name, "member_type": m.member_type,
                               "is_manager": m.is_manager, "employment_end_date": m.employment_end_date} for m in member_result]
        return project


@router.post("/projects")
def create_project(data: ProjectCreate):
    with get_connection() as conn:
        result = conn.execute(text("""
            INSERT INTO research_projects (name, project_type, project_source, collaboration_form, 
            expected_output, technical_goal, start_date, end_date, current_stage, budget_2026, total_budget, description)
            VALUES (:name, :project_type, :project_source, :collaboration_form, :expected_output,
                    :technical_goal, :start_date, :end_date, :current_stage, :budget_2026, :total_budget, :description)
            RETURNING id
        """), data.dict())
        project_id = result.fetchone().id
        conn.commit()
        return {"id": project_id, "message": "创建成功"}


@router.put("/projects/{project_id}")
def update_project(project_id: int, data: ProjectCreate):
    with get_connection() as conn:
        conn.execute(text("""
            UPDATE research_projects SET name=:name, project_type=:project_type, project_source=:project_source,
            collaboration_form=:collaboration_form, expected_output=:expected_output, technical_goal=:technical_goal,
            start_date=:start_date, end_date=:end_date, current_stage=:current_stage, 
            budget_2026=:budget_2026, total_budget=:total_budget, description=:description, updated_at=NOW()
            WHERE id = :id
        """), {**data.dict(), "id": project_id})
        conn.commit()
        return {"message": "更新成功"}


@router.delete("/projects/{project_id}")
def delete_project(project_id: int):
    with get_connection() as conn:
        result = conn.execute(text("SELECT status FROM research_projects WHERE id = :id"), {"id": project_id})
        row = result.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="项目不存在")
        if row.status in ['ongoing', 'completed']:
            raise HTTPException(status_code=400, detail="进行中或已完成的项目不能删除")
        conn.execute(text("DELETE FROM research_projects WHERE id = :id"), {"id": project_id})
        conn.commit()
        return {"message": "删除成功"}


@router.post("/projects/{project_id}/members")
def add_member(project_id: int, data: MemberCreate):
    with get_connection() as conn:
        conn.execute(text("""
            INSERT INTO research_project_members (project_id, member_name, member_type, is_manager, employment_end_date)
            VALUES (:pid, :name, :type, :manager, :end_date)
        """), {"pid": project_id, "name": data.member_name, "type": data.member_type,
               "manager": data.is_manager, "end_date": data.employment_end_date})
        conn.commit()
        return {"message": "添加成功"}


def get_project_summary(year: int, month: int = None):
    """获取项目汇总数据"""
    with get_connection() as conn:
        projects_info = conn.execute(text("""
            SELECT id, name, start_date, end_date, budget_2026
            FROM research_projects WHERE status = 'ongoing' ORDER BY id
        """))
        
        projects_base = {p.id: {
            "id": p.id, "name": p.name,
            "start_date": str(p.start_date), "end_date": str(p.end_date),
            "budget_2026": float(p.budget_2026 or 0),
            "members": {}, "monthly_total": {}, "daily_total": {}, "annual_total": 0
        } for p in projects_info}
        
        if not projects_base:
            return {"year": year, "month": month, "projects": []}
        
        if month is None:
            result = conn.execute(text("""
                SELECT wha.project_id, wha.member_name, rpm.is_manager,
                       rpm.employment_end_date::text as employment_end,
                       EXTRACT(MONTH FROM wha.allocation_date) as month,
                       SUM(wha.hours) as total_hours
                FROM work_hour_allocation wha
                LEFT JOIN research_project_members rpm 
                    ON rpm.project_id = wha.project_id AND rpm.member_name = wha.member_name
                WHERE EXTRACT(YEAR FROM wha.allocation_date) = :year
                GROUP BY wha.project_id, wha.member_name, rpm.is_manager, rpm.employment_end_date, EXTRACT(MONTH FROM wha.allocation_date)
                ORDER BY wha.project_id, wha.member_name, month
            """), {"year": year})
            
            for row in result:
                proj_id = row.project_id
                if proj_id not in projects_base: continue
                member_name = row.member_name
                month_num = int(row.month)
                
                if member_name not in projects_base[proj_id]["members"]:
                    projects_base[proj_id]["members"][member_name] = {
                        "is_manager": row.is_manager or False, "employment_end": row.employment_end,
                        "monthly_hours": {}, "daily_hours": {}, "total_hours": 0
                    }
                
                projects_base[proj_id]["members"][member_name]["monthly_hours"][month_num] = int(row.total_hours)
                projects_base[proj_id]["members"][member_name]["total_hours"] += int(row.total_hours)
                
                if month_num not in projects_base[proj_id]["monthly_total"]:
                    projects_base[proj_id]["monthly_total"][month_num] = 0
                projects_base[proj_id]["monthly_total"][month_num] += int(row.total_hours)
        else:
            workdays = get_workdays_in_month(year, month)
            
            result = conn.execute(text("""
                SELECT wha.project_id, wha.member_name, rpm.is_manager,
                       rpm.employment_end_date::text as employment_end,
                       EXTRACT(DAY FROM wha.allocation_date) as day,
                       SUM(wha.hours) as hours
                FROM work_hour_allocation wha
                LEFT JOIN research_project_members rpm 
                    ON rpm.project_id = wha.project_id AND rpm.member_name = wha.member_name
                WHERE EXTRACT(YEAR FROM wha.allocation_date) = :year
                  AND EXTRACT(MONTH FROM wha.allocation_date) = :month
                GROUP BY wha.project_id, wha.member_name, rpm.is_manager, rpm.employment_end_date, EXTRACT(DAY FROM wha.allocation_date)
                ORDER BY wha.project_id, wha.member_name, day
            """), {"year": year, "month": month})
            
            for row in result:
                proj_id = row.project_id
                if proj_id not in projects_base: continue
                member_name = row.member_name
                day_num = int(row.day)
                
                if member_name not in projects_base[proj_id]["members"]:
                    projects_base[proj_id]["members"][member_name] = {
                        "is_manager": row.is_manager or False, "employment_end": row.employment_end,
                        "monthly_hours": {}, "daily_hours": {}, "total_hours": 0
                    }
                
                projects_base[proj_id]["members"][member_name]["daily_hours"][day_num] = int(row.hours)
                projects_base[proj_id]["members"][member_name]["total_hours"] += int(row.hours)
                
                if day_num not in projects_base[proj_id]["daily_total"]:
                    projects_base[proj_id]["daily_total"][day_num] = 0
                projects_base[proj_id]["daily_total"][day_num] += int(row.hours)
            
            for proj in projects_base.values():
                proj["workdays"] = workdays
        
        for proj in projects_base.values():
            if month is None:
                proj["annual_total"] = sum(proj["monthly_total"].values())
            else:
                proj["monthly_total_sum"] = sum(proj["daily_total"].values())
        
        return {"year": year, "month": month, "projects": list(projects_base.values())}
