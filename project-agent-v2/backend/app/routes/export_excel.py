@router.get("/allocations/export")
def export_allocations(year: int = 2026, month: int = None):
    """导出工时表为Excel（多sheet，带样式）"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    import io
    import base64
    
    # 获取项目汇总数据
    summary = get_project_summary(year, month)
    wb = Workbook()
    
    # 样式定义
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    total_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    month_names = ['一月', '二月', '三月', '四月', '五月', '六月', 
                   '七月', '八月', '九月', '十月', '十一月', '十二月']
    
    # ============ Sheet 1: 合计 ============
    ws = wb.active
    ws.title = '合计'
    ws['A1'] = f'{year}年研发项目工时汇总'
    ws['A1'].font = Font(bold=True, size=14)
    
    # 表头
    if month:
        headers = ['项目名称', '人员数', f'{month}月工时(h)']
    else:
        headers = ['项目名称', '人员数'] + month_names + ['年度合计(h)']
    
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_align
    
    # 数据行
    total_all = 0
    for row_idx, proj in enumerate(summary['projects'], 4):
        ws.cell(row=row_idx, column=1, value=proj['name']).border = border
        ws.cell(row=row_idx, column=2, value=len(proj['members'])).border = border
        
        if month:
            h = proj.get('monthly_total_sum', 0)
            c = ws.cell(row=row_idx, column=3, value=h)
            c.border = border
            c.alignment = center_align
            total_all += h
        else:
            for col_idx, m in enumerate(range(1, 13), 3):
                c = ws.cell(row=row_idx, column=col_idx, value=proj['monthly_total'].get(m, 0) or '')
                c.border = border
                c.alignment = center_align
            tc = ws.cell(row=row_idx, column=15, value=proj['annual_total'])
            tc.border = border
            tc.alignment = center_align
            tc.fill = total_fill
            total_all += proj['annual_total']
    
    # 合计行
    total_row = len(summary['projects']) + 4
    ws.cell(row=total_row, column=1, value='合计').font = Font(bold=True)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=total_row, column=c)
        cell.border = border
        cell.fill = total_fill
        cell.font = Font(bold=True)
    ws.cell(row=total_row, column=len(headers), value=total_all)
    
    # 调整列宽
    ws.column_dimensions['A'].width = 40
    for c in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 12
    
    # ============ 各项目Sheet ============
    for proj in summary['projects']:
        if not proj['members']:
            continue
        
        pws = wb.create_sheet(title=proj['name'][:31])
        pws['A1'] = proj['name']
        pws['A1'].font = Font(bold=True, size=12)
        pws['A2'] = f"工期: {proj['start_date']} ~ {proj['end_date']}"
        
        # 表头
        if month:
            workdays = proj.get('workdays', [])
            ph = ['人员', '管理'] + [f'{d}日' for d in workdays] + ['合计(h)']
        else:
            ph = ['人员', '管理'] + month_names + ['合计(h)']
        
        for ci, h in enumerate(ph, 1):
            c = pws.cell(row=4, column=ci, value=h)
            c.font = header_font
            c.fill = header_fill
            c.border = border
            c.alignment = center_align
        
        # 数据行
        for row_idx, (member_name, member_data) in enumerate(proj['members'].items(), 5):
            pws.cell(row=row_idx, column=1, value=member_name).border = border
            pws.cell(row=row_idx, column=2, value='是' if member_data['is_manager'] else '').border = border
            pws.cell(row=row_idx, column=2).alignment = center_align
            
            if month:
                for ci, d in enumerate(workdays, 3):
                    c = pws.cell(row=row_idx, column=ci, value=member_data['daily_hours'].get(d, 0) or '')
                    c.border = border
                    c.alignment = center_align
                tc = pws.cell(row=row_idx, column=len(ph), value=member_data['total_hours'])
            else:
                for ci, m in enumerate(range(1, 13), 3):
                    c = pws.cell(row=row_idx, column=ci, value=member_data['monthly_hours'].get(m, 0) or '')
                    c.border = border
                    c.alignment = center_align
                tc = pws.cell(row=row_idx, column=len(ph), value=member_data['total_hours'])
            
            tc.border = border
            tc.fill = total_fill
            tc.alignment = center_align
        
        # 合计行
        sr = len(proj['members']) + 5
        pws.cell(row=sr, column=1, value='合计').font = Font(bold=True)
        pws.cell(row=sr, column=1).border = border
        
        if month:
            for ci, d in enumerate(workdays, 3):
                c = pws.cell(row=sr, column=ci, value=proj['daily_total'].get(d, 0) or '')
                c.border = border
                c.fill = total_fill
                c.font = Font(bold=True)
            pws.cell(row=sr, column=len(ph), value=proj.get('monthly_total_sum', 0))
        else:
            for ci, m in enumerate(range(1, 13), 3):
                c = pws.cell(row=sr, column=ci, value=proj['monthly_total'].get(m, 0) or '')
                c.border = border
                c.fill = total_fill
                c.font = Font(bold=True)
            pws.cell(row=sr, column=len(ph), value=proj['annual_total'])
        
        pws.cell(row=sr, column=len(ph)).border = border
        pws.cell(row=sr, column=len(ph)).fill = total_fill
        pws.cell(row=sr, column=len(ph)).font = Font(bold=True)
        
        # 调整列宽
        pws.column_dimensions['A'].width = 15
        pws.column_dimensions['B'].width = 8
        for c in range(3, len(ph) + 1):
            pws.column_dimensions[get_column_letter(c)].width = 10
    
    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # 转换为base64
    excel_base64 = base64.b64encode(output.read()).decode('utf-8')
    
    filename = f"研发工时_{year}年"
    if month:
        filename += f"{month}月"
    filename += ".xlsx"
    
    return {
        "excel_data": excel_base64,
        "filename": filename,
        "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    }
