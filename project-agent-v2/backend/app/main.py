# 项目管理智能体 - 后端服务
# FastAPI + LangChain/LangGraph + DeepSeek

from fastapi import FastAPI, Request, HTTPException, UploadFile, File, Depends, Form, BackgroundTasks
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import List, Optional, Dict, Any
import json
import os
from datetime import datetime, timedelta
import httpx
import asyncio
from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded

# 路由模块（拆分计划）- 使用相对导入
from .routes.health import router as health_router
from .routes.stats import router as stats_router
from .routes.dashboard import router as dashboard_router
from .routes.research import router as research_router

# LangChain imports
from langchain_openai import ChatOpenAI
from langgraph.graph import StateGraph, END
from langchain_core.messages import SystemMessage, HumanMessage, AIMessage

# AI 线程池工具（防止AI调用阻塞Worker）
try:
    from .task_auto import run_in_thread, AI_EXECUTOR
except ImportError:
    from task_auto import run_in_thread, AI_EXECUTOR

# API限流配置
limiter = Limiter(key_func=get_remote_address)

app = FastAPI(
    title="项目管理智能体",
    description="基于LangChain/LangGraph的项目管理AI服务",
    version="0.1.0"
)

# 添加限流处理器
app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)

# CORS配置
# ============== API限流配置 ==============
limiter = Limiter(key_func=get_remote_address)

# ============== FastAPI应用 ==============
app = FastAPI(
    title="项目智能体API",
    description="项目管理智能助手后端服务",
    version="1.0.0"
)

# 添加限流处理器
app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "https://yjypro.online",      # 生产环境
        "http://localhost:5173",      # 本地开发
        "http://127.0.0.1:5173",      # 本地开发（IP）
        "https://open.feishu.cn",     # 飞书机器人回调
    ],
    allow_credentials=True,
    allow_methods=["GET", "POST", "PUT", "DELETE", "OPTIONS"],
    allow_headers=["Authorization", "Content-Type", "Accept", "Origin", "X-Requested-With"],
)

# 加载环境变量
from dotenv import load_dotenv
load_dotenv()

# 注册路由模块（拆分计划）
# 注意：overview、projects、insight 端点保留在 main.py 中（包含完整数据）
app.include_router(health_router)  # health路由
app.include_router(stats_router)   # stats路由
app.include_router(research_router) # 研发项目工时归集路由
# dashboard路由暂时禁用（避免覆盖完整版端点）
# app.include_router(dashboard_router)

# 配置
class Settings:
    DEEPSEEK_API_KEY = os.getenv("DEEPSEEK_API_KEY", "")
    DEEPSEEK_BASE_URL = os.getenv("DEEPSEEK_BASE_URL", "https://api.deepseek.com/v1")
    DEEPSEEK_MODEL = os.getenv("DEEPSEEK_MODEL", "deepseek-chat")
    REDIS_URL = os.getenv("REDIS_URL", "redis://localhost:6379")
    BACKEND_API_URL = os.getenv("BACKEND_API_URL", "http://127.0.0.1:8000")
    UPLOAD_DIR = "/tmp/project-agent/uploads"
    # JWT配置（与现有后端一致）
    SECRET_KEY = os.getenv("SECRET_KEY", "")
    ALGORITHM = os.getenv("ALGORITHM", "HS256")
    ACCESS_TOKEN_EXPIRE_MINUTES = int(os.getenv("ACCESS_TOKEN_EXPIRE_MINUTES", "480"))
    # 工作时间配置
    WORK_TIME_MORNING_START = os.getenv("WORK_TIME_MORNING_START", "08:15")
    WORK_TIME_MORNING_END = os.getenv("WORK_TIME_MORNING_END", "12:00")
    WORK_TIME_AFTERNOON_START = os.getenv("WORK_TIME_AFTERNOON_START", "13:45")
    WORK_TIME_AFTERNOON_END = os.getenv("WORK_TIME_AFTERNOON_END", "18:00")
    WORK_HOURS_PER_DAY = float(os.getenv("WORK_HOURS_PER_DAY", "8.0"))

settings = Settings()
os.makedirs(settings.UPLOAD_DIR, exist_ok=True)

# ============== 数据库连接池（单例） ==============
from .database import get_engine, get_connection, text, dispose_engine

# ============== 日志框架 ==============
from .logger import get_logger
logger = get_logger(__name__)

# ============== 缓存管理（带 TTL） ==============
from .cache import cache_manager, store_user_token, get_user_token, get_user_info_cache

# ============== 定时任务 ==============

from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.cron import CronTrigger

scheduler = AsyncIOScheduler()

async def daily_alert_detection_job():
    """每日预警检测任务（凌晨1点执行）"""
    try:
        from .dashboard_service import run_daily_alert_detection
        count = run_daily_alert_detection()
        logger.info(f" 完成 {count} 个项目的预警检测")

        # 推送每日摘要到微信
        from .dashboard_service import get_dashboard_overview
        from .push_service import push_daily_summary_to_wechat

        overview = get_dashboard_overview()
        push_daily_summary_to_wechat(overview['stats'])

    except Exception as e:
        logger.error(f" {e}")

# ============== 认证相关 ==============

from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from jose import JWTError, jwt
from passlib.context import CryptContext

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/agent/api/agent/auth/login")


def verify_token(token: str) -> Optional[Dict]:
    """验证JWT token"""
    try:
        payload = jwt.decode(token, settings.SECRET_KEY, algorithms=[settings.ALGORITHM])
        username: str = payload.get("sub")
        if username is None:
            return None
        # 返回 sub 字段，保持一致性
        return {"sub": username, "user_id": payload.get("user_id")}
    except JWTError:
        return None


def create_access_token(data: dict, expires_delta: timedelta = None) -> str:
    """创建 JWT access token"""
    to_encode = data.copy()
    if expires_delta:
        expire = datetime.utcnow() + expires_delta
    else:
        expire = datetime.utcnow() + timedelta(hours=8)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, settings.SECRET_KEY, algorithm=settings.ALGORITHM)
    return encoded_jwt

async def get_current_user(token: str = Depends(oauth2_scheme)) -> Dict:
    """获取当前登录用户
    
    注意：JWT 本身是自包含的，不需要依赖内存缓存。
    每次请求都从 JWT 解析用户信息，避免多 worker 缓存不同步问题。
    """
    credentials_exception = HTTPException(
        status_code=401,
        detail="认证失败",
        headers={"WWW-Authenticate": "Bearer"},
    )

    # 直接验证 JWT（不依赖缓存，避免多 worker 不同步）
    payload = verify_token(token)
    if payload is None:
        raise credentials_exception
    
    logger.debug(f"verify_token 返回: {payload}")
    
    # 获取用户信息，补充 username, employee_id, name, role_id
    username = payload.get("sub")
    if username:
        # 重要：将 username 和原始 token 添加到 payload，确保后续 API 可以获取
        payload["username"] = username
        payload["_raw_token"] = token  # 保存原始 token，避免依赖内存缓存
        
        # 从数据库查询用户信息（绕过缓存）
        try:
            from dotenv import load_dotenv
            load_dotenv()
            with get_connection() as conn:
                logger.debug(f"查询用户信息: username={username}")
                result = conn.execute(text("""
                    SELECT p.employee_id, p.name, p.department, p.position, 
                           COALESCE(p.role_id, u.role_id, 13) as role_id
                    FROM personnel p
                    LEFT JOIN users u ON u.username = p.employee_id
                    WHERE p.employee_id = :username
                """), {"username": username}).fetchone()
                
                logger.debug(f"数据库查询结果: {result}")
                
                if result:
                    payload["employee_id"] = result[0] or username
                    payload["name"] = result[1] or ""
                    payload["department"] = result[2] or ""
                    payload["position"] = result[3] or ""
                    payload["role_id"] = result[4] or 13
                else:
                    logger.warning(f"用户 {username} 不存在于 personnel 表")
                    # 用户不存在于 personnel 表，使用默认值
                    payload["employee_id"] = username
                    payload["name"] = username
                    payload["role_id"] = 13
        except Exception as e:
            logger.error(f"获取用户信息失败: {e}")
            payload["employee_id"] = username
            payload["role_id"] = 13

    # 确保 employee_id 存在
    if "employee_id" not in payload:
        payload["employee_id"] = username
    
    if "role_id" not in payload:
        payload["role_id"] = 13

    logger.debug(f"get_current_user 返回: {payload}")
    return payload


def get_token_from_request(request: Request, username: str) -> Optional[str]:
    """从请求 header 或内存缓存获取 token
    
    优先从请求 header 获取（避免服务重启后内存缓存丢失）
    """
    # 优先从请求 header 获取
    auth_header = request.headers.get("authorization")
    if auth_header and auth_header.startswith("Bearer "):
        return auth_header[7:]
    
    # 兜底：从内存缓存获取
    return get_user_token(username)


async def get_current_user_with_token(
    request: Request,
    token: str = Depends(oauth2_scheme)
) -> Dict:
    """获取当前用户 + 从请求 header 提取 token
    
    返回的 payload 包含一个 `_raw_token` 字段，供后续 API 使用
    """
    user_info = await get_current_user(token)
    # 将原始 token 添加到 payload（方便后续 API 使用）
    user_info["_raw_token"] = token
    return user_info

async def get_user_info(token: str) -> Dict:
    """获取用户详细信息（包含角色、部门、岗位）"""
    try:
        # 获取当前用户信息
        response = await http_client.get(
            f"{settings.BACKEND_API_URL}/api/v1/auth/users/me",
            headers={"Authorization": f"Bearer {token}"},
            timeout=10.0
        )
        if response.status_code == 200:
            data = response.json()
            user_data = data.get("data", data)
            
            # 从 personnel 表补充部门、岗位信息
            # text 已从 database 模块导入
            from dotenv import load_dotenv
            load_dotenv()            
            employee_id = user_data.get("employee_id") or user_data.get("username")
            if employee_id:
                with get_connection() as conn:
                    # 从 users 表获取 role_id
                    user_result = conn.execute(text("""
                        SELECT role_id FROM users WHERE username = :employee_id
                    """), {"employee_id": employee_id}).fetchone()
                    
                    if user_result and user_result[0]:
                        user_data["role_id"] = user_result[0]
                    else:
                        user_data["role_id"] = 13  # 默认普通用户
                    
                    # 从 personnel 表获取其他信息
                    person_result = conn.execute(text("""
                        SELECT name, department, position, phone, email
                        FROM personnel
                        WHERE employee_id = :employee_id
                    """), {"employee_id": employee_id}).fetchone()
                    
                    if person_result:
                        user_data["name"] = person_result[0] or user_data.get("name", employee_id)
                        user_data["department"] = person_result[1] or ""
                        user_data["position"] = person_result[2] or ""
                        user_data["phone"] = person_result[3] or ""
                        user_data["email"] = person_result[4] or ""
            
            logger.info(f": employee_id={user_data.get('employee_id')}, name={user_data.get('name')}, role_id={user_data.get('role_id')}")
            return user_data
        return {}
    except Exception as e:
        logger.error(f" {e}")
        return {}

async def check_project_edit_permission(project_id: int, user_info: Dict) -> bool:
    """检查用户是否有项目编辑权限（项目负责人或管理员）"""
    role_id = user_info.get("role_id")
    employee_name = user_info.get("name", "")
    
    logger.debug(f"权限检查: project_id={project_id}, role_id={role_id}, name={employee_name}")
    
    # 管理员有权限
    if role_id == 11:
        logger.debug("管理员权限通过")
        return True
    
    # 检查是否是项目负责人
    with get_connection() as conn:
        result = conn.execute(text("""
            SELECT leader FROM projects WHERE id = :pid AND is_deleted = false
        """), {"pid": project_id})
        row = result.fetchone()
        if row:
            leader = row[0]
            logger.debug(f"项目leader={leader}, user={employee_name}")
            if leader == employee_name:
                logger.debug("负责人权限通过")
                return True
    
    logger.debug("权限检查失败")
    return False


def get_token_from_request(request: Request, username: str) -> Optional[str]:
    """从请求 header 或内存缓存获取 token
    
    优先从请求 header 获取（避免服务重启后内存缓存丢失）
    """
    # 优先从请求 header 获取
    auth_header = request.headers.get("authorization")
    if auth_header and auth_header.startswith("Bearer "):
        return auth_header[7:]
    
    # 兜底：从内存缓存获取
    return get_user_token(username)


async def get_projects_with_auth(token: str, user_info: Dict = None) -> List[Dict]:
    """获取项目列表（全员可见），并计算进度"""
    # text 已从 database 模块导入
    from dotenv import load_dotenv
    load_dotenv()
    with get_connection() as conn:
        # 全员可见所有项目
        result = conn.execute(text("""
            SELECT id, name, leader, status FROM projects
            WHERE is_deleted = false ORDER BY id
        """))

        # 计算每个项目的进度
        projects = []
        for row in result:
            project_id = row[0]
            try:
                task_stats = conn.execute(text("""
                    SELECT
                        COUNT(*) as total_tasks,
                        SUM(CASE WHEN status = '已完成' THEN 1 ELSE 0 END) as completed_tasks,
                        AVG(progress) as avg_progress
                    FROM project_tasks
                    WHERE project_id::integer = :pid
                      AND is_deleted = false
                      AND is_latest = true
                """), {"pid": project_id})
                ts = task_stats.fetchone()
                progress = round((ts[1] / ts[0] * 100 + float(ts[2] or 0)) / 2, 1) if ts and ts[0] else 0
            except:
                progress = 0

            projects.append({
                "id": project_id,
                "name": row[1],
                "leader": row[2],
                "status": row[3] or "进行中",
                "progress": progress
            })

        logger.debug(f"返回项目数: {len(projects)}")
        return projects

async def get_all_projects_for_matching() -> List[Dict]:
    """获取所有项目用于日报匹配（不受权限限制）"""
    # text 已从 database 模块导入
    from dotenv import load_dotenv
    load_dotenv()
    with get_connection() as conn:
        result = conn.execute(text("""
            SELECT id, name, leader, status FROM projects
            WHERE is_deleted = false ORDER BY id
        """))

        projects = []
        for row in result:
            project_id = row[0]
            try:
                task_stats = conn.execute(text("""
                    SELECT
                        COUNT(*) as total_tasks,
                        SUM(CASE WHEN status = '已完成' THEN 1 ELSE 0 END) as completed_tasks,
                        AVG(progress) as avg_progress
                    FROM project_tasks
                    WHERE project_id::integer = :pid
                      AND is_deleted = false
                      AND is_latest = true
                """), {"pid": project_id})
                ts = task_stats.fetchone()
                progress = round((ts[1] / ts[0] * 100 + float(ts[2] or 0)) / 2, 1) if ts and ts[0] else 0
            except:
                progress = 0

            projects.append({
                "id": project_id,
                "name": row[1],
                "leader": row[2],
                "status": row[3] or "进行中",
                "progress": progress
            })

        logger.info(f" 返回所有项目数: {len(projects)}")
        return projects

async def get_tasks_with_auth(project_id: int, token: str) -> List[Dict]:
    """使用认证token获取任务列表"""
    try:
        response = await http_client.get(
            f"{settings.BACKEND_API_URL}/api/v1/projects/{project_id}/tasks/",
            headers={"Authorization": f"Bearer {token}"},
            timeout=10.0
        )
        if response.status_code == 200:
            data = response.json()
            # 处理可能的嵌套结构
            if isinstance(data, dict):
                return data.get("data", data)
            return data
        return []
    except Exception as e:
        logger.error(f" {e}")
        return []

# 全局HTTP客户端（会在 startup/shutdown 中管理）
http_client: Optional[httpx.AsyncClient] = None

# LLM初始化 (DeepSeek)
llm = ChatOpenAI(
    model=settings.DEEPSEEK_MODEL,
    api_key=settings.DEEPSEEK_API_KEY,
    base_url=settings.DEEPSEEK_BASE_URL,
    temperature=0.2
)


def _llm_invoke_sync(llm, messages):
    """同步LLM调用（在线程池中执行）"""
    return llm.invoke(messages)


async def llm_invoke_threaded(messages):
    """
    异步LLM调用接口：在线程池中执行，不阻塞Worker
    
    用法：response = await llm_invoke_threaded(messages)
    """
    loop = asyncio.get_event_loop()
    return await loop.run_in_executor(AI_EXECUTOR, _llm_invoke_sync, llm, messages)

# ============== 数据模型 ==============

class DailyEntry(BaseModel):
    start_time: str
    end_time: str
    location: Optional[str] = None
    content: str
    project_hint: Optional[str] = None
    hours: float = 0
    # 智能匹配结果
    matched_project_id: Optional[int] = None
    matched_project_name: Optional[str] = None
    matched_task_id: Optional[str] = None
    matched_task_name: Optional[str] = None
    match_confidence: float = 0.0

class ParseDailyRequest(BaseModel):
    text: str
    user_id: Optional[str] = None

class ParseDailyResponse(BaseModel):
    entries: List[DailyEntry]
    confidence: float
    issues: List[str] = []

class ProjectInfo(BaseModel):
    id: int
    name: str
    leader: str
    status: str
    progress: float
    project_year: Optional[int] = None

class TaskInfo(BaseModel):
    task_id: str
    task_name: str
    assignee: Optional[str] = None
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    status: str

# ============== 现有后端API对接 ==============

async def get_projects_from_backend() -> List[Dict]:
    """从现有后端获取项目列表"""
    try:
        # 注意：现有后端需要认证，这里简化处理
        # 实际应该使用 service token 或用户 token
        response = await http_client.get(f"{settings.BACKEND_API_URL}/api/projects")
        if response.status_code == 200:
            data = response.json()
            # 根据实际返回结构调整
            return data.get("data", {}).get("list", [])
        return []
    except Exception as e:
        logger.error(f" {e}")
        return []

async def get_tasks_from_backend(project_id: int) -> List[Dict]:
    """从现有后端获取项目任务列表"""
    try:
        response = await http_client.get(
            f"{settings.BACKEND_API_URL}/api/projects/{project_id}/tasks"
        )
        if response.status_code == 200:
            return response.json()
        return []
    except Exception as e:
        logger.error(f" {e}")
        return []

# 缓存项目列表（简化版，生产环境用Redis）
_projects_cache: List[Dict] = []
_projects_cache_time: Optional[datetime] = None

async def get_cached_projects() -> List[Dict]:
    """获取缓存的项目列表"""
    global _projects_cache, _projects_cache_time

    # 缓存5分钟
    if (_projects_cache_time is None or
        datetime.now() - _projects_cache_time > timedelta(minutes=5) or
        not _projects_cache):
        _projects_cache = await get_projects_from_backend()
        _projects_cache_time = datetime.now()

    return _projects_cache

# ============== AI智能匹配 ==============

def match_project_by_name(project_hint: str, projects: List[Dict]) -> Optional[Dict]:
    """根据项目名称关键词匹配项目（支持模糊匹配）"""
    if not project_hint or not projects:
        return None

    import re
    
    # 提取关键词（去掉通用词）
    def extract_keywords(text):
        # 去掉通用词
        text = text.replace("项目", "").replace("工程", "").replace("系统", "").replace("研究", "").replace("开发", "").replace("协调", "")
        # 提取中文词组（2-10个字）
        keywords = re.findall(r'[\u4e00-\u9fa5]{2,10}', text)
        return [k.lower() for k in keywords]
    
    # 计算字符级别的相似度（处理"转换"vs"转化"这类情况）
    def char_similarity(s1, s2):
        if not s1 or not s2:
            return 0
        s1_set = set(s1)
        s2_set = set(s2)
        intersection = len(s1_set & s2_set)
        union = len(s1_set | s2_set)
        return intersection / union if union > 0 else 0
    
    hint_keywords = extract_keywords(project_hint)
    hint = project_hint.lower()
    best_match = None
    best_score = 0

    for project in projects:
        name = project.get("name", "")
        name_lower = name.lower()
        name_keywords = extract_keywords(name)
        
        score = 0
        
        # 完全匹配
        if hint == name_lower:
            return project
        
        # 包含匹配
        if hint in name_lower:
            score = max(score, len(hint) / len(name_lower))
        
        # 关键词匹配（模糊匹配）
        for hk in hint_keywords:
            for nk in name_keywords:
                if hk == nk:
                    score = max(score, 0.8)
                elif hk in nk or nk in hk:
                    score = max(score, 0.5)
                else:
                    # 字符级别相似度（处理同音字/形近字）
                    sim = char_similarity(hk, nk)
                    if sim > 0.7:  # 70%字符相同
                        score = max(score, 0.6)
        
        if score > best_score:
            best_score = score
            best_match = project

    # 阈值0.3以上认为是匹配
    if best_score >= 0.3:
        return best_match
    return None

# ============== 任务匹配增强（关键词提取 + 加权匹配） ==============

def extract_keywords(text: str, top_k: int = 5) -> List[str]:
    """
    提取关键词（jieba分词 + 词性过滤）
    
    解决向量稀释问题：长查询中非核心词汇干扰核心词汇匹配
    """
    try:
        import jieba.posseg as pseg
    except ImportError:
        logger.warning("jieba 未安装，使用简单关键词提取")
        # 降级方案：简单分词
        return [w for w in text if len(w) >= 2][:top_k]
    
    # 工程领域停用词
    stopwords = {
        "完成", "进行", "工作", "任务", "项目", "工程", "系统",
        "相关", "等", "及", "和", "的", "了", "在", "到", "对",
        "开展", "组织", "编写", "整理", "准备", "处理", "跟进",
        # 地名（避免地名匹配到任务名）
        "德保", "田阳", "隆林", "百色", "南宁"
    }
    
    words = pseg.cut(text)
    keywords = [
        word for word, flag in words
        if (flag.startswith('n') or flag.startswith('v'))  # 名词/动词
        and len(word) >= 2
        and word not in stopwords
    ]
    
    return keywords[:top_k]


def calculate_match_score(keywords: List[str], task_name: str, original_content: str) -> float:
    """
    计算匹配分数（加权策略）
    
    权重设计：
    - 关键词命中：+0.5分/词
    - 关键词在边界（开头/结尾）：+0.3分额外
    - 原有关键词表匹配：+0.3分
    """
    score = 0
    task_name_lower = task_name.lower()
    content_lower = original_content.lower()
    
    # 权重1：关键词命中（核心改进）
    for kw in keywords:
        if kw in task_name_lower:
            score += 0.5
            # 边界匹配加分（关键词在任务名开头或结尾）
            if task_name_lower.startswith(kw) or task_name_lower.endswith(kw):
                score += 0.3
    
    # 权重2：原有关键词表匹配（保持兼容）
    legacy_keywords = ["图纸", "审查", "设计", "招标", "采购", "施工", "勘察", "会议", "协调"]
    for kw in legacy_keywords:
        if kw in content_lower and kw in task_name_lower:
            score += 0.3
    
    # 权重3：完全包含加分
    if task_name_lower in content_lower:
        score += 0.8
    
    return min(score, 2.0)  # 上限2.0


def match_task_by_content(content: str, tasks: List[Dict]) -> Optional[Dict]:
    """
    根据工作内容匹配任务（增强版：关键词提取 + 加权匹配）
    
    解决问题：
    - 长查询"完成德保铝厂空压机图纸设计"匹配错误
    - 原因：向量稀释，非核心词汇干扰核心词汇匹配
    
    改进方案：
    - 提取关键词（jieba分词 + 词性过滤）
    - 加权匹配（关键词命中 + 边界匹配 + 原有逻辑）
    """
    if not content or not tasks:
        return None

    # 阶段1：关键词提取
    keywords = extract_keywords(content)
    logger.debug(f"[任务匹配] 提取关键词: {keywords}")
    
    # 阶段2：加权匹配
    scored_tasks = []
    for task in tasks:
        task_name = task.get("task_name", "")
        score = calculate_match_score(keywords, task_name, content)
        if score > 0:
            scored_tasks.append((task, score))
            logger.debug(f"[任务匹配] {task_name}: {score:.2f}")
    
    # 阶段3：排序并返回最佳匹配
    scored_tasks.sort(key=lambda x: x[1], reverse=True)
    
    if scored_tasks and scored_tasks[0][1] >= 0.5:  # 阈值0.5（平衡模式）
        best_task, best_score = scored_tasks[0]
        logger.info(f"[任务匹配] 匹配成功: '{content[:30]}...' -> {best_task.get('task_name')} (得分: {best_score:.2f})")
        return best_task
    
    logger.debug(f"[任务匹配] 未找到匹配: '{content[:30]}...'")
    return None

# ============== 日报解析（智能版） ==============

async def parse_daily_text_smart(text: str, projects: List[Dict], current_date: str = None) -> Dict[str, Any]:
    """
    智能解析日报文本，自动匹配项目和任务

    Args:
        text: 日报文本
        projects: 项目列表（用于匹配）
        current_date: 当前日期
    """
    if current_date is None:
        current_date = datetime.now().strftime("%Y-%m-%d")

    # 构建项目提示信息（包含项目名称关键词）
    project_list = "\n".join([
        f"- {p.get('id')}: {p.get('name')} (关键词: {p.get('name').replace('项目', '').replace('工程', '').strip()[:10]})"
        for p in projects[:20]
    ])

    system_prompt = f"""你是项目管理助手，专门解析工程人员的日报文本。

可匹配的项目列表：
{project_list}

## 解析规则

### 1. 时间识别（重要！）
支持多种时间格式：
- 纯时间："9点"、"09:00"、"14:30"
- 时间段："9:00-12:00"、"13:45-18:00"
- 混合格式："上午:9:00-12:00"、"下午：16:00-20:00"（注意冒号可能是中英文混用）
- 时段+时间："上午9点"、"下午2点半"

**标准工作时间**：
- 上午：08:15-12:00（含午休前）
- 下午：13:45-18:00
- **加班**：18:00之后的时间

**⚠️ 时间格式错误处理（默认开始时间）**：
- 如果用户描述的时间格式不符合标准（如"8:45到15:30"写成了"8:45-15:30 xxx"但未正确分隔）
- 或者开始时间异常（如早于08:15），**默认按08:15开始工作**
- 结束时间保持用户描述的值
- 例如：用户写"8:45到15:30 xxx项目"，但实际识别困难，默认从08:15开始计算

**时间段分隔符识别**：
- 分号（;或；）：分隔不同的工作时间段
- 例如："上午:9:00-12:00编写xxx；下午：16:00-20:00到yyy"
  - 这表示两个独立的工作时间段，应生成两个条目

**加班自动拆分**：
- 如果时间段跨越18:00，必须拆分为"标准时间"和"加班"两个条目
- 例如："下午16:00-20:00到田阳铝厂..."
  - 标准时间：16:00-18:00（2小时）
  - 加班时间：18:00-20:00（2小时）
  - 生成两个条目，内容相同但时间不同

### 2. 时间段共享
- 如果用户说"下午13:45-18:00做了4件事"，表示共享时间段
- 为每个任务生成独立条目，时间相同
- hours 字段填 0，系统自动计算

### 3. 序号内容分组（重要！）
当用户使用序号（1. 2. 3. 4.）列出多项工作时，要正确识别每项的边界：

**识别规则**：
- 序号后到下一个序号之前的内容，都属于该项
- 如果某项后面有逗号分隔的内容，仍属于该项
- 例如："4.隆林铝厂空压机项目研究，合同线下审批"
  - "隆林铝厂空压机项目研究"和"合同线下审批"都属于第4项
  - 应该合并为一个条目："推进隆林铝厂空压机项目研究，完成合同线下审批"
  
**错误示例**：
```
输入："1.xxx，2.xxx，3.xxx，4.xxx，yyy"
错误：把"yyy"单独分成一个条目
正确：第4项是"xxx，yyy"，合并为一个条目
```

**正确示例**：
```
输入："下午协调1.隆林铝厂除尘器研究，2.田林铝厂供电项目，3.隆林铝厂整流改造，4.隆林铝厂空压机项目，合同线下审批"

正确解析为4个条目：
1. 隆林铝厂除尘器研究
2. 田林铝厂供电项目
3. 隆林铝厂整流改造
4. 隆林铝厂空压机项目研究，完成合同线下审批（合并！）
```

### 4. 加班识别
- "额外X小时"、"加班X小时"、"晚上X小时" 表示加班
- 加班时间从 18:00 开始计算
- 必须生成独立的加班条目

### 4. 项目匹配

**已知项目别名映射表（优先匹配）**：

| 用户常用别名 | 项目ID | 正式项目名 |
|------------|-------|-----------|
| 炭渣项目 | 32 | 铝电解碳渣低温氧化处理技术 |
| 炭渣试验 | 32 | 铝电解碳渣低温氧化处理技术 |
| 田阳铝厂电解质炭渣处理 | 32 | 铝电解碳渣低温氧化处理技术 |
| 锰渣无害化 | 33 | 电解锰渣无害化处理项目 |
| 锰渣专题 | 33 | 电解锰渣无害化处理项目 |
| 锰锭试制 | 34 | 落地锰转化锰锭项目 |
| 田林铝厂供电整流 | 19 | 田林铝厂供电整流PLC控制系统稳定性研发项目 |
| 隆林铝厂除尘器 | 18 | 隆林铝厂除尘器布袋脉冲精准控制研究 |
| 隆林铝厂空压机 | 20 | 隆林铝厂空压机集中控制项目研究 |
| 田林铝厂空压机 | 23 | 田林铝厂空压机集中控制项目研究 |
| 德保铝厂空压机 | 22 | 德保铝厂空压机集中控制项目研究 |
| 田阳铝厂空压机 | 21 | 田阳铝厂空压机集中控制项目研究 |
| 电解槽新烟管 | 12 | 600KA槽上部烟气治理的技术研究 |
| 新烟管 | 12 | 600KA槽上部烟气治理的技术研究 |
| 新烟管软连接 | 12 | 600KA槽上部烟气治理的技术研究 |
| 烟管软连接 | 12 | 600KA槽上部烟气治理的技术研究 |
| 600KA槽烟气 | 12 | 600KA槽上部烟气治理的技术研究 |
| 槽上部烟气 | 12 | 600KA槽上部烟气治理的技术研究 |
| 德保铝厂化锰筑炉 | 34 | 落地锰转化锰锭项目 |
| 铁锭模 | 34 | 落地锰转化锰锭项目 |
| 锰锭试制 | 34 | 落地锰转化锰锭项目 |
| 德保铝厂化锰铸锰锭 | 34 | 落地锰转化锰锭项目 |
| 电解铝多功能天车抓斗 | 14 | 一种新型电解铝多功能天车抓斗结构的设计及产业化 |
| 田林电解天车抓斗 | 14 | 一种新型电解铝多功能天车抓斗结构的设计及产业化 |
| 天车抓斗改进 | 14 | 一种新型电解铝多功能天车抓斗结构的设计及产业化 |
| 抓斗产业化 | 14 | 一种新型电解铝多功能天车抓斗结构的设计及产业化 |
| 隆林铝厂除尘器 | 18 | 隆林铝厂除尘器布袋脉冲精准控制研究 |
| 田林铝厂供电整流 | 19 | 田林铝厂供电整流PLC控制系统稳定性研发项目 |
| 隆林铝厂整流系统 | 24 | 隆林铝厂整流系统总调PLC升级改造项目 |
| 隆林铝厂空压机 | 20 | 隆林铝厂空压机集中控制项目研究 |
| 田阳铝厂电解质炭渣处理 | 32 | 铝电解碳渣低温氧化处理技术 |
| 炭渣项目 | 32 | 铝电解碳渣低温氧化处理技术 |
| 炭渣试验 | 32 | 铝电解碳渣低温氧化处理技术 |
| 锰渣专题 | 33 | 电解锰渣无害化处理项目 |
| 锰渣固化 | 33 | 电解锰渣无害化处理项目 |
| 锰渣无害化 | 33 | 电解锰渣无害化处理项目 |

**匹配优先级**：
1. 别名映射表精确匹配 → 直接输出项目ID
2. 模糊匹配 → 检查关键词是否包含在正式项目名中
3. 无法匹配 → project_hint填"其他工作"，matched_project_id填null

**模糊匹配规则**：
- 关键词提取：去掉"项目"、"工程"、"研究"、"系统"等通用词
- 例如："隆林铝厂除尘器" → 匹配 "隆林铝厂除尘器布袋脉冲精准控制研究项目"

### 5. 内容润色（重要！按STAR原则改写）
每条工作内容需要润色为规范的结果汇报格式：

**STAR原则**：
- **S**pecific（具体）：使用动词开头，明确动作
- **T**ime-bound（时效）：体现当期进展
- **A**chievable（成果）：强调产出和结果
- **R**elevant（相关）：关联项目背景

**润色规则**：
1. 使用动词开头：完成、协调、审核、编制、讨论、推进、优化、落实
2. 量化成果：如有数据、文档数、进度百分比，务必保留
3. 去掉冗余：删除"协调"、"处理"等模糊词，改为具体动作
4. 控制长度：15-40字，简洁有力
5. 结果导向：强调"完成"、"提交"、"通过"等结果状态

**润色示例**：
- 原文："协调4个铝厂一种新型电解铝多功能天车抓斗结构的设计及产业化项目审核技术文件"
- 润色："审核电解铝多功能天车抓斗产业化项目技术文件，完成4个铝厂技术评审"

- 原文："隆林铝厂除尘器布袋脉冲精准控制研究"
- 润色："推进隆林铝厂除尘器布袋脉冲精准控制研究，完成技术方案讨论"

- 原文："合同线下审批"
- 润色："完成合同线下审批流程"

### 6. 工时计算
- hours 字段统一填 0，由系统自动计算
- 系统会自动扣除午休时间（12:00-13:45）

### 7. 多时间段+加班示例

**示例输入**：
"上午:9:00-12:00编写田林铝厂供电整流PLC控制系统稳定性研发项目的招标技术条件；下午：16:00-20:00到田阳铝厂开展田阳铝厂阳极组装提质增效技术服务项目的现场诊断工作"

**正确解析（跨越加班时间，必须拆分）**：
{{
  "entries": [
    {{
      "start_time": "09:00",
      "end_time": "12:00",
      "location": "办公室",
      "content": "编写田林铝厂供电整流PLC控制系统稳定性研发项目的招标技术条件",
      "project_hint": "田林铝厂供电整流PLC",
      "matched_project_id": null,
      "matched_project_name": "",
      "hours": 0
    }},
    {{
      "start_time": "16:00",
      "end_time": "18:00",
      "location": "田阳铝厂",
      "content": "开展田阳铝厂阳极组装提质增效技术服务项目的现场诊断工作",
      "project_hint": "田阳铝厂阳极组装提质增效",
      "matched_project_id": null,
      "matched_project_name": "",
      "hours": 0
    }},
    {{
      "start_time": "18:00",
      "end_time": "20:00",
      "location": "田阳铝厂",
      "content": "开展田阳铝厂阳极组装提质增效技术服务项目的现场诊断工作（加班）",
      "project_hint": "田阳铝厂阳极组装提质增效",
      "matched_project_id": null,
      "matched_project_name": "",
      "hours": 0
    }}
  ],
  "confidence": 0.9,
  "issues": []
}}
注意：下午16:00-20:00跨越18:00，拆分为两个条目。
---

### 8. 序号内容分组示例

示例输入：
"上午8:15-12:00协调4个铝厂一种新型电解铝多功能天车抓斗结构的设计及产业化项目审核技术文件；下午13:45-18:00协调1.隆林铝厂除尘器布袋脉冲精准控制研究，2.田林铝厂供电整流PLC控制系统稳定性研发项目，3.隆林铝厂整流系统总调PLC升级改造项目，4.隆林铝厂空压机集中控制项目研究，合同线下审批"

正确输出（注意：第4项"空压机项目研究，合同线下审批"合并为一个条目）：
{{
  "entries": [
    {{
      "start_time": "08:15",
      "end_time": "12:00",
      "location": "办公室",
      "content": "审核电解铝多功能天车抓斗产业化项目技术文件，完成4个铝厂技术评审",
      "project_hint": "电解铝多功能天车抓斗",
      "matched_project_id": null,
      "matched_project_name": "",
      "hours": 0
    }},
    {{
      "start_time": "13:45",
      "end_time": "18:00",
      "location": "办公室",
      "content": "推进隆林铝厂除尘器布袋脉冲精准控制研究，完成技术方案讨论",
      "project_hint": "隆林铝厂除尘器",
      "matched_project_id": null,
      "matched_project_name": "",
      "hours": 0
    }},
    {{
      "start_time": "13:45",
      "end_time": "18:00",
      "location": "办公室",
      "content": "推进田林铝厂供电整流PLC控制系统稳定性研发，完成需求对接",
      "project_hint": "田林铝厂供电整流",
      "matched_project_id": null,
      "matched_project_name": "",
      "hours": 0
    }},
    {{
      "start_time": "13:45",
      "end_time": "18:00",
      "location": "办公室",
      "content": "推进隆林铝厂整流系统总调PLC升级改造，完成方案评审",
      "project_hint": "隆林铝厂整流系统",
      "matched_project_id": null,
      "matched_project_name": "",
      "hours": 0
    }},
    {{
      "start_time": "13:45",
      "end_time": "18:00",
      "location": "办公室",
      "content": "推进隆林铝厂空压机集中控制项目研究，完成合同线下审批",
      "project_hint": "隆林铝厂空压机",
      "matched_project_id": null,
      "matched_project_name": "",
      "hours": 0
    }}
  ],
  "confidence": 0.95,
  "issues": []
}}

输出格式（严格JSON）：
{{
  "entries": [
    {{
      "start_time": "HH:MM",
      "end_time": "HH:MM",
      "location": "地点",
      "content": "工作内容",
      "project_hint": "项目关键词",
      "matched_project_id": 项目ID或null,
      "matched_project_name": "项目名或空",
      "hours": 0
    }}
  ],
  "confidence": 0.95,
  "issues": []
}}"""

    user_prompt = f"当前日期：{current_date}\n\n日报文本：{text}\n\n请解析并返回JSON格式结果："

    try:
        messages = [
            SystemMessage(content=system_prompt),
            HumanMessage(content=user_prompt)
        ]

        logger.debug(f" 调用 DeepSeek API...")
        response = await llm_invoke_threaded(messages)
        logger.debug(f" API 返回: {response.content[:200]}...")

        # 清理响应内容
        content = response.content.strip()
        if "```json" in content:
            content = content.split("```json")[1].split("```")[0].strip()
        elif "```" in content:
            content = content.split("```")[1].split("```")[0].strip()

        result = json.loads(content)

        # 后处理：验证和补充
        entries = result.get("entries", [])

        # 如果没有解析出条目，尝试简单解析
        if not entries:
            logger.debug(f" 未解析出条目，尝试兜底解析...")
            entries = simple_parse_fallback(text, projects)

        # 先识别共享时间段，再计算工时
        # 按时间段分组
        time_groups = {}
        for i, entry in enumerate(entries):
            if entry.get("start_time") and entry.get("end_time"):
                time_key = f"{entry['start_time']}-{entry['end_time']}"
                if time_key not in time_groups:
                    time_groups[time_key] = []
                time_groups[time_key].append(i)

        # 计算工时（共享时间段平均分配）
        for time_key, indices in time_groups.items():
            start_time, end_time = time_key.split("-")
            try:
                from app.work_time_config import calculate_work_hours
                total_hours = calculate_work_hours(start_time, end_time)
                # 平均分配
                avg_hours = total_hours / len(indices)
                for idx in indices:
                    entries[idx]["hours"] = round(avg_hours, 2)
            except:
                pass

        # 项目匹配
        for entry in entries:
            if not entry.get("matched_project_id") and entry.get("project_hint"):
                matched = match_project_by_name(entry["project_hint"], projects)
                if matched:
                    entry["matched_project_id"] = matched.get("id")
                    entry["matched_project_name"] = matched.get("name")
                    entry["match_confidence"] = 0.7

        result["entries"] = entries
        return result

    except Exception as e:
        logger.error(f" {e}")
        import traceback
        traceback.print_exc()

        # 兜底：简单解析
        entries = simple_parse_fallback(text, projects)

        return {
            "entries": entries,
            "confidence": 0.5,
            "issues": [f"AI解析失败，已使用基础解析: {str(e)}"]
        }

def simple_parse_fallback(text: str, projects: List[Dict]) -> List[Dict]:
    """
    简单解析兜底方案：当AI解析失败时使用
    """
    entries = []
    import re
    
    # 提取时间信息
    time_pattern = r'(\d{1,2}[：:]\d{2})\s*[-~至到]+\s*(\d{1,2}[：:]\d{2})'
    time_match = re.search(time_pattern, text)
    # ⚠️ 默认开始时间：08:15（标准工作时间）
    start_time = "08:15"
    end_time = "18:00"
    
    if time_match:
        start_time = time_match.group(1).replace('：', ':')
        end_time = time_match.group(2).replace('：', ':')
        # ⚠️ 如果开始时间早于08:15，默认设为08:15
        try:
            s_h, s_m = int(start_time[:2]), int(start_time[3:5])
            if s_h < 8 or (s_h == 8 and s_m < 15):
                start_time = "08:15"
        except:
            start_time = "08:15"

    # 提取工时信息（如果用户明确说了"做了X小时"）
    hours_pattern = r'(\d+(?:\.\d+)?)\s*小时'
    hours_matches = re.findall(hours_pattern, text)
    explicit_hours = float(hours_matches[0]) if hours_matches else None

    # 提取项目关键词（增加模糊匹配）
    best_match = None
    best_score = 0
    
    for project in projects:
        project_name = project.get("name", "")
        # 去掉"项目"、"工程"等通用词
        keywords = project_name.replace("项目", "").replace("工程", "").replace("系统", "").strip()
        
        # 完全匹配
        if keywords and keywords in text:
            best_match = (project, project_name, keywords)
            break
            
        # 模糊匹配（相似度检查）
        if keywords:
            # 计算关键词在文本中的覆盖率
            keyword_chars = set(keywords)
            text_chars = set(text)
            overlap = len(keyword_chars & text_chars)
            score = overlap / len(keyword_chars) if keyword_chars else 0
            
            if score > best_score and score > 0.6:  # 60% 相似度阈值
                best_score = score
                best_match = (project, project_name, keywords)

    if best_match:
        project, project_name, keywords = best_match
        # 找到匹配的项目
        entries.append({
            "start_time": start_time,
            "end_time": end_time,
            "location": "办公室",
            "content": text,  # 保留完整内容，不再截断
            "project_hint": keywords,
            "matched_project_id": project.get("id"),
            "matched_project_name": project_name,
            "hours": explicit_hours if explicit_hours else 0
        })
    else:
        # 没有匹配到项目
        entries.append({
            "start_time": start_time,
            "end_time": end_time,
            "location": "办公室",
            "content": text,  # 保留完整内容
            "project_hint": "",
            "matched_project_id": None,
            "matched_project_name": "",
            "hours": 0
        })

    return entries

async def enrich_with_tasks(entries: List[Dict]) -> List[Dict]:
    """为每个条目匹配任务"""
    for entry in entries:
        project_id = entry.get("matched_project_id")
        if project_id:
            tasks = await get_tasks_from_backend(project_id)
            matched_task = match_task_by_content(entry.get("content", ""), tasks)
            if matched_task:
                entry["matched_task_id"] = matched_task.get("task_id")
                entry["matched_task_name"] = matched_task.get("task_name")
                entry["match_confidence"] = max(entry.get("match_confidence", 0), 0.6)
    return entries

# ============== LangGraph工作流 ==============

class DailyParseState(dict):
    text: str
    user_id: Optional[str]
    projects: List[Dict]
    parsed_entries: List[Dict]
    confidence: float
    issues: List[str]

async def parse_node(state: DailyParseState):
    """解析节点"""
    result = await parse_daily_text_smart(
        state["text"],
        state.get("projects", []),
        datetime.now().strftime("%Y-%m-%d")
    )
    return {
        "parsed_entries": result.get("entries", []),
        "confidence": result.get("confidence", 0),
        "issues": result.get("issues", [])
    }

async def match_tasks_node(state: DailyParseState):
    """匹配任务节点"""
    entries = state.get("parsed_entries", [])
    enriched = await enrich_with_tasks(entries)
    return {"parsed_entries": enriched}

# 构建工作流（顺序执行）
from langgraph.graph import StateGraph, END

daily_workflow = StateGraph(DailyParseState)
daily_workflow.add_node("parse", parse_node)
daily_workflow.add_node("match_tasks", match_tasks_node)
daily_workflow.set_entry_point("parse")
daily_workflow.add_edge("parse", "match_tasks")
daily_workflow.add_edge("match_tasks", END)
daily_agent = daily_workflow.compile()

# ============== API路由 ==============
# 注：health 端点已迁移至 routes/health.py

# ============== 新增：智能解析代理接口 ==============

class SmartParseRequest(BaseModel):
    """智能解析请求"""
    text: str
    report_date: Optional[str] = None

@app.post("/agent/api/agent/daily/smart-parse")
async def smart_parse_daily(
    request: SmartParseRequest,
    current_user: Dict = Depends(get_current_user)
):
    """
    智能解析日报文本 - 一次 AI 调用完成项目+任务+时间解析

    支持多次输入，每次解析会覆盖之前的内容

    返回：
    - matched_projects: 匹配到的项目列表
    - unmatched_projects: 未匹配的项目名称
    - entries: 解析出的工作事项（含匹配的 task_id）
    - warnings: 警告信息
    """
    try:
        # 导入线程池版本的解析函数，AI调用不阻塞Worker
        import sys
        sys.path.insert(0, os.path.dirname(__file__))
        from task_auto import parse_daily_all_in_one_threaded
        from .ai_usage_tracker import check_usage_limit, log_ai_usage
        from datetime import datetime

        # httpx 已在文件顶部导入，这里不需要重复导入

        user_id = current_user.get("employee_id", current_user.get("username"))
        username = current_user.get("name", current_user.get("username"))
        
        # ========== 检查调用限制 ==========
        if not await check_usage_limit(user_id, "daily_parse"):
            raise HTTPException(
                status_code=429,
                detail=f"已达日报解析上限（{50}次/天），请明天再试"
            )

        logger.info(f"用户 {current_user.get('username')} 开始解析日报: {request.text[:50]}...")
        
        # 记录开始时间
        start_time = datetime.now()
        
        # 在线程池中执行 AI 解析，主事件循环释放处理其他请求
        try:
            result = await parse_daily_all_in_one_threaded(request.text, request.report_date)
        except httpx.ReadTimeout:
            logger.error("AI解析超时")
            raise HTTPException(status_code=504, detail="AI解析超时，请稍后再试")
        except Exception as e:
            logger.error(f"AI解析异常: {e}")
            raise HTTPException(status_code=500, detail=f"AI解析异常: {str(e)}")
        
        if not result.get("success"):
            error_msg = result.get("error", "AI解析失败")
            logger.error(f"AI解析失败: {error_msg}")
            raise HTTPException(status_code=500, detail=error_msg)
        
        # 转换为前端期望的格式
        entries = []
        matched_projects_list = []
        matched_project_ids = set()
        
        for entry in result.get("entries", []):
            project = entry.get("project")
            task = entry.get("task")
            time_info = entry.get("time", {})
            
            # 构建条目
            formatted_entry = {
                "content": entry.get("content", ""),
                "matched_project_id": project.get("id") if project else None,
                "matched_project_name": project.get("name") if project else None,
                "matched_task_id": task.get("id") if task else None,
                "matched_task_name": task.get("name") if task else None,
                "start_time": time_info.get("start"),
                "end_time": time_info.get("end"),
                "hours": time_info.get("hours", 4.0),
                "is_overtime": time_info.get("is_overtime", False),
                "confidence": entry.get("confidence", 0.8),
                # 共享时间段相关
                "shared_period": time_info.get("shared_period"),
                "period_total_hours": time_info.get("period_total_hours")
            }
            
            entries.append(formatted_entry)
            
            # 收集匹配的项目
            if project and project.get("id") not in matched_project_ids:
                matched_projects_list.append({
                    "id": project.get("id"),
                    "name": project.get("name")
                })
                matched_project_ids.add(project.get("id"))
        
        # ========== 记录AI调用日志 ==========
        # 估算tokens（实际应从API响应获取，这里用估算）
        input_tokens = 10500  # 系统提示词(400) + 项目列表(10000) + 用户输入(100)
        output_tokens = len(entries) * 300  # 每个事项约300 tokens
        await log_ai_usage(
            user_id=user_id,
            username=username,
            purpose="daily_parse",
            model="deepseek-v4-flash",
            input_tokens=input_tokens,
            output_tokens=output_tokens,
            success=True
        )
        
        return {
            "entries": entries,
            "matched_projects": matched_projects_list,
            "unmatched_projects": [],
            "warnings": result.get("warnings", []),
            "confidence": sum(e.get("confidence", 0.8) for e in result.get("entries", [])) / max(len(result.get("entries", [])), 1),
            "issues": result.get("warnings", [])
        }

    except HTTPException:
        raise  # 重新抛出已处理的HTTP异常
    except httpx.ReadTimeout:
        logger.error("AI解析超时")
        # 记录失败日志
        try:
            await log_ai_usage(
                user_id=user_id,
                username=username,
                purpose="daily_parse",
                model="deepseek-v4-flash",
                success=False,
                error_message="ReadTimeout"
            )
        except:
            pass
        raise HTTPException(status_code=504, detail="AI解析超时，请稍后再试")
    except Exception as e:
        logger.error(f"解析失败: {e}")
        import traceback
        traceback.print_exc()
        # 记录失败日志
        try:
            await log_ai_usage(
                user_id=user_id,
                username=username,
                purpose="daily_parse",
                model="deepseek-v4-flash",
                success=False,
                error_message=str(e)[:200]
            )
        except:
            pass
        raise HTTPException(status_code=500, detail=f"解析失败: {str(e)}")


@app.post("/agent/api/agent/daily/local-parse")
async def local_parse_daily(
    request: SmartParseRequest,
    current_user: Dict = Depends(get_current_user)
):
    """
    本地解析日报文本 - 使用本地Ollama 7B模型（快速响应）
    
    特点：
    - 模型：qwen2.5:7B（端口8001）
    - 响应速度：首次加载约50秒，常驻后约1-2秒
    - 无向量检索，直接LLM生成
    - 适合快速响应场景
    
    返回：
    - matched_projects: 匹配到的项目列表
    - unmatched_projects: 未匹配的项目名称
    - entries: 解析出的工作事项
    - warnings: 警告信息
    """
    try:
        from .task_auto import parse_daily_with_7b
        from datetime import datetime

        user_id = current_user.get("employee_id", current_user.get("username"))
        username = current_user.get("name", current_user.get("username"))
        
        logger.info(f"用户 {current_user.get('username')} 开始本地7B解析日报: {request.text[:50]}...")
        
        start_time = datetime.now()

        # 获取项目列表（用于匹配）
        projects = []
        with get_connection() as conn:
            result = conn.execute(text("""
                SELECT id, name, leader FROM projects
                WHERE status = '进行中' AND is_deleted = false
                ORDER BY id
                LIMIT 50
            """))
            projects = [
                {"id": row[0], "name": row[1], "leader": row[2]}
                for row in result.fetchall()
            ]

        # 调用7B模型解析
        try:
            result = await parse_daily_with_7b(request.text, request.report_date, projects)
        except httpx.ConnectError:
            logger.error("本地Ollama连接失败")
            raise HTTPException(status_code=503, detail="本地模型服务不可用，请检查Ollama是否运行")
        except httpx.ReadTimeout:
            logger.error("7B解析超时")
            raise HTTPException(status_code=504, detail="7B解析超时，请稍后再试")
        except Exception as e:
            logger.error(f"7B解析异常: {e}")
            raise HTTPException(status_code=500, detail=f"7B解析异常: {str(e)}")
        
        if not result.get("success"):
            error_msg = result.get("error", "7B解析失败")
            logger.error(f"7B解析失败: {error_msg}")
            raise HTTPException(status_code=500, detail=error_msg)
        
        entries = result.get("entries", [])
        matched_projects_list = result.get("matched_projects", [])
        
        end_time = datetime.now()
        duration_ms = int((end_time - start_time).total_seconds() * 1000)
        
        logger.info(f"7B解析成功，耗时 {duration_ms}ms，条目数: {len(entries)}")

        return {
            "success": True,
            "matched_projects": matched_projects_list,
            "unmatched_projects": [],
            "entries": entries,
            "warnings": result.get("warnings", []),
            "original_text": request.text,
            "parse_method": "local-7b",
            "duration_ms": duration_ms
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"本地解析失败: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"本地解析失败: {str(e)}")


@app.post("/agent/api/agent/daily/smart-parse-7b")
async def parse_daily_with_7b(
    request: SmartParseRequest,
    current_user: Dict = Depends(get_current_user)
):
    """
    本地解析日报文本 - 使用本地Ollama 7B模型（快速响应版）

    特点：
    - 模型：qwen2.5:7B（端口8001）
    - 响应速度：首次加载约50秒，后续约1-2秒
    - 无向量检索，直接LLM生成
    - 适合快速响应场景

    返回：
    - matched_projects: 匹配到的项目列表
    - unmatched_projects: 未匹配的项目名称
    - entries: 解析出的工作事项
    - warnings: 警告信息
    """
    try:
        from .task_auto import parse_daily_with_7b
        from datetime import datetime

        user_id = current_user.get("employee_id", current_user.get("username"))
        username = current_user.get("name", current_user.get("username"))

        logger.info(f"用户 {current_user.get('username')} 开始7B模型解析日报: {request.text[:50]}...")

        start_time = datetime.now()

        # 获取项目列表（用于匹配）
        projects = []
        with get_connection() as conn:
            result = conn.execute(text("""
                SELECT id, name, leader FROM projects
                WHERE status = '进行中' AND is_deleted = false
                ORDER BY id
                LIMIT 50
            """))
            projects = [
                {"id": row[0], "name": row[1], "leader": row[2]}
                for row in result.fetchall()
            ]

        # 调用7B模型解析
        try:
            result = await parse_daily_with_7b(request.text, request.report_date, projects)
        except httpx.ConnectError:
            logger.error("本地Ollama连接失败")
            raise HTTPException(status_code=503, detail="本地模型服务不可用，请检查Ollama是否运行")
        except httpx.ReadTimeout:
            logger.error("7B解析超时")
            raise HTTPException(status_code=504, detail="7B解析超时，请稍后再试")
        except Exception as e:
            logger.error(f"7B解析异常: {e}")
            raise HTTPException(status_code=500, detail=f"7B解析异常: {str(e)}")

        if not result.get("success"):
            error_msg = result.get("error", "7B解析失败")
            logger.error(f"7B解析失败: {error_msg}")
            raise HTTPException(status_code=500, detail=error_msg)

        entries = result.get("entries", [])
        matched_projects_list = result.get("matched_projects", [])

        end_time = datetime.now()
        duration_ms = int((end_time - start_time).total_seconds() * 1000)

        logger.info(f"7B解析成功，耗时 {duration_ms}ms，条目数: {len(entries)}")

        return {
            "success": True,
            "matched_projects": matched_projects_list,
            "unmatched_projects": [],
            "entries": entries,
            "warnings": result.get("warnings", []),
            "original_text": request.text,
            "parse_method": "local-7b",  # 标记解析方式
            "duration_ms": duration_ms
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"7B解析失败: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"7B解析失败: {str(e)}")


class CreateReportRequest(BaseModel):
    """创建日报请求"""
    report_date: str
    work_items: List[Dict[str, Any]]
    work_target: Optional[str] = None
    tomorrow_plan: Optional[str] = None
    original_input: Optional[str] = None  # 原始自然语言输入
    ai_parsed_data: Optional[Dict[str, Any]] = None  # AI解析结果

@app.post("/agent/api/agent/daily/create")
async def create_daily_report(
    request: CreateReportRequest,
    current_user: Dict = Depends(get_current_user)
):
    """
    创建日报 - 智能体专用（支持覆盖已有日报）

    如果该日期已存在日报，先删除旧日报，再创建新日报
    """
    try:
        # text 已从 database 模块导入
        from dotenv import load_dotenv
        import json
        load_dotenv()
        username = current_user.get("username") or current_user.get("sub")
        employee_id = current_user.get("employee_id")
        
        # 优先从 payload 获取 token（避免依赖内存缓存）
        token = current_user.get("_raw_token") or get_user_token(username)

        # 安全检查：确保 employee_id 存在
        if not employee_id:
            logger.warning(f" 用户 {username} 缺少 employee_id，使用 username 作为标识")
            employee_id = username

        if not token:
            raise HTTPException(status_code=401, detail="未找到用户认证信息")

        # 先删除该日期的旧日报（智能体专用：支持覆盖）
        # 安全检查：同时匹配 employee_id 和 employee_name
        with get_connection() as conn:
            # 获取用户姓名
            user_name = current_user.get("name", "")
            
            # 查找旧日报（同时匹配 employee_id 和 employee_name）
            result = conn.execute(text("""
                SELECT id, employee_id, employee_name FROM daily_reports
                WHERE employee_id = :eid AND report_date = :date AND is_deleted = false
            """), {"eid": employee_id, "date": request.report_date})

            old_report = result.fetchone()

            if old_report:
                # 额外安全检查：确认日报属于当前用户
                if old_report[1] != employee_id:
                    logger.error(f" 日报归属检查失败：期望 {employee_id}，实际 {old_report[1]}")
                    raise HTTPException(status_code=403, detail="无权删除此日报")
                
                # 删除旧日报的工作项
                conn.execute(text("""
                    DELETE FROM daily_work_items WHERE report_id = :rid
                """), {"rid": old_report[0]})

                # 删除旧日报
                conn.execute(text("""
                    DELETE FROM daily_reports WHERE id = :rid
                """), {"rid": old_report[0]})

                conn.commit()
                logger.info(f" 已删除 {request.report_date} 的旧日报 (ID: {old_report[0]}, 用户: {old_report[2]})")

        # 调用主后端创建接口
        response = await http_client.post(
            f"{settings.BACKEND_API_URL}/api/v1/ai-daily/create-from-parse",
            json={
                "report_date": request.report_date,
                "work_items": request.work_items,
                "work_target": request.work_target,
                "tomorrow_plan": request.tomorrow_plan
            },
            headers={"Authorization": f"Bearer {token}"},
            timeout=10.0
        )

        if response.status_code == 200:
            result = response.json()
            data = result.get("data", result)
            report_id = data.get("report_id")

            # 保存原始输入和AI解析结果
            if report_id and (request.original_input or request.ai_parsed_data):
                with get_connection() as conn:
                    conn.execute(text("""
                        UPDATE daily_reports
                        SET original_input = :input,
                            ai_parsed_data = :parsed,
                            parse_mode = 'free',
                            status = '已提交'
                        WHERE id = :rid
                    """), {
                        "input": request.original_input,
                        "parsed": json.dumps(request.ai_parsed_data) if request.ai_parsed_data else None,
                        "rid": report_id
                    })
                    conn.commit()
            
            # 更新工作项的时间字段（从 ai_parsed_data 中读取）
            if report_id and request.ai_parsed_data:
                entries = request.ai_parsed_data.get('entries', [])
                if entries:
                    with get_connection() as conn:
                        # 获取该日报的所有工作项
                        work_items_result = conn.execute(text("""
                            SELECT id FROM daily_work_items 
                            WHERE report_id = :rid 
                            ORDER BY id
                        """), {"rid": report_id})
                        work_item_ids = [row[0] for row in work_items_result.fetchall()]
                        
                        # 按顺序更新时间
                        for idx, entry in enumerate(entries):
                            if idx < len(work_item_ids) and entry.get('start_time') and entry.get('end_time'):
                                conn.execute(text("""
                                    UPDATE daily_work_items 
                                    SET start_time = :start_time, end_time = :end_time
                                    WHERE id = :wid
                                """), {
                                    "start_time": entry['start_time'],
                                    "end_time": entry['end_time'],
                                    "wid": work_item_ids[idx]
                                })
                        conn.commit()
                        logger.info(f" 已更新 {min(len(entries), len(work_item_ids))} 个工作项的时间字段")

            # 更新任务进度
            try:
                import sys
                sys.path.insert(0, os.path.dirname(__file__))
                from task_auto import update_task_progress_from_daily
                updated_tasks = update_task_progress_from_daily(request.work_items)
                if updated_tasks:
                    logger.info(f"已更新 {len(updated_tasks)} 个任务进度: {updated_tasks}")
            except Exception as e:
                logger.error(f"更新任务进度失败（不影响日报保存）: {e}")
                import traceback
                traceback.print_exc()

            return {
                "success": True,
                "message": "日报创建成功",
                "report_id": report_id,
                "updated_tasks": len(updated_tasks) if 'updated_tasks' in locals() else 0
            }
        else:
            logger.error(f" {response.status_code} - {response.text}")
            raise HTTPException(
                status_code=response.status_code,
                detail=f"创建失败: {response.text}"
            )

    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f" {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"创建失败: {str(e)}")


@app.get("/agent/api/agent/daily/my-reports")
async def get_my_daily_reports(
    page: int = 1,
    size: int = 10,
    current_user: Dict = Depends(get_current_user)
):
    """
    获取我的日报列表 - 从本地数据库直接查询
    """
    try:
        # text 已从 database 模块导入
        from dotenv import load_dotenv
        load_dotenv()
        username = current_user.get("username") or current_user.get("sub")
        employee_id = current_user.get("employee_id") or username

        with get_connection() as conn:
            # 获取日报列表
            offset = (page - 1) * size
            result = conn.execute(text("""
                SELECT dr.id, dr.report_date, dr.status,
                       to_char(dr.create_time, 'YYYY-MM-DD HH24:MI:SS') as created_at,
                       COUNT(dwi.id) as item_count,
                       COALESCE(SUM(dwi.hours_spent), 0) as total_hours
                FROM daily_reports dr
                LEFT JOIN daily_work_items dwi ON dwi.report_id = dr.id
                WHERE dr.employee_id = :eid
                  AND dr.is_deleted = false
                GROUP BY dr.id, dr.report_date, dr.status, dr.create_time
                ORDER BY dr.report_date DESC
                LIMIT :size OFFSET :offset
            """), {"eid": employee_id, "size": size, "offset": offset})

            reports = []
            for row in result:
                report_id = row[0]

                # 获取工作项
                items_result = conn.execute(text("""
                    SELECT work_content, project_name, start_time, end_time,
                           hours_spent, task_id, task_name
                    FROM daily_work_items
                    WHERE report_id = :rid
                    ORDER BY project_name, id
                """), {"rid": report_id})

                items = []
                for item in items_result:
                    items.append({
                        "work_content": item[0] or "",
                        "project_name": item[1] or "",
                        "start_time": item[2] or "",  # 空值显示为空，不显示假时间
                        "end_time": item[3] or "",    # 空值显示为空
                        "hours_spent": float(item[4] or 0),
                        "task_id": item[5],
                        "task_name": item[6]
                    })

                # 获取原始输入和AI解析数据
                meta_result = conn.execute(text("""
                    SELECT original_input, ai_parsed_data
                    FROM daily_reports
                    WHERE id = :rid
                """), {"rid": report_id})
                meta_row = meta_result.fetchone()

                original_input = meta_row[0] if meta_row else None
                ai_parsed_data = meta_row[1] if meta_row and meta_row[1] else None

                reports.append({
                    "id": report_id,
                    "report_date": str(row[1]),
                    "total_hours": float(row[5] or 0),
                    "status": row[2] or "已提交",
                    "created_at": row[3],  # 已格式化为 'YYYY-MM-DD HH24:MI:SS'
                    "items": items,
                    "original_input": original_input,
                    "ai_parsed_data": ai_parsed_data,
                    "ai_parsed": len(items) > 0 and any(item.get("task_id") for item in items)
                })

            # 获取总数
            count_result = conn.execute(text("""
                SELECT COUNT(DISTINCT id) FROM daily_reports
                WHERE employee_id = :eid AND is_deleted = false
            """), {"eid": employee_id})
            total = count_result.fetchone()[0]

            return {
                "items": reports,
                "total": total,
                "page": page,
                "size": size
            }

    except Exception as e:
        logger.exception(f" {e}")
        import traceback
        traceback.print_exc()
        return {"items": [], "total": 0, "page": page, "size": size}


@app.get("/agent/api/agent/daily/monthly-summary")
async def get_daily_monthly_summary(
    year: int,
    month: int,
    current_user: Dict = Depends(get_current_user)
):
    """
    获取某月的日报摘要 - 用于日历视图
    返回每日是否有日报及工时
    同时返回月度统计：工作日、总日报数、总工时
    """
    try:
        username = current_user.get("username") or current_user.get("sub")
        employee_id = current_user.get("employee_id") or username

        # 计算该月起止日期
        start_date = f"{year}-{month:02d}-01"
        if month == 12:
            end_date = f"{year + 1}-01-01"
        else:
            end_date = f"{year}-{month + 1:02d}-01"

        # 计算当月工作日数（广西节假日）
        guangxi_holidays_2026 = {
            (2026, 4, 4), (2026, 4, 5), (2026, 4, 6),  # 清明
            (2026, 4, 17), (2026, 4, 20),  # 三月三
            (2026, 5, 1), (2026, 5, 2), (2026, 5, 3), (2026, 5, 4), (2026, 5, 5),  # 劳动节
            (2026, 6, 20), (2026, 6, 21), (2026, 6, 22),  # 端午
            (2026, 9, 25), (2026, 9, 26), (2026, 9, 27),  # 中秋
            (2026, 10, 1), (2026, 10, 2), (2026, 10, 3), (2026, 10, 4),
            (2026, 10, 5), (2026, 10, 6), (2026, 10, 7), (2026, 10, 8),  # 国庆
        }
        guangxi_workdays_2026 = {
            (2026, 5, 9),  # 劳动节补班
        }
        
        month_start = datetime(year, month, 1).date()
        month_end_date = (month_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        working_days = 0
        
        current = month_start
        while current <= month_end_date:
            weekday = current.weekday()
            date_tuple = (current.year, current.month, current.day)
            
            if date_tuple in guangxi_workdays_2026:
                working_days += 1
            elif date_tuple in guangxi_holidays_2026:
                pass
            elif weekday < 5:
                working_days += 1
            
            current += timedelta(days=1)

        with get_connection() as conn:
            # 查询日报数据
            result = conn.execute(text("""
                SELECT 
                    EXTRACT(DAY FROM dr.report_date)::int as day,
                    dr.id as report_id,
                    COALESCE(SUM(dwi.hours_spent), 0) as total_hours
                FROM daily_reports dr
                LEFT JOIN daily_work_items dwi ON dwi.report_id = dr.id
                WHERE dr.employee_id = :eid
                  AND dr.report_date >= :start_date
                  AND dr.report_date < :end_date
                  AND dr.is_deleted = false
                GROUP BY dr.report_date, dr.id
                ORDER BY dr.report_date
            """), {"eid": employee_id, "start_date": start_date, "end_date": end_date})

            days = {}
            total_hours = 0
            report_count = 0
            
            for row in result:
                day = row[0]
                hours = float(row[2] or 0)
                days[day] = {
                    "has_report": True,
                    "total_hours": hours,
                    "report_id": row[1]
                }
                total_hours += hours
                report_count += 1

            # 查询请假记录
            leave_result = conn.execute(text("""
                SELECT 
                    EXTRACT(DAY FROM leave_date)::int as day,
                    leave_type,
                    reason
                FROM leave_records
                WHERE employee_id = :eid
                  AND leave_date >= :start_date
                  AND leave_date < :end_date
                  AND is_deleted = false
            """), {"eid": employee_id, "start_date": start_date, "end_date": end_date})
            
            leave_days = {}
            for row in leave_result:
                day = row[0]
                leave_days[day] = {
                    "is_leave": True,
                    "leave_type": row[1],
                    "reason": row[2]
                }
                # 如果这一天没有日报，标记为请假
                if day not in days:
                    days[day] = {
                        "has_report": False,
                        "is_leave": True,
                        "leave_type": row[1],
                        "reason": row[2]
                    }
                else:
                    # 有日报但也请假了（可能是半天假）
                    days[day]["is_leave"] = True
                    days[day]["leave_type"] = row[1]
                    days[day]["reason"] = row[2]

            # 计算缺失天数：工作日数 - 日报数 - 请假天数（无日报的请假日）
            leave_without_report = len([d for d in leave_days if d not in [k for k in days if days[k].get("has_report")]])
            missing_days = working_days - report_count - leave_without_report

            return {
                "year": year,
                "month": month,
                "days": days,
                "working_days": working_days,  # 当月工作日数
                "total_hours": round(total_hours, 1),  # 总工时
                "report_count": report_count,  # 日报数
                "leave_days": len(leave_days),  # 请假天数
                "missing_days": max(0, missing_days)  # 缺失天数（排除请假）
            }

    except Exception as e:
        logger.exception(f"获取月度日报摘要失败: {e}")
        return {"year": year, "month": month, "days": {}, "working_days": 0, "total_hours": 0, "report_count": 0, "missing_days": 0}


@app.get("/agent/api/agent/daily/by-date")
async def get_daily_by_date(
    date: str,
    current_user: Dict = Depends(get_current_user)
):
    """
    按日期获取日报详情
    """
    try:
        username = current_user.get("username") or current_user.get("sub")
        employee_id = current_user.get("employee_id") or username

        with get_connection() as conn:
            result = conn.execute(text("""
                SELECT dr.id, dr.report_date, dr.status,
                       to_char(dr.create_time, 'YYYY-MM-DD HH24:MI:SS') as created_at,
                       COUNT(dwi.id) as item_count,
                       COALESCE(SUM(dwi.hours_spent), 0) as total_hours
                FROM daily_reports dr
                LEFT JOIN daily_work_items dwi ON dwi.report_id = dr.id
                WHERE dr.employee_id = :eid
                  AND dr.report_date = :date
                  AND dr.is_deleted = false
                GROUP BY dr.id, dr.report_date, dr.status, dr.create_time
            """), {"eid": employee_id, "date": date})

            row = result.fetchone()
            if not row:
                return {"has_report": False}

            report_id = row[0]

            # 获取工作项
            items_result = conn.execute(text("""
                SELECT work_content, project_name, start_time, end_time,
                       hours_spent, task_id, task_name
                FROM daily_work_items
                WHERE report_id = :rid
                ORDER BY project_name, id
            """), {"rid": report_id})

            items = []
            for item in items_result:
                items.append({
                    "work_content": item[0] or "",
                    "project_name": item[1] or "",
                    "start_time": item[2] or "",
                    "end_time": item[3] or "",
                    "hours_spent": float(item[4] or 0),
                    "task_id": item[5],
                    "task_name": item[6]
                })

            # 获取原始输入
            meta_result = conn.execute(text("""
                SELECT original_input, ai_parsed_data
                FROM daily_reports
                WHERE id = :rid
            """), {"rid": report_id})
            meta_row = meta_result.fetchone()

            return {
                "has_report": True,
                "id": report_id,
                "report_date": str(row[1]),
                "total_hours": float(row[5] or 0),
                "status": row[2] or "已提交",
                "created_at": row[3],
                "items": items,
                "original_input": meta_row[0] if meta_row else None,
                "ai_parsed_data": meta_row[1] if meta_row and meta_row[1] else None,
                "ai_parsed": len(items) > 0 and any(item.get("task_id") for item in items)
            }

    except Exception as e:
        logger.exception(f"按日期获取日报失败: {e}")
        return {"has_report": False}


@app.post("/agent/api/agent/auth/login")
@limiter.limit("5/minute")  # 防暴力破解：每分钟最多5次
async def login(request: Request, form_data: OAuth2PasswordRequestForm = Depends()):
    """
    登录接口 - 代理到现有后端认证

    用户名/密码与现有管理系统一致
    """
    try:
        # 调用现有后端登录接口
        response = await http_client.post(
            f"{settings.BACKEND_API_URL}/api/v1/auth/login",
            data={
                "username": form_data.username,
                "password": form_data.password
            },
            headers={"Content-Type": "application/x-www-form-urlencoded"},
            timeout=10.0
        )

        if response.status_code == 200:
            data = response.json()
            logger.debug(f"后端返回: {data}")

            # 处理不同可能的返回格式
            response_data = data.get("data", data)

            token = (response_data.get("access_token") or
                    response_data.get("token"))

            user = response_data.get("user")

            if token:
                # 从 JWT 中解析用户标识作为 key
                try:
                    payload = jwt.decode(token, settings.SECRET_KEY, algorithms=[settings.ALGORITHM])
                    user_key = payload.get("sub") or form_data.username
                except:
                    user_key = form_data.username

                # 获取用户详细信息（包含角色）
                user_info = await get_user_info(token)
                logger.debug(f"用户信息: {user_info}")

                # 存储用户token和信息用于后续请求
                store_user_token(user_key, token, user_info)
                logger.debug(f"存储token: key={user_key}")

                return {
                    "access_token": token,
                    "token_type": "bearer",
                    "user": {
                        "id": user_info.get("employee_id"),
                        "name": user_info.get("name"),
                        "username": user_key,
                        "role_id": user_info.get("role_id")
                    }
                }
            else:
                return data
        else:
            error_data = response.json()
            detail = error_data.get("detail") or error_data.get("message") or "用户名或密码错误"
            raise HTTPException(
                status_code=401,
                detail=detail
            )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f" {e}")
        raise HTTPException(status_code=500, detail="登录服务异常")


@app.get("/agent/api/agent/auth/me")
async def get_current_user_info(current_user: Dict = Depends(get_current_user)):
    """
    获取当前用户详细信息（含角色、部门、岗位）
    """
    try:
        username = current_user.get("username")

        # text 已从 database 模块导入
        from dotenv import load_dotenv
        load_dotenv()
        with get_connection() as conn:
            # 从 users 表获取基本信息（包含 role_id）
            result = conn.execute(text("""
                SELECT id, username, role, role_id FROM users WHERE username = :username
            """), {"username": username}).fetchone()

            if result:
                current_user["id"] = result[0]
                current_user["role"] = result[2] or "user"
                current_user["role_id"] = result[3] or 13  # 默认13（普通用户）

            # 从 personnel 表获取部门、岗位信息
            person_result = conn.execute(text("""
                SELECT name, department, position, phone, email
                FROM personnel
                WHERE employee_id = :username
            """), {"username": username}).fetchone()

            if person_result:
                current_user["name"] = person_result[0] or username
                current_user["department"] = person_result[1] or ""
                current_user["position"] = person_result[2] or ""
                current_user["phone"] = person_result[3] or ""
                current_user["email"] = person_result[4] or ""

        return current_user
    except Exception as e:
        logger.error(f" {e}")
        current_user["role"] = "user"
        return current_user


@app.post("/agent/api/agent/auth/refresh")
async def refresh_token(current_user: Dict = Depends(get_current_user)):
    """
    刷新 Token - 基于当前JWT生成新token
    
    前端检测到 token 即将过期时自动调用
    不依赖内存缓存，避免服务重启后丢失状态
    """
    try:
        username = current_user.get("username") or current_user.get("sub")
        user_id = current_user.get("user_id") or current_user.get("employee_id")
        
        if not username:
            raise HTTPException(status_code=401, detail="无效的用户信息")
        
        # 直接生成新的 JWT token（不依赖缓存）
        access_token_expires = timedelta(hours=8)
        new_token = create_access_token(
            data={"sub": username, "user_id": user_id},
            expires_delta=access_token_expires
        )
        
        # 更新缓存（如果缓存存在）
        try:
            store_user_token(username, new_token)
        except:
            pass  # 缓存失败不影响token生成
        
        # 获取用户信息
        user_info = get_user_info_cache(username)
        if not user_info:
            try:
                with get_connection() as conn:
                    result = conn.execute(text("""
                        SELECT employee_id, name, department, position
                        FROM personnel WHERE employee_id = :username
                    """), {"username": username}).fetchone()
                    
                    if result:
                        user_info = {
                            "employee_id": result[0],
                            "name": result[1],
                            "department": result[2],
                            "position": result[3]
                        }
                        # 更新缓存
                        cache_manager.store_user_info(username, user_info)
            except Exception as e:
                logger.warning(f"获取用户信息失败: {e}")
        
        return {
            "access_token": new_token,
            "token_type": "bearer",
            "user": {
                "id": user_id,
                "name": user_info.get("name") if user_info else username,
                "username": username
            }
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"刷新token失败: {e}")
        raise HTTPException(status_code=401, detail="刷新token失败")


@app.put("/agent/api/agent/auth/push-token")
async def update_push_token(
    push_token: str = None,
    current_user: Dict = Depends(get_current_user)
):
    """
    更新用户的微信推送Token
    
    请求体：
    {
        "push_token": "your_pushplus_token"
    }
    """
    try:
        # text 已从 database 模块导入
        from dotenv import load_dotenv
        from pydantic import BaseModel
        load_dotenv()
        
        # 定义请求体模型
        class PushTokenRequest(BaseModel):
            push_token: str        
        username = current_user.get("username")
        
        with get_connection() as conn:
            conn.execute(text("""
                UPDATE users SET push_token = :token WHERE username = :username
            """), {"token": push_token, "username": username})
            conn.commit()
        
        return {"success": True, "message": "推送Token已更新"}
    
    except Exception as e:
        logger.error(f" {e}")
        raise HTTPException(status_code=500, detail=f"更新失败: {str(e)}")


def require_role(allowed_roles: List[str]):
    """角色权限检查装饰器"""
    async def role_checker(current_user: Dict = Depends(get_current_user)):
        # 先从数据库获取用户角色
        username = current_user.get("username")

        # text 已从 database 模块导入
        from dotenv import load_dotenv
        load_dotenv()
        with get_connection() as conn:
            result = conn.execute(text("""
                SELECT role FROM users WHERE username = :username
            """), {"username": username}).fetchone()

            user_role = result[0] if result else "user"

        # 同时检查 role_id（兼容旧系统）
        user_info = get_user_info_cache(username)
        role_id = user_info.get("role_id") if user_info else None

        # admin 判断：role=admin 或 role_id=11
        if "admin" in allowed_roles and (user_role == "admin" or role_id == 11):
            return current_user

        if user_role not in allowed_roles:
            raise HTTPException(
                status_code=403,
                detail=f"权限不足：需要 {allowed_roles} 角色"
            )
        return current_user
    return role_checker


@app.get("/agent/api/agent/work-hours/stats")
async def get_work_hours_stats(current_user: Dict = Depends(get_current_user), request: Request = None):
    """
    获取工时统计数据
    
    【修复】不再从内存缓存获取 token，直接从请求 header 获取
    
    返回：今日、本周、本月工时，项目工时分布
    """
    username = current_user.get("username") or current_user.get("sub")
    
    # 【修复】从请求 header 获取 token（不再依赖内存缓存）
    token = None
    if request and request.headers.get("authorization"):
        auth_header = request.headers.get("authorization")
        if auth_header.startswith("Bearer "):
            token = auth_header[7:]
    
    if not token:
        token = current_user.get("_raw_token") or get_user_token(username)
    
    if not token:
        raise HTTPException(status_code=401, detail="未找到用户认证信息")

    try:
        # 获取用户信息
        user_info = get_user_info_cache(username)
        employee_id = user_info.get("employee_id") if user_info else username

        # 计算日期范围
        today = datetime.now().date()
        week_start = today - timedelta(days=today.weekday())  # 本周一
        month_start = today.replace(day=1)  # 本月1号

        # 直接从数据库查询（更准确）
        # text 已从 database 模块导入
        from dotenv import load_dotenv
        load_dotenv()
        with get_connection() as conn:
            # 今日工时
            result = conn.execute(text("""
                SELECT COALESCE(SUM(hours_spent), 0) as hours
                FROM daily_work_items dwi
                JOIN daily_reports dr ON dwi.report_id = dr.id
                WHERE dr.employee_id = :emp_id
                AND dr.report_date = :today
                AND dr.is_deleted = false
            """), {"emp_id": employee_id, "today": today})
            today_hours = float(result.fetchone()[0] or 0)

            # 本周工时
            result = conn.execute(text("""
                SELECT COALESCE(SUM(hours_spent), 0) as hours
                FROM daily_work_items dwi
                JOIN daily_reports dr ON dwi.report_id = dr.id
                WHERE dr.employee_id = :emp_id
                AND dr.report_date >= :week_start
                AND dr.report_date <= :today
                AND dr.is_deleted = false
            """), {"emp_id": employee_id, "week_start": week_start, "today": today})
            week_hours = float(result.fetchone()[0] or 0)

            # 本月工时及项目分布
            # 使用 NULLIF 将空字符串转为 NULL，再用 COALESCE 替换为"其他工作"
            # "其他工作"按顺序放在最后一位
            result = conn.execute(text("""
                SELECT
                    COALESCE(NULLIF(TRIM(dwi.project_name), ''), '其他工作') as project_name,
                    SUM(dwi.hours_spent) as hours
                FROM daily_work_items dwi
                JOIN daily_reports dr ON dwi.report_id = dr.id
                WHERE dr.employee_id = :emp_id
                AND dr.report_date >= :month_start
                AND dr.report_date <= :today
                AND dr.is_deleted = false
                GROUP BY COALESCE(NULLIF(TRIM(dwi.project_name), ''), '其他工作')
                ORDER BY 
                    CASE WHEN COALESCE(NULLIF(TRIM(dwi.project_name), ''), '其他工作') = '其他工作' THEN 1 ELSE 0 END,
                    hours DESC
                LIMIT 5
            """), {"emp_id": employee_id, "month_start": month_start, "today": today})

            project_hours = {}
            month_total = 0
            for row in result:
                name = row[0]
                hours = float(row[1] or 0)
                project_hours[name] = hours
                month_total += hours

        # 计算项目工时分布
        project_distribution = []
        for name, hours in sorted(project_hours.items(), key=lambda x: -x[1]):
            percent = round(hours / month_total * 100) if month_total > 0 else 0
            project_distribution.append({
                "name": name,
                "hours": round(hours, 1),
                "percent": percent
            })

        return {
            "today": round(today_hours, 1),
            "week": round(week_hours, 1),
            "month": round(month_total, 1),
            "projects": project_distribution
        }

    except Exception as e:
        logger.error(f" {e}")
        import traceback
        traceback.print_exc()
        return {
            "today": 0,
            "week": 0,
            "month": 0,
            "projects": []
        }


# ============== 今日聚焦看板 API ==============

@app.get("/agent/api/agent/dashboard/today-focus")
async def get_today_focus(current_user: Dict = Depends(get_current_user), request: Request = None):
    """
    获取今日聚焦数据（组织隔离）
    
    权限规则：
    - 系统管理员：看全部项目的任务
    - 财务/看板：看研究院项目的任务
    - 普通用户：看本组织及下级项目的任务
    
    返回：
    - today_tasks: 今日待办任务
    - delayed_tasks: 延期任务
    - month_goals: 本月目标进度
    - daily_report_status: 日报填报状态
    """
    username = current_user.get("username") or current_user.get("sub")
    
    # 从请求 header 获取 token
    token = None
    if request and request.headers.get("authorization"):
        auth_header = request.headers.get("authorization")
        if auth_header.startswith("Bearer "):
            token = auth_header[7:]
    
    if not token:
        token = current_user.get("_raw_token") or get_user_token(username)
    
    if not token:
        raise HTTPException(status_code=401, detail="未找到用户认证信息")

    try:
        # 获取用户信息
        user_info = get_user_info_cache(username)
        employee_id = user_info.get("employee_id") if user_info else username
        employee_name = user_info.get("name") if user_info else username

        from dotenv import load_dotenv
        load_dotenv()
        today = datetime.now().date()
        current_month = today.strftime("%Y-%m")

        with get_connection() as conn:
            # 获取用户角色和组织
            user_result = conn.execute(text("""
                SELECT p.role_id, p.org_id
                FROM personnel p
                WHERE p.employee_id = :username AND p.is_deleted = false
            """), {"username": username}).fetchone()
            
            if not user_result:
                raise HTTPException(status_code=401, detail="用户不存在")
            
            role_id, org_id = user_result
            
            # 构建项目过滤条件
            if role_id == 11:  # 系统管理员
                org_filter = ""
            elif role_id in [15, 17]:  # 财务、看板：研究院项目
                org_filter = "AND p.org_id IN (SELECT id FROM organizations WHERE id = 2 OR parent_id = 2)"
            elif org_id:  # 普通用户：本组织及下级项目
                org_filter = f"AND p.org_id IN (SELECT id FROM organizations WHERE id = {org_id} OR parent_id = {org_id})"
            else:
                org_filter = "AND 1=0"  # 无组织用户，不显示任何项目

            # 1. 今日待办任务（本组织项目的今日截止任务）
            result = conn.execute(text(f"""
                SELECT pt.task_id, pt.task_name, pt.project_id, p.name as project_name,
                       pt.start_date, pt.end_date, pt.status, pt.progress, pt.assignee_id
                FROM project_tasks pt
                JOIN projects p ON CAST(pt.project_id AS INTEGER) = p.id
                WHERE pt.is_deleted = false
                  AND p.is_deleted = false
                  {org_filter}
                  AND pt.end_date = :today
                  AND pt.actual_end_date IS NULL
                ORDER BY pt.end_date
                LIMIT 10
            """), {"today": today})

            today_tasks = []
            for row in result:
                today_tasks.append({
                    "task_id": row[0],
                    "task_name": row[1],
                    "project_id": row[2],
                    "project_name": row[3],
                    "start_date": str(row[4]) if row[4] else None,
                    "end_date": str(row[5]) if row[5] else None,
                    "status": row[6],
                    "progress": float(row[7] or 0)
                })

            # 2. 延期任务（本组织项目的已超期未完成任务）
            result = conn.execute(text(f"""
                SELECT pt.task_id, pt.task_name, pt.project_id, p.name as project_name,
                       pt.start_date, pt.end_date,
                       CURRENT_DATE - pt.end_date as delay_days,
                       pt.status, pt.progress
                FROM project_tasks pt
                JOIN projects p ON CAST(pt.project_id AS INTEGER) = p.id
                WHERE pt.is_deleted = false
                  AND p.is_deleted = false
                  {org_filter}
                  AND pt.end_date < CURRENT_DATE
                  AND pt.actual_end_date IS NULL
                ORDER BY delay_days DESC
                LIMIT 10
            """))

            delayed_tasks = []
            for row in result:
                delayed_tasks.append({
                    "task_id": row[0],
                    "task_name": row[1],
                    "project_id": row[2],
                    "project_name": row[3],
                    "start_date": str(row[4]) if row[4] else None,
                    "end_date": str(row[5]) if row[5] else None,
                    "delay_days": row[6],
                    "status": row[7],
                    "progress": float(row[8] or 0)
                })

            # 3. 本月目标进度（本组织项目的任务）
            from datetime import timedelta
            month_start = today.replace(day=1)
            month_end = (month_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)

            month_goals = []

            # 获取本月要开始的任务
            month_start_result = conn.execute(text(f"""
                WITH latest_tasks AS (
                    SELECT pt.*,
                           CAST(SUBSTRING(pt.task_id FROM 'V([0-9]+)') AS INTEGER) as version,
                           MAX(CAST(SUBSTRING(pt.task_id FROM 'V([0-9]+)') AS INTEGER)) OVER (
                               PARTITION BY pt.project_id, SUBSTRING(pt.task_id FROM 'T[0-9]+$')
                           ) as max_version
                    FROM project_tasks pt
                    JOIN projects p ON CAST(pt.project_id AS INTEGER) = p.id
                    WHERE pt.is_deleted = false
                      AND p.is_deleted = false
                      {org_filter}
                )
                SELECT task_id, task_name, start_date, end_date, progress, status
                FROM latest_tasks
                WHERE version = max_version
                  AND start_date >= :month_start
                  AND start_date <= :month_end
                ORDER BY start_date
            """), {"month_start": month_start, "month_end": month_end})

            for row in month_start_result:
                month_goals.append({
                    "id": f"start_{row[0]}",
                    "title": f"🚀 {row[1]}",
                    "progress_rate": float(row[4] or 0),
                    "status": row[5],
                    "type": "本月启动",
                    "date": str(row[2]) if row[2] else None,
                    "end_date": str(row[3]) if row[3] else None
                })

            # 获取本月要完成的任务
            month_end_result = conn.execute(text(f"""
                WITH latest_tasks AS (
                    SELECT pt.*,
                           CAST(SUBSTRING(pt.task_id FROM 'V([0-9]+)') AS INTEGER) as version,
                           MAX(CAST(SUBSTRING(pt.task_id FROM 'V([0-9]+)') AS INTEGER)) OVER (
                               PARTITION BY pt.project_id, SUBSTRING(pt.task_id FROM 'T[0-9]+$')
                           ) as max_version
                    FROM project_tasks pt
                    JOIN projects p ON CAST(pt.project_id AS INTEGER) = p.id
                    WHERE pt.is_deleted = false
                      AND p.is_deleted = false
                      {org_filter}
                )
                SELECT task_id, task_name, start_date, end_date, progress, status
                FROM latest_tasks
                WHERE version = max_version
                  AND end_date >= :month_start
                  AND end_date <= :month_end
                  AND status != '已完成'
                ORDER BY end_date
            """), {"month_start": month_start, "month_end": month_end})

            for row in month_end_result:
                existing = next((g for g in month_goals if g["id"] == f"end_{row[0]}"), None)
                if not existing:
                    month_goals.append({
                        "id": f"end_{row[0]}",
                        "title": f"🎯 {row[1]}",
                        "progress_rate": float(row[4] or 0),
                        "status": row[5],
                        "type": "本月完成",
                        "date": str(row[2]) if row[2] else None,
                        "end_date": str(row[3]) if row[3] else None
                    })

            # 4. 今日日报填报状态（仍看自己的）
            result = conn.execute(text("""
                SELECT id, report_date, status
                FROM daily_reports
                WHERE is_deleted = false
                  AND employee_id = :emp_id
                  AND report_date = :today
                LIMIT 1
            """), {"emp_id": employee_id, "today": today})

            daily_report_row = result.fetchone()
            daily_report_status = {
                "submitted": daily_report_row is not None,
                "report_id": daily_report_row[0] if daily_report_row else None,
                "status": daily_report_row[2] if daily_report_row else None
            }

            # 5. 本周工作概览（仍看自己的）
            week_start = today - timedelta(days=today.weekday())
            result = conn.execute(text("""
                SELECT
                    COUNT(DISTINCT dr.id) as report_count,
                    COALESCE(SUM(dwi.hours_spent), 0) as total_hours,
                    COUNT(DISTINCT dwi.project_id) as project_count
                FROM daily_reports dr
                LEFT JOIN daily_work_items dwi ON dr.id = dwi.report_id
                WHERE dr.is_deleted = false
                  AND dr.employee_id = :emp_id
                  AND dr.report_date >= :week_start
                  AND dr.report_date <= :today
            """), {"emp_id": employee_id, "week_start": week_start, "today": today})

            week_row = result.fetchone()
            week_overview = {
                "report_count": week_row[0] if week_row else 0,
                "total_hours": float(week_row[1] or 0) if week_row else 0,
                "project_count": week_row[2] if week_row else 0
            }

        return {
            "today_tasks": today_tasks,
            "delayed_tasks": delayed_tasks,
            "month_goals": month_goals,
            "daily_report_status": daily_report_status,
            "week_overview": week_overview,
            "date": today.isoformat(),
            "employee_name": employee_name
        }

    except Exception as e:
        logger.error(f" {e}")
        import traceback
        traceback.print_exc()
        return {
            "today_tasks": [],
            "delayed_tasks": [],
            "month_goals": [],
            "daily_report_status": {"submitted": False},
            "week_overview": {"report_count": 0, "total_hours": 0, "project_count": 0},
            "date": datetime.now().date().isoformat(),
            "employee_name": ""
        }


@app.get("/agent/api/agent/dashboard/risk-alerts")
async def get_risk_alerts(current_user: Dict = Depends(get_current_user)):
    """
    获取风险预警数据（管理员视角）

    返回：
    - delayed_projects: 延期项目列表
    - unreported_users: 今日未填报人员
    - high_risk_projects: 高风险项目
    """
    username = current_user.get("username")
    token = current_user.get("_raw_token") or get_user_token(username)

    if not token:
        raise HTTPException(status_code=401, detail="未找到用户认证信息")

    try:
        user_info = get_user_info_cache(username)
        role_id = user_info.get("role_id") if user_info else None

        # text 已从 database 模块导入
        from dotenv import load_dotenv
        load_dotenv()
        today = datetime.now().date()

        with get_connection() as conn:
            # 1. 延期任务统计（按项目）
            result = conn.execute(text("""
                SELECT p.id, p.name, p.leader,
                       COUNT(*) as delayed_count,
                       MAX(CURRENT_DATE - pt.end_date) as max_delay_days
                FROM projects p
                JOIN project_tasks pt ON CAST(pt.project_id AS INTEGER) = p.id
                WHERE p.is_deleted = false
                  AND pt.is_deleted = false
                  AND pt.end_date < CURRENT_DATE
                  AND pt.actual_end_date IS NULL
                GROUP BY p.id, p.name, p.leader
                ORDER BY delayed_count DESC
                LIMIT 10
            """))

            delayed_projects = []
            for row in result:
                delayed_projects.append({
                    "project_id": row[0],
                    "project_name": row[1],
                    "leader": row[2],
                    "delayed_count": row[3],
                    "max_delay_days": row[4]
                })

            # 2. 今日未填报人员（仅管理员可见）
            unreported_users = []
            if role_id == 11:  # 系统管理员
                result = conn.execute(text("""
                    SELECT p.employee_id, p.name, p.department
                    FROM personnel p
                    WHERE p.is_deleted = false
                      AND p.role_id IN (13, 14)
                      AND p.employee_id NOT IN (
                          SELECT DISTINCT employee_id
                          FROM daily_reports
                          WHERE report_date = :today AND is_deleted = false
                      )
                    ORDER BY p.name
                    LIMIT 20
                """), {"today": today})

                for row in result:
                    unreported_users.append({
                        "employee_id": row[0],
                        "name": row[1],
                        "department": row[2]
                    })

            # 3. 高风险项目（延期率>30%）
            result = conn.execute(text("""
                SELECT p.id, p.name, p.leader,
                       COUNT(pt.task_id) as total_tasks,
                       SUM(CASE WHEN pt.end_date < CURRENT_DATE
                                AND pt.actual_end_date IS NULL THEN 1 ELSE 0 END) as delayed_tasks,
                       ROUND(100.0 * SUM(CASE WHEN pt.end_date < CURRENT_DATE
                                              AND pt.actual_end_date IS NULL THEN 1 ELSE 0 END)
                             / NULLIF(COUNT(pt.task_id), 0), 1) as delay_rate
                FROM projects p
                JOIN project_tasks pt ON CAST(pt.project_id AS INTEGER) = p.id
                WHERE p.is_deleted = false
                  AND pt.is_deleted = false
                GROUP BY p.id, p.name, p.leader
                HAVING 100.0 * SUM(CASE WHEN pt.end_date < CURRENT_DATE
                                        AND pt.actual_end_date IS NULL THEN 1 ELSE 0 END)
                       / NULLIF(COUNT(pt.task_id), 0) > 30
                ORDER BY delay_rate DESC
                LIMIT 5
            """))

            high_risk_projects = []
            for row in result:
                high_risk_projects.append({
                    "project_id": row[0],
                    "project_name": row[1],
                    "leader": row[2],
                    "total_tasks": row[3],
                    "delayed_tasks": row[4],
                    "delay_rate": float(row[5] or 0)
                })

        return {
            "delayed_projects": delayed_projects,
            "unreported_users": unreported_users,
            "high_risk_projects": high_risk_projects,
            "is_admin": role_id == 11
        }

    except Exception as e:
        logger.error(f" {e}")
        import traceback
        traceback.print_exc()
        return {
            "delayed_projects": [],
            "unreported_users": [],
            "high_risk_projects": [],
            "is_admin": False
        }


@app.get("/agent/api/agent/dashboard/my-project-risks")
async def get_my_project_risks(current_user: Dict = Depends(get_current_user)):
    """
    获取我负责的项目风险预警

    返回用户作为项目负责人的所有项目的风险信息
    """
    username = current_user.get("username") or current_user.get("sub")
    employee_id = current_user.get("employee_id") or username

    if not employee_id:
        return []

    # text 已从 database 模块导入
    from dotenv import load_dotenv
    load_dotenv()
    with get_connection() as conn:
        # 先通过 employee_id 查询员工信息（包括内部 id）
        emp_result = conn.execute(text("""
            SELECT id, name FROM personnel 
            WHERE employee_id = :emp_id AND is_deleted = false
            LIMIT 1
        """), {"emp_id": employee_id})
        emp_row = emp_result.fetchone()
        employee_name = emp_row[1] if emp_row else None
        personnel_id = emp_row[0] if emp_row else None  # personnel 表的内部 id

        # 查询用户负责的项目（通过 leader 字段匹配姓名或 leader_id 匹配 personnel.id）
        if personnel_id:
            result = conn.execute(text("""
                SELECT
                    p.id as project_id,
                    p.name as project_name,
                    p.leader
                FROM projects p
                WHERE p.is_deleted = false
                  AND (p.leader = :emp_name OR p.leader_id = :pid)
                ORDER BY p.id
            """), {"emp_name": employee_name or "", "pid": personnel_id})
        else:
            result = conn.execute(text("""
                SELECT
                    p.id as project_id,
                    p.name as project_name,
                    p.leader
                FROM projects p
                WHERE p.is_deleted = false
                  AND p.leader = :emp_name
                ORDER BY p.id
            """), {"emp_name": employee_name or ""})

        risks = []
        for row in result:
            project_id = row[0]

            # 查询最新版本的任务列表
            # 分类逻辑：
            # - completed: 按时完成（actual_end_date <= end_date）
            # - delayed_completed: 延期完成（actual_end_date > end_date）
            # - ongoing: 进行中（start_date <= 今天 <= end_date，未完成）
            # - delayed: 延期未完成（end_date < 今天，未完成）
            # - not_started: 未开始（start_date > 今天）
            tasks_result = conn.execute(text("""
                WITH max_ver AS (
                    SELECT MAX(CAST(SUBSTRING(task_id FROM 'V([0-9]+)') AS INTEGER)) as mv
                    FROM project_tasks
                    WHERE CAST(project_id AS INTEGER) = :project_id AND is_deleted = false
                )
                SELECT
                    task_id, task_name, status, progress, start_date, end_date, actual_end_date, assignee_id,
                    CASE
                        WHEN actual_end_date IS NOT NULL AND actual_end_date > end_date THEN actual_end_date - end_date
                        WHEN end_date < CURRENT_DATE AND actual_end_date IS NULL THEN CURRENT_DATE - end_date
                        ELSE 0
                    END as delay_days,
                    CASE
                        WHEN actual_end_date IS NOT NULL AND actual_end_date > end_date THEN 'delayed_completed'
                        WHEN actual_end_date IS NOT NULL AND actual_end_date <= end_date THEN 'completed'
                        WHEN end_date < CURRENT_DATE AND actual_end_date IS NULL THEN 'delayed'
                        WHEN start_date IS NOT NULL AND start_date <= CURRENT_DATE AND actual_end_date IS NULL THEN 'ongoing'
                        ELSE 'not_started'
                    END as task_status
                FROM project_tasks, max_ver
                WHERE CAST(project_id AS INTEGER) = :project_id
                  AND is_deleted = false
                  AND COALESCE(CAST(SUBSTRING(task_id FROM 'V([0-9]+)') AS INTEGER), 0) = COALESCE(max_ver.mv, 0)
            """), {"project_id": project_id})

            tasks = {"completed": [], "delayed_completed": [], "ongoing": [], "delayed": [], "not_started": []}

            for task in tasks_result:
                task_data = {
                    "task_id": task[0], "task_name": task[1], "status": task[2],
                    "progress": float(task[3] or 0),
                    "start_date": str(task[4]) if task[4] else None,
                    "end_date": str(task[5]) if task[5] else None,
                    "actual_end_date": str(task[6]) if task[6] else None,
                    "assignee_id": task[7], "delay_days": task[8] or 0,
                    "task_status": task[9] or 'ongoing'
                }
                status = task_data["task_status"]
                if status in tasks:
                    tasks[status].append(task_data)

            total_tasks = sum(len(tasks[k]) for k in tasks)
            delayed_count = len(tasks["delayed"]) + len(tasks["delayed_completed"])
            completed_count = len(tasks["completed"]) + len(tasks["delayed_completed"])
            all_tasks = [t for k in tasks for t in tasks[k]]
            avg_progress = sum(t["progress"] for t in all_tasks) / total_tasks if total_tasks > 0 else 0
            project_progress = round((completed_count / total_tasks * 100 + avg_progress) / 2, 1) if total_tasks > 0 else 0

            risks.append({
                "project_id": row[0], "project_name": row[1], "leader": row[2],
                "total_tasks": total_tasks, "progress": project_progress,
                "delayed_tasks": len(tasks["delayed"]), "delayed_count": delayed_count,
                "tasks": tasks
            })

        return risks


@app.get("/agent/api/agent/stats/team-work-hours")
async def get_team_work_hours(current_user: Dict = Depends(get_current_user)):
    """
    获取团队工时统计（组织隔离）

    权限规则：
    - 系统管理员：看全部项目工时
    - 财务/看板：看研究院项目工时
    - 普通用户：看本组织及下级项目工时
    """
    username = current_user.get("username") or current_user.get("sub")

    from dotenv import load_dotenv
    load_dotenv()
    
    today = datetime.now().date()
    month_start = today.replace(day=1)
    month_end = (month_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)

    with get_connection() as conn:
        # 获取用户角色和组织
        user_result = conn.execute(text("""
            SELECT p.role_id, p.org_id
            FROM personnel p
            WHERE p.employee_id = :username AND p.is_deleted = false
        """), {"username": username}).fetchone()
        
        if not user_result:
            return []
        
        role_id, org_id = user_result
        
        # 构建项目过滤条件
        if role_id == 11:  # 系统管理员
            org_filter = ""
        elif role_id in [15, 17]:  # 财务、看板
            org_filter = "AND p.org_id IN (SELECT id FROM organizations WHERE id = 2 OR parent_id = 2)"
        elif org_id:  # 普通用户
            org_filter = f"AND p.org_id IN (SELECT id FROM organizations WHERE id = {org_id} OR parent_id = {org_id})"
        else:
            return []  # 无组织用户

        # 查询本组织项目下所有成员的工时
        result = conn.execute(text(f"""
            SELECT
                p.name as project_name,
                per.name as member_name,
                SUM(dwi.hours_spent) as total_hours
            FROM daily_work_items dwi
            JOIN daily_reports dr ON dr.id = dwi.report_id
            JOIN personnel per ON per.employee_id = dr.employee_id AND per.is_deleted = false
            JOIN projects p ON dwi.project_id = p.id
            WHERE p.is_deleted = false
              {org_filter}
              AND dr.report_date >= :month_start
              AND dr.report_date <= :month_end
              AND dr.is_deleted = false
            GROUP BY p.name, per.name
            ORDER BY p.name, total_hours DESC
        """), {
            "month_start": month_start,
            "month_end": month_end
        })

        # 按项目分组
        project_hours = {}
        for row in result:
            project_name = row[0]
            member_name = row[1]
            hours = float(row[2] or 0)

            if project_name not in project_hours:
                project_hours[project_name] = {
                    "project_name": project_name,
                    "members": [],
                    "total_hours": 0
                }

            project_hours[project_name]["members"].append({
                "name": member_name,
                "hours": round(hours, 1),
                "percent": 0  # 稍后计算
            })
            project_hours[project_name]["total_hours"] += hours

        # 计算百分比
        result_list = []
        for project_data in project_hours.values():
            total = project_data["total_hours"]
            for member in project_data["members"]:
                member["percent"] = round(100 * member["hours"] / total, 1) if total > 0 else 0
            project_data["total_hours"] = round(total, 1)
            result_list.append(project_data)

        return result_list


# 导入节假日计算模块
from .holidays import calculate_working_days

@app.get("/agent/api/agent/stats/monthly-employee-hours")
@app.get("/agent/api/agent/plan-versions/pending-evaluation")
async def get_pending_evaluation_versions(
    current_user: Dict = Depends(get_current_user)
):
    """
    获取待评估的计划版本列表
    
    条件：
    1. 非初始计划（change_type != '初始计划'）
    2. 无效果评估（effect_note IS NULL）
    3. 上传超过7天
    """
    try:
        with get_connection() as conn:
            result = conn.execute(text("""
                SELECT 
                    pv.id,
                    pv.project_id,
                    pv.version_number,
                    pv.version_name,
                    pv.change_type,
                    pv.change_reason,
                    pv.upload_time,
                    pv.upload_by,
                    p.name as project_name,
                    pv.previous_status
                FROM project_plan_versions pv
                JOIN projects p ON p.id = pv.project_id
                WHERE pv.change_type != '初始计划'
                  AND pv.effect_note IS NULL
                  AND pv.upload_time < NOW() - INTERVAL '7 days'
                  AND p.is_deleted = false
                ORDER BY pv.upload_time DESC
            """))
            
            versions = []
            for row in result:
                # 计算距今天数
                upload_time = row[6]
                days_ago = (datetime.now() - upload_time).days if upload_time else 0
                
                versions.append({
                    "id": row[0],
                    "project_id": row[1],
                    "version_number": row[2],
                    "version_name": row[3],
                    "change_type": row[4],
                    "change_reason": row[5],
                    "upload_time": str(row[6]),
                    "upload_by": row[7],
                    "project_name": row[8],
                    "previous_status": row[9],
                    "days_ago": days_ago
                })
            
            return {"versions": versions}
    
    except Exception as e:
        logger.exception(f"获取待评估版本失败: {e}")
        return {"versions": []}


async def get_monthly_employee_hours(
    year: int = None,
    month: int = None,
    current_user: Dict = Depends(get_current_user)
):
    """
    获取月度人员工时统计（仅统计正式项目工时）

    参数:
    - year: 年份（默认当前年）
    - month: 月份（默认当前月）

    返回:
    - 所有员工在指定月份的工时统计
    - 按项目分组，显示每个员工在各项目的工时
    - 包含应填日报数、实填日报数、差异
    - 注意：总工时只统计正式项目，不包含其他工作（基础工作等）
    """
    # 默认当月
    today = datetime.now().date()
    year = year or today.year
    month = month or today.month

    # 计算月份起止日期
    month_start = datetime(year, month, 1).date()
    month_end = (month_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)

    # 动态计算工作日（使用holidays模块）
    working_days = calculate_working_days(year, month)

    with get_connection() as conn:
        # 查询每个员工实际填写的日报天数（去重）
        days_result = conn.execute(text("""
            SELECT 
                dr.employee_name,
                COUNT(DISTINCT dr.report_date) as days_filled
            FROM daily_reports dr
            WHERE dr.report_date >= :month_start
              AND dr.report_date <= :month_end
              AND dr.is_deleted = false
              AND LOWER(dr.employee_name) != 'admin'
            GROUP BY dr.employee_name
        """), {
            "month_start": month_start,
            "month_end": month_end
        })
        
        # 员工实际填报天数
        employee_days = {row[0]: row[1] for row in days_result}

        # 查询正式项目的月度工时（仅统计有project_id且匹配projects表的）
        result = conn.execute(text("""
            SELECT
                dr.employee_name,
                p.name as project_name,
                SUM(dwi.hours_spent) as total_hours,
                COUNT(DISTINCT dr.id) as report_count
            FROM daily_work_items dwi
            JOIN daily_reports dr ON dr.id = dwi.report_id
            JOIN projects p ON p.id::text = dwi.project_id
            WHERE dr.report_date >= :month_start
              AND dr.report_date <= :month_end
              AND dr.is_deleted = false
              AND LOWER(dr.employee_name) != 'admin'
              AND dwi.project_id IS NOT NULL AND dwi.project_id != ''
            GROUP BY dr.employee_name, p.name
            ORDER BY dr.employee_name, total_hours DESC
        """), {
            "month_start": month_start,
            "month_end": month_end
        })

        # 按员工分组
        employee_hours = {}
        for row in result:
            emp_name = row[0]
            project_name = row[1]
            hours = float(row[2] or 0)
            report_count = row[3] or 0

            if emp_name not in employee_hours:
                filled = employee_days.get(emp_name, 0)
                employee_hours[emp_name] = {
                    "employee_name": emp_name,
                    "projects": [],
                    "total_hours": 0,
                    "total_hours_raw": 0,  # 原始精度累加
                    "report_count": 0,
                    "required_days": working_days,
                    "filled_days": filled,
                    "missing_days": max(0, working_days - filled)  # 加班时缺失为0，不为负数
                }

            employee_hours[emp_name]["projects"].append({
                "project_name": project_name,
                "hours": round(hours, 1)
            })
            # 统一：累加原始值，最后统一round
            employee_hours[emp_name]["total_hours_raw"] += hours  # 累加原始精度
            employee_hours[emp_name]["report_count"] += report_count

        # 转为列表并计算百分比
        result_list = []
        total_hours_raw = 0  # 累加原始精度工时
        for emp_data in employee_hours.values():
            total_raw = emp_data["total_hours_raw"]
            total = round(total_raw, 1)  # 每人总工时round
            for proj in emp_data["projects"]:
                proj["percent"] = round(100 * proj["hours"] / total, 1) if total > 0 else 0
            emp_data["total_hours"] = total
            total_hours_raw += total_raw  # 累加原始精度
            result_list.append(emp_data)

        # 按总工时排序
        result_list.sort(key=lambda x: x["total_hours"], reverse=True)

        return {
            "year": year,
            "month": month,
            "month_start": month_start.isoformat(),
            "month_end": month_end.isoformat(),
            "working_days": working_days,  # 当月工作日数
            "employee_count": len(result_list),  # 参与人数
            "employees": result_list,
            "total_hours": round(total_hours_raw, 1),  # 总工时统一round（仅正式项目）
            "total_reports": sum(e["report_count"] for e in result_list)
        }


def classify_other_work(work_content: str, project_name: str) -> str:
    """
    分类基础工作：项目类、行政类、会议类、日常类
    
    优先级：
    1. 项目类：提到具体项目名或项目相关关键词（方案、编制、设计、研发等）
    2. 行政类：审批、签字、财务、采购等行政事务
    3. 会议类：会议、早会、培训、评审会等（汇报单独判断）
    4. 日常类：填报、整理、检查、录入等日常事务
    """
    content = (work_content or '') + ' ' + (project_name or '')
    
    # 项目类关键词（优先判断）
    project_keywords = ['项目', '方案编制', '方案设计', '技术方案', '可行性分析', 
                        '立项', '研发', '调研', '前期', '现场调研', '协调', 
                        '跟进', '推进', '落实', '编写', '编制', '修改', '完善',
                        '技术交流', '供应商交流', '设备选型']
    # 判断是否提到项目名或项目相关工作
    for kw in project_keywords:
        if kw in content:
            # 但如果是汇报会议，归入会议类
            if '汇报' in content and '会' in content:
                continue
            return '项目类'
    
    # 行政类关键词
    admin_keywords = ['审批', '签字', '盖章', '报销', '发票', '合同', 
                      '采购', '财务', '付款', '资金计划', '招标', '订价',
                      '流程', '申请', '审核', '质保金', '开票']
    for kw in admin_keywords:
        if kw in content:
            return '行政类'
    
    # 会议类关键词（精确匹配，避免误判）
    meeting_keywords = ['会议', '早会', '晚会', '评审会', '分析会', '讨论会', 
                        '培训', '参加', '交流会', '立项评审']
    for kw in meeting_keywords:
        if kw in content:
            # 排除：汇报工作内容（不含"会"字）
            if '汇报' in content and '会' not in content:
                return '项目类'  # 向领导汇报归入项目类
            return '会议类'
    
    # 日常类关键词
    daily_keywords = ['检查', '整理', '任务清单', 'KPI', 
                      '绩效', '督办', '填写', '填报', '台账', '报表',
                      '录入', '数据', '资料', '电脑', '设备维护',
                      '安全检查', '隐患', '梳理', '汇总', '统计']
    for kw in daily_keywords:
        if kw in content:
            return '日常类'
    
    # 其余归入项目类
    return '项目类'


def extract_category_name(work_content: str) -> str:
    """
    从工作内容中智能提取分类名称
    
    规则：
    1. 如果内容明确指向某个项目（如"XXX项目技术方案"），提取项目名
    2. 如果内容描述具体工作类型（如"会议"、"审核"），提取工作类型
    3. 如果用户明确写"其他工作"，归为"其他工作"
    4. 否则，提取关键实体作为分类名
    """
    import re
    
    work_content = work_content.strip()
    if not work_content:
        return "其他工作"
    
    # 用户明确提到"其他工作"且无具体内容
    if work_content == "其他工作" or work_content.lower() == "其他工作":
        return "其他工作"
    
    # 提取项目类关键词（如"XXX项目"、"XXX系统"）
    project_patterns = [
        r'([^\s]+项目)[^\s]*',  # XXX项目
        r'([^\s]+系统)[^\s]*(?:技术方案|设计|研发|调试)',  # XXX系统技术方案
        r'([^\s]+研发项目)',  # XXX研发项目
        r'([^\s]+监控系统)',  # XXX监控系统
        r'([^\s]+控制系统)',  # XXX控制系统
        r'([^\s]+自动化控制)',  # XXX自动化控制
    ]
    
    for pattern in project_patterns:
        match = re.search(pattern, work_content)
        if match:
            category = match.group(1).strip()
            # 清理前缀词（如"前往"、"编写"等）
            category = re.sub(r'^前往|^编写|^完成|^开展|^参加|^在|^到', '', category).strip()
            return category
    
    # 提取地点+设备类（如"田阳新材料熔炼炉"）
    location_device_pattern = r'(田阳|隆林|德保|田林|靖锰|精铝)[^\s]*(?:熔炼炉|铝厂|车间|新材料)[^\s]*(?:激光测距仪|在线监测|设备调试)'
    match = re.search(location_device_pattern, work_content)
    if match:
        return match.group(0).strip()
    
    # 提取工作类型关键词
    work_type_keywords = {
        '会议': ['会议', '早会', '晚会', '评审会', '讨论会', '立项评审会', '审查会', '培训', '学习', '交流', '调研'],
        '行政审批': ['审核', '盖章', '签字', '审批', '提交', '督办', '流程', 'KPI', '绩效', '任务清单', '关键任务'],
        '采购财务': ['采购', '付款', '报销', '发票', '合同', '订价', '招标', '物资', '材料', '财务'],
        '安全检查': ['安全检查', '实验室安全', '隐患', '消防', '安全培训'],
        '设备调试': ['调试', '安装', '维修', '维护', '检修', '故障处理', '设备检查'],
        '技术方案': ['技术方案', '设计方案', '立项报告', '可行性分析', '成本分析', '试验方案'],
    }
    
    for work_type, keywords in work_type_keywords.items():
        for kw in keywords:
            if kw in work_content:
                # 如果找到了具体的项目/地点，结合显示
                location_match = re.search(r'(田阳|隆林|德保|田林|靖锰|精铝|新材料)', work_content)
                if location_match:
                    return f"{location_match.group(1)}{work_type}"
                return work_type
    
    # 提取地点/车间关键词
    location_patterns = [
        r'(田阳新材料[^\s]*)',
        r'(精铝车间)',
        r'(田阳铝厂)',
        r'(隆林铝厂)',
        r'(德保铝厂)',
        r'(田林铝厂)',
        r'(靖锰公司)',
    ]
    
    for pattern in location_patterns:
        match = re.search(pattern, work_content)
        if match:
            return match.group(1).strip()
    
    # 最后：使用工作内容的前30个字符作为分类名（去掉动词前缀）
    short_content = work_content[:50]
    short_content = re.sub(r'^前往|^编写|^完成|^开展|^参加|^在|^到|^根据|^整理|^处理|^制作|^核对|^重新|^提交|^出差', '', short_content).strip()
    return short_content if short_content else "其他工作"


@app.get("/agent/api/agent/stats/monthly-project-hours")
async def get_monthly_project_hours(
    year: int = None,
    month: int = None,
    current_user: Dict = Depends(get_current_user)
):
    """
    获取月度项目工时统计（项目维度）
    
    判断逻辑：
    1. project_id有值 → 通过project_id匹配正式项目
    2. project_id为空 → project_name精确匹配正式项目表
    3. 都匹配不上 → 归类为其他工作
    """
    today = datetime.now().date()
    year = year or today.year
    month = month or today.month

    month_start = datetime(year, month, 1).date()
    month_end = (month_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)

    # 动态计算工作日（使用holidays模块）
    working_days = calculate_working_days(year, month)
    month_end = (month_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)

    with get_connection() as conn:
        # 获取正式项目的ID->名称映射
        official_projects_result = conn.execute(text("""
            SELECT id::text, name FROM projects 
            WHERE is_deleted = false 
            ORDER BY name
        """))
        project_id_to_name = {row[0]: row[1] for row in official_projects_result}
        project_name_to_id = {row[1]: row[0] for row in official_projects_result}
        
        # 查询正式项目的工时数据
        official_result = conn.execute(text("""
            SELECT 
                dwi.project_id,
                p.name as project_name,
                dr.employee_name,
                SUM(dwi.hours_spent) as total_hours
            FROM daily_work_items dwi
            JOIN daily_reports dr ON dr.id = dwi.report_id
            JOIN projects p ON p.id::text = dwi.project_id
            WHERE dr.report_date >= :month_start
              AND dr.report_date <= :month_end
              AND dr.is_deleted = false
              AND LOWER(dr.employee_name) != 'admin'
              AND dwi.project_id IS NOT NULL AND dwi.project_id != ''
            GROUP BY dwi.project_id, p.name, dr.employee_name
            ORDER BY p.name, total_hours DESC
        """), {
            "month_start": month_start,
            "month_end": month_end
        })
        
        # 查询未匹配项目的工时数据
        other_result = conn.execute(text("""
            SELECT 
                dwi.project_name,
                dwi.work_content,
                dr.employee_name,
                SUM(dwi.hours_spent) as total_hours
            FROM daily_work_items dwi
            JOIN daily_reports dr ON dr.id = dwi.report_id
            WHERE dr.report_date >= :month_start
              AND dr.report_date <= :month_end
              AND dr.is_deleted = false
              AND LOWER(dr.employee_name) != 'admin'
              AND (dwi.project_id IS NULL OR dwi.project_id = '')
            GROUP BY dwi.project_name, dwi.work_content, dr.employee_name
            ORDER BY total_hours DESC
        """), {
            "month_start": month_start,
            "month_end": month_end
        })

        # 分类：正式项目 vs 其他工作（通过project_id判断）
        official_project_hours = {}
        other_work_hours = {}
        all_employees = set()
        
        # 处理正式项目数据
        for row in official_result:
            project_id = row[0]
            project_name = row[1]
            emp_name = row[2]
            hours = float(row[3] or 0)
            
            all_employees.add(emp_name)
            
            if project_name not in official_project_hours:
                official_project_hours[project_name] = {
                    "project_name": project_name,
                    "members": {},
                    "total_hours": 0
                }
            
            official_project_hours[project_name]["members"][emp_name] = official_project_hours[project_name]["members"].get(emp_name, 0) + hours
            official_project_hours[project_name]["total_hours"] += hours
        
        # 处理未匹配项目数据 - 按四类分类
        for row in other_result:
            project_name_from_db = row[0] or ""
            work_content_from_db = row[1] or ""
            emp_name = row[2]
            hours = float(row[3] or 0)
            
            all_employees.add(emp_name)
            
            # 使用新的分类函数
            category_name = classify_other_work(work_content_from_db, project_name_from_db)
            
            if category_name not in other_work_hours:
                other_work_hours[category_name] = {
                    "project_name": category_name,
                    "members": {},
                    "total_hours": 0
                }
            
            other_work_hours[category_name]["members"][emp_name] = other_work_hours[category_name]["members"].get(emp_name, 0) + hours
            other_work_hours[category_name]["total_hours"] += hours

        # 转为列表并排序（保留原始精度）
        official_list = []
        official_employee_totals_raw: Dict[str, float] = {}
        
        for proj_name, proj_data in official_project_hours.items():
            # 累加原始精度工时到员工汇总
            for emp, hours in proj_data["members"].items():
                official_employee_totals_raw[emp] = official_employee_totals_raw.get(emp, 0) + hours
            # 累加原始精度到项目总工时
            proj_total_raw = proj_data["total_hours"]
            proj_data["total_hours"] = round(proj_total_raw, 1)
            proj_data["members"] = {k: round(v, 1) for k, v in proj_data["members"].items()}
            official_list.append(proj_data)
        
        official_list.sort(key=lambda x: x["total_hours"], reverse=True)
        # 使用原始精度计算员工总工时，最后统一round
        official_employee_totals = {k: round(v, 1) for k, v in sorted(official_employee_totals_raw.items(), key=lambda x: x[1], reverse=True)}
        official_grand_total = round(sum(official_employee_totals_raw.values()), 1)
        
        # 其他工作列表（保留原始精度）
        other_list = []
        other_employee_totals_raw: Dict[str, float] = {}
        
        for proj_name, proj_data in other_work_hours.items():
            # 累加原始精度工时到员工汇总
            for emp, hours in proj_data["members"].items():
                other_employee_totals_raw[emp] = other_employee_totals_raw.get(emp, 0) + hours
            # 累加原始精度到项目总工时
            proj_total_raw = proj_data["total_hours"]
            proj_data["total_hours"] = round(proj_total_raw, 1)
            proj_data["members"] = {k: round(v, 1) for k, v in proj_data["members"].items()}
            other_list.append(proj_data)
        
        other_list.sort(key=lambda x: {
            "项目类": 0,
            "行政类": 1,
            "会议类": 2,
            "日常类": 3
        }.get(x["project_name"], 4))
        # 使用原始精度计算员工总工时，最后统一round
        other_employee_totals = {k: round(v, 1) for k, v in sorted(other_employee_totals_raw.items(), key=lambda x: x[1], reverse=True)}
        other_grand_total = round(sum(other_employee_totals_raw.values()), 1)
        
        # 全部员工列表（合并两部分）
        all_employee_totals: Dict[str, float] = {}
        for emp, hours in official_employee_totals_raw.items():
            all_employee_totals[emp] = hours
        for emp, hours in other_employee_totals_raw.items():
            all_employee_totals[emp] = all_employee_totals.get(emp, 0) + hours
        all_employee_totals = {k: round(v, 1) for k, v in sorted(all_employee_totals.items(), key=lambda x: x[1], reverse=True)}
        
        grand_total = round(sum(official_employee_totals_raw.values()) + sum(other_employee_totals_raw.values()), 1)
        
        return {
            "year": year,
            "month": month,
            "working_days": working_days,  # 月份总工作日（动态计算）
            "employee_count": len(all_employees),  # 参与人数
            "official_projects": official_list,
            "official_employee_totals": official_employee_totals,
            "official_grand_total": official_grand_total,
            "other_works": other_list,
            "other_employee_totals": other_employee_totals,
            "other_grand_total": other_grand_total,
            "all_employees": list(all_employee_totals.keys()),
            "all_employee_totals": all_employee_totals,
            "grand_total": grand_total,
            "official_project_count": len(official_list),
            "other_work_count": len(other_list)
        }


@app.get("/agent/api/agent/stats/person-project-analysis")
async def get_person_project_analysis(
    employee_name: str = None,
    year: int = None,
    current_user: Dict = Depends(get_current_user)
):
    """
    获取人员项目投入分析（单人员维度）
    
    只统计正式项目（projects表中的项目），不含早会等基础工作
    
    参数:
    - employee_name: 人员姓名（必填）
    - year: 年份（默认当前年）
    """
    if not employee_name:
        raise HTTPException(status_code=400, detail="请指定人员姓名")
    
    today = datetime.now().date()
    year = year or today.year
    year_start = datetime(year, 1, 1).date()
    year_end = datetime(year, 12, 31).date()
    
    with get_connection() as conn:
        # 1. 查询该人员在正式项目的工时（只统计projects表中的项目）
        result = conn.execute(text("""
            SELECT 
                p.name as project_name,
                SUM(dwi.hours_spent) as total_hours,
                COUNT(DISTINCT dwi.report_id) as report_count
            FROM daily_work_items dwi
            JOIN daily_reports dr ON dwi.report_id = dr.id
            JOIN projects p ON dwi.project_id = p.id::text
            WHERE dr.employee_name = :employee_name
              AND dr.report_date >= :year_start
              AND dr.report_date <= :year_end
              AND dr.is_deleted = false
              AND p.is_deleted = false
            GROUP BY p.name
            ORDER BY total_hours DESC
        """), {
            "employee_name": employee_name,
            "year_start": year_start,
            "year_end": year_end
        })
        
        projects = []
        total_hours = 0
        for row in result:
            project_hours = float(row[1] or 0)
            projects.append({
                "project_name": row[0],
                "hours": round(project_hours, 1),
                "report_count": row[2]
            })
            total_hours += project_hours
        
        # 2. 计算占比（基于正式项目总工时）
        for p in projects:
            p["percent"] = round(p["hours"] / total_hours * 100, 1) if total_hours > 0 else 0
        
        # 3. 查询月度工时趋势（只统计正式项目）
        monthly_result = conn.execute(text("""
            SELECT 
                EXTRACT(MONTH FROM dr.report_date) as month,
                SUM(dwi.hours_spent) as hours
            FROM daily_work_items dwi
            JOIN daily_reports dr ON dwi.report_id = dr.id
            JOIN projects p ON dwi.project_id = p.id::text
            WHERE dr.employee_name = :employee_name
              AND dr.report_date >= :year_start
              AND dr.report_date <= :year_end
              AND dr.is_deleted = false
              AND p.is_deleted = false
            GROUP BY EXTRACT(MONTH FROM dr.report_date)
            ORDER BY month
        """), {
            "employee_name": employee_name,
            "year_start": year_start,
            "year_end": year_end
        })
        
        monthly_trend = []
        for row in monthly_result:
            monthly_trend.append({
                "month": int(row[0]),
                "hours": round(float(row[1] or 0), 1)
            })
        
        # 4. 获取所有员工列表（用于下拉选择）
        employees_result = conn.execute(text("""
            SELECT DISTINCT employee_name
            FROM daily_reports
            WHERE is_deleted = false
              AND LOWER(employee_name) != 'admin'
            ORDER BY employee_name
        """))
        all_employees = [row[0] for row in employees_result]
        
        return {
            "employee_name": employee_name,
            "year": year,
            "total_hours": round(total_hours, 1),
            "project_count": len(projects),
            "projects": projects,
            "monthly_trend": monthly_trend,
            "all_employees": all_employees
        }


@app.get("/agent/api/agent/stats/project-employee-details")
async def get_project_employee_details(
    project_name: str,
    employee_name: str,
    year: int,
    month: int,
    current_user: Dict = Depends(get_current_user)
):
    """
    获取指定项目、人员、月份的日报详情列表
    
    project_name 可能是：
    - 正式项目名
    - 基础工作分类：会议类、行政类、日常类、项目类
    """
    from urllib.parse import unquote
    
    project_name = unquote(project_name)
    employee_name = unquote(employee_name)
    
    month_start = datetime(year, month, 1).date()
    month_end = (month_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)
    
    with get_connection() as conn:
        # 判断是基础工作分类还是正式项目
        basic_categories = ['会议类', '行政类', '日常类', '项目类']
        
        if project_name in basic_categories:
            # 基础工作分类：查询所有未匹配项目的日报，再按分类过滤
            result = conn.execute(text("""
                SELECT 
                    dr.report_date,
                    dwi.project_name,
                    dwi.work_content,
                    dwi.hours_spent,
                    dwi.start_time,
                    dwi.end_time
                FROM daily_work_items dwi
                JOIN daily_reports dr ON dr.id = dwi.report_id
                JOIN personnel p ON dr.employee_id = p.employee_id
                WHERE p.name = :emp_name
                  AND dr.report_date >= :month_start
                  AND dr.report_date <= :month_end
                  AND dr.is_deleted = false
                  AND dwi.is_deleted = false
                  AND (dwi.project_id IS NULL OR dwi.project_id = '')
                ORDER BY dr.report_date DESC
            """), {
                "emp_name": employee_name,
                "month_start": month_start,
                "month_end": month_end
            })
            
            # 在 Python 中过滤分类
            details = []
            total_hours = 0
            for row in result:
                work_content = row[2] or ""
                proj_name = row[1] or ""
                # 使用分类函数判断
                if classify_other_work(work_content, proj_name) == project_name:
                    details.append({
                        "date": str(row[0]),
                        "project": row[1] or "基础工作",
                        "content": row[2],
                        "hours": float(row[3] or 0),
                        "time_range": f"{row[4] or ''}-{row[5] or ''}" if row[4] and row[5] else ""
                    })
                    total_hours += float(row[3] or 0)
        else:
            # 正式项目：按 project_name 匹配
            result = conn.execute(text("""
                SELECT 
                    dr.report_date,
                    dwi.project_name,
                    dwi.work_content,
                    dwi.hours_spent,
                    dwi.start_time,
                    dwi.end_time
                FROM daily_work_items dwi
                JOIN daily_reports dr ON dr.id = dwi.report_id
                JOIN personnel p ON dr.employee_id = p.employee_id
                WHERE p.name = :emp_name
                  AND dr.report_date >= :month_start
                  AND dr.report_date <= :month_end
                  AND dr.is_deleted = false
                  AND dwi.is_deleted = false
                  AND dwi.project_name = :project_name
                ORDER BY dr.report_date DESC
            """), {
                "emp_name": employee_name,
                "project_name": project_name,
                "month_start": month_start,
                "month_end": month_end
            })
            
            details = []
            total_hours = 0
            for row in result:
                details.append({
                    "date": str(row[0]),
                    "project": row[1],
                    "content": row[2],
                    "hours": float(row[3] or 0),
                    "time_range": f"{row[4] or ''}-{row[5] or ''}" if row[4] and row[5] else ""
                })
                total_hours += float(row[3] or 0)
        
        return {
            "project_name": project_name,
            "employee_name": employee_name,
            "year": year,
            "month": month,
            "details": details,
            "total_hours": round(total_hours, 2),
            "count": len(details)
        }



@app.get("/agent/api/agent/stats/monthly-employee-hours/export")
async def export_monthly_employee_hours(
    year: int = None,
    month: int = None,
    current_user: Dict = Depends(get_current_user)
):
    """
    导出月度人员工时统计为Excel（包含人员维度和正式项目维度两个sheet）
    注意：只统计正式项目工时，排除其他工作（含基础工作的项目类）
    """
    from fastapi.responses import StreamingResponse
    import io
    import pandas as pd
    from urllib.parse import quote
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    # 调用统计API获取数据
    employee_data = await get_monthly_employee_hours(year, month, current_user)
    project_data = await get_monthly_project_hours(year, month, current_user)

    output = io.BytesIO()
    
    # 样式定义
    header_font = Font(bold=True, size=12, color='FFFFFF')
    header_fill = PatternFill(start_color='3B82F6', end_color='3B82F6', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    subtotal_fill = PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid')
    subtotal_font = Font(bold=True)
    
    grand_total_fill = PatternFill(start_color='E0E7FF', end_color='E0E7FF', fill_type='solid')
    grand_total_font = Font(bold=True, size=12, color='6366F1')
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Sheet1: 人员维度汇总
        summary_rows = []
        for emp in employee_data["employees"]:
            if emp["employee_name"].lower() == "admin":
                continue
            summary_rows.append({
                "员工姓名": emp["employee_name"],
                "应填日报": emp["required_days"],
                "实填日报": emp["filled_days"],
                "缺失": emp["missing_days"],
                "总工时(h)": emp["total_hours"],
                "填报率(%)": round(100 * emp["filled_days"] / emp["required_days"], 1) if emp["required_days"] > 0 else 0
            })
        
        summary_sheet = f'{employee_data["year"]}年{employee_data["month"]}月人员汇总'
        pd.DataFrame(summary_rows).to_excel(writer, index=False, sheet_name=summary_sheet)
        ws1 = writer.sheets[summary_sheet]
        
        # 设置列宽
        ws1.column_dimensions['A'].width = 12
        ws1.column_dimensions['B'].width = 10
        ws1.column_dimensions['C'].width = 10
        ws1.column_dimensions['D'].width = 8
        ws1.column_dimensions['E'].width = 12
        ws1.column_dimensions['F'].width = 12
        
        # 应用表头样式
        for cell in ws1[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 应用数据行样式
        for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row, min_col=1, max_col=6):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Sheet2: 正式项目维度（只包含正式项目，排除其他工作）
        official_rows = []
        for proj in project_data["official_projects"]:
            row_data = {"项目名称": proj["project_name"]}
            for emp in project_data["all_employees"]:
                row_data[emp] = proj["members"].get(emp, "")
            row_data["项目小计"] = proj["total_hours"]
            official_rows.append(row_data)
        
        # 正式项目小计行
        official_subtotal = {"项目名称": "【正式项目小计】"}
        for emp in project_data["all_employees"]:
            official_subtotal[emp] = project_data["official_employee_totals"].get(emp, "")
        official_subtotal["项目小计"] = project_data["official_grand_total"]
        official_rows.append(official_subtotal)
        
        project_sheet = f'{project_data["year"]}年{project_data["month"]}月正式项目'
        pd.DataFrame(official_rows).to_excel(writer, index=False, sheet_name=project_sheet)
        ws2 = writer.sheets[project_sheet]
        
        # 设置列宽
        ws2.column_dimensions['A'].width = 35
        for i, emp in enumerate(project_data["all_employees"]):
            col_letter = chr(66 + i)  # B, C, D...
            ws2.column_dimensions[col_letter].width = 10
        last_col = chr(66 + len(project_data["all_employees"]))
        ws2.column_dimensions[last_col].width = 12
        
        # 应用表头样式
        for cell in ws2[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 应用数据行样式
        subtotal_row_num = len(project_data["official_projects"]) + 2
        
        for row_num, row in enumerate(ws2.iter_rows(min_row=2, max_row=ws2.max_row, min_col=1, max_col=len(project_data["all_employees"])+2), start=2):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center' if cell.column > 1 else 'left', vertical='center')
            
            # 小计行样式
            if row_num == subtotal_row_num:
                for cell in row:
                    cell.fill = subtotal_fill
                    cell.font = subtotal_font

    output.seek(0)
    
    # 修改文件名为"正式项目工时统计"
    filename = f'正式项目工时统计_{employee_data["year"]}年{employee_data["month"]}月.xlsx'
    encoded_filename = quote(filename)
    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f"attachment; filename*=UTF-8''{encoded_filename}"}
    )

    # 导出到Excel（两个sheet）
    output = io.BytesIO()
    summary_sheet = f'{data["year"]}年{data["month"]}月汇总'
    detail_sheet = f'{data["year"]}年{data["month"]}月明细'
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # 汇总表
        pd.DataFrame(summary_rows).to_excel(writer, index=False, sheet_name=summary_sheet)
        ws_summary = writer.sheets[summary_sheet]
        ws_summary.column_dimensions['A'].width = 12
        ws_summary.column_dimensions['B'].width = 10
        ws_summary.column_dimensions['C'].width = 10
        ws_summary.column_dimensions['D'].width = 8
        ws_summary.column_dimensions['E'].width = 12
        ws_summary.column_dimensions['F'].width = 12
        
        # 明细表
        pd.DataFrame(detail_rows).to_excel(writer, index=False, sheet_name=detail_sheet)
        ws_detail = writer.sheets[detail_sheet]
        ws_detail.column_dimensions['A'].width = 12
        ws_detail.column_dimensions['B'].width = 30
        ws_detail.column_dimensions['C'].width = 12
        ws_detail.column_dimensions['D'].width = 10

    output.seek(0)

    # 返回文件流（URL编码文件名）
    filename = f'月度工时统计_{data["year"]}年{data["month"]}月.xlsx'
    encoded_filename = quote(filename)
    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f"attachment; filename*=UTF-8''{encoded_filename}"}
    )


# ============== Phase 14: 数据可视化 API ==============

@app.get("/agent/api/agent/dashboard/project-board")
async def get_project_board(current_user: Dict = Depends(get_current_user)):
    """
    获取项目看板数据

    返回所有项目的进度、风险等级、延期任务数
    """
    username = current_user.get("username")
    token = current_user.get("_raw_token") or get_user_token(username)

    if not token:
        raise HTTPException(status_code=401, detail="未找到用户认证信息")

    try:
        user_info = get_user_info_cache(username)

        # text 已从 database 模块导入
        from dotenv import load_dotenv
        load_dotenv()
        with get_connection() as conn:
            # 获取项目列表及风险评分
            result = conn.execute(text("""
                SELECT p.id, p.name, p.leader, p.status, p.progress,
                       COUNT(pt.task_id) as total_tasks,
                       SUM(CASE WHEN pt.end_date < CURRENT_DATE
                                AND pt.actual_end_date IS NULL THEN 1 ELSE 0 END) as delayed_tasks
                FROM projects p
                LEFT JOIN project_tasks pt ON CAST(pt.project_id AS INTEGER) = p.id
                    AND pt.is_deleted = false
                WHERE p.is_deleted = false
                GROUP BY p.id, p.name, p.leader, p.status, p.progress
                ORDER BY p.id DESC
            """))

            projects = []
            for row in result:
                total = row[5] or 0
                delayed = row[6] or 0

                # 计算风险等级
                if total > 0:
                    delay_rate = delayed / total * 100
                    if delay_rate > 30:
                        risk_level = "high"
                    elif delay_rate > 15:
                        risk_level = "medium"
                    else:
                        risk_level = "low"
                else:
                    risk_level = "low"

                projects.append({
                    "id": row[0],
                    "name": row[1],
                    "leader": row[2],
                    "status": row[3] or "进行中",
                    "progress": float(row[4] or 0),
                    "risk_level": risk_level,
                    "delayed_tasks": delayed,
                    "total_tasks": total
                })

            return {"projects": projects}

    except Exception as e:
        logger.error(f" {e}")
        import traceback
        traceback.print_exc()
        return {"projects": []}


@app.get("/agent/api/agent/dashboard/risk-matrix")
async def get_risk_matrix(current_user: Dict = Depends(get_current_user)):
    """
    获取风险矩阵数据

    返回所有项目的进度风险、资源风险、综合风险
    """
    username = current_user.get("username")
    token = current_user.get("_raw_token") or get_user_token(username)

    if not token:
        raise HTTPException(status_code=401, detail="未找到用户认证信息")

    try:
        # text 已从 database 模块导入
        from dotenv import load_dotenv
        load_dotenv()
        today = datetime.now().date()

        with get_connection() as conn:
            result = conn.execute(text("""
                SELECT p.id, p.name,
                       COUNT(pt.task_id) as total_tasks,
                       SUM(CASE WHEN pt.end_date < :today AND pt.actual_end_date IS NULL THEN 1 ELSE 0 END) as delayed_tasks,
                       COUNT(DISTINCT pt.assignee_id) as team_size,
                       SUM(COALESCE(pt.planned_hours, 0)) as planned_hours
                FROM projects p
                LEFT JOIN project_tasks pt ON CAST(pt.project_id AS INTEGER) = p.id
                    AND pt.is_deleted = false
                WHERE p.is_deleted = false
                GROUP BY p.id, p.name
                ORDER BY p.id
            """), {"today": today})

            projects = []
            for row in result:
                total = row[2] or 0
                delayed = row[3] or 0
                team_size = row[4] or 1
                planned_hours = float(row[5] or 0)

                # 进度风险
                if total > 0:
                    schedule_risk = min(100, delayed / total * 100)
                else:
                    schedule_risk = 0

                # 资源风险（简单估算：团队规模 vs 任务量）
                if team_size > 0 and total > 0:
                    tasks_per_person = total / team_size
                    resource_risk = min(100, tasks_per_person * 10)  # 每人超过10个任务开始计风险
                else:
                    resource_risk = 0

                # 综合风险
                overall_risk = (schedule_risk * 0.6 + resource_risk * 0.4)

                projects.append({
                    "project_id": row[0],
                    "project_name": row[1],
                    "schedule_risk": round(schedule_risk, 1),
                    "resource_risk": round(resource_risk, 1),
                    "overall_risk": round(overall_risk, 1)
                })

            return {"projects": projects}

    except Exception as e:
        logger.error(f" {e}")
        return {"projects": []}


# ============== Phase 15: 智能推荐 API ==============

@app.get("/agent/api/agent/dashboard/smart-assistant")
async def get_smart_assistant(current_user: Dict = Depends(get_current_user)):
    """
    智能助手 - 整合今日优先任务 + 延期预警 + 工时预测 + 智能建议

    返回用户登录后应该看到的所有关键信息
    """
    # 直接从 current_user 获取 employee_id
    employee_id = current_user.get("username")

    try:
        # text 已从 database 模块导入
        from dotenv import load_dotenv
        load_dotenv()
        today = datetime.now().date()

        result_data = {
            "priority_tasks": [],      # 今日优先任务（排序后）
            "delayed_warnings": [],    # 延期预警
            "hours_prediction": {},    # 工时预测
            "suggestions": [],         # 智能建议
            "daily_report_status": {}  # 日报状态
        }

        with get_connection() as conn:
            # 1. 今日优先任务（按紧急程度排序）
            result = conn.execute(text("""
                SELECT pt.task_id, pt.task_name, pt.project_id, p.name as project_name,
                       pt.end_date, pt.status, pt.progress
                FROM project_tasks pt
                JOIN projects p ON CAST(pt.project_id AS INTEGER) = p.id
                WHERE pt.is_deleted = false
                  AND p.is_deleted = false
                  AND pt.assignee_id = :emp_id
                  AND pt.actual_end_date IS NULL
                  AND pt.end_date <= CURRENT_DATE + INTERVAL '3 days'
                ORDER BY
                    CASE WHEN pt.end_date < CURRENT_DATE THEN 0
                         WHEN pt.end_date = CURRENT_DATE THEN 1
                         ELSE 2 END,
                    pt.end_date
                LIMIT 5
            """), {"emp_id": employee_id})

            for row in result:
                end_date = row[4]
                is_delayed = end_date < today if end_date else False
                is_today = end_date == today if end_date else False

                # 计算紧急程度
                if is_delayed:
                    urgency = "urgent"
                    urgency_label = "🔴 延期"
                elif is_today:
                    urgency = "high"
                    urgency_label = "🟠 今日截止"
                else:
                    urgency = "medium"
                    urgency_label = "🟡 即将到期"

                result_data["priority_tasks"].append({
                    "task_id": row[0],
                    "task_name": row[1],
                    "project_id": row[2],
                    "project_name": row[3],
                    "end_date": str(end_date) if end_date else None,
                    "status": row[5],
                    "progress": float(row[6] or 0),
                    "urgency": urgency,
                    "urgency_label": urgency_label,
                    "suggestion": _get_task_suggestion(urgency, row[6] or 0)
                })

            # 2. 延期预警（详细）
            result = conn.execute(text("""
                SELECT pt.task_id, pt.task_name, pt.project_id, p.name as project_name,
                       CURRENT_DATE - pt.end_date as delay_days,
                       pt.status, pt.progress
                FROM project_tasks pt
                JOIN projects p ON CAST(pt.project_id AS INTEGER) = p.id
                WHERE pt.is_deleted = false
                  AND p.is_deleted = false
                  AND pt.assignee_id = :emp_id
                  AND pt.end_date < CURRENT_DATE
                  AND pt.actual_end_date IS NULL
                ORDER BY delay_days DESC
                LIMIT 5
            """), {"emp_id": employee_id})

            for row in result:
                result_data["delayed_warnings"].append({
                    "task_id": row[0],
                    "task_name": row[1],
                    "project_name": row[3],
                    "delay_days": row[4],
                    "progress": float(row[6] or 0),
                    "suggestion": f"已延期{row[4]}天，建议立即处理或申请延期"
                })

            # 移除工时预测
            result_data["hours_prediction"] = {}

            # 4. 日报状态
            result = conn.execute(text("""
                SELECT id, status FROM daily_reports
                WHERE employee_id = :emp_id
                  AND is_deleted = false
                  AND report_date = :today
                LIMIT 1
            """), {"emp_id": employee_id, "today": today})

            report_row = result.fetchone()
            if report_row:
                result_data["daily_report_status"] = {
                    "submitted": True,
                    "report_id": report_row[0],
                    "status": report_row[1]
                }
            else:
                result_data["daily_report_status"] = {
                    "submitted": False,
                    "suggestion": "今日日报尚未填报"
                }

            # 5. 智能建议（综合）
            suggestions = []

            if result_data["delayed_warnings"]:
                suggestions.append({
                    "type": "delayed",
                    "priority": 1,
                    "message": f"您有 {len(result_data['delayed_warnings'])} 项任务延期，建议优先处理"
                })

            if not result_data["daily_report_status"].get("submitted"):
                suggestions.append({
                    "type": "report",
                    "priority": 2,
                    "message": "今日日报尚未填报，建议下午 5 点前完成"
                })

            # 移除工时预警建议

            if result_data["priority_tasks"]:
                urgent_count = sum(1 for t in result_data["priority_tasks"] if t["urgency"] == "urgent")
                if urgent_count > 0:
                    suggestions.append({
                        "type": "urgent",
                        "priority": 0,
                        "message": f"有 {urgent_count} 项紧急任务，建议立即处理"
                    })

            result_data["suggestions"] = sorted(suggestions, key=lambda x: x["priority"])

        return result_data

    except Exception as e:
        logger.error(f" {e}")
        import traceback
        traceback.print_exc()
        return {
            "priority_tasks": [],
            "delayed_warnings": [],
            "hours_prediction": {},
            "suggestions": [],
            "daily_report_status": {}
        }


def _get_task_suggestion(urgency: str, progress: float) -> str:
    """根据紧急程度和进度生成建议"""
    if urgency == "urgent":
        if progress < 30:
            return "建议立即启动，必要时申请资源支持"
        else:
            return "建议优先完成，如遇阻塞及时上报"
    elif urgency == "high":
        if progress < 50:
            return "建议上午完成，预留下午评审时间"
        else:
            return "继续保持，今日可完成"
    else:
        return "合理安排时间，按计划推进"


@app.get("/agent/api/agent/dashboard/smart-assistant")
async def get_smart_assistant(current_user: Dict = Depends(get_current_user)):
    """
    智能助手 - 整合今日优先任务 + 延期预警 + 工时预测 + 智能建议

    返回用户登录后应该看到的所有关键信息
    """
    # 直接从 current_user 获取 employee_id
    employee_id = current_user.get("username")

    try:
        # text 已从 database 模块导入
        from dotenv import load_dotenv
        load_dotenv()
        today = datetime.now().date()

        result_data = {
            "priority_tasks": [],      # 今日优先任务（排序后）
            "delayed_warnings": [],    # 延期预警
            "hours_prediction": {},    # 工时预测
            "suggestions": [],         # 智能建议
            "daily_report_status": {}  # 日报状态
        }

        with get_connection() as conn:
            # 1. 今日优先任务（按紧急程度排序）
            result = conn.execute(text("""
                SELECT pt.task_id, pt.task_name, pt.project_id, p.name as project_name,
                       pt.end_date, pt.status, pt.progress
                FROM project_tasks pt
                JOIN projects p ON CAST(pt.project_id AS INTEGER) = p.id
                WHERE pt.is_deleted = false
                  AND p.is_deleted = false
                  AND pt.assignee_id = :emp_id
                  AND pt.actual_end_date IS NULL
                  AND pt.end_date <= CURRENT_DATE + INTERVAL '3 days'
                ORDER BY
                    CASE WHEN pt.end_date < CURRENT_DATE THEN 0
                         WHEN pt.end_date = CURRENT_DATE THEN 1
                         ELSE 2 END,
                    pt.end_date
                LIMIT 5
            """), {"emp_id": employee_id})

            for row in result:
                end_date = row[4]
                is_delayed = end_date < today if end_date else False
                is_today = end_date == today if end_date else False

                # 计算紧急程度
                if is_delayed:
                    urgency = "urgent"
                    urgency_label = "🔴 延期"
                elif is_today:
                    urgency = "high"
                    urgency_label = "🟠 今日截止"
                else:
                    urgency = "medium"
                    urgency_label = "🟡 即将到期"

                result_data["priority_tasks"].append({
                    "task_id": row[0],
                    "task_name": row[1],
                    "project_id": row[2],
                    "project_name": row[3],
                    "end_date": str(end_date) if end_date else None,
                    "status": row[5],
                    "progress": float(row[6] or 0),
                    "urgency": urgency,
                    "urgency_label": urgency_label,
                    "suggestion": _get_task_suggestion(urgency, row[6] or 0)
                })

            # 2. 延期预警（详细）
            result = conn.execute(text("""
                SELECT pt.task_id, pt.task_name, pt.project_id, p.name as project_name,
                       CURRENT_DATE - pt.end_date as delay_days,
                       pt.status, pt.progress
                FROM project_tasks pt
                JOIN projects p ON CAST(pt.project_id AS INTEGER) = p.id
                WHERE pt.is_deleted = false
                  AND p.is_deleted = false
                  AND pt.assignee_id = :emp_id
                  AND pt.end_date < CURRENT_DATE
                  AND pt.actual_end_date IS NULL
                ORDER BY delay_days DESC
                LIMIT 5
            """), {"emp_id": employee_id})

            for row in result:
                result_data["delayed_warnings"].append({
                    "task_id": row[0],
                    "task_name": row[1],
                    "project_name": row[3],
                    "delay_days": row[4],
                    "progress": float(row[6] or 0),
                    "suggestion": f"已延期{row[4]}天，建议立即处理或申请延期"
                })

            # 移除工时预测
            result_data["hours_prediction"] = {}
            # 4. 日报状态
            result = conn.execute(text("""
                SELECT id, status FROM daily_reports
                WHERE employee_id = :emp_id
                  AND is_deleted = false
                  AND report_date = :today
                LIMIT 1
            """), {"emp_id": employee_id, "today": today})

            report_row = result.fetchone()
            if report_row:
                result_data["daily_report_status"] = {
                    "submitted": True,
                    "report_id": report_row[0],
                    "status": report_row[1]
                }
            else:
                result_data["daily_report_status"] = {
                    "submitted": False,
                    "suggestion": "今日日报尚未填报"
                }

            # 5. 智能建议（综合）
            suggestions = []

            if result_data["delayed_warnings"]:
                suggestions.append({
                    "type": "delayed",
                    "priority": 1,
                    "message": f"您有 {len(result_data['delayed_warnings'])} 项任务延期，建议优先处理"
                })

            if not result_data["daily_report_status"].get("submitted"):
                suggestions.append({
                    "type": "report",
                    "priority": 2,
                    "message": "今日日报尚未填报，建议下午 5 点前完成"
                })

            # 移除工时预警
            # if is_warning:
            #     suggestions.append({
            #         "type": "hours",
            #         "priority": 3,
            #         "message": f"本月工时预计 {int(predicted_hours)}h，接近预警线"
            #     })

            if result_data["priority_tasks"]:
                urgent_count = sum(1 for t in result_data["priority_tasks"] if t["urgency"] == "urgent")
                if urgent_count > 0:
                    suggestions.append({
                        "type": "urgent",
                        "priority": 0,
                        "message": f"有 {urgent_count} 项紧急任务，建议立即处理"
                    })

            # 6. 进度和成本预警
            # 查询用户负责的项目进度和成本情况
            result = conn.execute(text("""
                SELECT p.id, p.name, p.progress, p.start_date, p.end_date,
                       p.budget_total_cost, p.actual_total_cost
                FROM projects p
                WHERE p.is_deleted = false
                  AND p.leader_id = :emp_id
                  AND p.status = '进行中'
            """), {"emp_id": employee_id})

            for row in result:
                pid = row[0]
                pname = row[1]
                pprogress = float(row[2] or 0)
                pstart = row[3]
                pend = row[4]
                pbudget = float(row[5] or 0)
                pactual = float(row[6] or 0)

                # 计算计划进度
                if pstart and pend:
                    from datetime import datetime as dt_cls
                    start_dt = pstart if isinstance(pstart, date) else dt_cls.strptime(str(pstart), '%Y-%m-%d').date()
                    end_dt = pend if isinstance(pend, date) else dt_cls.strptime(str(pend), '%Y-%m-%d').date()
                    total_d = (end_dt - start_dt).days
                    elapsed_d = (today - start_dt).days
                    planned_pct = round(elapsed_d / total_d * 100, 1) if total_d > 0 else 0
                else:
                    planned_pct = 0

                # 进度滞后预警
                if pprogress < planned_pct - 10:
                    lag_amt = planned_pct - pprogress
                    suggestions.append({
                        "type": "progress",
                        "priority": 2,
                        "message": f"【{pname}】进度滞后 {lag_amt:.1f}%，建议加快推进"
                    })

                # 成本超支预警
                if pbudget > 0 and pactual > pbudget:
                    over_pct = (pactual - pbudget) / pbudget * 100
                    suggestions.append({
                        "type": "cost",
                        "priority": 2,
                        "message": f"【{pname}】成本超支 {over_pct:.1f}%，请注意控制"
                    })

            result_data["suggestions"] = sorted(suggestions, key=lambda x: x["priority"])

        return result_data

    except Exception as e:
        logger.error(f" {e}")
        import traceback
        traceback.print_exc()
        return {
            "priority_tasks": [],
            "delayed_warnings": [],
            "hours_prediction": {},
            "suggestions": [],
            "daily_report_status": {}
        }


@app.get("/agent/api/agent/stats/hours-trend")
async def get_hours_trend(
    time_range: str = "week",
    current_user: Dict = Depends(get_current_user)
):
    """
    获取工时趋势数据

    Args:
        time_range: week 或 month
    """
    # 直接从 current_user 获取 employee_id
    employee_id = current_user.get("username")  # username 就是 employee_id
    logger.debug(f" hours_trend: employee_id={employee_id}, time_range={time_range}")

    try:
        # text 已从 database 模块导入
        from dotenv import load_dotenv
        load_dotenv()
        today = datetime.now().date()

        if time_range == "week":
            days = 7
        else:
            days = 30

        start_date = today - timedelta(days=days-1)
        logger.debug(f" today={today}, start_date={start_date}")

        with get_connection() as conn:
            result = conn.execute(text("""
                SELECT dr.report_date, SUM(dwi.hours_spent) as hours
                FROM daily_reports dr
                JOIN daily_work_items dwi ON dr.id = dwi.report_id
                WHERE dr.employee_id = :emp_id
                  AND dr.is_deleted = false
                  AND dr.report_date >= :start
                  AND dr.report_date <= :today
                GROUP BY dr.report_date
                ORDER BY dr.report_date
            """), {"emp_id": employee_id, "start": start_date, "today": today})

            # 构建日期序列
            dates = []
            actual = []
            data_map = {}

            for row in result:
                data_map[str(row[0])] = float(row[1] or 0)
                logger.debug(f" Found data: {row[0]} -> {row[1]}")

            logger.debug(f" data_map: {data_map}")

            for i in range(days):
                d = start_date + timedelta(days=i)
                date_str = d.strftime("%Y-%m-%d")
                dates.append(d.strftime("%m-%d"))
                actual.append(data_map.get(date_str, 0))

            # 预测线（简单移动平均）
            predicted = []
            window = 3
            for i in range(len(actual)):
                if i < window:
                    predicted.append(actual[i])
                else:
                    avg = sum(actual[i-window:i]) / window
                    predicted.append(round(avg, 1))

            return {
                "dates": dates,
                "actual": actual,
                "predicted": predicted
            }

    except Exception as e:
        logger.error(f" {e}")
        import traceback
        traceback.print_exc()
        return {"dates": [], "actual": [], "predicted": []}


@app.get("/agent/api/agent/stats/project-distribution")
async def get_project_distribution(current_user: Dict = Depends(get_current_user)):
    """
    获取项目工时分布（饼图）
    """
    username = current_user.get("username")
    user_info = get_user_info_cache(username)
    employee_id = user_info.get("employee_id") if user_info else username

    try:
        # text 已从 database 模块导入
        from dotenv import load_dotenv
        load_dotenv()
        today = datetime.now().date()
        month_start = today.replace(day=1)

        with get_connection() as conn:
            result = conn.execute(text("""
                SELECT COALESCE(p.name, '其他') as name, SUM(dwi.hours_spent) as value
                FROM daily_work_items dwi
                JOIN daily_reports dr ON dwi.report_id = dr.id
                LEFT JOIN projects p ON CAST(dwi.project_id AS INTEGER) = p.id
                WHERE dr.employee_id = :emp_id
                  AND dr.is_deleted = false
                  AND dr.report_date >= :start
                  AND dr.report_date <= :today
                GROUP BY p.name
                ORDER BY value DESC
                LIMIT 10
            """), {"emp_id": employee_id, "start": month_start, "today": today})

            distribution = []
            for row in result:
                distribution.append({
                    "name": row[0],
                    "value": float(row[1] or 0)
                })

            return distribution

    except Exception as e:
        logger.error(f" {e}")
        return []


# ============== 项目风险雷达 API ==============

@app.get("/agent/api/agent/projects/{project_id}/risk-radar")
async def get_project_risk_radar(
    project_id: int,
    current_user: Dict = Depends(get_current_user)
):
    """
    获取项目风险雷达数据

    返回五个维度的风险评分（0-100，分数越高风险越大）：
    - schedule_risk: 进度风险（延期任务比例）
    - material_risk: 材料成本风险（材料成本超支率）
    - outsourcing_risk: 外包成本风险（外包成本超支率）
    - labor_risk: 人工成本风险（人工成本超支率）
    - indirect_risk: 间接成本风险（间接成本超支率）
    """
    # text 已从 database 模块导入
    from dotenv import load_dotenv
    load_dotenv()
    today = datetime.now().date()

    with get_connection() as conn:
        # 1. 进度风险：延期任务比例
        result = conn.execute(text("""
            SELECT
                COUNT(*) as total_tasks,
                SUM(CASE WHEN pt.end_date < :today AND pt.actual_end_date IS NULL THEN 1 ELSE 0 END) as delayed_tasks,
                SUM(CASE WHEN pt.actual_end_date IS NOT NULL THEN 1 ELSE 0 END) as completed_tasks
            FROM project_tasks pt
            WHERE pt.project_id = :project_id
              AND pt.is_deleted = false
        """), {"project_id": str(project_id), "today": today})

        task_row = result.fetchone()
        total_tasks = task_row[0] or 0
        delayed_tasks = task_row[1] or 0
        completed_tasks = task_row[2] or 0

        # 进度风险评分：延期率 × 100
        schedule_risk = round((delayed_tasks / total_tasks * 100) if total_tasks > 0 else 0, 1)

        # 2. 四大成本风险：从项目表获取
        result = conn.execute(text("""
            SELECT
                material_budget, material_cost,
                outsourcing_budget, outsourcing_cost,
                labor_budget, labor_cost,
                indirect_budget, indirect_cost,
                p.end_date, p.start_date
            FROM projects p
            WHERE p.id = :project_id
        """), {"project_id": project_id})

        cost_row = result.fetchone()

        if cost_row:
            # 材料成本风险
            material_budget = float(cost_row[0] or 0)
            material_cost = float(cost_row[1] or 0)
            if material_budget > 0:
                material_risk = max(0, round((material_cost - material_budget) / material_budget * 100, 1))
            else:
                material_risk = 0

            # 外包成本风险
            outsourcing_budget = float(cost_row[2] or 0)
            outsourcing_cost = float(cost_row[3] or 0)
            if outsourcing_budget > 0:
                outsourcing_risk = max(0, round((outsourcing_cost - outsourcing_budget) / outsourcing_budget * 100, 1))
            else:
                outsourcing_risk = 0

            # 人工成本风险
            labor_budget = float(cost_row[4] or 0)
            labor_cost = float(cost_row[5] or 0)
            if labor_budget > 0:
                labor_risk = max(0, round((labor_cost - labor_budget) / labor_budget * 100, 1))
            else:
                labor_risk = 0

            # 间接成本风险
            indirect_budget = float(cost_row[6] or 0)
            indirect_cost = float(cost_row[7] or 0)
            if indirect_budget > 0:
                indirect_risk = max(0, round((indirect_cost - indirect_budget) / indirect_budget * 100, 1))
            else:
                indirect_risk = 0

            project_end_date = cost_row[8]
        else:
            material_budget = material_cost = 0
            outsourcing_budget = outsourcing_cost = 0
            labor_budget = labor_cost = 0
            indirect_budget = indirect_cost = 0
            material_risk = outsourcing_risk = labor_risk = indirect_risk = 0
            project_end_date = None

        # 计算剩余任务数
        result = conn.execute(text("""
            SELECT COUNT(pt.task_id) as remaining_tasks
            FROM project_tasks pt
            WHERE pt.project_id = :project_id
              AND pt.actual_end_date IS NULL
              AND pt.is_deleted = false
        """), {"project_id": str(project_id)})

        remaining_tasks = result.fetchone()[0] or 0

        # 计算剩余天数
        if project_end_date:
            days_remaining = (project_end_date - today).days
        else:
            days_remaining = None

        # 综合风险评分：进度40% + 四大成本平均60%
        cost_avg_risk = (material_risk + outsourcing_risk + labor_risk + indirect_risk) / 4
        overall_risk = round(schedule_risk * 0.4 + cost_avg_risk * 0.6, 1)

        # 风险等级
        if overall_risk >= 70:
            risk_level = "high"
            risk_label = "高风险"
        elif overall_risk >= 40:
            risk_level = "medium"
            risk_label = "中风险"
        else:
            risk_level = "low"
            risk_label = "低风险"

        return {
            "project_id": project_id,
            "radar": {
                "schedule_risk": schedule_risk,
                "material_risk": material_risk,
                "outsourcing_risk": outsourcing_risk,
                "labor_risk": labor_risk,
                "indirect_risk": indirect_risk
            },
            "overall_risk": overall_risk,
            "risk_level": risk_level,
            "risk_label": risk_label,
            "details": {
                "total_tasks": total_tasks,
                "delayed_tasks": delayed_tasks,
                "completed_tasks": completed_tasks,
                "days_remaining": days_remaining,
                "remaining_tasks": remaining_tasks,
                "cost_details": {
                    "material": {"budget": material_budget, "actual": material_cost},
                    "outsourcing": {"budget": outsourcing_budget, "actual": outsourcing_cost},
                    "labor": {"budget": labor_budget, "actual": labor_cost},
                    "indirect": {"budget": indirect_budget, "actual": indirect_cost}
                }
            }
        }


@app.get("/agent/api/agent/projects/{project_id}/task-risks")
async def get_project_task_risks(
    project_id: int,
    current_user: Dict = Depends(get_current_user)
):
    """
    获取项目任务风险预警

    检查项：
    - 延期风险：计划结束时间已过，进度 < 100%
    - 即将到期风险：3天内到期，进度 < 80%
    - 未报告风险：已启动但无日报记录
    - 即将启动提醒：3天内开始
    """
    # 导入风险检查函数
    import sys
    sys.path.insert(0, os.path.dirname(__file__))
    from task_auto import check_task_risks
    risks = check_task_risks(project_id)

    # 按风险等级排序
    risk_order = {"high": 0, "medium": 1, "low": 2}
    risks.sort(key=lambda x: risk_order.get(x["risk_level"], 3))

    return {
        "project_id": project_id,
        "risks": risks,
        "risk_count": len(risks),
        "high_risk_count": sum(1 for r in risks if r["risk_level"] == "high"),
        "medium_risk_count": sum(1 for r in risks if r["risk_level"] == "medium"),
        "low_risk_count": sum(1 for r in risks if r["risk_level"] == "low")
    }


@app.post("/agent/api/agent/projects/{project_id}/update-task-status")
async def update_project_task_status(
    project_id: int,
    current_user: Dict = Depends(get_current_user)
):
    """
    更新项目所有任务状态（仅项目负责人或管理员可操作）

    根据进度和时间自动计算任务状态：
    - 未开始：计划开始时间未到
    - 进行中：在计划周期内
    - 延期：计划结束时间已过，进度 < 100%
    - 已完成：进度 >= 100%
    """
    # 权限检查：只有项目负责人或管理员可以更新
    if not await check_project_edit_permission(project_id, current_user):
        raise HTTPException(status_code=403, detail="只有项目负责人或管理员可以更新任务状态")

    try:
        from .task_auto import get_latest_version_tasks, calculate_task_status
        # text 已从 database 模块导入
        from dotenv import load_dotenv
        load_dotenv()
        tasks = get_latest_version_tasks(project_id)
        updated_tasks = []

        with get_connection() as conn:
            for task in tasks:
                new_status, changed = calculate_task_status(task)

                if changed:
                    # 更新状态
                    update_fields = ["status = :status", "update_time = CURRENT_TIMESTAMP"]
                    params = {"tid": task["task_id"], "status": new_status}

                    # 如果完成，设置实际完成时间
                    if new_status == "已完成" and not task["actual_end_date"]:
                        update_fields.append("actual_end_date = CURRENT_DATE")

                    # 安全构建SET子句（白名单验证字段名）
                    allowed_fields = {"status", "progress", "actual_end_date", "updated_at"}
                    set_clause = ", ".join(f"{field} = :{field}" for field in update_fields if field.split("=")[0].strip() in allowed_fields)
                    if set_clause:
                        conn.execute(text(f"""
                            UPDATE project_tasks
                            SET {set_clause}
                            WHERE task_id = :tid
                        """), params)

                    updated_tasks.append({
                        "task_id": task["task_id"],
                        "task_name": task["task_name"],
                        "old_status": task["status"],
                        "new_status": new_status
                    })

            conn.commit()

        return {
            "project_id": project_id,
            "updated_count": len(updated_tasks),
            "updated_tasks": updated_tasks
        }

    except Exception as e:
        logger.error(f" {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"更新失败: {str(e)}")


@app.get("/agent/api/agent/projects", response_model=List[ProjectInfo])
async def get_projects(current_user: Dict = Depends(get_current_user)):
    """
    获取项目列表（需要认证，组织隔离）

    权限规则：
    - 系统管理员(role_id=11)：看全部项目
    - 财务(role_id=15)：看研究院所有项目
    - 看板(role_id=17)：看研究院所有项目
    - 院领导(role_id=16)：看本组织及下级项目
    - 普通用户：看本组织及下级项目 + 自己负责的项目
    """
    username = current_user.get("username")
    
    from dotenv import load_dotenv
    load_dotenv()
    
    with get_connection() as conn:
        # 查询用户信息（含组织）
        user_result = conn.execute(text("""
            SELECT p.employee_id, p.name, p.department, p.role_id, p.org_id
            FROM personnel p
            WHERE p.employee_id = :username AND p.is_deleted = false
        """), {"username": username}).fetchone()
        
        if not user_result:
            raise HTTPException(status_code=401, detail="用户不存在")
        
        employee_id, employee_name, department, role_id, org_id = user_result
        
        logger.debug(f"用户 {employee_name}(role={role_id}, org={org_id}) 查询项目")
        
        # 构建权限过滤条件
        if role_id == 11:  # 系统管理员：看全部
            result = conn.execute(text("""
                SELECT id, name, leader, status, project_year FROM projects
                WHERE is_deleted = false ORDER BY project_year DESC, id
            """))
        elif role_id in [15, 17]:  # 财务、看板：看研究院所有项目
            result = conn.execute(text("""
                SELECT id, name, leader, status, project_year FROM projects
                WHERE is_deleted = false 
                  AND org_id IN (SELECT id FROM organizations WHERE id = 2 OR parent_id = 2)
                ORDER BY project_year DESC, id
            """))
        elif role_id == 16:  # 院领导：看本组织及下级
            result = conn.execute(text("""
                SELECT id, name, leader, status, project_year FROM projects
                WHERE is_deleted = false 
                  AND org_id IN (SELECT id FROM organizations WHERE id = :org_id OR parent_id = :org_id)
                ORDER BY project_year DESC, id
            """), {"org_id": org_id})
        else:  # 普通用户：看自己负责的 + 自己参与的（有任务或填过日报）
            result = conn.execute(text("""
                SELECT DISTINCT id, name, leader, status, project_year FROM projects p
                WHERE p.is_deleted = false 
                  AND (
                    p.leader = :name
                    OR p.id IN (
                      SELECT DISTINCT CAST(pt.project_id AS INTEGER) 
                      FROM project_tasks pt 
                      WHERE pt.is_deleted = false 
                        AND pt.assignee_id = :emp_id
                        AND pt.project_id ~ '^[0-9]+$'
                    )
                    OR CAST(p.id AS VARCHAR) IN (
                      SELECT DISTINCT dwi.project_id 
                      FROM daily_work_items dwi
                      JOIN daily_reports dr ON dr.id = dwi.report_id
                      WHERE dr.is_deleted = false 
                        AND dr.employee_id = :emp_id
                        AND dwi.project_id IS NOT NULL
                        AND dwi.project_id ~ '^[0-9]+$'
                    )
                  )
                ORDER BY p.project_year DESC, p.id
            """), {"name": employee_name, "emp_id": employee_id})
        
        # 计算每个项目的进度（工期加权，与详情页和看板统一）
        projects = []
        for row in result:
            project_id = row[0]
            try:
                # 只统计叶子任务，使用工期加权计算
                task_stats = conn.execute(text("""
                    SELECT
                        COUNT(*) as total_tasks,
                        SUM(CASE WHEN progress >= 100 THEN 1 ELSE 0 END) as completed_tasks,
                        -- 总工期天数
                        SUM(CASE WHEN end_date IS NOT NULL AND start_date IS NOT NULL 
                            THEN end_date - start_date + 1 ELSE 5 END) as total_work_days,
                        -- 已完成工期天数
                        SUM(CASE 
                            WHEN progress >= 100 AND end_date IS NOT NULL AND start_date IS NOT NULL 
                            THEN end_date - start_date + 1 
                            WHEN end_date IS NOT NULL AND start_date IS NOT NULL AND end_date < CURRENT_DATE
                            THEN (end_date - start_date + 1) * LEAST(progress / 100.0, 0.5)
                            WHEN end_date IS NOT NULL AND start_date IS NOT NULL
                            THEN (end_date - start_date + 1) * progress / 100.0
                            ELSE 0 
                        END) as completed_work_days
                    FROM project_tasks
                    WHERE project_id = CAST(:pid AS VARCHAR)
                      AND is_deleted = false
                      AND is_latest = true
                      AND ("isNode" = false OR "isNode" IS NULL)
                """), {"pid": project_id})
                ts = task_stats.fetchone()
                
                total_tasks = int(ts[0] or 0)
                completed_tasks = int(ts[1] or 0)
                total_work_days = float(ts[2] or 0)
                completed_work_days = float(ts[3] or 0)
                
                # 进度 = 已完成工期天数 / 总工期天数
                if total_work_days > 0:
                    progress = round(completed_work_days / total_work_days * 100, 1)
                else:
                    progress = 0
                
                projects.append({
                    "id": row[0],
                    "name": row[1],
                    "leader": row[2],
                    "status": row[3],
                    "progress": progress,
                    "project_year": row[4] if len(row) > 4 else None
                })
            except Exception as e:
                logger.warning(f"计算项目进度失败: {e}")
                projects.append({
                    "id": row[0],
                    "name": row[1],
                    "leader": row[2],
                    "status": row[3],
                    "progress": 0,
                    "project_year": row[4] if len(row) > 4 else None
                })
        
        logger.debug(f"返回项目数: {len(projects)}")
        return [ProjectInfo(**p) for p in projects]


@app.get("/agent/api/agent/projects/{project_id}")
async def get_project_detail(
    project_id: int,
    request: Request,
    current_user: Dict = Depends(get_current_user)
):
    """
    获取项目详情（包含工时统计、成本数据、进度计算）
    权限检查：用户只能访问有权限的项目
    """
    username = current_user.get("username")
    
    # 从请求 header 获取 token（优先），避免依赖内存缓存
    token = get_token_from_request(request, username)

    if not token:
        raise HTTPException(status_code=401, detail="未找到用户认证信息")

    try:
        from dotenv import load_dotenv
        load_dotenv()
        
        with get_connection() as conn:
            # 1. 查询用户信息
            user_result = conn.execute(text("""
                SELECT p.employee_id, p.name, p.role_id, p.org_id
                FROM personnel p
                WHERE p.employee_id = :username AND p.is_deleted = false
            """), {"username": username}).fetchone()
            
            if not user_result:
                raise HTTPException(status_code=401, detail="用户不存在")
            
            _, employee_name, role_id, user_org_id = user_result
            
            # 2. 查询项目所属组织
            project_org_result = conn.execute(text("""
                SELECT org_id, leader FROM projects WHERE id = :pid AND is_deleted = false
            """), {"pid": project_id}).fetchone()
            
            if not project_org_result:
                raise HTTPException(status_code=404, detail="项目不存在")
            
            project_org_id, project_leader = project_org_result
            
            # 3. 权限检查
            has_access = False
            if role_id == 11:  # 系统管理员
                has_access = True
            elif role_id in [15, 17]:  # 财务、看板：看研究院项目
                has_access = project_org_id in [2, 5, 6, 7, 8, 9]
            elif role_id == 16 and user_org_id:  # 院领导
                org_ids = [user_org_id] + [r[0] for r in conn.execute(text(
                    "SELECT id FROM organizations WHERE parent_id = :oid"
                ), {"oid": user_org_id}).fetchall()]
                has_access = project_org_id in org_ids
            elif user_org_id:  # 普通用户
                org_ids = [user_org_id] + [r[0] for r in conn.execute(text(
                    "SELECT id FROM organizations WHERE parent_id = :oid"
                ), {"oid": user_org_id}).fetchall()]
                has_access = project_org_id in org_ids or project_leader == employee_name
            
            if not has_access:
                raise HTTPException(status_code=403, detail="无权访问该项目")
            
            # 4. 项目基本信息
            project_result = conn.execute(text("""
                SELECT id, name, leader, status,
                       start_date, end_date,
                       budget_total_cost, contract_amount,
                       material_budget, material_cost,
                       outsourcing_budget, outsourcing_cost,
                       labor_budget, labor_cost,
                       indirect_budget, indirect_cost,
                       project_category, project_subject,
                       implementation_mode, project_level
                FROM projects
                WHERE id = :pid
            """), {"pid": project_id})

            project_row = project_result.fetchone()
            if not project_row:
                raise HTTPException(status_code=404, detail="项目不存在")

            # 获取最新版本叶子任务的时间范围和进度（排除分组父节点）
            task_stats = conn.execute(text("""
                WITH latest_version AS (
                    SELECT MAX(CAST(SUBSTRING(task_id FROM 'V([0-9]+)') AS INTEGER)) as max_ver
                    FROM project_tasks
                    WHERE project_id::integer = :pid AND is_deleted = false
                )
                SELECT
                    MIN(start_date) as plan_start,
                    MAX(end_date) as plan_end,
                    COUNT(*) as total_tasks,
                    SUM(CASE WHEN progress >= 100 THEN 1 ELSE 0 END) as completed_tasks,
                    AVG(progress) as avg_progress,
                    -- 计算总工期天数
                    SUM(CASE WHEN end_date IS NOT NULL AND start_date IS NOT NULL 
                        THEN end_date - start_date + 1 ELSE 5 END) as total_work_days,
                    -- 计算已完成工期天数
                    SUM(CASE 
                        WHEN progress >= 100 AND end_date IS NOT NULL AND start_date IS NOT NULL 
                        THEN end_date - start_date + 1 
                        WHEN end_date IS NOT NULL AND start_date IS NOT NULL AND end_date < CURRENT_DATE
                        THEN (end_date - start_date + 1) * LEAST(progress / 100.0, 0.5)
                        WHEN end_date IS NOT NULL AND start_date IS NOT NULL
                        THEN (end_date - start_date + 1) * progress / 100.0
                        ELSE 0 
                    END) as completed_work_days
                FROM project_tasks pt, latest_version lv
                WHERE pt.project_id::integer = :pid
                  AND pt.is_deleted = false
                  AND (pt."isNode" = false OR pt."isNode" IS NULL)
                  AND COALESCE(CAST(SUBSTRING(pt.task_id FROM 'V([0-9]+)') AS INTEGER), 0) = COALESCE(lv.max_ver, 0)
            """), {"pid": project_id})

            task_row = task_stats.fetchone()

            # 计算项目进度（按工期加权，与看板统一）
            total_tasks = task_row[2] or 0
            completed_tasks = task_row[3] or 0
            total_work_days = float(task_row[5] or 0)
            completed_work_days = float(task_row[6] or 0)
            
            if total_work_days > 0:
                project_progress = completed_work_days / total_work_days * 100
            else:
                project_progress = 0

            # 项目总工时
            hours_result = conn.execute(text("""
                SELECT COALESCE(SUM(hours_spent), 0) as hours
                FROM daily_work_items
                WHERE project_id = :pid
            """), {"pid": str(project_id)})
            total_hours = float(hours_result.fetchone()[0] or 0)

            # 计算人力成本（基于费率表）
            labor_cost_result = conn.execute(text("""
                SELECT COALESCE(SUM(
                    dwi.hours_spent * COALESCE(pr.hourly_rate, 0)
                ), 0) as labor_cost
                FROM daily_work_items dwi
                JOIN daily_reports dr ON dwi.report_id = dr.id
                LEFT JOIN personnel p ON dr.employee_id = p.employee_id
                LEFT JOIN personnel_rates pr ON p.id = pr.personnel_id
                    AND pr.year = TO_CHAR(dr.report_date, 'YYYY')
                    AND pr.month = TO_CHAR(dr.report_date, 'MM')
                    AND pr.is_deleted = false
                WHERE dwi.project_id = :pid
            """), {"pid": str(project_id)})

            calculated_labor_cost = float(labor_cost_result.fetchone()[0] or 0)

            # 如果计算的人力成本大于数据库中的，使用计算的值
            labor_cost_from_db = float(project_row[13] or 0)
            labor_cost = calculated_labor_cost if calculated_labor_cost > labor_cost_from_db else labor_cost_from_db

            # 各人员工时
            worker_result = conn.execute(text("""
                SELECT
                    dr.employee_name,
                    SUM(dwi.hours_spent) as hours
                FROM daily_work_items dwi
                JOIN daily_reports dr ON dwi.report_id = dr.id
                WHERE dwi.project_id = :pid
                GROUP BY dr.employee_name
                ORDER BY hours DESC
                LIMIT 5
            """), {"pid": str(project_id)})

            worker_hours = []
            for row in worker_result:
                worker_hours.append({
                    "name": row[0],
                    "hours": float(row[1] or 0)
                })

        return {
            "id": project_row[0],
            "name": project_row[1],
            "leader": project_row[2],
            "status": project_row[3],
            "start_date": str(project_row[4]) if project_row[4] else None,
            "end_date": str(project_row[5]) if project_row[5] else None,
            "budget": float(project_row[6] or 0),
            "contract_amount": float(project_row[7] or 0),
            # 成本数据
            "material_budget": float(project_row[8] or 0),
            "material_cost": float(project_row[9] or 0),
            "outsourcing_budget": float(project_row[10] or 0),
            "outsourcing_cost": float(project_row[11] or 0),
            "labor_budget": float(project_row[12] or 0),
            "labor_cost": labor_cost,  # 使用计算的人力成本
            "indirect_budget": float(project_row[14] or 0),
            "indirect_cost": float(project_row[15] or 0),
            # 分类信息
            "project_category": project_row[16],
            "project_subject": project_row[17],
            "implementation_mode": project_row[18],
            "project_level": project_row[19],
            # 计算数据
            "plan_start_date": str(task_row[0]) if task_row[0] else None,
            "plan_end_date": str(task_row[1]) if task_row[1] else None,
            "progress": round(project_progress, 1),
            "progress_formula": f"工期加权: {completed_work_days:.0f}/{total_work_days:.0f}天 = {round(project_progress, 1)}%",
            "total_tasks": total_tasks,
            "completed_tasks": completed_tasks,
            "total_hours": round(total_hours, 1),
            "worker_hours": worker_hours,
            "description": None
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f" {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"获取项目详情失败: {str(e)}")


@app.get("/agent/api/agent/projects/{project_id}/tasks")
async def get_project_tasks(
    project_id: int,
    current_user: Dict = Depends(get_current_user)
):
    """
    获取项目任务列表（从本地数据库，只返回最新版本）
    版本规则：task_id 包含 V{版本号}，返回最大版本号的任务
    """
    # text 已从 database 模块导入
    from dotenv import load_dotenv
    load_dotenv()
    with get_connection() as conn:
        # 获取最新版本号
        version_result = conn.execute(text("""
            SELECT MAX(CAST(SUBSTRING(task_id FROM 'V([0-9]+)') AS INTEGER)) as max_version
            FROM project_tasks
            WHERE project_id::integer = :pid
              AND is_deleted = false
        """), {"pid": project_id})

        max_version_row = version_result.fetchone()
        max_version = max_version_row[0] if max_version_row and max_version_row[0] else 1

        # 获取指定版本的任务
        result = conn.execute(text("""
            SELECT pt.task_id, pt.task_name, pt.assignee, pt.start_date, pt.end_date,
                   pt.status, pt.progress, pt.planned_hours,
                   pt.parent_task_id, pt.task_level, pt.actual_end_date,
                   pt."isNode", pt.leaf_node,
                   COALESCE(
                       json_agg(
                           json_build_object(
                               'report_date', dr.report_date,
                               'work_content', dwi.work_content,
                               'hours_spent', dwi.hours_spent
                           )
                           ORDER BY dr.report_date
                       ) FILTER (WHERE dwi.id IS NOT NULL),
                       '[]'::json
                   ) as daily_reports
            FROM project_tasks pt
            LEFT JOIN daily_work_items dwi ON dwi.task_id = pt.task_id
            LEFT JOIN daily_reports dr ON dr.id = dwi.report_id
            WHERE pt.project_id::integer = :pid
              AND pt.is_deleted = false
              AND CAST(SUBSTRING(pt.task_id FROM 'V([0-9]+)') AS INTEGER) = :max_version
            GROUP BY pt.task_id, pt.task_name, pt.assignee, pt.start_date, pt.end_date,
                     pt.status, pt.progress, pt.planned_hours,
                     pt.parent_task_id, pt.task_level, pt.actual_end_date,
                     pt."isNode", pt.leaf_node
            ORDER BY CAST(SUBSTRING(pt.task_id FROM 'T([0-9]+)$') AS INTEGER)
        """), {"pid": project_id, "max_version": max_version})

        tasks = []
        for row in result:
            tasks.append({
                "task_id": row[0],
                "task_name": row[1],
                "assignee": row[2],
                "start_date": str(row[3]) if row[3] else None,
                "end_date": str(row[4]) if row[4] else None,
                "status": row[5] or "未开始",
                "progress": float(row[6] or 0),
                "planned_hours": float(row[7] or 0),
                "parent_task_id": row[8],
                "task_level": row[9],
                "actual_end_date": str(row[10]) if row[10] else None,
                "is_node": row[11],
                "leaf_node": row[12],
                "daily_reports": row[13] if row[13] else []
            })

        return tasks

@app.post("/agent/api/agent/daily/parse", response_model=ParseDailyResponse)
async def parse_daily(
    request: ParseDailyRequest,
    current_user: Dict = Depends(get_current_user)
):
    """
    智能解析日报文本（需要认证）

    - 提取时间、地点、内容
    - 自动匹配项目
    - 推荐关联任务
    """
    try:
        username = current_user.get("username")
        token = current_user.get("_raw_token") or get_user_token(username)

        # 获取项目列表（用于匹配）
        if token:
            projects = await get_projects_with_auth(token)
        else:
            projects = await get_cached_projects()

        logger.debug(f"获取到 {len(projects)} 个项目用于匹配")

        # 执行工作流
        result = await daily_agent.ainvoke({
            "text": request.text,
            "user_id": username or request.user_id,
            "projects": projects
        })

        entries = [DailyEntry(**e) for e in result.get("parsed_entries", [])]

        return ParseDailyResponse(
            entries=entries,
            confidence=result.get("confidence", 0),
            issues=result.get("issues", [])
        )
    except Exception as e:
        logger.exception(f" {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

class SubmitDailyRequest(BaseModel):
    """提交日报请求"""
    date: str  # YYYY-MM-DD格式
    entries: List[DailyEntry]
    employee_id: str
    employee_name: str


@app.post("/agent/api/agent/daily/submit")
async def submit_daily(
    request: SubmitDailyRequest,
    current_user: Dict = Depends(get_current_user)
):
    """
    提交日报到现有后端

    使用当前用户的token，调用现有后端API创建日报
    """
    try:
        username = current_user.get("username")
        token = current_user.get("_raw_token") or get_user_token(username)

        if not token:
            raise HTTPException(status_code=401, detail="未找到用户认证信息")

        # 获取员工信息
        # 简化：使用username作为employee_id，实际应该查数据库
        employee_id = request.employee_id or username
        employee_name = request.employee_name or username

        # 构建工作事项列表
        work_items = []
        total_hours = 0

        for entry in request.entries:
            hours = entry.hours or 0
            total_hours += hours

            work_items.append({
                "work_content": entry.content,
                "project_id": str(entry.matched_project_id) if entry.matched_project_id else None,
                "project_name": entry.matched_project_name or entry.project_hint,
                "task_id": entry.matched_task_id,
                "task_name": entry.matched_task_name,
                "start_time": entry.start_time,
                "end_time": entry.end_time,
                "hours_spent": hours,
                "progress_status": "正常",
                "progress_percentage": 0
            })

        # 构建日报数据
        daily_report_data = {
            "report": {
                "report_date": request.date,
                "employee_id": employee_id,
                "employee_name": employee_name,
                "work_target": "完成日常工作",
                "planned_hours": total_hours,
                "key_work_tracking": None,
                "tomorrow_plan": None
            },
            "work_items": work_items
        }

        logger.debug(f"提交日报数据: {json.dumps(daily_report_data, ensure_ascii=False)}")

        # 调用现有后端API (/api/v1/daily-report 而非 /api/v1/daily)
        response = await http_client.post(
            f"{settings.BACKEND_API_URL}/api/v1/daily-report/my-reports/with-items",
            json=daily_report_data,
            headers={"Authorization": f"Bearer {token}"},
            timeout=10.0
        )

        if response.status_code == 200:
            result = response.json()
            # 后端直接返回日报对象，不是嵌套在 data 中
            report_id = result.get("id")
            return {
                "success": True,
                "message": "日报提交成功",
                "report_id": report_id
            }
        else:
            logger.error(f" {response.status_code} - {response.text}")
            raise HTTPException(
                status_code=response.status_code,
                detail=f"提交失败: {response.text}"
            )

    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f" {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"提交失败: {str(e)}")

# ============== 计划版本管理 API ==============

@app.post("/agent/api/agent/plans/upload/{project_id}")
@limiter.limit("20/hour")  # 上传限流：每小时最多20次
async def upload_plan_excel(
    request: Request,
    project_id: int,
    file: UploadFile = File(...),
    version_name: Optional[str] = None,
    description: Optional[str] = None,
    change_type: Optional[str] = Form(None),
    change_reason: Optional[str] = Form(None),
    current_user: Dict = Depends(get_current_user)
):
    """
    上传Excel计划并解析导入（仅项目负责人或管理员可操作）

    Excel格式要求：
    - 第一行为表头
    - 必须包含列：任务名称
    - 可选列：负责人、开始日期、结束日期、工时、状态、备注
    
    变更追踪参数：
    - change_type: 变更类型（初始计划/目标调整/路径调整/偏差纠正/资源调整/其他）
    - change_reason: 变更原因说明
    """
    username = current_user.get("username")
    
    # 调试日志：打印 current_user 结构
    logger.debug(f"upload_plan_excel current_user: {current_user}")

    # 权限检查：只有项目负责人或管理员可以上传（先检查权限，再获取token）
    if not await check_project_edit_permission(project_id, current_user):
        raise HTTPException(status_code=403, detail="只有项目负责人或管理员可以上传计划")
    
    # 记录调整前状态快照（如果有历史版本）
    previous_status = None
    with get_connection() as conn:
        # 获取当前最新版本的状态
        result = conn.execute(text("""
            SELECT pv.id, pv.version_number, pv.upload_time
            FROM project_plan_versions pv
            WHERE pv.project_id = :pid
            ORDER BY pv.upload_time DESC
            LIMIT 1
        """), {"pid": project_id})
        prev_version = result.fetchone()
        
        if prev_version:
            # 获取当前项目状态快照
            status_result = conn.execute(text("""
                SELECT 
                    COUNT(*) FILTER (WHERE progress >= 100) as completed_tasks,
                    COUNT(*) FILTER (WHERE progress < 100 AND end_date < CURRENT_DATE) as delayed_tasks,
                    COUNT(*) FILTER (WHERE progress > 0 AND progress < 100) as ongoing_tasks,
                    COUNT(*) FILTER (WHERE progress = 0 OR progress IS NULL) as pending_tasks,
                    COALESCE(AVG(progress), 0) as avg_progress
                FROM project_tasks
                WHERE project_id = CAST(:pid AS VARCHAR) AND is_latest = true AND is_deleted = false
            """), {"pid": project_id})
            status_row = status_result.fetchone()
            
            previous_status = {
                "snapshot_date": datetime.now().strftime("%Y-%m-%d"),
                "previous_version": prev_version[1],
                "completed_tasks": status_row[0] or 0,
                "delayed_tasks": status_row[1] or 0,
                "ongoing_tasks": status_row[2] or 0,
                "pending_tasks": status_row[3] or 0,
                "avg_progress": round(float(status_row[4] or 0), 1)
            }
    
    # 从请求 header 获取 token（优先），避免依赖内存缓存
    token = None
    auth_header = request.headers.get("authorization")
    if auth_header and auth_header.startswith("Bearer "):
        token = auth_header[7:]
    
    # 兜底：从内存缓存获取
    if not token:
        token = current_user.get("_raw_token") or get_user_token(username)
    
    if not token:
        raise HTTPException(status_code=401, detail="未找到用户认证信息")

    # 检查文件类型
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="只支持Excel文件(.xlsx, .xls)")

    try:
        # 读取文件内容
        content = await file.read()

        # 构建multipart/form-data请求
        files = {
            "file": (file.filename, content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        }

        params = {}
        if version_name:
            params["version_name"] = version_name
        if description:
            params["description"] = description

        # 调用主后端上传接口
        response = await http_client.post(
            f"{settings.BACKEND_API_URL}/api/v1/plan-versions/upload-excel/{project_id}",
            files=files,
            params=params,
            headers={"Authorization": f"Bearer {token}"},
            timeout=30.0
        )

        if response.status_code == 200:
            result = response.json()
            data = result.get("data", result)
            
            # 上传成功后，更新本地 project_plan_versions 表
            version_id = data.get("version_id")
            if version_id:
                uploader_name = current_user.get("name") or current_user.get("username")
                try:
                    with get_connection() as conn:
                        # 更新 upload_by 和变更追踪信息
                        conn.execute(text("""
                            UPDATE project_plan_versions 
                            SET upload_by = :name,
                                change_type = :change_type,
                                change_reason = :change_reason,
                                previous_status = :previous_status
                            WHERE id = :vid
                        """), {
                            "name": uploader_name, 
                            "vid": version_id,
                            "change_type": change_type or "初始计划",
                            "change_reason": change_reason,
                            "previous_status": json.dumps(previous_status) if previous_status else None
                        })
                        conn.commit()
                        logger.info(f"更新版本 {version_id} 的上传者为 {uploader_name}")
                except Exception as e:
                    logger.warning(f"更新上传者失败: {e}")  # 不影响主流程
            
            return {
                "success": True,
                "message": f"成功导入 {data.get('task_count', 0)} 个任务",
                "version_id": version_id,
                "version_number": data.get("version_number"),
                "version_name": data.get("version_name"),
                "task_count": data.get("task_count"),
                "tasks": data.get("tasks", [])
            }
        else:
            error_msg = "上传失败"
            try:
                error_data = response.json()
                error_msg = error_data.get("detail") or error_data.get("message") or error_msg
            except:
                pass
            raise HTTPException(status_code=response.status_code, detail=error_msg)

    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f" {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"上传失败: {str(e)}")


@app.get("/agent/api/agent/plans/versions/{project_id}")
async def get_plan_versions(
    project_id: int,
    current_user: Dict = Depends(get_current_user)
):
    """获取项目的计划版本列表（从本地数据库读取）"""
    try:
        with get_connection() as conn:
            result = conn.execute(text("""
                SELECT 
                    id, 
                    project_id,
                    version_number,
                    version_name,
                    description,
                    upload_by,
                    upload_time,
                    file_name,
                    task_count,
                    is_current,
                    created_at
                FROM project_plan_versions
                WHERE project_id = :project_id
                ORDER BY created_at DESC
            """), {"project_id": project_id})
            
            versions = []
            for row in result:
                versions.append({
                    "id": row[0],
                    "project_id": row[1],
                    "version_number": row[2],
                    "version_name": row[3] or "",
                    "description": row[4] or "",
                    "upload_by": row[5],  # 已经是名字（初始版本是负责人名字，上传版本是上传者名字）
                    "upload_time": row[6].isoformat() if row[6] else None,
                    "file_name": row[7] or "",
                    "task_count": row[8] or 0,
                    "is_current": row[9] or False,
                    "created_at": row[10].isoformat() if row[10] else None
                })
            
            return versions
    except Exception as e:
        logger.error(f"获取版本列表失败: {e}")
        return []


@app.get("/agent/api/agent/plans/compare/{version_id1}/{version_id2}")
async def compare_plan_versions(
    version_id1: int,
    version_id2: int,
    current_user: Dict = Depends(get_current_user)
):
    """对比两个计划版本"""
    username = current_user.get("username")
    token = current_user.get("_raw_token") or get_user_token(username)

    if not token:
        raise HTTPException(status_code=401, detail="未找到用户认证信息")

    try:
        response = await http_client.get(
            f"{settings.BACKEND_API_URL}/api/v1/plan-versions/compare/{version_id1}/{version_id2}",
            headers={"Authorization": f"Bearer {token}"},
            timeout=10.0
        )

        if response.status_code == 200:
            result = response.json()
            return result.get("data", result)
        raise HTTPException(status_code=response.status_code, detail="对比失败")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f" {e}")
        raise HTTPException(status_code=500, detail=f"对比失败: {str(e)}")

# ============== 智能问答工具 ==============

# 全局数据库引擎（懒加载）
_db_engine = None

# 数据库引擎单例（已废弃，改用 database 模块的全局单例）
# 保留 _db_engine 变量以兼容可能的引用
_db_engine = None

def get_db_engine():
    """获取数据库引擎（使用 database 模块的全局单例）"""
    return get_engine()


# 定义查询工具（简化版，不用LangChain tools）
def execute_query(tool_name: str, params: dict) -> str:
    """执行查询工具"""
    try:
        engine = get_db_engine()
        from sqlalchemy import text

        with get_connection() as conn:
            if tool_name == "query_projects":
                sql = """
                    SELECT id, name, leader, status, progress
                    FROM projects WHERE is_deleted = false
                """
                conditions = []
                if params.get("keyword"):
                    conditions.append(f"name LIKE '%{params['keyword']}%'")
                if params.get("leader"):
                    conditions.append(f"leader LIKE '%{params['leader']}%'")
                if conditions:
                    sql += " AND " + " AND ".join(conditions)
                sql += " LIMIT 20"

                result = conn.execute(text(sql))
                return json.dumps([{
                    "id": r[0], "name": r[1], "leader": r[2],
                    "status": r[3], "progress": r[4]
                } for r in result], ensure_ascii=False)

            elif tool_name == "query_tasks":
                today = datetime.now().date()
                sql = """
                    SELECT pt.task_name, pt.project_id, p.name as project_name,
                           pt.assignee, pt.end_date, pt.status
                    FROM project_tasks pt
                    LEFT JOIN projects p ON CAST(pt.project_id AS INTEGER) = p.id
                    WHERE pt.is_deleted = false
                """
                conditions = []
                if params.get("assignee"):
                    conditions.append(f"pt.assignee LIKE '%{params['assignee']}%'")
                if params.get("days"):
                    end_date = today + timedelta(days=params["days"])
                    conditions.append(f"pt.end_date >= '{today}' AND pt.end_date <= '{end_date}'")
                if conditions:
                    sql += " AND " + " AND ".join(conditions)
                sql += " LIMIT 50"

                result = conn.execute(text(sql))
                return json.dumps([{
                    "task_name": r[0], "project_id": r[1], "project_name": r[2],
                    "assignee": r[3], "end_date": str(r[4]) if r[4] else None, "status": r[5]
                } for r in result], ensure_ascii=False)

            elif tool_name == "query_project_tasks_by_id":
                project_id = params.get("project_id")
                if not project_id:
                    return json.dumps({"error": "缺少 project_id 参数"}, ensure_ascii=False)
                
                sql = """
                    SELECT pt.task_id, pt.task_name, pt.assignee, pt.status, 
                           pt.progress, pt.start_date, pt.end_date
                    FROM project_tasks pt
                    WHERE pt.is_deleted = false
                      AND CAST(pt.project_id AS INTEGER) = :pid
                """
                conditions = []
                if params.get("status"):
                    status = params["status"]
                    today = datetime.now().date()
                    if status == "延期":
                        conditions.append("pt.end_date < CURRENT_DATE AND (pt.progress < 100 OR pt.progress IS NULL)")
                    elif status == "已完成":
                        conditions.append("pt.progress >= 100")
                    elif status == "进行中":
                        conditions.append("(pt.progress > 0 AND pt.progress < 100)")
                    elif status == "未开始":
                        conditions.append("(pt.progress = 0 OR pt.progress IS NULL)")
                if conditions:
                    sql += " AND " + " AND ".join(conditions)
                sql += " ORDER BY pt.task_id LIMIT 100"

                result = conn.execute(text(sql), {"pid": int(project_id)})
                return json.dumps([{
                    "task_id": r[0], "task_name": r[1], "assignee": r[2],
                    "status": r[3], "progress": float(r[4] or 0),
                    "start_date": str(r[5]) if r[5] else None,
                    "end_date": str(r[6]) if r[6] else None
                } for r in result], ensure_ascii=False)

            elif tool_name == "query_risks":
                sql = """
                    SELECT p.name, p.leader, COUNT(*) as delayed_count
                    FROM projects p
                    JOIN project_tasks pt ON CAST(pt.project_id AS INTEGER) = p.id
                    WHERE p.is_deleted = false AND pt.is_deleted = false
                      AND pt.end_date < CURRENT_DATE AND pt.actual_end_date IS NULL
                    GROUP BY p.id, p.name, p.leader
                    ORDER BY delayed_count DESC LIMIT 10
                """
                result = conn.execute(text(sql))
                return json.dumps([{
                    "project_name": r[0], "leader": r[1], "delayed_count": r[2]
                } for r in result], ensure_ascii=False)

            elif tool_name == "query_work_hours":
                today = datetime.now().date()
                month = params.get("month", today.strftime("%Y-%m"))
                month_start = datetime.strptime(month + "-01", "%Y-%m-%d").date()

                sql = f"""
                    SELECT dr.employee_name, SUM(dwi.hours_spent) as total_hours
                    FROM daily_reports dr
                    JOIN daily_work_items dwi ON dr.id = dwi.report_id
                    WHERE dr.is_deleted = false
                      AND dr.report_date >= '{month_start}'
                      AND dr.report_date <= '{today}'
                """
                if params.get("employee_name"):
                    sql += f" AND dr.employee_name LIKE '%{params['employee_name']}%'"
                sql += " GROUP BY dr.employee_name ORDER BY total_hours DESC LIMIT 10"

                result = conn.execute(text(sql))
                return json.dumps([{
                    "employee_name": r[0], "total_hours": float(r[1] or 0)
                } for r in result], ensure_ascii=False)

            elif tool_name == "query_goals":
                # 方案：根据本月任务自动生成月度目标（不再依赖 monthly_goals 表）
                month = datetime.now().strftime("%Y-%m")
                month_start = datetime.now().replace(day=1)
                month_end = (month_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)

                # 查询用户本月负责的任务
                sql = f"""
                    SELECT pt.task_id, pt.task_name, pt.status, pt.progress,
                           pt.start_date, pt.end_date, p.name as project_name
                    FROM project_tasks pt
                    JOIN projects p ON CAST(pt.project_id AS INTEGER) = p.id
                    WHERE pt.is_deleted = false
                      AND p.is_deleted = false
                      AND (
                        (pt.start_date >= '{month_start}' AND pt.start_date <= '{month_end}')
                        OR (pt.end_date >= '{month_start}' AND pt.end_date <= '{month_end}')
                        OR (pt.start_date < '{month_start}' AND pt.end_date > '{month_end}')
                      )
                """

                if params.get("employee_name"):
                    sql += f" AND pt.assignee LIKE '%{params['employee_name']}%'"

                sql += " ORDER BY pt.end_date LIMIT 20"

                result = conn.execute(text(sql))
                tasks = []
                for r in result:
                    tasks.append({
                        "task_id": r[0],
                        "task_name": r[1],
                        "status": r[2],
                        "progress": float(r[3] or 0),
                        "start_date": str(r[4]) if r[4] else None,
                        "end_date": str(r[5]) if r[5] else None,
                        "project_name": r[6]
                    })

                # 根据任务自动推断月度目标
                if tasks:
                    # 按项目分组
                    project_tasks = {}
                    for task in tasks:
                        pname = task["project_name"]
                        if pname not in project_tasks:
                            project_tasks[pname] = []
                        project_tasks[pname].append(task)

                    # 生成月度目标描述
                    goals = []
                    for pname, ptasks in project_tasks.items():
                        ongoing = [t for t in ptasks if t["status"] in ["进行中", "未开始"]]
                        completed = [t for t in ptasks if t["status"] == "已完成"]
                        delayed = [t for t in ptasks if t["status"] == "延期"]

                        if ongoing:
                            goals.append(f"{pname}：推进{len(ongoing)}个任务")
                        if completed:
                            goals.append(f"{pname}：已完成{len(completed)}个任务")
                        if delayed:
                            goals.append(f"{pname}：{len(delayed)}个任务延期")

                    return json.dumps({
                        "month": month,
                        "tasks": tasks,
                        "goals_summary": goals,
                        "total_tasks": len(tasks),
                        "ongoing_count": len([t for t in tasks if t["status"] in ["进行中", "未开始"]]),
                        "completed_count": len([t for t in tasks if t["status"] == "已完成"]),
                        "delayed_count": len([t for t in tasks if t["status"] == "延期"])
                    }, ensure_ascii=False)
                else:
                    return json.dumps({
                        "month": month,
                        "tasks": [],
                        "goals_summary": ["本月暂无分配任务"],
                        "total_tasks": 0
                    }, ensure_ascii=False)

            else:
                return json.dumps({"error": f"未知工具: {tool_name}"})

    except Exception as e:
        return json.dumps({"error": str(e)})


# 工具描述（用于提示LLM）
TOOL_DESCRIPTIONS = """
可用工具：
1. query_projects(keyword, leader) - 查询项目列表
   参数：keyword(项目名关键词，如"600KA"、"烟气治理"), leader(负责人姓名)
   用途：根据项目名称或负责人查询项目信息

2. query_tasks(assignee, days) - 查询任务
   参数：assignee(负责人姓名), days(未来N天)

3. query_project_tasks_by_id(project_id, status) - 查询指定项目的任务
   参数：project_id(项目ID，整数), status(任务状态：进行中/已完成/延期/未开始)
   用途：查询某个具体项目的任务列表、负责人、进度等

4. query_risks() - 查询延期风险项目

5. query_work_hours(employee_name, month) - 查询工时
   参数：employee_name(员工姓名), month(月份YYYY-MM)

6. query_goals(employee_name) - 查询月度目标
   参数：employee_name(员工姓名)
"""


# ============== 智能周报生成 API ==============

@app.get("/agent/api/agent/reports/weekly")
async def generate_weekly_report(
    week_start: Optional[str] = None,
    current_user: Dict = Depends(get_current_user)
):
    """
    生成智能周报

    Args:
        week_start: 周起始日期（YYYY-MM-DD），默认本周一

    Returns:
        周报内容（工时、项目、任务、风险）
    """
    username = current_user.get("username")
    user_info = get_user_info_cache(username)
    employee_id = user_info.get("employee_id") if user_info else username
    employee_name = user_info.get("name") if user_info else username

    try:
        # text 已从 database 模块导入
        from dotenv import load_dotenv
        load_dotenv()
        today = datetime.now().date()

        # 计算本周起始日期（周一）
        if week_start:
            start_date = datetime.strptime(week_start, "%Y-%m-%d").date()
        else:
            start_date = today - timedelta(days=today.weekday())

        end_date = start_date + timedelta(days=6)

        with get_connection() as conn:
            # 1. 工时统计
            result = conn.execute(text("""
                SELECT
                    COUNT(DISTINCT dr.id) as report_count,
                    COALESCE(SUM(dwi.hours_spent), 0) as total_hours,
                    COUNT(DISTINCT dwi.project_id) as project_count
                FROM daily_reports dr
                LEFT JOIN daily_work_items dwi ON dr.id = dwi.report_id
                WHERE dr.employee_id = :emp_id
                  AND dr.is_deleted = false
                  AND dr.report_date >= :start_date
                  AND dr.report_date <= :end_date
            """), {"emp_id": employee_id, "start_date": start_date, "end_date": end_date})

            hours_row = result.fetchone()
            report_count = hours_row[0] or 0
            total_hours = float(hours_row[1] or 0)
            project_count = hours_row[2] or 0

            # 2. 项目工时分布
            result = conn.execute(text("""
                SELECT
                    COALESCE(p.name, '其他') as project_name,
                    SUM(dwi.hours_spent) as hours,
                    COUNT(*) as item_count
                FROM daily_work_items dwi
                JOIN daily_reports dr ON dwi.report_id = dr.id
                LEFT JOIN projects p ON CAST(dwi.project_id AS INTEGER) = p.id
                WHERE dr.employee_id = :emp_id
                  AND dr.is_deleted = false
                  AND dr.report_date >= :start_date
                  AND dr.report_date <= :end_date
                GROUP BY p.name
                ORDER BY hours DESC
            """), {"emp_id": employee_id, "start_date": start_date, "end_date": end_date})

            project_hours = []
            for row in result:
                project_hours.append({
                    "project_name": row[0],
                    "hours": float(row[1] or 0),
                    "item_count": row[2]
                })

            # 3. 工作事项汇总（按项目分组）
            result = conn.execute(text("""
                SELECT
                    COALESCE(p.name, '其他') as project_name,
                    string_agg(DISTINCT dwi.work_content, '；') as contents,
                    SUM(dwi.hours_spent) as total_hours
                FROM daily_work_items dwi
                JOIN daily_reports dr ON dwi.report_id = dr.id
                LEFT JOIN projects p ON CAST(dwi.project_id AS INTEGER) = p.id
                WHERE dr.employee_id = :emp_id
                  AND dr.is_deleted = false
                  AND dr.report_date >= :start_date
                  AND dr.report_date <= :end_date
                GROUP BY p.name
                ORDER BY total_hours DESC
            """), {"emp_id": employee_id, "start_date": start_date, "end_date": end_date})

            work_summary = []
            for row in result:
                contents = row[1][:300] if row[1] and len(row[1]) > 300 else row[1]  # 截取前300字
                work_summary.append({
                    "project_name": row[0],
                    "contents": contents,
                    "hours": float(row[2] or 0)
                })

            # 4. 任务完成情况
            result = conn.execute(text("""
                SELECT
                    pt.task_name,
                    p.name as project_name,
                    pt.status,
                    pt.progress
                FROM project_tasks pt
                LEFT JOIN projects p ON CAST(pt.project_id AS INTEGER) = p.id
                WHERE pt.assignee_id = :emp_id
                  AND pt.is_deleted = false
                  AND (
                    (pt.actual_end_date >= :start_date AND pt.actual_end_date <= :end_date)
                    OR (pt.end_date >= :start_date AND pt.end_date <= :end_date)
                  )
                ORDER BY pt.end_date
            """), {"emp_id": employee_id, "start_date": start_date, "end_date": end_date})

            tasks = []
            completed_count = 0
            for row in result:
                tasks.append({
                    "task_name": row[0],
                    "project_name": row[1],
                    "status": row[2],
                    "progress": float(row[3] or 0)
                })
                if row[2] == "已完成":
                    completed_count += 1

            # 5. 延期任务
            result = conn.execute(text("""
                SELECT
                    pt.task_name,
                    p.name as project_name,
                    CURRENT_DATE - pt.end_date as delay_days
                FROM project_tasks pt
                LEFT JOIN projects p ON CAST(pt.project_id AS INTEGER) = p.id
                WHERE pt.assignee_id = :emp_id
                  AND pt.is_deleted = false
                  AND pt.end_date < CURRENT_DATE
                  AND pt.actual_end_date IS NULL
                ORDER BY delay_days DESC
                LIMIT 5
            """), {"emp_id": employee_id})

            delayed_tasks = []
            for row in result:
                delayed_tasks.append({
                    "task_name": row[0],
                    "project_name": row[1],
                    "delay_days": row[2]
                })

            # 6. 月度目标进度
            result = conn.execute(text("""
                SELECT title, progress_rate
                FROM monthly_goals
                WHERE user_id = :emp_id
                  AND is_deleted = false
                  AND month = :month
            """), {"emp_id": employee_id, "month": today.strftime("%Y-%m")})

            goals = []
            for row in result:
                goals.append({
                    "title": row[0],
                    "progress_rate": float(row[1] or 0)
                })

        # 生成周报文本
        week_number = start_date.isocalendar()[1]

        report_markdown = f"""# 周报 ({start_date.strftime('%m.%d')}-{end_date.strftime('%m.%d')})

## 基本信息
- **姓名**：{employee_name}
- **周次**：第{week_number}周
- **填报天数**：{report_count}天
- **累计工时**：{total_hours}小时
- **涉及项目**：{project_count}个

## 本周工作内容

"""

        if work_summary:
            for i, work in enumerate(work_summary, 1):
                report_markdown += f"### {i}. {work['project_name']}（{work['hours']}h）\n"
                report_markdown += f"{work['contents']}\n\n"
        else:
            report_markdown += "_暂无工作记录_\n\n"

        if tasks:
            report_markdown += f"## 任务完成情况\n\n"
            report_markdown += f"- 本周任务：{len(tasks)}项\n"
            report_markdown += f"- 已完成：{completed_count}项\n\n"

            if completed_count > 0:
                report_markdown += "**已完成任务**：\n"
                for t in tasks:
                    if t['status'] == '已完成':
                        report_markdown += f"- ✅ {t['task_name']}（{t['project_name']}）\n"
                report_markdown += "\n"

        if delayed_tasks:
            report_markdown += "## ⚠️ 延期预警\n\n"
            for t in delayed_tasks:
                report_markdown += f"- {t['task_name']}（{t['project_name']}）延期{t['delay_days']}天\n"
            report_markdown += "\n"

        if goals:
            report_markdown += "## 月度目标进度\n\n"
            for g in goals:
                status = "🟢" if g['progress_rate'] >= 80 else "🟡" if g['progress_rate'] >= 50 else "🔴"
                report_markdown += f"- {status} {g['title']}：{g['progress_rate']}%\n"

        report_markdown += "\n---\n*本报告由项目管家智能生成*"

        return {
            "success": True,
            "employee_name": employee_name,
            "week_range": {
                "start": start_date.isoformat(),
                "end": end_date.isoformat(),
                "week_number": week_number
            },
            "statistics": {
                "report_count": report_count,
                "total_hours": total_hours,
                "project_count": project_count,
                "task_count": len(tasks),
                "completed_count": completed_count,
                "delayed_count": len(delayed_tasks)
            },
            "project_hours": project_hours,
            "work_summary": work_summary,
            "tasks": tasks,
            "delayed_tasks": delayed_tasks,
            "goals": goals,
            "report_markdown": report_markdown
        }

    except Exception as e:
        logger.error(f" {e}")
        import traceback
        traceback.print_exc()
        return {
            "success": False,
            "message": str(e)
        }


# ============== 系统通知 API ==============

@app.get("/agent/api/agent/notifications")
async def get_my_notifications(
    unread_only: bool = False,
    limit: int = 20,
    current_user: Dict = Depends(get_current_user)
):
    """
    获取我的通知列表

    Args:
        unread_only: 仅未读
        limit: 返回数量
    """
    username = current_user.get("username")
    user_info = get_user_info_cache(username)
    employee_id = user_info.get("employee_id") if user_info else username

    try:
        # text 已从 database 模块导入
        from dotenv import load_dotenv
        load_dotenv()
        with get_connection() as conn:
            sql = """
                SELECT id, notification_type, priority_level, title, content,
                       is_read, create_time, related_task_id
                FROM tracking_notifications
                WHERE recipient_id = :emp_id AND is_deleted = false
            """
            params = {"emp_id": employee_id}

            if unread_only:
                sql += " AND is_read = false"

            sql += " ORDER BY create_time DESC LIMIT :limit"
            params["limit"] = limit

            result = conn.execute(text(sql), params)
            notifications = []
            for row in result:
                notifications.append({
                    "id": row[0],
                    "type": row[1],
                    "priority": row[2],
                    "title": row[3],
                    "content": row[4],
                    "is_read": row[5],
                    "create_time": row[6].isoformat() if row[6] else None,
                    "related_task_id": row[7]
                })

            # 获取未读数量
            count_result = conn.execute(text("""
                SELECT COUNT(*) FROM tracking_notifications
                WHERE recipient_id = :emp_id AND is_deleted = false AND is_read = false
            """), {"emp_id": employee_id})
            unread_count = count_result.fetchone()[0]

            return {
                "notifications": notifications,
                "unread_count": unread_count
            }

    except Exception as e:
        logger.error(f" {e}")
        return {"notifications": [], "unread_count": 0}


@app.post("/agent/api/agent/notifications/{notification_id}/read")
async def mark_notification_read(
    notification_id: int,
    current_user: Dict = Depends(get_current_user)
):
    """标记通知为已读"""
    try:
        # text 已从 database 模块导入
        from dotenv import load_dotenv
        load_dotenv()
        with get_connection() as conn:
            conn.execute(text("""
                UPDATE tracking_notifications
                SET is_read = true, read_time = CURRENT_TIMESTAMP
                WHERE id = :id
            """), {"id": notification_id})
            conn.commit()

        return {"success": True, "message": "已标记为已读"}

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"操作失败: {str(e)}")


@app.post("/agent/api/agent/notifications/read-all")
async def mark_all_notifications_read(
    current_user: Dict = Depends(get_current_user)
):
    """标记所有通知为已读"""
    username = current_user.get("username")
    user_info = get_user_info_cache(username)
    employee_id = user_info.get("employee_id") if user_info else username

    try:
        # text 已从 database 模块导入
        from dotenv import load_dotenv
        load_dotenv()
        with get_connection() as conn:
            conn.execute(text("""
                UPDATE tracking_notifications
                SET is_read = true, read_time = CURRENT_TIMESTAMP
                WHERE recipient_id = :emp_id AND is_read = false AND is_deleted = false
            """), {"emp_id": employee_id})
            conn.commit()

        return {"success": True, "message": "已全部标记为已读"}

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"操作失败: {str(e)}")


@app.post("/agent/api/agent/notifications/generate")
async def generate_smart_notifications(
    current_user: Dict = Depends(get_current_user)
):
    """
    生成智能通知

    - 今日待办提醒
    - 延期任务预警
    - 日报填报提醒
    """
    username = current_user.get("username")
    user_info = get_user_info_cache(username)
    employee_id = user_info.get("employee_id") if user_info else username
    employee_name = user_info.get("name") if user_info else username

    try:
        # text 已从 database 模块导入
        from dotenv import load_dotenv
        load_dotenv()
        today = datetime.now().date()
        notifications_created = []

        with get_connection() as conn:
            # 1. 今日待办提醒
            result = conn.execute(text("""
                SELECT pt.task_name, p.name as project_name
                FROM project_tasks pt
                JOIN projects p ON CAST(pt.project_id AS INTEGER) = p.id
                WHERE pt.assignee_id = :emp_id
                  AND pt.is_deleted = false
                  AND pt.end_date = :today
                  AND pt.actual_end_date IS NULL
            """), {"emp_id": employee_id, "today": today})

            today_tasks = list(result)
            if today_tasks:
                content = "今日待办任务：\n" + "\n".join([f"• {t[0]}（{t[1]}）" for t in today_tasks[:5]])
                conn.execute(text("""
                    INSERT INTO tracking_notifications
                    (tracking_id, notification_type, priority_level, title, content,
                     recipient_id, recipient_name, is_sent, create_time)
                    VALUES (0, 'task_reminder', '高', '📋 今日待办提醒', :content,
                            :emp_id, :emp_name, true, CURRENT_TIMESTAMP)
                """), {"content": content, "emp_id": employee_id, "emp_name": employee_name})
                notifications_created.append("今日待办提醒")

            # 2. 延期任务预警
            result = conn.execute(text("""
                SELECT pt.task_name, p.name as project_name,
                       CURRENT_DATE - pt.end_date as delay_days
                FROM project_tasks pt
                JOIN projects p ON CAST(pt.project_id AS INTEGER) = p.id
                WHERE pt.assignee_id = :emp_id
                  AND pt.is_deleted = false
                  AND pt.end_date < CURRENT_DATE
                  AND pt.actual_end_date IS NULL
                ORDER BY delay_days DESC
                LIMIT 5
            """), {"emp_id": employee_id})

            delayed_tasks = list(result)
            if delayed_tasks:
                content = "延期任务预警：\n" + "\n".join([f"• {t[0]}（{t[1]}）延期{t[2]}天" for t in delayed_tasks])
                conn.execute(text("""
                    INSERT INTO tracking_notifications
                    (tracking_id, notification_type, priority_level, title, content,
                     recipient_id, recipient_name, is_sent, create_time)
                    VALUES (0, 'delay_warning', '紧急', '⚠️ 延期任务预警', :content,
                            :emp_id, :emp_name, true, CURRENT_TIMESTAMP)
                """), {"content": content, "emp_id": employee_id, "emp_name": employee_name})
                notifications_created.append("延期任务预警")

            # 3. 日报填报提醒（检查今日是否已提交）
            result = conn.execute(text("""
                SELECT id FROM daily_reports
                WHERE employee_id = :emp_id AND report_date = :today AND is_deleted = false
            """), {"emp_id": employee_id, "today": today})

            if not result.fetchone():
                conn.execute(text("""
                    INSERT INTO tracking_notifications
                    (tracking_id, notification_type, priority_level, title, content,
                     recipient_id, recipient_name, is_sent, create_time)
                    VALUES (0, 'daily_reminder', '中', '📝 日报填报提醒',
                            '今日日报尚未提交，请及时填报。',
                            :emp_id, :emp_name, true, CURRENT_TIMESTAMP)
                """), {"emp_id": employee_id, "emp_name": employee_name})
                notifications_created.append("日报填报提醒")

            conn.commit()

        return {
            "success": True,
            "notifications_created": notifications_created,
            "count": len(notifications_created)
        }

    except Exception as e:
        logger.error(f" {e}")
        import traceback
        traceback.print_exc()
        return {"success": False, "message": str(e)}


# ============== RAG文档问答（简化版 - 使用全文搜索） ==============

import re
from typing import List, Optional, Tuple

def chunk_text_smart(text: str, max_length: int = 500, min_length: int = 100) -> List[Tuple[str, dict]]:
    """
    智能文本切分 - 按标题和段落切分

    返回: [(chunk_text, metadata), ...]
    """
    chunks = []

    # 识别标题模式
    title_patterns = [
        r'^第[一二三四五六七八九十]+[章节]',  # 第一章、第二节
        r'^[一二三四五六七八九十]+[、.]',     # 一、二、
        r'^\d+\.[\d\s]',                     # 1.1、1.2
        r'^#{1,3}\s',                        # # ## ###
        r'^【[^】]+】',                       # 【标题】
    ]

    # 按段落分割
    paragraphs = re.split(r'\n\s*\n', text)

    current_chunk = ""
    current_title = ""
    chunk_start = 0

    for i, para in enumerate(paragraphs):
        para = para.strip()
        if not para:
            continue

        # 检测是否为标题
        is_title = any(re.match(p, para) for p in title_patterns)

        if is_title:
            # 如果当前块有内容，先保存
            if current_chunk and len(current_chunk) >= min_length:
                chunks.append((
                    current_chunk.strip(),
                    {"title": current_title, "para_start": chunk_start}
                ))
            current_chunk = para + "\n\n"
            current_title = para[:50]  # 记录标题
            chunk_start = i
        else:
            # 检查是否需要切分
            if len(current_chunk) + len(para) > max_length and len(current_chunk) >= min_length:
                chunks.append((
                    current_chunk.strip(),
                    {"title": current_title, "para_start": chunk_start}
                ))
                current_chunk = para + "\n\n"
                chunk_start = i
            else:
                current_chunk += para + "\n\n"

    # 保存最后一块
    if current_chunk and len(current_chunk) >= min_length:
        chunks.append((
            current_chunk.strip(),
            {"title": current_title, "para_start": chunk_start}
        ))

    # 如果没有切分出任何块，返回整个文本
    if not chunks:
        chunks.append((text.strip(), {"title": "全文", "para_start": 0}))

    return chunks


# 嵌入模型（懒加载）
_embedding_model = None

def get_embedding_model():
    """获取嵌入模型（单例）- 使用 BGE-base-zh 中文向量模型"""
    global _embedding_model
    if _embedding_model is None:
        # 配置 HuggingFace 镜像（解决国内网络问题）
        os.environ['HF_ENDPOINT'] = 'https://hf-mirror.com'

        from sentence_transformers import SentenceTransformer
        # BGE-base-zh: 中文语义向量模型，768维
        _embedding_model = SentenceTransformer('BAAI/bge-base-zh-v1.5')
        logger.info(" 已加载 BAAI/bge-base-zh-v1.5")
    return _embedding_model


def generate_embedding(text: str) -> Optional[List[float]]:
    """生成文本嵌入向量（使用 BGE-base-zh）"""
    try:
        model = get_embedding_model()
        # 截断过长文本（BGE 最大 512 tokens）
        if len(text) > 2000:
            text = text[:2000]
        # BGE 推荐添加指令前缀（但用于检索时不需要）
        embedding = model.encode(text, normalize_embeddings=True, convert_to_numpy=True)
        return embedding.tolist()
    except Exception as e:
        logger.error(f" {e}")
        import traceback
        traceback.print_exc()
        return None

@app.post("/agent/api/agent/documents/upload")
async def upload_document(
    file: UploadFile = File(...),
    project_id: Optional[int] = None,
    current_user: Dict = Depends(get_current_user)
):
    """上传文档（支持智能切分和向量嵌入）"""
    username = current_user.get("username")

    try:
        content = await file.read()
        filename = file.filename
        file_type = filename.split('.')[-1].lower()

        text_content = ""
        if file_type == 'txt':
            text_content = content.decode('utf-8')
        elif file_type == 'pdf':
            import PyPDF2
            import io
            pdf_reader = PyPDF2.PdfReader(io.BytesIO(content))
            for page in pdf_reader.pages:
                text_content += page.extract_text() + "\n"
        elif file_type in ['doc', 'docx']:
            import docx
            import io
            doc = docx.Document(io.BytesIO(content))
            for para in doc.paragraphs:
                text_content += para.text + "\n"
        else:
            return {"success": False, "message": f"不支持的文件类型: {file_type}"}

        if not text_content.strip():
            return {"success": False, "message": "文档内容为空"}

        engine = get_db_engine()

        with get_connection() as conn:
            result = conn.execute(text("""
                INSERT INTO documents (filename, file_type, file_size, project_id, uploaded_by)
                VALUES (:filename, :file_type, :file_size, :project_id, :uploaded_by)
                RETURNING id
            """), {
                "filename": filename,
                "file_type": file_type,
                "file_size": len(content),
                "project_id": project_id,
                "uploaded_by": username
            })
            doc_id = result.fetchone()[0]
            conn.commit()

        # 智能切分
        chunks_with_meta = chunk_text_smart(text_content)

        # 批量生成嵌入（提升效率）
        logger.info(f" 开始为 {len(chunks_with_meta)} 个片段生成嵌入...")
        chunk_texts = [c[0] for c in chunks_with_meta]

        try:
            model = get_embedding_model()
            embeddings = model.encode(chunk_texts, convert_to_numpy=True, show_progress_bar=False)
        except Exception as e:
            logger.info(f" 嵌入生成失败，使用空嵌入: {e}")
            embeddings = [None] * len(chunks_with_meta)

        # 插入数据库（使用原生 psycopg2 绕过 SQLAlchemy text() 的类型转换限制）
        import psycopg2
        db_url = os.getenv("DATABASE_URL")
        if not db_url:
            raise RuntimeError("DATABASE_URL environment variable not set")
        conn = psycopg2.connect(db_url)
        cursor = conn.cursor()

        for i, (chunk_text, meta) in enumerate(chunks_with_meta):
            embedding = embeddings[i].tolist() if embeddings[i] is not None else None

            if embedding:
                cursor.execute("""
                    INSERT INTO document_chunks (document_id, chunk_text, chunk_index, embedding, metadata)
                    VALUES (%s, %s, %s, %s::vector, %s::jsonb)
                """, (doc_id, chunk_text, i, str(embedding), json.dumps(meta)))
            else:
                cursor.execute("""
                    INSERT INTO document_chunks (document_id, chunk_text, chunk_index, metadata)
                    VALUES (%s, %s, %s, %s::jsonb)
                """, (doc_id, chunk_text, i, json.dumps(meta)))

        conn.commit()
        cursor.close()
        conn.close()

        return {
            "success": True,
            "document_id": doc_id,
            "filename": filename,
            "chunks_count": len(chunks_with_meta),
            "message": f"文档上传成功，已切分为{len(chunks_with_meta)}个片段并生成向量嵌入"
        }

    except Exception as e:
        import traceback
        traceback.print_exc()
        return {"success": False, "message": str(e)}

@app.get("/agent/api/agent/documents")
async def list_documents(
    project_id: Optional[int] = None,
    current_user: Dict = Depends(get_current_user)
):
    """列出文档"""
    try:
        # text 已从 database 模块导入
        engine = get_db_engine()

        with get_connection() as conn:
            if project_id:
                result = conn.execute(text("""
                    SELECT id, filename, file_type, file_size, uploaded_by, created_at
                    FROM documents WHERE project_id = :pid ORDER BY created_at DESC
                """), {"pid": project_id})
            else:
                result = conn.execute(text("""
                    SELECT id, filename, file_type, file_size, uploaded_by, created_at
                    FROM documents ORDER BY created_at DESC LIMIT 50
                """))

            docs = [{"id": r[0], "filename": r[1], "file_type": r[2],
                     "file_size": r[3], "uploaded_by": r[4], "created_at": str(r[5])}
                    for r in result]

            return {"success": True, "documents": docs}

    except Exception as e:
        return {"success": False, "message": str(e)}

@app.delete("/agent/api/agent/documents/{doc_id}")
async def delete_document(doc_id: int, current_user: Dict = Depends(get_current_user)):
    """删除文档"""
    try:
        # text 已从 database 模块导入
        engine = get_db_engine()

        with get_connection() as conn:
            conn.execute(text("DELETE FROM documents WHERE id = :id"), {"id": doc_id})
            conn.commit()

        return {"success": True, "message": "文档已删除"}

    except Exception as e:
        return {"success": False, "message": str(e)}

@app.post("/agent/api/agent/documents/search")
async def search_documents(request: Dict, current_user: Dict = Depends(get_current_user)):
    """搜索文档（向量语义搜索 + 关键词匹配）"""
    query = request.get("query", "")
    top_k = request.get("top_k", 5)
    use_semantic = request.get("use_semantic", True)

    if not query:
        return {"success": False, "message": "请输入查询内容"}

    try:
        import psycopg2
        db_url = os.getenv("DATABASE_URL")
        if not db_url:
            raise RuntimeError("DATABASE_URL environment variable not set")

        results = []

        # 1. 语义搜索
        if use_semantic:
            try:
                query_embedding = generate_embedding(query)
                if query_embedding:
                    conn = psycopg2.connect(db_url)
                    cursor = conn.cursor()

                    cursor.execute("""
                        SELECT dc.id, dc.document_id, dc.chunk_text, d.filename,
                               dc.metadata,
                               1 - (dc.embedding <=> %s::vector) as similarity
                        FROM document_chunks dc
                        JOIN documents d ON dc.document_id = d.id
                        WHERE dc.embedding IS NOT NULL
                        ORDER BY dc.embedding <=> %s::vector
                        LIMIT %s
                    """, (str(query_embedding), str(query_embedding), top_k))

                    for r in cursor.fetchall():
                        meta = r[4] if r[4] else {}
                        results.append({
                            "chunk_id": r[0],
                            "document_id": r[1],
                            "content": r[2][:500] if r[2] else "",
                            "filename": r[3],
                            "title": meta.get("title", ""),
                            "similarity": float(r[5]) if r[5] else 0,
                            "search_type": "semantic"
                        })

                    cursor.close()
                    conn.close()
            except Exception as e:
                logger.error(f" {e}")

        # 2. 关键词搜索（补充）
        if not results or len(results) < top_k:
            from sqlalchemy import text as text
            engine = get_db_engine()

            with get_connection() as conn:
                result = conn.execute(text("""
                    SELECT dc.id, dc.document_id, dc.chunk_text, d.filename, dc.metadata
                    FROM document_chunks dc
                    JOIN documents d ON dc.document_id = d.id
                    WHERE dc.chunk_text ILIKE :query
                    LIMIT :limit
                """), {"query": f"%{query}%", "limit": top_k - len(results)})

                for r in result:
                    if not any(res["chunk_id"] == r[0] for res in results):
                        meta = r[4] if r[4] else {}
                        results.append({
                            "chunk_id": r[0],
                            "document_id": r[1],
                            "content": r[2][:500] if r[2] else "",
                            "filename": r[3],
                            "title": meta.get("title", ""),
                            "similarity": 0.5,
                            "search_type": "keyword"
                        })

        return {
            "success": True,
            "results": results,
            "query": query,
            "search_type": "semantic" if results and results[0].get("search_type") == "semantic" else "keyword"
        }

    except Exception as e:
        import traceback
        traceback.print_exc()
        return {"success": False, "message": str(e)}
    query = request.get("query", "")
    top_k = request.get("top_k", 5)

    if not query:
        return {"success": False, "message": "请输入查询内容"}

    try:
        # text 已从 database 模块导入
        engine = get_db_engine()

        with get_connection() as conn:
            result = conn.execute(text("""
                SELECT dc.id, dc.document_id, dc.chunk_text, d.filename
                FROM document_chunks dc
                JOIN documents d ON dc.document_id = d.id
                WHERE dc.chunk_text ILIKE :query
                LIMIT :limit
            """), {"query": f"%{query}%", "limit": top_k})

            results = [{"chunk_id": r[0], "document_id": r[1],
                        "content": r[2][:500], "filename": r[3]} for r in result]

            return {"success": True, "results": results, "query": query}

    except Exception as e:
        return {"success": False, "message": str(e)}


# ============== 数据导出 API ==============

@app.get("/agent/api/agent/export/hours-excel")
async def export_hours_excel(
    month: Optional[str] = None,
    current_user: Dict = Depends(get_current_user)
):
    """导出工时Excel"""
    username = current_user.get("username")
    user_info = get_user_info_cache(username)
    employee_id = user_info.get("employee_id") if user_info else username

    try:
        # text 已从 database 模块导入
        engine = get_db_engine()

        today = datetime.now().date()
        if not month:
            month = today.strftime("%Y-%m")
        month_start = datetime.strptime(month + "-01", "%Y-%m-%d").date()

        with get_connection() as conn:
            result = conn.execute(text("""
                SELECT dr.report_date, p.name as project_name, dwi.work_content,
                       dwi.hours_spent, dr.employee_name
                FROM daily_reports dr
                JOIN daily_work_items dwi ON dr.id = dwi.report_id
                LEFT JOIN projects p ON CAST(dwi.project_id AS INTEGER) = p.id
                WHERE dr.employee_id = :emp_id
                  AND dr.is_deleted = false
                  AND dr.report_date >= :start
                  AND dr.report_date <= :today
                ORDER BY dr.report_date DESC
            """), {"emp_id": employee_id, "start": month_start, "today": today})

            rows = []
            for r in result:
                rows.append({
                    "date": str(r[0]),
                    "project": r[1] or "其他",
                    "content": r[2],
                    "hours": float(r[3] or 0),
                    "employee": r[4]
                })

            return {"success": True, "data": rows, "month": month, "count": len(rows)}

    except Exception as e:
        return {"success": False, "message": str(e)}


# ============== 预测分析 API ==============

@app.get("/agent/api/agent/predict/hours")
async def predict_month_hours(current_user: Dict = Depends(get_current_user)):
    """预测本月工时"""
    username = current_user.get("username")
    user_info = get_user_info_cache(username)
    employee_id = user_info.get("employee_id") if user_info else username

    try:
        # text 已从 database 模块导入
        engine = get_db_engine()

        today = datetime.now().date()
        month_start = today.replace(day=1)
        days_passed = today.day
        days_in_month = (today.replace(month=today.month % 12 + 1, day=1) - timedelta(days=1)).day

        with get_connection() as conn:
            result = conn.execute(text("""
                SELECT COALESCE(SUM(dwi.hours_spent), 0)
                FROM daily_reports dr
                JOIN daily_work_items dwi ON dr.id = dwi.report_id
                WHERE dr.employee_id = :emp_id
                  AND dr.report_date >= :start
                  AND dr.report_date <= :today
            """), {"emp_id": employee_id, "start": month_start, "today": today})

            current_hours = float(result.fetchone()[0] or 0)

            # 预测：当前工时 / 已过天数 * 总天数
            predicted_hours = (current_hours / days_passed * days_in_month) if days_passed > 0 else 0
            daily_avg = current_hours / days_passed if days_passed > 0 else 0

            return {
                "success": True,
                "current_hours": round(current_hours, 1),
                "predicted_hours": round(predicted_hours, 1),
                "daily_avg": round(daily_avg, 1),
                "days_passed": days_passed,
                "days_in_month": days_in_month,
                "status": "normal" if predicted_hours < 160 else "warning"
            }

    except Exception as e:
        return {"success": False, "message": str(e)}


# ============== 团队看板 API ==============

@app.get("/agent/api/agent/team/hours-ranking")
async def get_team_hours_ranking(
    month: Optional[str] = None,
    current_user: Dict = Depends(get_current_user)
):
    """团队工时排名"""
    username = current_user.get("username")
    user_info = get_user_info_cache(username)

    # 仅管理员可访问
    if not user_info or user_info.get("role_id") != 11:
        raise HTTPException(status_code=403, detail="无权限访问")

    try:
        # text 已从 database 模块导入
        engine = get_db_engine()

        today = datetime.now().date()
        if not month:
            month = today.strftime("%Y-%m")
        month_start = datetime.strptime(month + "-01", "%Y-%m-%d").date()

        with get_connection() as conn:
            result = conn.execute(text("""
                SELECT dr.employee_name, p.department,
                       SUM(dwi.hours_spent) as total_hours,
                       COUNT(DISTINCT dr.id) as report_count
                FROM daily_reports dr
                JOIN daily_work_items dwi ON dr.id = dwi.report_id
                LEFT JOIN personnel p ON dr.employee_id = p.employee_id
                WHERE dr.is_deleted = false
                  AND dr.report_date >= :start
                  AND dr.report_date <= :today
                GROUP BY dr.employee_name, p.department
                ORDER BY total_hours DESC
                LIMIT 20
            """), {"start": month_start, "today": today})

            ranking = []
            for i, r in enumerate(result, 1):
                ranking.append({
                    "rank": i,
                    "name": r[0],
                    "department": r[1] or "未知",
                    "hours": float(r[2] or 0),
                    "report_count": r[3]
                })

            return {"success": True, "month": month, "ranking": ranking}

    except Exception as e:
        return {"success": False, "message": str(e)}


@app.get("/agent/api/agent/team/goals-progress")
async def get_team_goals_progress(current_user: Dict = Depends(get_current_user)):
    """团队目标进度"""
    username = current_user.get("username")
    user_info = get_user_info_cache(username)

    if not user_info or user_info.get("role_id") != 11:
        raise HTTPException(status_code=403, detail="无权限访问")

    try:
        # text 已从 database 模块导入
        engine = get_db_engine()

        month = datetime.now().strftime("%Y-%m")

        with get_connection() as conn:
            result = conn.execute(text("""
                SELECT user_name, title, progress_rate, status
                FROM monthly_goals
                WHERE is_deleted = false AND month = :month
                ORDER BY progress_rate ASC
            """), {"month": month})

            goals = []
            for r in result:
                goals.append({
                    "name": r[0],
                    "title": r[1],
                    "progress": float(r[2] or 0),
                    "status": r[3]
                })

            return {"success": True, "month": month, "goals": goals}

    except Exception as e:
        return {"success": False, "message": str(e)}


# ============== 会话存储 ==============

# 内存会话存储（作为二级缓存）
_session_store: Dict[str, List] = {}

def get_session_history(session_id: str) -> List:
    """获取会话历史（优先数据库）"""
    # 先查内存缓存
    if session_id in _session_store:
        return _session_store[session_id]
    
    # 内存未命中，从数据库加载
    try:
        with get_connection() as conn:
            result = conn.execute(text("""
                SELECT messages FROM chat_sessions 
                WHERE session_key = :key
            """), {"key": session_id})
            row = result.fetchone()
            if row and row[0]:
                messages_data = row[0] if isinstance(row[0], list) else json.loads(row[0])
                history = [
                    HumanMessage(content=m["content"]) if m["type"] == "human"
                    else AIMessage(content=m["content"])
                    for m in messages_data
                ]
                _session_store[session_id] = history
                return history
    except Exception as e:
        logger.error(f"获取会话历史失败: {e}")
    
    return []

def save_session_history(session_id: str, messages: List):
    """保存会话历史（内存 + 数据库）"""
    # 只保留最近10轮对话
    history = messages[-20:]
    _session_store[session_id] = history
    
    # 持久化到数据库
    try:
        messages_data = [
            {"type": "human" if isinstance(m, HumanMessage) else "ai", "content": m.content}
            for m in history
        ]
        
        with get_connection() as conn:
            # 使用 CAST 而非 ::jsonb，避免 SQLAlchemy text() 解析问题
            conn.execute(text("""
                INSERT INTO chat_sessions (session_key, session_type, user_id, messages, updated_at)
                VALUES (:key, 'general', :key, CAST(:messages AS jsonb), NOW())
                ON CONFLICT (session_key) 
                DO UPDATE SET messages = CAST(:messages AS jsonb), updated_at = NOW()
            """), {"key": session_id, "messages": json.dumps(messages_data)})
            conn.commit()
    except Exception as e:
        logger.error(f"保存会话历史失败: {e}")


def generate_project_context(project_id: int, engine) -> str:
    """
    根据项目ID实时生成项目背景MD（方案C：查询时动态生成）

    Args:
        project_id: 项目ID
        engine: 数据库引擎

    Returns:
        项目背景MD字符串
    """
    from sqlalchemy import text as text

    context = ""

    with get_connection() as conn:
        # 1. 获取项目基本信息
        result = conn.execute(text("""
            SELECT id, name, leader, status, progress, start_date, end_date
            FROM projects
            WHERE id = :pid
        """), {"pid": project_id})

        project = result.fetchone()
        if not project:
            return ""

        # 2. 获取项目任务列表（只取最新版本）
        tasks_result = conn.execute(text("""
            WITH latest_version AS (
                SELECT MAX(CAST(SUBSTRING(task_id FROM 'V([0-9]+)') AS INTEGER)) as max_ver
                FROM project_tasks
                WHERE project_id::integer = :pid AND is_deleted = false
            )
            SELECT task_id, task_name, assignee, status, progress, end_date, actual_end_date
            FROM project_tasks pt, latest_version lv
            WHERE pt.project_id::integer = :pid
              AND pt.is_deleted = false
              AND COALESCE(CAST(SUBSTRING(pt.task_id FROM 'V([0-9]+)') AS INTEGER), 0) = COALESCE(lv.max_ver, 0)
            ORDER BY task_id
        """), {"pid": project_id})

        tasks = []
        for row in tasks_result:
            tasks.append({
                "task_id": row[0],
                "task_name": row[1],
                "assignee": row[2],
                "status": row[3],
                "progress": row[4],
                "end_date": str(row[5]) if row[5] else None,
                "actual_end_date": str(row[6]) if row[6] else None
            })

        # 3. 统计任务状态
        completed = [t for t in tasks if t["status"] == "已完成"]
        ongoing = [t for t in tasks if t["status"] == "进行中"]
        delayed = [t for t in tasks if t["status"] == "延期"]

        # 4. 动态计算项目进度（与前端一致）
        total_tasks = len(tasks)
        completed_count = len(completed)
        avg_progress = sum(float(t["progress"] or 0) for t in tasks) / total_tasks if total_tasks > 0 else 0
        calculated_progress = round((completed_count / total_tasks * 100 + avg_progress) / 2, 1) if total_tasks > 0 else 0

        # 5. 生成MD
        context = f"""# 项目：{project[1]}

## 基本信息
- 项目ID: {project[0]}
- 项目名称: {project[1]}
- 负责人: **{project[2]}**
- 状态: {project[3]}
- 进度: {calculated_progress}%
- 开始日期: {project[5] or '未设置'}
- 结束日期: {project[6] or '未设置'}

## 任务统计
- 总任务数: {len(tasks)}
- 已完成: {len(completed)}
- 进行中: {len(ongoing)}
- 延期: {len(delayed)}

## 任务列表
"""
        for task in tasks[:10]:  # 只显示前10个任务
            context += f"- {task['task_id']}: {task['task_name']}（{task['assignee'] or '未分配'}，{task['status']}，{task['progress']}%）\n"

        if len(tasks) > 10:
            context += f"... 还有 {len(tasks) - 10} 个任务\n"

        if delayed:
            context += "\n## 延期任务\n"
            for task in delayed:
                context += f"- **{task['task_name']}**：截止日期 {task['end_date']}，当前进度 {task['progress']}%\n"

    return context


@app.post("/agent/api/agent/chat")
@limiter.limit("10/minute")  # AI接口限流：每分钟最多10次
async def chat(
    request: Request,
    req: Dict,
    current_user: Dict = Depends(get_current_user)
):
    """
    智能问答接口（支持多轮对话 + RAG文档检索）

    支持自然语言查询项目、任务、风险、工时、文档知识库等
    """
    message = req.get("message", "")
    session_id = req.get("session_id", "default")
    if not message:
        raise HTTPException(status_code=400, detail="请输入问题")

    try:
        # 获取用户信息
        username = current_user.get("username")
        user_info = get_user_info_cache(username)
        employee_name = user_info.get("name", username) if user_info else username

        # 获取会话历史
        session_key = f"{username}_{session_id}"
        history = get_session_history(session_key)

        # ========== 第一步：动态生成项目背景（方案C）==========
        # 根据问题关键词，实时查询相关项目信息
        project_context = ""
        project_names = []

        try:
            engine = get_db_engine()
            from sqlalchemy import text as text

            # 1. 提取项目关键词（智能提取）
            # 尝试多种关键词长度匹配
            keywords_to_try = []

            # 提取问题中可能的项目关键词（去掉常见问题词）
            question_words = ["负责人", "进度", "状态", "任务", "延期", "是谁", "是什么", "如何", "怎么样", "？", "?", "的"]
            cleaned_message = message
            for word in question_words:
                cleaned_message = cleaned_message.replace(word, "")

            # 尝试不同长度的关键词
            if len(cleaned_message) >= 10:
                keywords_to_try.append(cleaned_message[:10])  # 前10个字符
            if len(cleaned_message) >= 8:
                keywords_to_try.append(cleaned_message[:8])   # 前8个字符
            if len(cleaned_message) >= 5:
                keywords_to_try.append(cleaned_message[:5])   # 前5个字符

            # 2. 尝试匹配项目
            matched_projects = []
            with get_connection() as conn:
                for keyword in keywords_to_try:
                    result = conn.execute(text("""
                        SELECT id, name, leader, status, progress
                        FROM projects
                        WHERE is_deleted = false
                          AND name ILIKE :query
                        LIMIT 3
                    """), {"query": f"%{keyword}%"})

                    for row in result:
                        # 避免重复
                        if not any(p["id"] == row[0] for p in matched_projects):
                            matched_projects.append({
                                "id": row[0],
                                "name": row[1],
                                "leader": row[2],
                                "status": row[3],
                                "progress": row[4]
                            })
                            project_names.append(row[1])

                    if matched_projects:
                        break  # 找到匹配就停止

            # 3. 如果匹配到项目，动态生成项目背景
            for project in matched_projects:
                project_context += generate_project_context(project["id"], engine)

        except Exception as e:
            logger.error(f" {e}")

        # ========== 第二步：RAG文档检索（补充知识）==========
        rag_context = ""
        rag_sources = []

        try:
            with get_connection() as conn:
                # 从项目知识库检索上传的文档
                result = conn.execute(text("""
                    SELECT project_name, doc_name, content, summary
                    FROM project_knowledge_base
                    WHERE is_deleted = false
                      AND doc_name != '项目概况'
                      AND (content ILIKE :query OR summary ILIKE :query OR project_name ILIKE :query)
                    LIMIT 3
                """), {"query": f"%{message[:30]}%"})

                for row in result:
                    rag_context += f"\n【文档：{row[0]} - {row[1]}】\n{row[2][:500]}\n"
                    rag_sources.append(f"{row[0]} - {row[1]}")

        except Exception as e:
            logger.error(f" {e}")

        # ========== 第二步：意图识别与工具调用 ==========
        # 构建对话上下文摘要（帮助意图识别理解代词和指代）
        recent_context = ""
        if history:
            recent_context = "\n\n最近对话：\n" + "\n".join([
                f"{'用户' if isinstance(m, HumanMessage) else '助手'}: {m.content[:150]}"
                for m in history[-4:]
            ])

        # 如果匹配到了项目，告诉AI项目ID（方便追问时定位）
        project_hint = ""
        matched_project_ids = [p["id"] for p in matched_projects] if matched_projects else []
        if matched_project_ids:
            project_hint = f"\n\n当前讨论的项目ID: {matched_project_ids[0]}"
            if len(matched_project_ids) > 1:
                project_hint = f"\n\n当前讨论的项目ID列表: {matched_project_ids}"

        analysis_prompt = f"""你是一个项目管理助手的意图识别模块。

用户：{employee_name}
问题：{message}
{recent_context}
{project_hint}
{TOOL_DESCRIPTIONS}

请分析用户意图，以JSON格式返回：
{{
  "tool": "工具名称（如果问题与流程、规范、制度相关，返回'none'）",
  "params": {{参数对象}},
  "summary": "一句话说明你要查什么"
}}

注意：
- 如果问题包含代词（它、它们、这个等），结合最近对话理解指代对象
- 例如：上文讨论"Demo项目"后问"它的进度"，应查询项目进度
- 例如：上文讨论"任务列表"后问"它们谁负责"，应使用 query_project_tasks_by_id 并带上项目ID
- 如果知道项目ID，优先使用 query_project_tasks_by_id 而非 query_tasks

只返回JSON，不要其他内容。"""

        # 构建消息列表（包含历史）
        messages = []
        for msg in history[-6:]:  # 最近3轮（6条消息）
            messages.append(msg)
        messages.append(HumanMessage(content=analysis_prompt))

        analysis_response = await llm_invoke_threaded(messages)
        analysis_text = analysis_response.content.strip()

        # 解析JSON
        tool_name = "none"
        params = {}
        data = {}

        try:
            if "```json" in analysis_text:
                analysis_text = analysis_text.split("```json")[1].split("```")[0].strip()
            elif "```" in analysis_text:
                analysis_text = analysis_text.split("```")[1].split("```")[0].strip()

            analysis = json.loads(analysis_text)
            tool_name = analysis.get("tool", "none")
            params = analysis.get("params", {})
        except:
            pass

        # 执行工具（如果不是纯文档问题）
        if tool_name != "none":
            tool_result = execute_query(tool_name, params)
            data = json.loads(tool_result)

        # ========== 第三步：生成回答 ==========
        # 提取对话中的关键实体（用于上下文理解）
        entities_mentioned = []
        if history:
            # 从最近对话中提取可能的项目名、任务名等
            for m in history[-4:]:
                content = m.content
                # 简单提取：找到引号中的内容、特定项目名等
                import re
                # 匹配 "xxx项目" 或 "xxx任务"
                project_matches = re.findall(r'[""「]([^""」]+)[""」]|(\S+项目)', content)
                for match in project_matches:
                    entity = match[0] or match[1]
                    if entity and entity not in entities_mentioned:
                        entities_mentioned.append(entity)

        # 构建实体提示
        entity_hint = ""
        if entities_mentioned:
            entity_hint = f"\n\n对话中提到的关键实体：{', '.join(entities_mentioned[:3])}"

        history_context = ""
        if history:
            history_context = "\n\n历史对话摘要：\n" + "\n".join([
                f"{'用户' if isinstance(m, HumanMessage) else '助手'}: {m.content[:100]}..."
                for m in history[-4:]
            ])

        # 构建上下文部分（方案C：项目背景 + RAG文档 + 工具结果）
        context_parts = []

        # 1. 项目背景（动态生成）
        if project_context:
            context_parts.append(f"📊 **相关项目信息**：\n{project_context}")

        # 2. RAG文档（补充知识）
        if rag_context:
            context_parts.append(f"📚 **相关文档知识**：{rag_context}")

        # 3. 工具查询结果
        if data and isinstance(data, dict) and not data.get("error"):
            context_parts.append(f"🔎 **查询结果**：\n{json.dumps(data, ensure_ascii=False, indent=2)}")
        elif data and isinstance(data, list):
            context_parts.append(f"🔎 **查询结果**：\n{json.dumps(data, ensure_ascii=False, indent=2)}")

        context_str = "\n\n".join(context_parts) if context_parts else "（未找到相关数据）"

        answer_prompt = f"""你是项目管理助手，帮助用户解答问题。

用户：{employee_name}
问题：{message}
{history_context}
{entity_hint}

{context_str}

请用简洁的自然语言回答用户问题。要点：
1. **优先结合对话上下文**：如果问题包含"它""它们""这个"等代词，根据历史对话理解指代
2. 如果文档中有相关信息，优先基于文档回答
3. 补充数据库中的相关数据
4. 如有风险或异常，主动提示
5. 控制在3-5句话
6. 如果引用了文档，在回答末尾标注来源"""

        final_response = await llm_invoke_threaded([HumanMessage(content=answer_prompt)])

        # 保存会话历史
        history.append(HumanMessage(content=message))
        history.append(final_response)
        save_session_history(session_key, history)

        return {
            "response": final_response.content,
            "session_id": session_id,
            "sources": rag_sources if rag_sources else None
        }

    except Exception as e:
        logger.exception(f" {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"问答失败: {str(e)}")

# ============== 定时任务调度 ==============

from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.cron import CronTrigger

scheduler = AsyncIOScheduler()

async def scheduled_morning_reminder():
    """早间提醒 - 08:00"""
    logger.info(f" 早间提醒 - {datetime.now()}")
    try:
        # text 已从 database 模块导入
        engine = get_db_engine()

        with get_connection() as conn:
            # 查询今日待办和延期任务
            result = conn.execute(text("""
                SELECT assignee_id, assignee, COUNT(*) as task_count
                FROM project_tasks
                WHERE is_deleted = false
                  AND (end_date = CURRENT_DATE OR end_date < CURRENT_DATE)
                  AND actual_end_date IS NULL
                  AND assignee_id IS NOT NULL
                GROUP BY assignee_id, assignee
            """))

            reminders = []
            for row in result:
                reminders.append({
                    "employee_id": row[0],
                    "employee_name": row[1],
                    "task_count": row[2]
                })

            logger.info(f" 需提醒{len(reminders)}人")
            return {"success": True, "count": len(reminders), "reminders": reminders}

    except Exception as e:
        logger.error(f" {e}")
        return {"success": False, "error": str(e)}

async def scheduled_evening_reminder():
    """晚间提醒 - 18:00"""
    logger.info(f" 晚间提醒 - {datetime.now()}")
    try:
        # text 已从 database 模块导入
        engine = get_db_engine()

        with get_connection() as conn:
            # 查询今日未填日报的人员
            result = conn.execute(text("""
                SELECT employee_id, name
                FROM personnel
                WHERE employee_id NOT IN (
                    SELECT employee_id FROM daily_reports
                    WHERE report_date = CURRENT_DATE AND is_deleted = false
                )
                AND is_deleted = false
            """))

            unreported = []
            for row in result:
                unreported.append({
                    "employee_id": row[0],
                    "name": row[1]
                })

            logger.info(f" {len(unreported)}人未填日报")
            return {"success": True, "count": len(unreported), "unreported": unreported}

    except Exception as e:
        logger.error(f" {e}")
        return {"success": False, "error": str(e)}

async def scheduled_risk_alert():
    """风险预警 - 09:00"""
    logger.info(f" 风险预警 - {datetime.now()}")
    try:
        # text 已从 database 模块导入
        engine = get_db_engine()

        with get_connection() as conn:
            # 查询高风险项目
            result = conn.execute(text("""
                SELECT p.id, p.name, p.leader,
                       COUNT(pt.task_id) as delayed_count
                FROM projects p
                JOIN project_tasks pt ON CAST(pt.project_id AS INTEGER) = p.id
                WHERE pt.is_deleted = false
                  AND pt.end_date < CURRENT_DATE
                  AND pt.actual_end_date IS NULL
                GROUP BY p.id, p.name, p.leader
                HAVING COUNT(pt.task_id) >= 3
                ORDER BY delayed_count DESC
            """))

            risks = []
            for row in result:
                risks.append({
                    "project_id": row[0],
                    "project_name": row[1],
                    "leader": row[2],
                    "delayed_count": row[3]
                })

            logger.info(f"[风险预警] {len(risks)}个项目高风险")
            return {"success": True, "count": len(risks), "risks": risks}

    except Exception as e:
        logger.error(f" {e}")
        return {"success": False, "error": str(e)}

# 手动触发API（测试用）
@app.post("/agent/api/agent/scheduler/trigger/{task_name}")
async def trigger_scheduled_task(
    task_name: str,
    current_user: Dict = Depends(get_current_user)
):
    """手动触发定时任务（测试用）"""
    tasks = {
        "morning": scheduled_morning_reminder,
        "evening": scheduled_evening_reminder,
        "risk": scheduled_risk_alert
    }

    if task_name not in tasks:
        raise HTTPException(status_code=400, detail=f"未知任务: {task_name}")

    result = await tasks[task_name]()
    return result

@app.get("/agent/api/agent/scheduler/jobs")
async def list_scheduled_jobs(current_user: Dict = Depends(get_current_user)):
    """列出所有定时任务"""
    jobs = scheduler.get_jobs()
    return {
        "jobs": [
            {
                "id": job.id,
                "next_run": str(job.next_run_time),
                "trigger": str(job.trigger)
            }
            for job in jobs
        ]
    }


# ========== 项目智能问答API ==========

def get_chat_history(project_id: int, user_id: str) -> List:
    """从数据库获取对话历史"""
    session_key = f"project_{project_id}_{user_id}"
    try:
        with get_connection() as conn:
            result = conn.execute(text("""
                SELECT messages FROM chat_sessions 
                WHERE session_key = :key
            """), {"key": session_key})
            row = result.fetchone()
            if row and row[0]:
                # 从 JSON 反序列化为消息对象
                messages_data = row[0] if isinstance(row[0], list) else json.loads(row[0])
                return [
                    HumanMessage(content=m["content"]) if m["type"] == "human"
                    else AIMessage(content=m["content"])
                    for m in messages_data
                ]
    except Exception as e:
        logger.error(f"获取对话历史失败: {e}")
    return []


def save_chat_history(project_id: int, user_id: str, history: List):
    """保存对话历史到数据库"""
    session_key = f"project_{project_id}_{user_id}"
    try:
        # 序列化为 JSON
        messages_data = [
            {"type": "human" if isinstance(m, HumanMessage) else "ai", "content": m.content}
            for m in history
        ]
        
        with get_connection() as conn:
            # 使用 CAST 而非 ::jsonb，避免 SQLAlchemy text() 解析问题
            conn.execute(text("""
                INSERT INTO chat_sessions (session_key, session_type, project_id, user_id, messages, updated_at)
                VALUES (:key, 'project', :pid, :uid, CAST(:messages AS jsonb), NOW())
                ON CONFLICT (session_key) 
                DO UPDATE SET messages = CAST(:messages AS jsonb), updated_at = NOW()
            """), {"key": session_key, "pid": project_id, "uid": user_id, "messages": json.dumps(messages_data)})
            conn.commit()
    except Exception as e:
        logger.error(f"保存对话历史失败: {e}")


# 内存缓存（启动时为空，作为二级缓存）
_project_session_store: Dict[str, List] = {}


@app.post("/agent/api/agent/projects/{project_id}/chat")
async def project_chat(
    project_id: int,
    request: Dict,
    current_user: Dict = Depends(get_current_user)
):
    """
    项目智能问答接口（数据隔离）

    只能查询当前项目相关的数据，确保数据安全
    """
    message = request.get("message", "")
    session_id = request.get("session_id", "default")

    if not message:
        raise HTTPException(status_code=400, detail="请输入问题")

    try:
        # 导入追踪模块
        from .ai_usage_tracker import check_usage_limit, log_ai_usage
        
        # 获取用户信息
        username = current_user.get("username")
        user_id = current_user.get("employee_id", username)
        user_info = get_user_info_cache(username)
        employee_name = user_info.get("name", username) if user_info else username
        
        # ========== 检查调用限制 ==========
        if not await check_usage_limit(user_id, "chat"):
            raise HTTPException(
                status_code=429,
                detail=f"已达智能问答上限（{100}次/天），请明天再试"
            )

        # 获取项目对话历史（优先数据库，回退内存缓存）
        session_key = f"project_{project_id}_{username}_{session_id}"
        history = _project_session_store.get(session_key)
        if history is None:
            # 内存缓存未命中，从数据库加载
            history = get_chat_history(project_id, username)
            _project_session_store[session_key] = history

        # ========== 第一步：动态生成项目背景 ==========
        engine = get_db_engine()
        from sqlalchemy import text as text

        project_context = generate_project_context(project_id, engine)

        if not project_context:
            raise HTTPException(status_code=404, detail="项目不存在")

        # ========== 第二步：查询项目知识库 ==========
        rag_context = ""
        rag_sources = []

        try:
            with get_connection() as conn:
                result = conn.execute(text("""
                    SELECT doc_name, content
                    FROM project_knowledge_base
                    WHERE project_id = :pid
                      AND is_deleted = false
                      AND (content ILIKE :query OR doc_name ILIKE :query)
                    LIMIT 3
                """), {"pid": project_id, "query": f"%{message[:30]}%"})

                for row in result:
                    rag_context += f"\n【文档：{row[0]}】\n{row[1][:500]}\n"
                    rag_sources.append(row[0])
        except Exception as e:
            logger.error(f" {e}")

        # ========== 第三步：意图识别（限制在项目范围内）==========
        project_tool_descriptions = f"""
可用工具（仅限当前项目，project_id={project_id}）：
1. query_project_tasks(status) - 查询项目任务
   参数：status(任务状态：进行中/已完成/延期)

2. query_project_risks() - 查询项目风险

3. query_project_hours(employee_name) - 查询项目工时
   参数：employee_name(员工姓名)

4. none - 如果问题与项目数据无关，直接回答
"""

        analysis_prompt = f"""你是项目管理助手，专门回答关于项目【ID={project_id}】的问题。

用户：{employee_name}
问题：{message}

{project_tool_descriptions}

请分析用户意图，返回JSON：
{{
  "tool": "工具名称或none",
  "params": {{}}
}}
"""

        messages = history[-4:] + [HumanMessage(content=analysis_prompt)]
        analysis_response = await llm_invoke_threaded(messages)
        analysis_text = analysis_response.content.strip()

        # 解析JSON
        tool_name = "none"
        params = {}
        data = {}

        try:
            if "```json" in analysis_text:
                analysis_text = analysis_text.split("```json")[1].split("```")[0].strip()
            elif "```" in analysis_text:
                analysis_text = analysis_text.split("```")[1].split("```")[0].strip()

            analysis = json.loads(analysis_text)
            tool_name = analysis.get("tool", "none")
            params = analysis.get("params", {})
        except Exception as e:
            logger.error(f" {e}")
            tool_name = "none"

        # ========== 第四步：执行工具（项目范围限制）==========
        if tool_name == "query_project_tasks":
            # 查询项目任务
            status_filter = params.get("status", "")

            with get_connection() as conn:
                sql = f"""
                    SELECT task_id, task_name, assignee, status, progress, end_date, actual_end_date
                    FROM project_tasks
                    WHERE CAST(project_id AS INTEGER) = {project_id}
                      AND is_deleted = false
                """

                if status_filter:
                    sql += f" AND status = '{status_filter}'"

                sql += " ORDER BY task_id"

                result = conn.execute(text(sql))
                tasks = []
                for row in result:
                    tasks.append({
                        "task_id": row[0],
                        "task_name": row[1],
                        "assignee": row[2],
                        "status": row[3],
                        "progress": float(row[4] or 0),
                        "end_date": str(row[5]) if row[5] else None,
                        "actual_end_date": str(row[6]) if row[6] else None
                    })

                data = {"tasks": tasks, "total": len(tasks)}

        elif tool_name == "query_project_risks":
            # 查询项目风险
            with get_connection() as conn:
                result = conn.execute(text("""
                    SELECT task_id, task_name, assignee, end_date,
                           CURRENT_DATE - end_date as delay_days
                    FROM project_tasks
                    WHERE CAST(project_id AS INTEGER) = :project_id
                      AND is_deleted = false
                      AND end_date < CURRENT_DATE
                      AND actual_end_date IS NULL
                    ORDER BY delay_days DESC
                """), {"project_id": project_id})

                risks = []
                for row in result:
                    risks.append({
                        "task_id": row[0],
                        "task_name": row[1],
                        "assignee": row[2],
                        "end_date": str(row[3]),
                        "delay_days": row[4]
                    })

                data = {"risks": risks, "total": len(risks)}

        elif tool_name == "query_project_hours":
            # 查询项目工时
            employee_name_filter = params.get("employee_name", "")

            with get_connection() as conn:
                sql = f"""
                    SELECT dwi.assignee, SUM(dwi.hours_spent) as total_hours
                    FROM daily_work_items dwi
                    JOIN daily_reports dr ON dwi.report_id = dr.id
                    WHERE dwi.project_id = {project_id}
                      AND dr.is_deleted = false
                """

                if employee_name_filter:
                    sql += f" AND dwi.assignee LIKE '%{employee_name_filter}%'"

                sql += " GROUP BY dwi.assignee ORDER BY total_hours DESC"

                result = conn.execute(text(sql))
                hours = []
                for row in result:
                    hours.append({
                        "assignee": row[0],
                        "hours": float(row[1] or 0)
                    })

                data = {"hours": hours, "total": len(hours)}

        # ========== 第五步：生成回答 ==========
        context_parts = []

        if project_context:
            context_parts.append(f"📊 **项目背景**：\n{project_context}")

        if rag_context:
            context_parts.append(f"📚 **项目文档**：{rag_context}")

        if data:
            context_parts.append(f"🔎 **查询结果**：\n{json.dumps(data, ensure_ascii=False, indent=2)}")

        context_str = "\n\n".join(context_parts)

        answer_prompt = f"""你是项目管理助手，专门回答关于当前项目的问题。

用户：{employee_name}
问题：{message}

{context_str}

请用简洁的自然语言回答。如果涉及具体数据，请准确引用。"""

        messages = history[-4:] + [HumanMessage(content=answer_prompt)]
        final_response = await llm_invoke_threaded(messages)
        answer = final_response.content

        # 保存对话历史（内存 + 数据库持久化）
        history.append(HumanMessage(content=message))
        history.append(AIMessage(content=answer))
        history = history[-20:]  # 保留最近10轮
        _project_session_store[session_key] = history
        save_chat_history(project_id, username, history)  # 持久化到数据库
        
        # ========== 记录AI调用日志 ==========
        # 估算tokens（实际应从API响应获取）
        input_tokens = len(message) // 2 + 500  # 问题 + 上下文
        output_tokens = len(answer) // 2
        await log_ai_usage(
            user_id=user_id,
            username=employee_name,
            purpose="chat",
            model="deepseek-chat",
            input_tokens=input_tokens,
            output_tokens=output_tokens,
            success=True
        )

        return {
            "success": True,
            "answer": answer,
            "sources": rag_sources,
            "tool_used": tool_name
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f" {e}")
        import traceback
        traceback.print_exc()
        # 记录失败日志
        try:
            await log_ai_usage(
                user_id=user_id,
                username=employee_name,
                purpose="chat",
                model="deepseek-chat",
                success=False,
                error_message=str(e)[:200]
            )
        except:
            pass
        return {
            "success": False,
            "answer": "抱歉，查询出现问题，请稍后重试。",
            "error": str(e)
        }

# 启动时初始化调度器和HTTP客户端
@app.on_event("startup")
async def startup_event():
    global http_client
    
    # 初始化数据库连接池
    engine = get_engine()
    logger.info("[Database] 连接池已初始化")
    
    # 初始化AI调用追踪器
    from .ai_usage_tracker import init_tracker
    init_tracker(engine)
    logger.info("[AI追踪器] 已初始化")
    
    # 初始化HTTP客户端
    http_client = httpx.AsyncClient(timeout=30.0)
    
    # ========== 定时任务：所有 Worker 注册，推送函数内部去重 ==========
    # Gunicorn 3 个 Worker 都会触发 startup
    # 推送函数内部使用数据库锁保证同一时间只执行一次
    
    # 导入推送函数
    from .push_service import push_morning_alerts, push_afternoon_reminder

    # 早间高风险预警汇总 08:00
    scheduler.add_job(
        push_morning_alerts,
        CronTrigger(hour=8, minute=0),
        id="morning_alerts",
        replace_existing=True
    )

    # 下午日报提醒 16:00
    scheduler.add_job(
        push_afternoon_reminder,
        CronTrigger(hour=16, minute=0),
        id="afternoon_reminder",
        replace_existing=True
    )

    scheduler.start()
    logger.info("[调度器] 定时任务已启动（推送去重由数据库锁保证）")
    logger.info("[HTTP客户端] 已初始化")

@app.on_event("shutdown")
async def shutdown_event():
    # 关闭HTTP客户端
    global http_client
    if http_client:
        await http_client.aclose()
        http_client = None
        logger.info("[HTTP客户端] 已关闭")
    
    # 释放数据库连接池
    dispose_engine()
    logger.info("[Database] 连接池已释放")

    scheduler.shutdown()
    logger.info("[调度器] 定时任务已停止")


# ============== 项目知识库 API ==============

from fastapi import UploadFile, File, Form
import shutil

@app.get("/agent/api/agent/knowledge/stats")
async def get_knowledge_stats_api(
    project_id: Optional[int] = None,
    current_user: Dict = Depends(get_current_user)
):
    """
    获取知识库统计信息

    参数：
    - project_id: 项目ID（可选，不传则统计所有项目）
    """
    try:
        from app.knowledge_base import get_knowledge_stats
        stats = get_knowledge_stats(project_id)
        return {
            "success": True,
            "data": stats
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"获取统计信息失败: {str(e)}")


@app.get("/agent/api/agent/knowledge/list")
async def get_knowledge_list_api(
    project_id: Optional[int] = None,
    doc_type: Optional[str] = None,
    limit: int = 20,
    current_user: Dict = Depends(get_current_user)
):
    """
    获取知识库文档列表

    参数：
    - project_id: 项目ID（可选）
    - doc_type: 文档类型（可选）
    - limit: 返回数量（默认20）
    """
    try:
        from app.knowledge_base import get_knowledge_list
        docs = get_knowledge_list(project_id, doc_type, limit)
        return {
            "success": True,
            "data": docs,
            "total": len(docs)
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"获取文档列表失败: {str(e)}")


@app.post("/agent/api/agent/knowledge/upload")
async def upload_document_api(
    project_id: int = Form(...),
    project_name: str = Form(...),
    doc_name: str = Form(...),
    doc_type: str = Form(...),
    file: UploadFile = File(...),
    current_user: Dict = Depends(get_current_user)
):
    """
    上传文档到知识库

    参数：
    - project_id: 项目ID
    - project_name: 项目名称
    - doc_name: 文档名称
    - doc_type: 文档类型（需求文档/设计文档/会议纪要/技术方案）
    - file: 文件（支持PDF/Word/Txt）
    """
    try:
        # 获取文件扩展名
        file_ext = os.path.splitext(file.filename)[1].lower()

        if file_ext not in ['.pdf', '.docx', '.doc', '.txt', '.md']:
            raise HTTPException(status_code=400, detail="不支持的文件格式，仅支持 PDF/Word/Txt/Markdown")

        # 读取文件内容
        file_content = await file.read()

        # 上传者信息
        uploader_id = current_user.get("employee_id", "0001")
        uploader_name = current_user.get("name", "admin")

        # 调用上传函数
        from app.knowledge_base import upload_document
        result = await upload_document(
            project_id=project_id,
            project_name=project_name,
            doc_name=doc_name,
            doc_type=doc_type,
            file_content=file_content,
            file_ext=file_ext,
            uploader_id=uploader_id,
            uploader_name=uploader_name
        )

        return result

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f" {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"上传失败: {str(e)}")


@app.post("/agent/api/agent/knowledge/query")
async def query_knowledge_api(
    request: Dict[str, Any],
    current_user: Dict = Depends(get_current_user)
):
    """
    基于知识库的智能问答

    请求体：
    {
        "question": "需求调研的结论是什么？",
        "project_id": 35  // 可选，不传则查询所有项目
    }
    """
    try:
        question = request.get("question")
        project_id = request.get("project_id")

        if not question:
            raise HTTPException(status_code=400, detail="请输入问题")

        from app.knowledge_base import query_knowledge
        result = await query_knowledge(question, project_id)

        return {
            "success": True,
            "data": result
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f" {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"查询失败: {str(e)}")


@app.delete("/agent/api/agent/knowledge/{doc_id}")
async def delete_document_api(
    doc_id: int,
    current_user: Dict = Depends(get_current_user)
):
    """
    删除知识库文档（软删除）

    参数：
    - doc_id: 文档ID
    """
    try:
        # text 已从 database 模块导入
        from dotenv import load_dotenv
        load_dotenv()
        with get_connection() as conn:
            conn.execute(text("""
                UPDATE project_knowledge_base
                SET is_deleted = true
                WHERE id = :doc_id
            """), {"doc_id": doc_id})
            conn.commit()

        return {
            "success": True,
            "message": "文档已删除"
        }

    except Exception as e:
        logger.error(f" {e}")
        raise HTTPException(status_code=500, detail=f"删除失败: {str(e)}")


# ============== 项目追踪 API（三视图）===============

@app.get("/agent/api/agent/tracking/execution")
async def get_tracking_execution(
    current_user: Dict = Depends(get_current_user)
):
    """
    追踪-执行视图：任务驱动
    
    返回：
    - 我的任务（今日/本周/本月）
    - 进行中任务
    - 近期完成
    """
    try:
        from .tracking_service import get_execution_view
        
        user_id = current_user.get("employee_id", "")
        user_name = current_user.get("name", "")
        role_id = current_user.get("role_id", 0)
        
        data = get_execution_view(user_id, user_name, role_id)
        
        return {"code": 200, "data": data}
    except Exception as e:
        logger.exception(f"获取执行视图失败: {e}")
        raise HTTPException(status_code=500, detail=f"获取失败: {str(e)}")


@app.get("/agent/api/agent/tracking/health")
async def get_tracking_health(
    current_user: Dict = Depends(get_current_user)
):
    """
    追踪-健康视图：风险雷达
    
    返回：
    - 五维度风险雷达
    - 高风险项目 TOP5
    - 趋势预警
    """
    try:
        from .tracking_service import get_health_view
        
        user_id = current_user.get("employee_id", "")
        role_id = current_user.get("role_id", 0)
        
        data = get_health_view(user_id, role_id)
        
        return {"code": 200, "data": data}
    except Exception as e:
        logger.exception(f"获取健康视图失败: {e}")
        raise HTTPException(status_code=500, detail=f"获取失败: {str(e)}")


@app.get("/agent/api/agent/tracking/trace")
async def get_tracking_trace(
    current_user: Dict = Depends(get_current_user)
):
    """
    追踪-溯源视图：数据血缘
    
    返回：
    - 关联率统计
    - 项目关联排行
    - 不可追溯项目
    """
    try:
        from .tracking_service import get_trace_view
        
        user_id = current_user.get("employee_id", "")
        role_id = current_user.get("role_id", 0)
        
        data = get_trace_view(user_id, role_id)
        
        return {"code": 200, "data": data}
    except Exception as e:
        logger.exception(f"获取溯源视图失败: {e}")
        raise HTTPException(status_code=500, detail=f"获取失败: {str(e)}")


# ============== 公共看板 API（独立模块）===============

@app.get("/agent/api/agent/dashboard/overview")
async def get_dashboard_overview_api(
    current_user: Dict = Depends(get_current_user)
):
    """
    获取公共看板概览数据

    返回：
    - 统计数据
    - 健康度排名
    - 最近预警
    """
    try:
        from .dashboard_service import get_dashboard_overview

        role = current_user.get("role", "user")
        user_id = current_user.get("id")

        data = get_dashboard_overview(role=role, user_id=user_id)
        return data

    except Exception as e:
        logger.error(f" {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"获取数据失败: {str(e)}")


@app.get("/agent/api/agent/dashboard/projects")
async def get_dashboard_projects_api(
    current_user: Dict = Depends(get_current_user)
):
    """
    获取看板项目列表（含详细信息和任务数据）
    """
    try:
        # text 已从 database 模块导入
        from dotenv import load_dotenv
        load_dotenv()        
        with get_connection() as conn:
            # 获取项目基本信息
            projects = conn.execute(text("""
                SELECT 
                    p.id, p.name, p.leader, p.status, p.progress,
                    p.start_date, p.end_date,
                    p.contract_amount, p.budget_total_cost, p.actual_total_cost
                FROM projects p
                WHERE p.is_deleted = false
                ORDER BY p.id
            """)).fetchall()
            
            result = []
            for p in projects:
                project_id = p[0]
                
                # 获取项目的最新版本叶子任务（用于计算进度，排除分组父节点）
                tasks = conn.execute(text("""
                    SELECT 
                        task_id, task_name, start_date, end_date, 
                        actual_end_date, progress, status
                    FROM project_tasks
                    WHERE project_id = :pid
                    AND is_latest = true
                    AND is_deleted = false
                    AND end_date IS NOT NULL
                    AND ("isNode" = false OR "isNode" IS NULL)
                    ORDER BY end_date DESC
                """), {"pid": str(project_id)}).fetchall()
                
                # ⚠️ 判断是否有明确计划（有最新版本任务）
                has_plan = len(tasks) > 0
                
                # 确定项目结束时间：优先用项目 end_date，否则取最后一个任务的结束时间
                project_start_date = p[5]
                project_end_date = p[6]
                
                if not project_end_date and tasks:
                    # 取最后一个任务的结束时间
                    latest_task = tasks[0]  # 已按 end_date DESC 排序
                    project_end_date = latest_task[3]  # end_date
                
                # 确定项目开始时间：优先用项目 start_date，否则取第一个任务的开始时间
                if not project_start_date and tasks:
                    # 找最早的任务开始时间
                    task_starts = [t[2] for t in tasks if t[2]]
                    if task_starts:
                        project_start_date = min(task_starts)
                
                # 计算计划进度（基于时间：已过天数/总天数）
                today = datetime.now().date()
                
                if project_start_date and project_end_date:
                    start = datetime.strptime(str(project_start_date), '%Y-%m-%d').date() if isinstance(project_start_date, str) else project_start_date
                    end = datetime.strptime(str(project_end_date), '%Y-%m-%d').date() if isinstance(project_end_date, str) else project_end_date
                    
                    if today <= start:
                        planned_progress = 0.0
                    elif today >= end:
                        planned_progress = 100.0
                    else:
                        total_days = (end - start).days
                        elapsed_days = (today - start).days
                        planned_progress = round(elapsed_days / total_days * 100, 1) if total_days > 0 else 0
                else:
                    planned_progress = 0.0
                
                # 计算实际进度（按任务工期天数计算）
                # 与计划进度保持一致的时间维度
                total_tasks = len(tasks)
                
                if total_tasks > 0:
                    total_work_days = 0
                    completed_work_days = 0
                    
                    for t in tasks:
                        task_start = t[2]  # start_date
                        task_end = t[3]    # end_date
                        actual_end = t[4]  # actual_end_date
                        task_progress = float(t[5] or 0) / 100
                        
                        # 计算任务工期（天）- 含首尾日，与详情页统一
                        if task_start and task_end:
                            start_dt = task_start if isinstance(task_start, type(today)) else datetime.strptime(str(task_start), '%Y-%m-%d').date()
                            end_dt = task_end if isinstance(task_end, type(today)) else datetime.strptime(str(task_end), '%Y-%m-%d').date()
                            work_days = max((end_dt - start_dt).days + 1, 1)  # 含首日，至少1天
                        else:
                            work_days = 5  # 默认5天
                        
                        total_work_days += work_days
                        
                        if task_progress >= 1.0 or actual_end:
                            # 已完成：计入完整工期
                            completed_work_days += work_days
                        elif task_end and task_end < today:
                            # 延期未完成：最高计50%
                            completed_work_days += work_days * min(task_progress, 0.5)
                        else:
                            # 进行中：按进度计入
                            completed_work_days += work_days * task_progress
                    
                    actual_progress = round(completed_work_days / total_work_days * 100, 1) if total_work_days > 0 else 0
                else:
                    # 无任务时，使用项目进度字段
                    actual_progress = float(p[4] or 0)
                
                # 取前10个任务用于显示
                display_tasks = tasks[:10] if len(tasks) > 10 else tasks
                
                # 获取项目预警
                alerts = conn.execute(text("""
                    SELECT alert_type, severity, title, content
                    FROM project_alerts
                    WHERE project_id = :pid
                    AND NOT is_resolved
                    ORDER BY 
                        CASE severity 
                            WHEN 'high' THEN 1 
                            WHEN 'medium' THEN 2 
                            ELSE 3 
                        END
                    LIMIT 3
                """), {"pid": project_id}).fetchall()
                
                result.append({
                    "id": project_id,
                    "name": p[1],
                    "leader": p[2],
                    "status": p[3],
                    "progress": float(p[4] or 0),
                    "planned_progress": planned_progress,
                    "actual_progress": actual_progress,
                    "start_date": str(project_start_date) if project_start_date else None,
                    "end_date": str(project_end_date) if project_end_date else None,
                    "contract_amount": float(p[7] or 0),
                    "budget_total_cost": float(p[8] or 0),
                    "actual_total_cost": float(p[9] or 0),
                    "has_plan": has_plan,  # 新增：是否有明确计划
                    "tasks": [{
                        "task_id": t[0],
                        "task_name": t[1],
                        "start_date": str(t[2]) if t[2] else None,
                        "end_date": str(t[3]) if t[3] else None,
                        "actual_end_date": str(t[4]) if t[4] else None,
                        "progress": float(t[5] or 0),
                        "status": t[6]
                    } for t in display_tasks],
                    "alerts": [{
                        "type": a[0],
                        "severity": a[1],
                        "title": a[2],
                        "content": a[3]
                    } for a in alerts]
                })
            
            return result
    
    except Exception as e:
        logger.error(f" {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"获取失败: {str(e)}")


@app.get("/agent/api/agent/dashboard/projects-grouped")
async def get_dashboard_projects_grouped_api(
    current_user: Dict = Depends(get_current_user)
):
    """
    获取看板项目列表（分组版：有计划的优先）
    
    返回：{
        "with_plan": [...],  # 有明确计划的项目
        "without_plan": [...],  # 无计划的项目
        "total": 总数
    }
    """
    try:
        from .database import get_connection
        with get_connection() as conn:
            # 获取所有项目
            projects = conn.execute(text("""
                SELECT 
                    p.id, p.name, p.leader, p.status, p.progress,
                    p.start_date, p.end_date,
                    p.contract_amount, p.budget_total_cost, p.actual_total_cost
                FROM projects p
                WHERE p.is_deleted = false
                ORDER BY p.id
            """)).fetchall()
            
            with_plan = []
            without_plan = []
            
            for p in projects:
                project_id = p[0]
                
                # 检查是否有最新版本任务
                task_count = conn.execute(text("""
                    SELECT COUNT(*) FROM project_tasks
                    WHERE project_id = :pid
                    AND is_latest = true
                    AND is_deleted = false
                """), {"pid": str(project_id)}).fetchone()[0]
                
                has_plan = task_count > 0
                
                # 获取任务数据（用于计算进度）
                tasks = conn.execute(text("""
                    SELECT 
                        task_id, task_name, start_date, end_date, 
                        actual_end_date, progress, status
                    FROM project_tasks
                    WHERE project_id = :pid
                    AND is_latest = true
                    AND is_deleted = false
                    AND end_date IS NOT NULL
                    AND ("isNode" = false OR "isNode" IS NULL)
                    ORDER BY end_date DESC
                """), {"pid": str(project_id)}).fetchall()
                
                # 计算项目时间
                project_start_date = p[5]
                project_end_date = p[6]
                
                if not project_end_date and tasks:
                    project_end_date = tasks[0][3]
                
                if not project_start_date and tasks:
                    task_starts = [t[2] for t in tasks if t[2]]
                    if task_starts:
                        project_start_date = min(task_starts)
                
                # 计算计划进度
                today = datetime.now().date()
                if project_start_date and project_end_date:
                    start = datetime.strptime(str(project_start_date), '%Y-%m-%d').date() if isinstance(project_start_date, str) else project_start_date
                    end = datetime.strptime(str(project_end_date), '%Y-%m-%d').date() if isinstance(project_end_date, str) else project_end_date
                    
                    if today <= start:
                        planned_progress = 0.0
                    elif today >= end:
                        planned_progress = 100.0
                    else:
                        total_days = (end - start).days
                        elapsed_days = (today - start).days
                        planned_progress = round(elapsed_days / total_days * 100, 1) if total_days > 0 else 0
                else:
                    planned_progress = 0.0
                
                # 计算实际进度
                if len(tasks) > 0:
                    total_work_days = 0
                    completed_work_days = 0
                    
                    for t in tasks:
                        task_start = t[2]
                        task_end = t[3]
                        actual_end = t[4]
                        task_progress = float(t[5] or 0) / 100
                        
                        if task_start and task_end:
                            start_dt = task_start if isinstance(task_start, type(today)) else datetime.strptime(str(task_start), '%Y-%m-%d').date()
                            end_dt = task_end if isinstance(task_end, type(today)) else datetime.strptime(str(task_end), '%Y-%m-%d').date()
                            work_days = max((end_dt - start_dt).days + 1, 1)
                        else:
                            work_days = 5
                        
                        total_work_days += work_days
                        
                        if task_progress >= 1.0 or actual_end:
                            completed_work_days += work_days
                        elif task_end and task_end < today:
                            completed_work_days += work_days * min(task_progress, 0.5)
                        else:
                            completed_work_days += work_days * task_progress
                    
                    actual_progress = round(completed_work_days / total_work_days * 100, 1) if total_work_days > 0 else 0
                else:
                    actual_progress = float(p[4] or 0)
                
                # 构建项目数据
                project_data = {
                    "id": project_id,
                    "name": p[1],
                    "leader": p[2],
                    "status": p[3],
                    "progress": float(p[4] or 0),
                    "planned_progress": planned_progress,
                    "actual_progress": actual_progress,
                    "start_date": str(project_start_date) if project_start_date else None,
                    "end_date": str(project_end_date) if project_end_date else None,
                    "contract_amount": float(p[7] or 0),
                    "budget_total_cost": float(p[8] or 0),
                    "actual_total_cost": float(p[9] or 0),
                    "has_plan": has_plan,
                    "task_count": task_count
                }
                
                if has_plan:
                    with_plan.append(project_data)
                else:
                    without_plan.append(project_data)
            
            return {
                "with_plan": with_plan,
                "without_plan": without_plan,
                "total": len(with_plan) + len(without_plan),
                "tracked_count": len(with_plan),
                "untracked_count": len(without_plan)
            }
    
    except Exception as e:
        logger.error(f" {e}")
        raise HTTPException(status_code=500, detail=f"获取失败: {str(e)}")


@app.get("/agent/api/agent/dashboard/alerts")
async def get_dashboard_alerts_api(
    severity: str = None,
    project_id: int = None,
    current_user: Dict = Depends(get_current_user)
):
    """
    获取预警列表

    参数：
    - severity: 按严重程度过滤 (high/medium/low)
    - project_id: 按项目过滤
    """
    try:
        # text 已从 database 模块导入
        from dotenv import load_dotenv
        load_dotenv()
        sql = """
            SELECT
                a.id, a.project_id, p.name as project_name,
                a.alert_type, a.severity, a.title, a.content, a.details,
                a.created_at, a.is_resolved, a.resolved_at
            FROM project_alerts a
            JOIN projects p ON p.id = a.project_id
            WHERE NOT a.is_resolved
        """
        params = {}

        if severity:
            sql += " AND a.severity = :severity"
            params["severity"] = severity

        if project_id:
            sql += " AND a.project_id = :project_id"
            params["project_id"] = project_id

        sql += " ORDER BY CASE a.severity WHEN 'high' THEN 1 WHEN 'medium' THEN 2 ELSE 3 END, a.created_at DESC"

        with get_connection() as conn:
            alerts = conn.execute(text(sql), params).fetchall()

            return [{
                "id": a[0],
                "project_id": a[1],
                "project_name": a[2],
                "alert_type": a[3],
                "severity": a[4],
                "title": a[5],
                "content": a[6],
                "details": a[7],
                "created_at": str(a[8]),
                "is_resolved": a[9],
                "resolved_at": str(a[10]) if a[10] else None
            } for a in alerts]

    except Exception as e:
        logger.error(f" {e}")
        raise HTTPException(status_code=500, detail=f"获取失败: {str(e)}")


@app.post("/agent/api/agent/dashboard/alerts/{alert_id}/resolve")
async def resolve_alert_api(
    alert_id: int,
    current_user: Dict = Depends(get_current_user)
):
    """
    标记预警已处理
    """
    try:
        from .dashboard_service import resolve_alert

        user_id = current_user.get("id")
        resolve_alert(alert_id, user_id)

        return {"success": True, "message": "预警已标记为已处理"}

    except Exception as e:
        logger.error(f" {e}")
        raise HTTPException(status_code=500, detail=f"处理失败: {str(e)}")


@app.get("/agent/api/agent/dashboard/alert-rules")
async def get_alert_rules_api(
    current_user: Dict = Depends(require_role(["admin"]))
):
    """
    获取预警规则配置（仅 admin）
    """
    try:
        from .dashboard_service import get_alert_rules
        return get_alert_rules()

    except Exception as e:
        logger.error(f" {e}")
        raise HTTPException(status_code=500, detail=f"获取失败: {str(e)}")


@app.put("/agent/api/agent/dashboard/alert-rules/{rule_id}")
async def update_alert_rule_api(
    rule_id: int,
    enabled: bool = None,
    thresholds: Dict = None,
    current_user: Dict = Depends(require_role(["admin"]))
):
    """
    更新预警规则配置（仅 admin）
    """
    try:
        from .dashboard_service import update_alert_rule
        update_alert_rule(rule_id, enabled, thresholds)
        return {"success": True, "message": "规则已更新"}

    except Exception as e:
        logger.error(f" {e}")
        raise HTTPException(status_code=500, detail=f"更新失败: {str(e)}")


@app.get("/agent/api/agent/dashboard/health/{project_id}/trend")
async def get_health_trend_api(
    project_id: int,
    days: int = 30,
    current_user: Dict = Depends(get_current_user)
):
    """
    获取项目健康度趋势
    """
    try:
        from .dashboard_service import get_project_health_trend
        return get_project_health_trend(project_id, days)

    except Exception as e:
        logger.error(f" {e}")
        raise HTTPException(status_code=500, detail=f"获取失败: {str(e)}")


@app.get("/agent/api/agent/dashboard/insight")
async def get_ai_insight_api(
    current_user: Dict = Depends(get_current_user)
):
    """
    获取 AI 洞察（读取缓存）
    
    洞察由定时任务在凌晨和中午各生成一次，前端直接读取缓存
    """
    try:
        from .ai_insight_service import get_latest_insight_from_db
        from datetime import date
        
        # 从数据库获取最新洞察
        insight = get_latest_insight_from_db()
        
        if insight:
            return {
                "content": insight["content"],
                "period": insight["period"],
                "generated_at": insight["created_at"].isoformat() if insight["created_at"] else None,
                "cached": True
            }
        
        # 如果今天还没有洞察，返回提示
        return {
            "content": "今日洞察尚未生成，请稍后再试。",
            "cached": False
        }

    except Exception as e:
        logger.error(f"获取AI洞察失败: {e}")
        raise HTTPException(status_code=500, detail=f"获取失败: {str(e)}")


@app.post("/agent/api/agent/dashboard/insight/generate")
async def generate_insight_api(
    period: str = "morning",
    current_user: Dict = Depends(require_role(["admin"]))
):
    """
    手动触发洞察生成（定时任务调用）
    
    参数：
    - period: "morning" 或 "noon"
    
    仅管理员可调用
    
    注意：此接口会阻塞约40秒（调用本地模型润色）
    """
    try:
        from .ai_insight_service import generate_ai_insight_with_polish, save_insight_to_db
        
        # 生成洞察（同步执行，约40秒）
        insight_data = await generate_ai_insight_with_polish(period)
        
        # 保存到数据库
        record_id = save_insight_to_db(insight_data)
        
        logger.info(f"[定时任务] AI洞察生成完成，ID: {record_id}, 时段: {period}")
        
        return {
            "success": True,
            "id": record_id,
            "period": period,
            "content": insight_data["polished"],
            "raw_content": insight_data["raw"],
            "generated_at": insight_data["generated_at"]
        }

    except Exception as e:
        logger.error(f"生成AI洞察失败: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"生成失败: {str(e)}")


async def generate_ai_insight() -> str:
    """生成 AI 洞察（从进度、风险、成本三方面分析）"""
    # text 已从 database 模块导入
    from dotenv import load_dotenv
    from datetime import date
    
    load_dotenv()    
    with get_connection() as conn:
        # 获取项目进度统计
        progress_stats = conn.execute(text("""
            SELECT 
                COUNT(*) as total,
                COUNT(*) FILTER (WHERE status = '进行中') as ongoing,
                COUNT(*) FILTER (WHERE status = '已完成') as completed,
                AVG(progress) FILTER (WHERE status = '进行中') as avg_progress,
                COUNT(*) FILTER (WHERE progress < 50 AND status = '进行中') as low_progress
            FROM projects WHERE is_deleted = false
        """)).fetchone()
        
        # 获取风险统计
        risk_stats = conn.execute(text("""
            SELECT 
                COUNT(*) FILTER (WHERE severity = 'high' AND NOT is_resolved) as high,
                COUNT(*) FILTER (WHERE severity = 'medium' AND NOT is_resolved) as medium,
                COUNT(*) FILTER (WHERE severity = 'low' AND NOT is_resolved) as low
            FROM project_alerts
            WHERE created_at >= CURRENT_DATE - INTERVAL '7 days'
        """)).fetchone()
        
        # 获取成本统计
        cost_stats = conn.execute(text("""
            SELECT 
                SUM(budget_total_cost) as total_budget,
                SUM(actual_total_cost) as total_actual,
                COUNT(*) FILTER (WHERE actual_total_cost > budget_total_cost * 1.1) as overspent
            FROM projects WHERE is_deleted = false
        """)).fetchone()
        
        # 获取延期项目
        delayed_projects = conn.execute(text("""
            SELECT name, progress 
            FROM projects 
            WHERE is_deleted = false 
            AND status = '进行中' 
            AND progress < 100
            ORDER BY progress ASC
            LIMIT 3
        """)).fetchall()
        
        # 获取高成本超支项目
        overspent_projects = conn.execute(text("""
            SELECT name, 
                   (actual_total_cost - budget_total_cost) as overspent,
                   (actual_total_cost / NULLIF(budget_total_cost, 0) * 100) as overspent_pct
            FROM projects 
            WHERE is_deleted = false 
            AND actual_total_cost > budget_total_cost * 1.1
            ORDER BY (actual_total_cost / NULLIF(budget_total_cost, 0)) DESC
            LIMIT 3
        """)).fetchall()
        
        # 获取已开始的进行中项目数（用于成本分析）
        started_projects_count = conn.execute(text("""
            SELECT COUNT(*) FROM projects 
            WHERE is_deleted = false 
            AND status = '进行中' 
            AND start_date <= CURRENT_DATE
        """)).fetchone()
    
    # 构建分析结果
    total = progress_stats[0] or 0
    ongoing = progress_stats[1] or 0
    completed = progress_stats[2] or 0
    avg_progress = float(progress_stats[3] or 0)
    low_progress = progress_stats[4] or 0
    
    high_risk = risk_stats[0] or 0
    medium_risk = risk_stats[1] or 0
    
    total_budget = float(cost_stats[0] or 0)
    total_actual = float(cost_stats[1] or 0)
    overspent = cost_stats[2] or 0
    started_count = started_projects_count[0] if started_projects_count else 0
    
    # 生成洞察内容
    lines = []
    
    # 项目进度分析
    lines.append(f"📊 【项目进度】进行中 {ongoing} 个，平均进度 {avg_progress:.1f}%，已完成 {completed} 个")
    if low_progress > 0:
        lines.append(f"   ⚠️ {low_progress} 个项目进度低于50%，需要加快")
    if delayed_projects:
        lines.append(f"   📌 低进度项目：{', '.join([f'{p[0]}({p[1]}%)' for p in delayed_projects])}")
    
    lines.append("")
    
    # 风险预警分析
    lines.append(f"🚨 【风险预警】高风险 {high_risk} 个，中风险 {medium_risk} 个")
    if high_risk > 0:
        lines.append("   ⚠️ 存在高风险预警，建议立即处理")
    else:
        lines.append("   ✅ 暂无高风险预警")
    
    lines.append("")
    
    # 成本支出分析
    lines.append("")
    if total_budget > 0:
        cost_rate = (total_actual / total_budget * 100)
        lines.append(f"💰 【成本支出】预算 ¥{total_budget/10000:.1f}万，实际支出 ¥{total_actual/10000:.1f}万（{cost_rate:.1f}%）")
        
        # 深度分析
        if total_actual == 0:
            if started_count > 0:
                lines.append("   ⚠️ 有进行中项目但无成本记录，可能存在数据缺失或成本未及时录入")
                lines.append("   📌 建议：检查项目成本填报情况，确保数据完整")
            else:
                lines.append("   📊 暂无成本支出，项目可能处于筹备阶段")
                lines.append("   📌 建议：关注项目启动后的成本录入")
        elif overspent > 0:
            lines.append(f"   ⚠️ {overspent} 个项目超支10%以上")
            if overspent_projects:
                lines.append(f"   📌 超支项目：{', '.join([f'{p[0]}(+{(p[2] or 0)-100:.0f}%)' for p in overspent_projects])}")
            lines.append("   📌 建议：加强成本管控，防止进一步超支")
        elif cost_rate < 50:
            lines.append("   ✅ 成本支出低于预算50%，项目进展初期或预算充足")
        else:
            lines.append("   ✅ 成本控制良好，支出在预算范围内")
    else:
        lines.append(f"💰 【成本支出】总支出 ¥{total_actual/10000:.1f}万")
        if total_actual == 0:
            lines.append("   📊 暂无成本数据，可能是新项目或数据未录入")
    
    lines.append("")
    
    # 总结建议
    lines.append("💡 【建议】")
    if low_progress > 0 or high_risk > 0:
        lines.append("   1. 关注低进度项目，协调资源加快进展")
        lines.append("   2. 优先处理高风险预警，降低项目风险")
    else:
        lines.append("   1. 各项目进展正常，继续保持")
    
    if overspent > 0:
        lines.append("   3. 加强成本管控，防止进一步超支")
    
    return "\n".join(lines)


@app.post("/agent/api/agent/dashboard/run-detection")
async def run_detection_api(
    current_user: Dict = Depends(require_role(["admin"]))
):
    """
    手动触发预警检测（仅 admin）
    """
    try:
        from .dashboard_service import run_daily_alert_detection

        count = run_daily_alert_detection()

        return {
            "success": True,
            "message": f"已完成 {count} 个项目的预警检测"
        }

    except Exception as e:
        logger.error(f" {e}")
        raise HTTPException(status_code=500, detail=f"检测失败: {str(e)}")


@app.post("/agent/api/agent/dashboard/test-push")
async def test_push_api(
    current_user: Dict = Depends(require_role(["admin"]))
):
    """
    测试推送（仅 admin）
    """
    try:
        from .push_service import push_to_wechat

        result = push_to_wechat(
            title="🔔 测试推送",
            content="<h3>推送测试成功</h3><p>这是一条来自项目智能体的测试消息</p>"
        )

        return {
            "success": result,
            "message": "推送成功" if result else "推送失败"
        }

    except Exception as e:
        logger.error(f" {e}")
        raise HTTPException(status_code=500, detail=f"推送失败: {str(e)}")


@app.post("/agent/api/agent/dashboard/test-morning-push")
async def test_morning_push_api(
    current_user: Dict = Depends(require_role(["admin"]))
):
    """
    测试早上高风险预警推送（仅 admin）
    """
    try:
        from .push_service import push_morning_alerts

        result = push_morning_alerts()

        return {
            "success": result,
            "message": "推送成功" if result else "推送失败或无高风险预警"
        }

    except Exception as e:
        logger.error(f" {e}")
        raise HTTPException(status_code=500, detail=f"推送失败: {str(e)}")


@app.post("/agent/api/agent/dashboard/test-afternoon-push")
async def test_afternoon_push_api(
    current_user: Dict = Depends(require_role(["admin"]))
):
    """
    测试下午日报提醒推送（仅 admin）
    """
    try:
        from .push_service import push_afternoon_reminder

        result = push_afternoon_reminder()

        return {
            "success": result,
            "message": "推送成功" if result else "推送失败"
        }

    except Exception as e:
        logger.error(f" {e}")
        raise HTTPException(status_code=500, detail=f"推送失败: {str(e)}")


# ============== 成本数据智能导入 ==============

@app.post("/agent/api/agent/cost/import/analyze")
async def analyze_cost_excel(
    file: UploadFile = File(...),
    current_user: Dict = Depends(require_role(["admin", "project_manager"]))
):
    """
    分析Excel文件结构
    """
    try:
        from .cost_import import analyze_excel_structure
        
        content = await file.read()
        result = analyze_excel_structure(content, file.filename)
        
        return {
            "success": True,
            "data": result
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"分析失败: {str(e)}")


@app.post("/agent/api/agent/cost/import/identify")
async def identify_cost_columns(
    request: dict,
    current_user: Dict = Depends(require_role(["admin", "project_manager"]))
):
    """
    AI识别列含义
    """
    try:
        from .cost_import import ai_identify_columns
        
        columns = request.get("columns", [])
        sample_data = request.get("sample_data", [])
        
        result = ai_identify_columns(columns, sample_data)
        
        return {
            "success": True,
            "data": result
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"识别失败: {str(e)}")


@app.post("/agent/api/agent/cost/import/preview")
async def preview_cost_import(
    request: dict,
    current_user: Dict = Depends(require_role(["admin", "project_manager"]))
):
    """
    预览导入结果
    """
    try:
        from .cost_import import preview_import
        
        file_content = bytes(request.get("file_content", []))
        file_name = request.get("file_name", "")
        sheet_name = request.get("sheet_name", "")
        column_mapping = request.get("column_mapping", {})
        
        with get_connection() as conn:
            result = preview_import(file_content, file_name, sheet_name, column_mapping, conn)
        
        return {
            "success": True,
            "data": result
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"预览失败: {str(e)}")


@app.post("/agent/api/agent/cost/import/execute")
async def execute_cost_import(
    request: dict,
    current_user: Dict = Depends(require_role(["admin", "project_manager"]))
):
    """
    执行导入
    """
    try:
        from .cost_import import import_cost_data
        
        file_content = bytes(request.get("file_content", []))
        file_name = request.get("file_name", "")
        sheet_name = request.get("sheet_name", "")
        column_mapping = request.get("column_mapping", {})
        cost_type = request.get("cost_type", "")
        cost_subtype = request.get("cost_subtype", "")
        
        with get_connection() as conn:
            result = import_cost_data(
                file_content, file_name, sheet_name, 
                column_mapping, cost_type, cost_subtype, conn
            )
        
        return {
            "success": result["success"],
            "data": result
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"导入失败: {str(e)}")


@app.get("/agent/api/agent/cost/types")
async def get_cost_types(
    current_user: Dict = Depends(require_role(["admin", "project_manager"]))
):
    """
    获取成本类型列表
    """
    try:
        with get_connection() as conn:
            # 间接成本类型
            indirect_result = conn.execute(text("""
                SELECT id, type_name, description 
                FROM indirect_cost_types 
                WHERE is_deleted = false
                ORDER BY id
            """))
            indirect_types = [dict(row._mapping) for row in indirect_result]
            
            # 外包服务类型
            outsourcing_result = conn.execute(text("""
                SELECT id, type_name, description 
                FROM outsourcing_service_types 
                WHERE is_deleted = false
                ORDER BY id
            """))
            outsourcing_types = [dict(row._mapping) for row in outsourcing_result]
        
        return {
            "success": True,
            "data": {
                "indirect": indirect_types,
                "outsourcing": outsourcing_types
            }
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"获取失败: {str(e)}")


# ============== 智能周报生成 ==============

# 数据库连接辅助函数（已废弃，改用 database 模块）
def get_db():
    """获取数据库引擎（使用 database 模块的全局单例）"""
    return get_engine()

@app.get("/agent/api/agent/weekly-reports")
async def get_weekly_reports(
    page: int = 1,
    size: int = 10,
    project_id: str = None,
    current_user: Dict = Depends(get_current_user)
):
    """
    获取周报列表
    """
    try:
        # text 已从 database 模块导入        
        # 获取当前用户
        username = current_user.get("username") or current_user.get("sub")
        employee_id = current_user.get("employee_id") or username
        
        offset = (page - 1) * size
        
        with get_connection() as conn:
            # 构建查询条件（只查询当前用户的周报）
            where_clause = "WHERE is_deleted = false AND created_by = :created_by"
            params = {"created_by": employee_id}
            
            if project_id:
                where_clause += " AND project_id = :project_id"
                params["project_id"] = project_id
            
            # 安全构建WHERE子句（只允许预定义条件）
            safe_where = where_clause if where_clause and where_clause.startswith("WHERE") else ""
            
            # 查询总数
            count_result = conn.execute(text(f"""
                SELECT COUNT(*) FROM weekly_reports {safe_where}
            """), params)
            total = count_result.fetchone()[0]
            
            # 查询列表
            params["offset"] = offset
            params["size"] = size
            result = conn.execute(text(f"""
                SELECT id, project_id, project_name, week_start, week_end, 
                       total_hours, task_count, created_at, created_by
                FROM weekly_reports
                {safe_where}
                ORDER BY week_start DESC
                OFFSET :offset LIMIT :size
            """), params)
            
            reports = []
            for row in result:
                reports.append({
                    "id": row[0],
                    "project_id": row[1],
                    "project_name": row[2],
                    "week_start": str(row[3]),
                    "week_end": str(row[4]),
                    "total_hours": float(row[5]) if row[5] else 0,
                    "task_count": row[6],
                    "created_at": str(row[7]) if row[7] else None,
                    "created_by": row[8]
                })
        
        return {
            "success": True,
            "data": {
                "items": reports,
                "total": total,
                "page": page,
                "size": size
            }
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"获取周报列表失败: {str(e)}")


@app.post("/agent/api/agent/weekly-reports/generate")
async def generate_weekly_report(
    request: dict,
    current_user: Dict = Depends(require_role(["admin", "project_manager"]))
):
    """
    生成智能周报（个人周报，只包含当前用户的日报数据）
    
    Args:
        project_id: 项目ID（可选，不传则生成全部项目周报）
        week_start: 周开始日期（可选，默认上周一）
        week_end: 周结束日期（可选，默认上周日）
    """
    try:
        project_id = request.get("project_id")
        
        # 获取当前用户信息
        username = current_user.get("username") or current_user.get("sub")
        employee_id = current_user.get("employee_id") or username
        
        # 计算上周日期范围
        today = datetime.now()
        days_since_monday = today.weekday()
        this_monday = today - timedelta(days=days_since_monday)
        last_monday = this_monday - timedelta(days=7)
        last_sunday = last_monday + timedelta(days=6)
        
        week_start = request.get("week_start", str(last_monday.date()))
        week_end = request.get("week_end", str(last_sunday.date()))
        
        # text 已从 database 模块导入        
        with get_connection() as conn:
            # 获取当前用户的日报数据
            if project_id:
                # 单个项目
                result = conn.execute(text("""
                    SELECT dr.report_date, dr.employee_name, dwi.project_id, 
                           dwi.project_name, dwi.task_id, dwi.task_name,
                           dwi.work_content, dwi.hours_spent, dwi.progress_percentage
                    FROM daily_reports dr
                    JOIN daily_work_items dwi ON dr.id = dwi.report_id
                    WHERE dr.report_date BETWEEN :week_start AND :week_end
                    AND dr.is_deleted = false
                    AND dr.employee_id = :employee_id
                    AND dwi.project_id = :project_id
                    ORDER BY dr.report_date, dr.employee_name
                """), {"week_start": week_start, "week_end": week_end, "employee_id": employee_id, "project_id": project_id})
            else:
                # 当前用户全部项目
                result = conn.execute(text("""
                    SELECT dr.report_date, dr.employee_name, dwi.project_id, 
                           dwi.project_name, dwi.task_id, dwi.task_name,
                           dwi.work_content, dwi.hours_spent, dwi.progress_percentage
                    FROM daily_reports dr
                    JOIN daily_work_items dwi ON dr.id = dwi.report_id
                    WHERE dr.report_date BETWEEN :week_start AND :week_end
                    AND dr.is_deleted = false
                    AND dr.employee_id = :employee_id
                    ORDER BY dwi.project_name, dr.report_date
                """), {"week_start": week_start, "week_end": week_end, "employee_id": employee_id})
            
            # 整理数据
            daily_data = []
            project_stats = {}
            employee_name_from_data = None  # 从日报中获取用户姓名
            
            for row in result:
                if not employee_name_from_data and row[1]:
                    employee_name_from_data = row[1]
                    
                daily_data.append({
                    "date": str(row[0]),
                    "employee": row[1],
                    "project_id": row[2],
                    "project_name": row[3] or "未分配项目",
                    "task_id": row[4],
                    "task_name": row[5],
                    "content": row[6],
                    "hours": float(row[7]) if row[7] else 0,
                    "progress": float(row[8]) if row[8] else 0
                })
                
                # 统计项目数据
                pid = row[2] or "unknown"
                if pid not in project_stats:
                    project_stats[pid] = {
                        "name": row[3] or "未分配项目",
                        "total_hours": 0,
                        "task_count": 0
                    }
                project_stats[pid]["total_hours"] += float(row[7]) if row[7] else 0
                project_stats[pid]["task_count"] += 1
            
            if not daily_data:
                return {
                    "success": False,
                    "message": f"该时间段({week_start} ~ {week_end})没有日报数据"
                }
            
            # 使用 DeepSeek AI 生成周报摘要（线程池执行，不阻塞Worker）
            from langchain_deepseek import ChatDeepSeek
            
            api_key = os.getenv("DEEPSEEK_API_KEY", "")
            
            weekly_llm = ChatDeepSeek(
                model="deepseek-chat",
                api_key=api_key,
                temperature=0.3
            )
            
            # 准备提示词
            prompt = f"""根据以下日报数据，生成一份简洁专业的项目周报摘要。

时间范围：{week_start} 至 {week_end}

日报数据：
{json.dumps(daily_data[:50], ensure_ascii=False, indent=2)}
...（共 {len(daily_data)} 条记录）

项目统计：
{json.dumps(project_stats, ensure_ascii=False, indent=2)}

请生成周报，包含以下内容：
1. 本周工作概述（100字以内）
2. 各项目进展摘要（每个项目50字以内）
3. 下周重点关注事项（基于未完成任务和延期风险）
4. 整体工时统计

请直接返回JSON格式：
{{
    "summary": "本周工作概述...",
    "project_progress": [
        {{"name": "项目名", "progress": "进展描述", "hours": 工时}}
    ],
    "next_week_focus": ["事项1", "事项2"],
    "total_hours": 总工时,
    "highlights": ["亮点1", "亮点2"]
}}
"""

            response = await run_in_thread(weekly_llm.invoke, [HumanMessage(content=prompt)])
            content = response.content.strip()
            
            # 解析 AI 返回
            if "```json" in content:
                content = content.split("```json")[1].split("```")[0].strip()
            elif "```" in content:
                content = content.split("```")[1].split("```")[0].strip()
            
            ai_result = json.loads(content)
            
            # 保存到数据库
            username = current_user.get("username", "system")
            # 优先使用日报中的姓名，其次使用 current_user 中的信息
            employee_name = employee_name_from_data or current_user.get("name") or current_user.get("employee_name") or username
            saved_reports = []
            
            for pid, stats in project_stats.items():
                # 对于没有 project_id 的记录，使用 project_name 或 "个人工作" 作为标识
                save_pid = pid if pid != "unknown" else f"personal_{username}"
                save_pname = stats["name"] if stats["name"] != "未分配项目" else f"{employee_name} 个人工作"
                    
                # 检查是否已存在
                existing = conn.execute(text("""
                    SELECT id FROM weekly_reports
                    WHERE project_id = :project_id AND week_start = :week_start
                    AND is_deleted = false
                """), {"project_id": save_pid, "week_start": week_start})
                
                if existing.fetchone():
                    # 更新
                    conn.execute(text("""
                        UPDATE weekly_reports
                        SET summary = :summary, total_hours = :hours,
                            task_count = :count, updated_at = NOW()
                        WHERE project_id = :project_id AND week_start = :week_start
                    """), {
                        "summary": ai_result.get("summary", ""),
                        "hours": stats["total_hours"],
                        "count": stats["task_count"],
                        "project_id": save_pid,
                        "week_start": week_start
                    })
                else:
                    # 新增
                    conn.execute(text("""
                        INSERT INTO weekly_reports
                        (project_id, project_name, week_start, week_end, summary,
                         total_hours, task_count, ai_analysis, created_at, created_by, is_deleted)
                        VALUES (:project_id, :project_name, :week_start, :week_end, :summary,
                         :hours, :count, :analysis, NOW(), :created_by, false)
                    """), {
                        "project_id": save_pid,
                        "project_name": save_pname,
                        "week_start": week_start,
                        "week_end": week_end,
                        "summary": ai_result.get("summary", ""),
                        "hours": stats["total_hours"],
                        "count": stats["task_count"],
                        "analysis": json.dumps(ai_result, ensure_ascii=False),
                        "created_by": username
                    })
                
                saved_reports.append({
                    "project_id": save_pid,
                    "project_name": save_pname,
                    "week_start": week_start,
                    "week_end": week_end,
                    "total_hours": stats["total_hours"],
                    "task_count": stats["task_count"]
                })
            
            conn.commit()
        
        return {
            "success": True,
            "data": {
                "week_start": week_start,
                "week_end": week_end,
                "reports": saved_reports,
                "ai_analysis": ai_result,
                "daily_count": len(daily_data)
            }
        }
        
    except json.JSONDecodeError as e:
        raise HTTPException(status_code=500, detail=f"AI返回格式错误: {str(e)}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"生成周报失败: {str(e)}")


@app.get("/agent/api/agent/weekly-reports/{report_id}")
async def get_weekly_report_detail(
    report_id: int,
    current_user: Dict = Depends(get_current_user)
):
    """
    获取周报详情
    """
    try:
        # text 已从 database 模块导入        
        with get_connection() as conn:
            result = conn.execute(text("""
                SELECT id, project_id, project_name, week_start, week_end,
                       summary, total_hours, task_count, ai_analysis,
                       created_at, created_by
                FROM weekly_reports
                WHERE id = :id AND is_deleted = false
            """), {"id": report_id})
            
            row = result.fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="周报不存在")
            
            # 获取该周的日报明细
            daily_result = conn.execute(text("""
                SELECT dr.report_date, dr.employee_name, dwi.work_content, 
                       dwi.hours_spent, dwi.progress_percentage
                FROM daily_reports dr
                JOIN daily_work_items dwi ON dr.id = dwi.report_id
                WHERE dr.report_date BETWEEN :week_start AND :week_end
                AND dwi.project_id = :project_id
                AND dr.is_deleted = false
                ORDER BY dr.report_date
            """), {"week_start": str(row[3]), "week_end": str(row[4]), "project_id": row[1]})
            
            daily_items = []
            for d in daily_result:
                daily_items.append({
                    "date": str(d[0]),
                    "employee": d[1],
                    "content": d[2],
                    "hours": float(d[3]) if d[3] else 0,
                    "progress": float(d[4]) if d[4] else 0
                })
        
        return {
            "success": True,
            "data": {
                "id": row[0],
                "project_id": row[1],
                "project_name": row[2],
                "week_start": str(row[3]),
                "week_end": str(row[4]),
                "summary": row[5],
                "total_hours": float(row[6]) if row[6] else 0,
                "task_count": row[7],
                "ai_analysis": row[8] if row[8] else {},
                "created_at": str(row[9]) if row[9] else None,
                "created_by": row[10],
                "daily_items": daily_items
            }
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"获取周报详情失败: {str(e)}")


# ============== 启动 ==============

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=3000)


# ============== Excel文件下载接口 ==============

@app.get("/agent/api/agent/plans/file/{version_id}")
async def download_plan_file(version_id: int, current_user: Dict = Depends(get_current_user)):
    """
    下载/预览计划Excel文件
    
    返回文件流供前端SheetJS解析
    """
    with get_connection() as conn:
        result = conn.execute(text("""
            SELECT file_name, file_path 
            FROM project_plan_versions 
            WHERE id = :version_id
        """), {"version_id": version_id}).fetchone()
        
        if not result:
            raise HTTPException(status_code=404, detail="版本不存在")
        
        file_name = result[0]
        file_path = result[1]
        
        if not file_path or not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail="文件不存在或已删除")
        
        from fastapi.responses import FileResponse
        return FileResponse(
            path=file_path,
            filename=file_name,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

# ==================== 质量管理 API ====================

@app.get("/agent/api/agent/quality/overview")
async def get_quality_overview(current_user: dict = Depends(get_current_user)):
    """
    质量概览 - 六西格玛指标
    - DPMO 计算
    - 西格玛水平
    - 缺陷分布
    """
    from sqlalchemy import text
    try:
        with get_connection() as conn:
            # 1. 统计总任务数（机会数）- 动态计算状态
            task_sql = text("""
                SELECT 
                    COUNT(*) as total,
                    SUM(CASE WHEN progress >= 100 THEN 1 ELSE 0 END) as completed,
                    SUM(CASE WHEN progress > 0 AND progress < 100 THEN 1 ELSE 0 END) as ongoing,
                    SUM(CASE WHEN (progress = 0 OR progress IS NULL) AND (start_date IS NULL OR start_date > CURRENT_DATE) THEN 1 ELSE 0 END) as pending,
                    SUM(CASE WHEN end_date < CURRENT_DATE AND (progress < 100 OR progress IS NULL) THEN 1 ELSE 0 END) as delayed
                FROM project_tasks
                WHERE is_latest = true AND is_deleted = false
            """)
            task_result = conn.execute(task_sql).fetchone()
            total_tasks = task_result[0] or 1
            completed_tasks = task_result[1] or 0
            ongoing_tasks = task_result[2] or 0
            pending_tasks = task_result[3] or 0
            delayed_tasks = task_result[4] or 0
            
            # 2. 统计超支项目数
            cost_sql = text("""
                SELECT COUNT(*) as total,
                       SUM(CASE WHEN actual_total_cost > budget_total_cost * 1.1 THEN 1 ELSE 0 END) as overbudget
                FROM projects
                WHERE is_deleted = false
                  AND budget_total_cost > 0
                  AND actual_total_cost > 0
            """)
            cost_result = conn.execute(cost_sql).fetchone()
            cost_projects = cost_result[0] or 1
            overbudget_projects = cost_result[1] or 0
            
            # 3. 缺陷定义 - 动态计算严重延期
            # - 延期 > 3天的任务 = 缺陷
            defect_task_sql = text("""
                SELECT COUNT(*) 
                FROM project_tasks
                WHERE is_latest = true 
                  AND is_deleted = false
                  AND end_date < CURRENT_DATE - INTERVAL '3 days'
                  AND (progress < 100 OR progress IS NULL)
            """)
            severe_delayed = conn.execute(defect_task_sql).scalar() or 0
            
            # 4. 计算总机会数和缺陷数
            total_opportunities = total_tasks + cost_projects
            total_defects = severe_delayed + overbudget_projects
            
            # 5. 计算 DPMO
            dpmo = (total_defects / total_opportunities * 1000000) if total_opportunities > 0 else 0
            
            # 6. 西格玛水平（简化对照表）
            def get_sigma_level(dpmo: float) -> float:
                if dpmo <= 3.4:
                    return 6.0
                elif dpmo <= 233:
                    return 5.5
                elif dpmo <= 6210:
                    return 4.5
                elif dpmo <= 66800:
                    return 3.5
                elif dpmo <= 308000:
                    return 2.5
                elif dpmo <= 690000:
                    return 1.5
                else:
                    return 1.0
            
            sigma_level = get_sigma_level(dpmo)
            
            # 7. 按项目统计缺陷 - 动态计算延期
            project_defects_sql = text("""
                SELECT 
                    p.id as project_id,
                    p.name as project_name,
                    p.leader,
                    COUNT(t.task_id) as total_tasks,
                    SUM(CASE WHEN t.end_date < CURRENT_DATE AND (t.progress < 100 OR t.progress IS NULL) 
                             THEN 1 ELSE 0 END) as delayed_tasks,
                    SUM(CASE WHEN t.end_date < CURRENT_DATE - INTERVAL '3 days' 
                             AND (t.progress < 100 OR t.progress IS NULL)
                             THEN 1 ELSE 0 END) as severe_delayed,
                    CASE WHEN p.actual_total_cost > p.budget_total_cost * 1.1 THEN 1 ELSE 0 END as cost_overrun
                FROM projects p
                LEFT JOIN project_tasks t ON t.project_id = CAST(p.id AS VARCHAR) 
                    AND t.is_latest = true AND t.is_deleted = false
                WHERE p.is_deleted = false
                GROUP BY p.id, p.name, p.leader, p.budget_total_cost, p.actual_total_cost
                HAVING SUM(CASE WHEN t.end_date < CURRENT_DATE AND (t.progress < 100 OR t.progress IS NULL) THEN 1 ELSE 0 END) > 0
                   OR (p.actual_total_cost > p.budget_total_cost * 1.1 AND p.budget_total_cost > 0)
                ORDER BY (SUM(CASE WHEN t.end_date < CURRENT_DATE AND (t.progress < 100 OR t.progress IS NULL) THEN 1 ELSE 0 END) + 
                         CASE WHEN p.actual_total_cost > p.budget_total_cost * 1.1 THEN 1 ELSE 0 END) DESC
                LIMIT 10
            """)
            project_defects = []
            for row in conn.execute(project_defects_sql):
                total_project_defects = (row[4] or 0) + (row[6] or 0)  # 延期 + 超支
                if total_project_defects > 0:
                    project_defects.append({
                        "project_id": row[0],
                        "project_name": row[1],
                        "leader": row[2],
                        "total_tasks": row[3] or 0,
                        "delayed_tasks": row[4] or 0,
                        "severe_delayed": row[5] or 0,
                        "cost_overrun": row[6] or 0,
                        "total_defects": total_project_defects
                    })
            
            # 8. 缺陷趋势（近6周）
            trend_sql = text("""
                SELECT 
                    DATE_TRUNC('week', update_time)::date as week_start,
                    COUNT(*) as new_tasks,
                    SUM(CASE WHEN status = '延期' THEN 1 ELSE 0 END) as new_delayed
                FROM project_tasks
                WHERE update_time >= CURRENT_DATE - INTERVAL '6 weeks'
                  AND is_deleted = false
                GROUP BY DATE_TRUNC('week', update_time)::date
                ORDER BY week_start
            """)
            defect_trend = []
            for row in conn.execute(trend_sql):
                defect_trend.append({
                    "week": row[0].strftime("%m/%d") if row[0] else "-",
                    "new_tasks": row[1] or 0,
                    "new_delayed": row[2] or 0
                })
            
            return {
                "success": True,
                "data": {
                    "summary": {
                        "total_tasks": total_tasks,
                        "total_opportunities": total_opportunities,
                        "total_defects": total_defects,
                        "dpmo": round(dpmo, 1),
                        "sigma_level": sigma_level,
                        "defect_rate": round((total_defects / total_opportunities * 100), 2) if total_opportunities > 0 else 0
                    },
                    "breakdown": {
                        "delayed_defects": severe_delayed,
                        "total_delayed": delayed_tasks,
                        "overbudget_defects": overbudget_projects,
                        "total_cost_projects": cost_projects,
                        "severe_delayed": severe_delayed
                    },
                    "status_distribution": {
                        "completed": completed_tasks,
                        "ongoing": ongoing_tasks,
                        "pending": pending_tasks,
                        "delayed": delayed_tasks
                    },
                    "project_defects": project_defects,
                    "defect_trend": defect_trend,
                    "formulas": {
                        "dpmo": "DPMO = (缺陷数 / 总机会数) × 1,000,000",
                        "defect_definition": "缺陷标准：任务延期>3天 或 成本超支>10%",
                        "sigma_table": "1σ=690000, 2σ=308000, 3σ=66800, 4σ=6210, 5σ=233, 6σ=3.4"
                    }
                }
            }
    except Exception as e:
        logger.exception(f"获取质量概览失败: {e}")
        return {"success": False, "error": str(e)}


@app.get("/agent/api/agent/quality/project/{project_id}")
async def get_project_quality_detail(
    project_id: int,
    current_user: dict = Depends(get_current_user)
):
    """
    单项目质量详情
    """
    from sqlalchemy import text
    try:
        with get_connection() as conn:
            # 项目基本信息
            project_sql = text("""
                SELECT id, name, leader, progress, budget_total_cost, actual_total_cost,
                       start_date, end_date
                FROM projects
                WHERE id = :pid AND is_deleted = false
            """)
            project = conn.execute(project_sql, {"pid": project_id}).fetchone()
            if not project:
                return {"success": False, "error": "项目不存在"}
            
            # 任务缺陷分析 - 动态计算延期状态
            task_sql = text("""
                SELECT 
                    task_id, task_name, status, progress, start_date, end_date,
                    CASE 
                        WHEN end_date < CURRENT_DATE - INTERVAL '3 days' AND (progress < 100 OR progress IS NULL)
                        THEN 'severe'
                        WHEN end_date < CURRENT_DATE AND (progress < 100 OR progress IS NULL) THEN 'minor'
                        ELSE 'ok'
                    END as defect_level,
                    CASE WHEN end_date < CURRENT_DATE THEN CURRENT_DATE - end_date ELSE 0 END as delay_days
                FROM project_tasks
                WHERE project_id = CAST(:pid AS VARCHAR) AND is_latest = true AND is_deleted = false
                ORDER BY 
                    CASE WHEN end_date < CURRENT_DATE AND (progress < 100 OR progress IS NULL) THEN 0 ELSE 1 END,
                    end_date DESC
            """)
            tasks = []
            defects = []
            for row in conn.execute(task_sql, {"pid": project_id}):
                task_info = {
                    "task_id": row[0],
                    "task_name": row[1],
                    "status": row[2],
                    "progress": row[3],
                    "start_date": str(row[4]) if row[4] else None,
                    "end_date": str(row[5]) if row[5] else None,
                    "defect_level": row[6],
                    "delay_days": row[7] if row[7] and row[7] > 0 else None
                }
                tasks.append(task_info)
                if row[6] != 'ok':
                    defects.append(task_info)
            
            # 成本分析
            budget = float(project[4] or 0)
            actual = float(project[5] or 0)
            cost_overrun = actual > budget * 1.1 if budget > 0 else False
            cost_variance = ((actual - budget) / budget * 100) if budget > 0 else 0
            
            return {
                "success": True,
                "data": {
                    "project": {
                        "id": project[0],
                        "name": project[1],
                        "leader": project[2],
                        "progress": float(project[3] or 0),
                        "budget": budget,
                        "actual": actual,
                        "cost_variance": round(cost_variance, 2),
                        "cost_overrun": cost_overrun
                    },
                    "tasks": tasks,
                    "defects": defects,
                    "defect_count": len(defects),
                    "total_tasks": len(tasks),
                    "recommendations": generate_quality_recommendations(defects, cost_overrun, cost_variance)
                }
            }
    except Exception as e:
        logger.exception(f"获取项目质量详情失败: {e}")
        return {"success": False, "error": str(e)}


def generate_quality_recommendations(defects: list, cost_overrun: bool, cost_variance: float) -> list:
    """生成质量改进建议"""
    recommendations = []
    
    if len(defects) > 0:
        severe_count = sum(1 for d in defects if d.get('defect_level') == 'severe')
        if severe_count > 0:
            recommendations.append({
                "priority": "high",
                "type": "schedule",
                "message": f"有 {severe_count} 个任务严重延期（>3天），建议立即review项目计划"
            })
        
        recommendations.append({
            "priority": "medium",
            "type": "process",
            "message": "建议进行根因分析，识别延期原因（资源不足/估算偏差/需求变更）"
        })
    
    if cost_overrun:
        recommendations.append({
            "priority": "high",
            "type": "cost",
            "message": f"成本超支 {abs(cost_variance):.1f}%，建议审查支出明细"
        })
    
    return recommendations


# ==================== 质量帕累托分析 ====================

@app.get("/agent/api/agent/quality/pareto")
async def get_quality_pareto(current_user: dict = Depends(get_current_user)):
    """
    帕累托分析 - 80/20 规律
    
    返回：
    1. 项目缺陷帕累托 - 哪些项目贡献了大部分缺陷
    2. 延期原因帕累托 - 延期的主要原因分布
    3. 时间段帕累托 - 哪个时间段延期最多
    """
    from sqlalchemy import text
    try:
        with get_connection() as conn:
            # 1. 项目缺陷帕累托（TOP 10）
            project_pareto_sql = text("""
                WITH project_defects AS (
                    SELECT 
                        p.id,
                        p.name as project_name,
                        p.leader,
                        COUNT(t.task_id) as total_tasks,
                        SUM(CASE WHEN t.end_date < CURRENT_DATE - INTERVAL '3 days' 
                                 AND (t.progress < 100 OR t.progress IS NULL) THEN 1 ELSE 0 END) as severe_defects,
                        SUM(CASE WHEN t.end_date < CURRENT_DATE 
                                 AND (t.progress < 100 OR t.progress IS NULL) THEN 1 ELSE 0 END) as total_defects
                    FROM projects p
                    LEFT JOIN project_tasks t ON t.project_id = CAST(p.id AS VARCHAR)
                        AND t.is_latest = true AND t.is_deleted = false
                    WHERE p.is_deleted = false
                    GROUP BY p.id, p.name, p.leader
                ),
                ranked AS (
                    SELECT 
                        project_name,
                        leader,
                        total_tasks,
                        severe_defects,
                        total_defects,
                        ROW_NUMBER() OVER (ORDER BY total_defects DESC, project_name) as rn,
                        SUM(total_defects) OVER () as total_all_defects
                    FROM project_defects
                    WHERE total_defects > 0
                ),
                cumulative AS (
                    SELECT 
                        project_name,
                        leader,
                        total_tasks,
                        severe_defects,
                        total_defects,
                        SUM(total_defects) OVER (ORDER BY rn) as cum_defects,
                        total_all_defects,
                        rn
                    FROM ranked
                    ORDER BY rn
                    LIMIT 10
                )
                SELECT 
                    project_name,
                    leader,
                    total_tasks,
                    severe_defects,
                    total_defects,
                    cum_defects,
                    total_all_defects,
                    ROUND(cum_defects * 100.0 / NULLIF(total_all_defects, 0), 1) as cumulative_pct
                FROM cumulative
            """)
            
            project_pareto = []
            total_defects_count = 0
            for row in conn.execute(project_pareto_sql):
                total_defects_count = int(float(row[6] or 0))
                project_pareto.append({
                    "project_name": row[0],
                    "leader": row[1] or "未分配",
                    "total_tasks": int(float(row[2] or 0)),
                    "severe_defects": int(float(row[3] or 0)),
                    "total_defects": int(float(row[4] or 0)),
                    "cumulative": int(float(row[5] or 0)),
                    "cumulative_pct": float(row[7] or 0)
                })
            
            # 2. 延期时间段帕累托（按延期天数分段 - 统一用周口径）
            time_pareto_sql = text("""
                SELECT 
                    CASE 
                        WHEN CURRENT_DATE - end_date <= 7 THEN '1周内'
                        WHEN CURRENT_DATE - end_date <= 14 THEN '1-2周'
                        WHEN CURRENT_DATE - end_date <= 28 THEN '2-4周'
                        WHEN CURRENT_DATE - end_date <= 56 THEN '4-8周'
                        ELSE '超过8周'
                    END as delay_range,
                    COUNT(*) as defect_count,
                    ROUND(COUNT(*) * 100.0 / NULLIF(SUM(COUNT(*)) OVER(), 0), 1) as pct
                FROM project_tasks
                WHERE is_latest = true 
                  AND is_deleted = false
                  AND end_date < CURRENT_DATE 
                  AND (progress < 100 OR progress IS NULL)
                GROUP BY 
                    CASE 
                        WHEN CURRENT_DATE - end_date <= 7 THEN '1周内'
                        WHEN CURRENT_DATE - end_date <= 14 THEN '1-2周'
                        WHEN CURRENT_DATE - end_date <= 28 THEN '2-4周'
                        WHEN CURRENT_DATE - end_date <= 56 THEN '4-8周'
                        ELSE '超过8周'
                    END
                ORDER BY defect_count DESC
            """)
            
            time_pareto = []
            for row in conn.execute(time_pareto_sql):
                time_pareto.append({
                    "delay_range": row[0],
                    "defect_count": int(float(row[1] or 0)),
                    "percentage": float(row[2] or 0)
                })
            
            # 3. 计算帕累托关键点（80% 临界点）
            pareto_80_index = -1  # -1 表示未找到
            for i, p in enumerate(project_pareto):
                if p["cumulative_pct"] >= 80:
                    pareto_80_index = i
                    break
            
            # 4. 关键洞察
            insights = []
            if len(project_pareto) > 0:
                # 如果达到80%累计，用实际项目数；否则用所有项目
                if pareto_80_index >= 0:
                    top_project_count = pareto_80_index + 1
                    top_pct = project_pareto[pareto_80_index]["cumulative_pct"]
                    insights.append({
                        "type": "project_concentration",
                        "message": f"前 {top_project_count} 个项目贡献了 {top_pct}% 的缺陷",
                        "recommendation": "优先处理这些高风险项目可快速降低整体缺陷率"
                    })
                else:
                    # 未达到80%，说明缺陷分散
                    total_in_top10 = sum(int(p["total_defects"]) for p in project_pareto)
                    total_all = int(total_defects_count)
                    top10_pct = round(total_in_top10 * 100.0 / total_all, 1) if total_all > 0 else 0
                    insights.append({
                        "type": "project_concentration",
                        "message": f"缺陷分布较分散，前 {len(project_pareto)} 个项目共贡献 {top10_pct}% 缺陷",
                        "recommendation": "缺陷未集中，需全面关注各项目进度"
                    })
            
            if len(time_pareto) > 0:
                max_delay_range = time_pareto[0]["delay_range"]
                max_delay_count = time_pareto[0]["defect_count"]
                insights.append({
                    "type": "delay_pattern",
                    "message": f"延期主要集中在「{max_delay_range}」时段，共 {max_delay_count} 个任务",
                    "recommendation": "分析该时间段的任务特征，找出共性原因"
                })
            
            return {
                "success": True,
                "data": {
                    "project_pareto": project_pareto,
                    "time_pareto": time_pareto,
                    "pareto_80_index": pareto_80_index,
                    "total_defects": total_defects_count,
                    "insights": insights,
                    "formulas": {
                        "pareto_rule": "帕累托定律：20% 的原因贡献了 80% 的问题",
                        "project_defect": "严重延期(>3天) + 一般延期",
                        "focus_strategy": "优先解决 TOP 20% 项目 → 快速降低整体风险"
                    }
                }
            }
    except Exception as e:
        logger.exception(f"获取帕累托分析失败: {e}")
        return {"success": False, "error": str(e)}


# ==================== 根因分析 API ====================

@app.get("/agent/api/agent/quality/analysis/{project_id}")
async def analyze_project_defects(
    project_id: int,
    current_user: dict = Depends(get_current_user)
):
    """
    AI 根因分析 - 分析项目延期原因
    
    使用 DeepSeek AI 分析：
    1. 延期任务特征
    2. 可能的延期原因
    3. 改进建议
    """
    from sqlalchemy import text
    try:
        with get_connection() as conn:
            # 1. 获取项目基本信息
            project_sql = text("""
                SELECT id, name, leader, progress, start_date, end_date,
                       budget_total_cost, actual_total_cost
                FROM projects
                WHERE id = :pid AND is_deleted = false
            """)
            project = conn.execute(project_sql, {"pid": project_id}).fetchone()
            if not project:
                return {"success": False, "error": "项目不存在"}
            
            # 2. 获取延期任务详情
            delayed_tasks_sql = text("""
                SELECT 
                    task_id, task_name, start_date, end_date,
                    progress, assignee,
                    CURRENT_DATE - end_date as delay_days
                FROM project_tasks
                WHERE project_id = CAST(:pid AS VARCHAR)
                  AND is_latest = true
                  AND is_deleted = false
                  AND end_date < CURRENT_DATE
                  AND (progress < 100 OR progress IS NULL)
                ORDER BY end_date
            """)
            delayed_tasks = []
            for row in conn.execute(delayed_tasks_sql, {"pid": project_id}):
                delayed_tasks.append({
                    "task_id": row[0],
                    "task_name": row[1],
                    "start_date": str(row[2]) if row[2] else None,
                    "end_date": str(row[3]) if row[3] else None,
                    "progress": float(row[4] or 0),
                    "assignee": row[5],
                    "delay_days": int(row[6] or 0)
                })
            
            if len(delayed_tasks) == 0:
                return {
                    "success": True,
                    "data": {
                        "project_name": project[1],
                        "has_defects": False,
                        "message": "该项目当前无延期任务"
                    }
                }
            
            # 3. 统计延期特征
            total_delayed = len(delayed_tasks)
            avg_delay = sum(t["delay_days"] for t in delayed_tasks) / total_delayed
            max_delay = max(t["delay_days"] for t in delayed_tasks)
            assignees = list(set(t["assignee"] for t in delayed_tasks if t["assignee"]))
            
            # 4. AI 分析延期原因
            analysis_prompt = f"""分析以下项目延期情况，给出延期原因分类和改进建议。

项目名称：{project[1]}
负责人：{project[2]}
当前进度：{float(project[3] or 0):.1f}%

延期任务数量：{total_delayed}
平均延期天数：{avg_delay:.1f}天
最长延期天数：{max_delay}天

延期任务列表（前5个）：
{chr(10).join(f"- {t['task_name']}（延期{t['delay_days']}天，负责人：{t['assignee'] or '未知'}）" for t in delayed_tasks[:5])}

请分析：
1. 主要延期原因（从以下分类选择：资源不足、估算偏差、需求变更、外部依赖、技术难点、沟通问题、其他）
2. 每个原因的影响程度（高/中/低）
3. 改进建议（具体可执行）

请用JSON格式返回：
{{
  "reasons": [
    {{"type": "原因类型", "impact": "影响程度", "detail": "具体说明"}}
  ],
  "recommendations": [
    {{"action": "改进措施", "responsible": "责任人建议", "priority": "优先级"}}
  ]
}}
"""
            
            # 调用 DeepSeek AI（线程池执行，不阻塞Worker）
            try:
                def _call_delay_analysis():
                    from openai import OpenAI
                    client = OpenAI(
                        api_key=os.getenv("DEEPSEEK_API_KEY", ""),
                        base_url="https://api.deepseek.com"
                    )
                    return client.chat.completions.create(
                        model="deepseek-chat",
                        messages=[
                            {"role": "system", "content": "你是项目管理专家，擅长分析项目延期原因并给出改进建议。"},
                            {"role": "user", "content": analysis_prompt}
                        ],
                        temperature=0.7,
                        max_tokens=1000
                    )
                
                response = await run_in_thread(_call_delay_analysis)
                
                ai_response = response.choices[0].message.content
                
                # 解析 AI 返回
                import json
                import re
                
                # 提取 JSON 部分
                json_match = re.search(r'\{[\s\S]*\}', ai_response)
                if json_match:
                    analysis_result = json.loads(json_match.group())
                else:
                    analysis_result = {
                        "reasons": [{"type": "其他", "impact": "中", "detail": "AI 分析结果解析失败"}],
                        "recommendations": []
                    }
                
            except Exception as e:
                logger.error(f"AI 分析失败: {e}")
                # 降级：基于规则的简单分析
                analysis_result = {
                    "reasons": [
                        {"type": "估算偏差", "impact": "高" if avg_delay > 14 else "中", "detail": f"平均延期{avg_delay:.0f}天，可能存在估算偏差"}
                    ],
                    "recommendations": [
                        {"action": "重新评估剩余任务工期", "responsible": "项目经理", "priority": "高"},
                        {"action": "增加资源投入或调整项目计划", "responsible": "项目负责人", "priority": "高"}
                    ]
                }
            
            return {
                "success": True,
                "data": {
                    "project_id": project_id,
                    "project_name": project[1],
                    "leader": project[2],
                    "has_defects": True,
                    "delayed_tasks": delayed_tasks,
                    "statistics": {
                        "total_delayed": total_delayed,
                        "avg_delay_days": round(avg_delay, 1),
                        "max_delay_days": max_delay,
                        "assignees": assignees
                    },
                    "analysis": analysis_result
                }
            }
            
    except Exception as e:
        logger.exception(f"根因分析失败: {e}")
        return {"success": False, "error": str(e)}


@app.post("/agent/api/agent/quality/improvement")
async def create_improvement_action(
    request: dict,
    current_user: dict = Depends(get_current_user)
):
    """
    创建改进措施
    """
    from sqlalchemy import text
    try:
        project_id = request.get("project_id")
        action_type = request.get("action_type", "纠正")
        description = request.get("description")
        responsible = request.get("responsible")
        target_date = request.get("target_date")
        
        if not project_id or not description:
            return {"success": False, "error": "缺少必填字段"}
        
        with get_connection() as conn:
            result = conn.execute(text("""
                INSERT INTO improvement_actions 
                (project_id, action_type, description, responsible, target_date, created_at)
                VALUES (:pid, :type, :desc, :resp, :target, NOW())
                RETURNING id
            """), {
                "pid": project_id,
                "type": action_type,
                "desc": description,
                "resp": responsible,
                "target": target_date
            })
            conn.commit()
            
            action_id = result.fetchone()[0]
            
            return {
                "success": True,
                "data": {"id": action_id, "message": "改进措施已创建"}
            }
            
    except Exception as e:
        logger.exception(f"创建改进措施失败: {e}")
        return {"success": False, "error": str(e)}


@app.get("/agent/api/agent/quality/improvements/{project_id}")
async def get_project_improvements(
    project_id: int,
    current_user: dict = Depends(get_current_user)
):
    """
    获取项目改进措施列表
    """
    from sqlalchemy import text
    try:
        with get_connection() as conn:
            result = conn.execute(text("""
                SELECT id, action_type, description, responsible, 
                       target_date, status, effect_measure, 
                       created_at, completed_at
                FROM improvement_actions
                WHERE project_id = :pid
                ORDER BY created_at DESC
            """), {"pid": project_id})
            
            improvements = []
            for row in result:
                improvements.append({
                    "id": row[0],
                    "action_type": row[1],
                    "description": row[2],
                    "responsible": row[3],
                    "target_date": str(row[4]) if row[4] else None,
                    "status": row[5],
                    "effect_measure": float(row[6]) if row[6] else None,
                    "created_at": str(row[7]) if row[7] else None,
                    "completed_at": str(row[8]) if row[8] else None
                })
            
            return {"success": True, "data": improvements}
            
    except Exception as e:
        logger.exception(f"获取改进措施失败: {e}")
        return {"success": False, "error": str(e)}



# ============== 注册模块化路由 ==============
# 注意：这是拆分计划的一部分，路由模块独立于上面定义的端点
# 注册后会与原端点共存，待测试通过后再移除原端点

# from routes import auth
# app.include_router(auth.router)


# ============== AI调用统计管理API ==============


class AIUsageResponse(BaseModel):
    """AI调用统计响应"""
    date: str
    total_calls: int
    total_cost: float
    total_input_tokens: int
    total_output_tokens: int
    by_purpose: Dict[str, Dict[str, float]]
    by_user: List[Dict[str, Any]]


@app.get("/agent/api/agent/admin/ai-usage/stats", response_model=AIUsageResponse)
async def get_ai_usage_stats(
    target_date: Optional[str] = None,
    current_user: Dict = Depends(get_current_user)
):
    """
    获取AI调用统计（管理员专用）
    
    参数：
    - target_date: 目标日期，格式YYYY-MM-DD，默认今天
    
    返回：每日调用次数、费用、用户分布
    """
    # 检查权限（只有管理员和高级用户可以查看）
    role_id = current_user.get("role_id", 15)
    if role_id not in [11, 12]:  # 管理员、高级用户
        raise HTTPException(status_code=403, detail="无权限查看AI调用统计")
    
    try:
        from .ai_usage_tracker import get_ai_daily_stats
        from datetime import datetime
        
        if target_date:
            date_obj = datetime.strptime(target_date, "%Y-%m-%d").date()
        else:
            date_obj = None
        
        stats = get_ai_daily_stats(date_obj)
        
        if not stats:
            return {
                "date": target_date or str(date.today()),
                "total_calls": 0,
                "total_cost": 0,
                "total_input_tokens": 0,
                "total_output_tokens": 0,
                "by_purpose": {},
                "by_user": []
            }
        
        return stats
        
    except Exception as e:
        logger.error(f"获取AI调用统计失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/agent/api/agent/admin/ai-usage/users")
async def get_ai_usage_users(
    current_user: Dict = Depends(get_current_user)
):
    """
    获取今日用户调用排行（管理员专用）
    
    返回：按调用次数排序的用户列表
    """
    role_id = current_user.get("role_id", 15)
    if role_id not in [11, 12]:
        raise HTTPException(status_code=403, detail="无权限查看")
    
    try:
        from .ai_usage_tracker import get_ai_daily_stats
        stats = get_ai_daily_stats()
        return {"users": stats.get("by_user", [])}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/agent/api/agent/admin/ai-usage/limits")
async def get_ai_usage_limits(current_user: Dict = Depends(get_current_user)):
    """获取AI调用限制配置"""
    from .ai_usage_tracker import USAGE_LIMITS
    return {"limits": USAGE_LIMITS}


# ============== 人力成本导出（匹配Excel模板格式）==============

@app.get("/agent/api/agent/stats/human-cost-export")
async def export_human_cost_excel(
    year: int = None,
    month: int = None,
    current_user: Dict = Depends(get_current_user)
):
    """
    导出人力成本Excel（匹配研究院模板格式）
    
    Sheet1: 项目成本汇总表 - 项目、时长（天），保留所有列名和边框
    Sheet2: 项目人工成本汇总表 - 项目维度，每行显示项目+参与人工时
    
    时长计算：总工时（小时） / 8
    
    注意：仅统计正式项目工时
    """
    from fastapi.responses import StreamingResponse
    import io
    import pandas as pd
    from urllib.parse import quote
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl import Workbook
    from .holidays import calculate_working_days

    # 默认当月
    today = datetime.now().date()
    year = year or today.year
    month = month or today.month

    month_start = datetime(year, month, 1).date()
    month_end = (month_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)
    
    # 计算当月工作日数（动态）
    working_days = calculate_working_days(year, month)

    with get_connection() as conn:
        # 1. 查询员工工号映射
        emp_mapping_result = conn.execute(text("""
            SELECT employee_code, employee_name 
            FROM employee_group_relations 
            WHERE relation_type = 'member' AND employee_code IS NOT NULL
        """))
        name_to_code = {row[1]: row[0] for row in emp_mapping_result}
        
        # 2. 查询正式项目的工时数据（项目维度汇总）
        project_result = conn.execute(text("""
            SELECT 
                p.name as project_name,
                SUM(dwi.hours_spent) as total_hours
            FROM daily_work_items dwi
            JOIN daily_reports dr ON dr.id = dwi.report_id
            JOIN projects p ON p.id::text = dwi.project_id
            WHERE dr.report_date >= :month_start
              AND dr.report_date <= :month_end
              AND dr.is_deleted = false
              AND LOWER(dr.employee_name) != 'admin'
              AND dwi.project_id IS NOT NULL AND dwi.project_id != ''
            GROUP BY p.name
            ORDER BY p.name
        """), {"month_start": month_start, "month_end": month_end})
        
        project_rows = []
        for idx, row in enumerate(project_result, start=1):
            project_name = row[0]
            total_hours = float(row[1] or 0)
            days = round(total_hours / 8, 2)
            project_rows.append({
                "序号": idx,
                "项目": project_name,
                "时长（天）": days
            })
        
        # 3. 查询正式项目的工时数据（项目+人员维度）
        emp_project_result = conn.execute(text("""
            SELECT 
                p.name as project_name,
                dr.employee_name,
                SUM(dwi.hours_spent) as total_hours
            FROM daily_work_items dwi
            JOIN daily_reports dr ON dr.id = dwi.report_id
            JOIN projects p ON p.id::text = dwi.project_id
            WHERE dr.report_date >= :month_start
              AND dr.report_date <= :month_end
              AND dr.is_deleted = false
              AND LOWER(dr.employee_name) != 'admin'
              AND dwi.project_id IS NOT NULL AND dwi.project_id != ''
            GROUP BY p.name, dr.employee_name
            ORDER BY p.name, dr.employee_name
        """), {"month_start": month_start, "month_end": month_end})
        
        # 按项目分组人员工时
        project_emp_hours = {}
        for row in emp_project_result:
            project_name = row[0]
            emp_name = row[1]
            hours = float(row[2] or 0)
            days = round(hours / 8, 2)
            
            if project_name not in project_emp_hours:
                project_emp_hours[project_name] = []
            
            project_emp_hours[project_name].append({
                "工号": name_to_code.get(emp_name, ""),
                "姓名": emp_name,
                "时长（天）": days
            })

    # 创建Excel（使用openpyxl直接创建，保留所有列名和边框）
    wb = Workbook()
    
    # 边框样式
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 字体样式：微软雅黑
    font_yahei9 = Font(name='微软雅黑', size=9)
    font_yahei9_bold = Font(name='微软雅黑', size=9, bold=True)
    font_title = Font(name='微软雅黑', size=14, bold=True)
    
    center_align = Alignment(horizontal='center', vertical='center')
    wrap_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # ==================== Sheet1: 项目成本汇总表 ====================
    ws1 = wb.active
    ws1.title = f"{month}月项目成本汇总表"
    
    # 标题行
    ws1.merge_cells('A1:G1')
    ws1['A1'] = f"研究院{month}月项目人工成本"
    ws1['A1'].font = font_title
    ws1['A1'].alignment = center_align
    
    # 表头（保留所有列名，行高40支持折行）
    ws1.row_dimensions[2].height = 40
    headers1 = ["序号", "项目", "时长（天）", "人力成本\n（工资）", "人力成本\n（社保）", "公积金", "总额"]
    for col, header in enumerate(headers1, start=1):
        cell = ws1.cell(row=2, column=col, value=header)
        cell.font = font_yahei9_bold
        cell.alignment = wrap_align
        cell.border = thin_border
    
    # 数据行（只填充序号、项目、时长，其他留空但保留边框）
    for row_idx, proj in enumerate(project_rows, start=3):
        ws1.cell(row=row_idx, column=1, value=proj["序号"]).border = thin_border
        ws1.cell(row=row_idx, column=2, value=proj["项目"]).border = thin_border
        ws1.cell(row=row_idx, column=3, value=proj["时长（天）"]).border = thin_border
        # 其他列留空但有边框
        for col in range(4, 8):
            ws1.cell(row=row_idx, column=col, value="").border = thin_border
        
        # 居中对齐 + 微软雅黑 9号
        for col in range(1, 8):
            cell = ws1.cell(row=row_idx, column=col)
            cell.alignment = center_align
            cell.font = font_yahei9
    
    # 设置列宽
    ws1.column_dimensions['A'].width = 8
    ws1.column_dimensions['B'].width = 40
    ws1.column_dimensions['C'].width = 12
    ws1.column_dimensions['D'].width = 14
    ws1.column_dimensions['E'].width = 14
    ws1.column_dimensions['F'].width = 10
    ws1.column_dimensions['G'].width = 10
    
    # ==================== Sheet2: 项目人工成本汇总表 ====================
    ws2 = wb.create_sheet(f"{month}月项目人工成本汇总表")
    
    # 标题行
    ws2.merge_cells('A1:Q1')
    ws2['A1'] = "研究院项目人工成本明细"
    ws2['A1'].font = font_title
    ws2['A1'].alignment = center_align
    
    # 表头行1（高度设置为40，支持折行）
    ws2.row_dimensions[2].height = 40
    headers2_row1 = ["序号", "项目", "工号", "姓名", "时长（天）", "社保", "公积金", "应发薪资", 
                     f"工日单价\n（{month}月应出勤{working_days}天）", 
                     "项目单人成本", "", "", "", "项目总成本", "", "", ""]
    for col, header in enumerate(headers2_row1, start=1):
        cell = ws2.cell(row=2, column=col, value=header)
        cell.font = font_yahei9_bold
        cell.alignment = wrap_align
        cell.border = thin_border
    
    # 合并单元格：项目单人成本（J2:M2）、项目总成本（N2:Q2）
    ws2.merge_cells('J2:M2')
    ws2.merge_cells('N2:Q2')
    
    # 表头行2（高度设置为40，支持折行）
    ws2.row_dimensions[3].height = 40
    headers2_row2 = ["", "", "", "", "", "", "", "", "", 
                     "人力成本\n（工资）", "人力成本\n（社保）", "人力成本\n（公积金）", "总额",
                     "总人力成本\n（工资）", "总人力成本\n（社保）", "总人力成本\n（公积金）", "总额"]
    for col, header in enumerate(headers2_row2, start=1):
        cell = ws2.cell(row=3, column=col, value=header)
        cell.font = font_yahei9_bold
        cell.alignment = wrap_align
        cell.border = thin_border
    
    # 数据行：项目维度，每个项目下显示所有参与人
    row_idx = 4
    seq = 1
    
    # 按项目名排序
    sorted_projects = sorted(project_emp_hours.keys())
    
    for project_name in sorted_projects:
        emp_list = project_emp_hours[project_name]
        start_row = row_idx
        
        for i, emp_data in enumerate(emp_list):
            ws2.cell(row=row_idx, column=1, value=seq).border = thin_border
            ws2.cell(row=row_idx, column=2, value=project_name if i == 0 else "").border = thin_border
            ws2.cell(row=row_idx, column=3, value=emp_data["工号"]).border = thin_border
            ws2.cell(row=row_idx, column=4, value=emp_data["姓名"]).border = thin_border
            ws2.cell(row=row_idx, column=5, value=emp_data["时长（天）"]).border = thin_border
            
            # 其他列留空但有边框
            for col in range(6, 18):
                ws2.cell(row=row_idx, column=col, value="").border = thin_border
            
            # 居中对齐 + 微软雅黑 9号
            for col in range(1, 18):
                cell = ws2.cell(row=row_idx, column=col)
                cell.alignment = center_align
                cell.font = font_yahei9
            
            row_idx += 1
            seq += 1
        
        # 合并项目名称单元格（如果该项目有多人参与）
        if len(emp_list) > 1:
            ws2.merge_cells(f'B{start_row}:B{row_idx-1}')
            # 合并后的单元格样式
            merged_cell = ws2[f'B{start_row}']
            merged_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            merged_cell.font = font_yahei9
            
            # 同时合并项目总成本相关的列（N、O、P、Q）
            ws2.merge_cells(f'N{start_row}:N{row_idx-1}')
            ws2.merge_cells(f'O{start_row}:O{row_idx-1}')
            ws2.merge_cells(f'P{start_row}:P{row_idx-1}')
            ws2.merge_cells(f'Q{start_row}:Q{row_idx-1}')
            # 设置合并后单元格样式
            for col_letter in ['N', 'O', 'P', 'Q']:
                cell = ws2[f'{col_letter}{start_row}']
                cell.alignment = center_align
                cell.font = font_yahei9
                cell.border = thin_border
    
    # 添加合计行和计算行
    # 计算总工时
    total_days = sum(emp_data["时长（天）"] for emp_list in project_emp_hours.values() for emp_data in emp_list)
    calc_value = round(total_days / 21.75, 2) if total_days > 0 else 0
    
    # 合计行
    sum_row = row_idx
    ws2.cell(row=sum_row, column=5, value=round(total_days, 2)).border = thin_border
    ws2.cell(row=sum_row, column=5).font = font_yahei9_bold
    ws2.cell(row=sum_row, column=5).alignment = center_align
    
    # 计算行（合计/21.75）
    calc_row = row_idx + 1
    ws2.cell(row=calc_row, column=5, value=calc_value).border = thin_border
    ws2.cell(row=calc_row, column=5).font = font_yahei9_bold
    ws2.cell(row=calc_row, column=5).alignment = center_align
    
    # 冻结窗格：冻结前3行（标题行+表头两行）
    ws2.freeze_panes = 'A4'
    
    # 设置列宽
    ws2.column_dimensions['A'].width = 8
    ws2.column_dimensions['B'].width = 35
    ws2.column_dimensions['C'].width = 12
    ws2.column_dimensions['D'].width = 10
    ws2.column_dimensions['E'].width = 12
    for col in ['F', 'G', 'H', 'I']:
        ws2.column_dimensions[col].width = 10
    for col in ['J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q']:
        ws2.column_dimensions[col].width = 12

    # 保存到BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f'{year}年{month}月研究院人员项目成本归集.xlsx'
    encoded_filename = quote(filename)
    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f"attachment; filename*=UTF-8''{encoded_filename}"}
    )


# ============== 计划版本效果评估 ==============

class PlanEffectUpdate(BaseModel):
    """计划效果评估更新"""
    effect_note: str  # 效果评估说明


@app.put("/agent/api/agent/plan-versions/{version_id}/effect")
async def update_plan_effect(
    version_id: int,
    data: PlanEffectUpdate,
    current_user: Dict = Depends(get_current_user)
):
    """
    更新计划版本的效果评估
    
    在调整执行一段时间后，评估调整效果
    """
    try:
        with get_connection() as conn:
            # 检查版本是否存在
            result = conn.execute(text("""
                SELECT pv.project_id, p.leader FROM project_plan_versions pv
                JOIN projects p ON p.id = pv.project_id
                WHERE pv.id = :vid
            """), {"vid": version_id})
            row = result.fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="计划版本不存在")
            
            # 权限检查：项目负责人或管理员可以更新
            username = current_user.get("username") or current_user.get("sub")
            role_id = current_user.get("role_id", 15)
            leader = row[1]
            
            if role_id != 11 and leader != username:
                raise HTTPException(status_code=403, detail="无权限评估此计划版本")
            
            # 更新效果评估
            conn.execute(text("""
                UPDATE project_plan_versions SET effect_note = :note WHERE id = :vid
            """), {"note": data.effect_note, "vid": version_id})
            conn.commit()
            
            logger.info(f"更新计划版本效果评估: version_id={version_id}")
            return {"success": True, "message": "效果评估已更新"}
    
    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"更新效果评估失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ============== 请假记录管理 ==============

class LeaveRecordCreate(BaseModel):
    """请假记录创建"""
    employee_id: str
    leave_date: str  # YYYY-MM-DD
    leave_type: str  # 年假/病假/事假/调休/婚假/产假/其他
    reason: Optional[str] = None


class LeaveRecordUpdate(BaseModel):
    """请假记录更新"""
    leave_type: Optional[str] = None
    reason: Optional[str] = None


@app.get("/agent/api/agent/leave/records")
async def get_leave_records(
    employee_id: Optional[str] = None,
    year: Optional[int] = None,
    month: Optional[int] = None,
    current_user: Dict = Depends(get_current_user)
):
    """
    获取请假记录列表
    
    参数：
    - employee_id: 员工工号（管理员可查看所有，普通用户只能查看自己）
    - year: 年份
    - month: 月份
    """
    try:
        username = current_user.get("username") or current_user.get("sub")
        requester_id = current_user.get("employee_id") or username
        role_id = current_user.get("role_id", 15)
        
        # 权限检查：管理员可查看所有，普通用户只能查看自己
        if employee_id and employee_id != requester_id and role_id not in [11, 12]:
            raise HTTPException(status_code=403, detail="无权限查看他人请假记录")
        
        # 默认查询自己
        target_emp = employee_id or requester_id
        
        with get_connection() as conn:
            sql = """
                SELECT id, employee_id, leave_date, leave_type, reason, 
                       created_at, created_by
                FROM leave_records
                WHERE is_deleted = false
            """
            params = {}
            
            if target_emp:
                sql += " AND employee_id = :eid"
                params["eid"] = target_emp
            
            if year and month:
                start_date = f"{year}-{month:02d}-01"
                if month == 12:
                    end_date = f"{year + 1}-01-01"
                else:
                    end_date = f"{year}-{month + 1:02d}-01"
                sql += " AND leave_date >= :start_date AND leave_date < :end_date"
                params["start_date"] = start_date
                params["end_date"] = end_date
            
            sql += " ORDER BY leave_date DESC"
            
            result = conn.execute(text(sql), params)
            records = []
            for row in result:
                records.append({
                    "id": row[0],
                    "employee_id": row[1],
                    "leave_date": str(row[2]),
                    "leave_type": row[3],
                    "reason": row[4],
                    "created_at": str(row[5]),
                    "created_by": row[6]
                })
            
            return {"records": records}
    
    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"获取请假记录失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/agent/api/agent/leave/records")
async def create_leave_record(
    data: LeaveRecordCreate,
    current_user: Dict = Depends(get_current_user)
):
    """
    创建请假记录
    
    权限：管理员可创建所有，普通用户只能创建自己的
    """
    try:
        username = current_user.get("username") or current_user.get("sub")
        requester_id = current_user.get("employee_id") or username
        role_id = current_user.get("role_id", 15)
        
        # 权限检查
        if data.employee_id != requester_id and role_id not in [11, 12]:
            raise HTTPException(status_code=403, detail="无权限为他人创建请假记录")
        
        # 验证请假类型
        valid_types = ["年假", "病假", "事假", "调休", "婚假", "产假", "其他"]
        if data.leave_type not in valid_types:
            raise HTTPException(status_code=400, detail=f"无效的请假类型，可选：{', '.join(valid_types)}")
        
        # 验证日期格式
        try:
            leave_date = datetime.strptime(data.leave_date, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="日期格式错误，应为 YYYY-MM-DD")
        
        with get_connection() as conn:
            # 检查是否已存在
            existing = conn.execute(text("""
                SELECT id FROM leave_records 
                WHERE employee_id = :eid AND leave_date = :date AND is_deleted = false
            """), {"eid": data.employee_id, "date": data.leave_date})
            
            if existing.fetchone():
                raise HTTPException(status_code=400, detail="该日期已存在请假记录")
            
            # 插入记录
            result = conn.execute(text("""
                INSERT INTO leave_records (employee_id, leave_date, leave_type, reason, created_by)
                VALUES (:eid, :date, :type, :reason, :created_by)
                RETURNING id
            """), {
                "eid": data.employee_id,
                "date": data.leave_date,
                "type": data.leave_type,
                "reason": data.reason,
                "created_by": requester_id
            })
            conn.commit()
            
            record_id = result.fetchone()[0]
            logger.info(f"创建请假记录: id={record_id}, employee={data.employee_id}, date={data.leave_date}")
            
            return {"success": True, "id": record_id, "message": "请假记录创建成功"}
    
    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"创建请假记录失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/agent/api/agent/leave/records/{record_id}")
async def update_leave_record(
    record_id: int,
    data: LeaveRecordUpdate,
    current_user: Dict = Depends(get_current_user)
):
    """更新请假记录"""
    try:
        username = current_user.get("username") or current_user.get("sub")
        requester_id = current_user.get("employee_id") or username
        role_id = current_user.get("role_id", 15)
        
        with get_connection() as conn:
            # 检查记录是否存在
            existing = conn.execute(text("""
                SELECT employee_id FROM leave_records 
                WHERE id = :id AND is_deleted = false
            """), {"id": record_id})
            row = existing.fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="请假记录不存在")
            
            # 权限检查
            if row[0] != requester_id and role_id not in [11, 12]:
                raise HTTPException(status_code=403, detail="无权限修改他人请假记录")
            
            # 更新
            update_fields = []
            params = {"id": record_id}
            
            if data.leave_type:
                valid_types = ["年假", "病假", "事假", "调休", "婚假", "产假", "其他"]
                if data.leave_type not in valid_types:
                    raise HTTPException(status_code=400, detail=f"无效的请假类型")
                update_fields.append("leave_type = :type")
                params["type"] = data.leave_type
            
            if data.reason is not None:
                update_fields.append("reason = :reason")
                params["reason"] = data.reason
            
            if not update_fields:
                return {"success": True, "message": "无更新内容"}
            
            sql = f"UPDATE leave_records SET {', '.join(update_fields)} WHERE id = :id"
            conn.execute(text(sql), params)
            conn.commit()
            
            return {"success": True, "message": "更新成功"}
    
    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"更新请假记录失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/agent/api/agent/leave/records/{record_id}")
async def delete_leave_record(
    record_id: int,
    current_user: Dict = Depends(get_current_user)
):
    """删除请假记录（软删除）"""
    try:
        username = current_user.get("username") or current_user.get("sub")
        requester_id = current_user.get("employee_id") or username
        role_id = current_user.get("role_id", 15)
        
        with get_connection() as conn:
            # 检查记录是否存在
            existing = conn.execute(text("""
                SELECT employee_id FROM leave_records 
                WHERE id = :id AND is_deleted = false
            """), {"id": record_id})
            row = existing.fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="请假记录不存在")
            
            # 权限检查
            if row[0] != requester_id and role_id not in [11, 12]:
                raise HTTPException(status_code=403, detail="无权限删除他人请假记录")
            
            # 软删除
            conn.execute(text("""
                UPDATE leave_records SET is_deleted = true WHERE id = :id
            """), {"id": record_id})
            conn.commit()
            
            logger.info(f"删除请假记录: id={record_id}")
            return {"success": True, "message": "删除成功"}
    
    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"删除请假记录失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/agent/api/agent/leave/summary")
async def get_leave_summary(
    year: int,
    month: int,
    current_user: Dict = Depends(get_current_user)
):
    """
    获取请假统计摘要（管理员专用）
    
    返回：员工请假统计列表
    """
    try:
        role_id = current_user.get("role_id", 15)
        if role_id not in [11, 12]:
            raise HTTPException(status_code=403, detail="无权限查看请假统计")
        
        start_date = f"{year}-{month:02d}-01"
        if month == 12:
            end_date = f"{year + 1}-01-01"
        else:
            end_date = f"{year}-{month + 1:02d}-01"
        
        with get_connection() as conn:
            result = conn.execute(text("""
                SELECT 
                    lr.employee_id,
                    e.name as employee_name,
                    COUNT(*) as leave_count,
                    json_agg(json_build_object(
                        'date', lr.leave_date,
                        'type', lr.leave_type,
                        'reason', lr.reason
                    ) ORDER BY lr.leave_date) as leave_details
                FROM leave_records lr
                LEFT JOIN employees e ON e.employee_id = lr.employee_id
                WHERE lr.leave_date >= :start_date 
                  AND lr.leave_date < :end_date
                  AND lr.is_deleted = false
                GROUP BY lr.employee_id, e.name
                ORDER BY leave_count DESC
            """), {"start_date": start_date, "end_date": end_date})
            
            summaries = []
            for row in result:
                summaries.append({
                    "employee_id": row[0],
                    "employee_name": row[1] or row[0],
                    "leave_count": row[2],
                    "leave_details": row[3]
                })
            
            return {"summaries": summaries}
    
    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"获取请假统计失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# 文件末尾保留（勿删除）

# ============== 前后端分离：静态资源由 Nginx 直接服务 ==============
# 静态文件（/agent/assets/*, /agent/index.html 等）已由 Nginx 直接服务
# 后端只处理 /agent/api/* 路由
# 如需回退到后端服务静态文件，恢复以下代码：
#
# from fastapi.staticfiles import StaticFiles
# from fastapi.responses import FileResponse
# import pathlib
# STATIC_DIR = pathlib.Path(__file__).parent / "static"
# if (STATIC_DIR / "assets").exists():
#     app.mount("/agent/assets", StaticFiles(directory=str(STATIC_DIR / "assets")), name="assets")
# @app.get("/agent/{path:path}")
# async def serve_spa(path: str):
#     if path.startswith("api/"):
#         raise HTTPException(status_code=404, detail="API endpoint not found")
#     index_file = STATIC_DIR / "index.html"
#     if index_file.exists():
#         return FileResponse(str(index_file), headers={"Cache-Control": "no-cache"})
#     raise HTTPException(status_code=404, detail="Frontend not found")
# @app.get("/agent")
# async def root():
#     index_file = STATIC_DIR / "index.html"
#     if index_file.exists():
#         return FileResponse(str(index_file), headers={"Cache-Control": "no-cache"})
#     raise HTTPException(status_code=404, detail="Frontend not found")

