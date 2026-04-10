#!/usr/bin/env python3
"""
导入甘特图格式 Excel 到数据库（作为新版本）

用法：
    python import_gantt_plan.py <excel_path> <project_id> <version_name>

示例：
    python import_gantt_plan.py /home/ubuntu/info/隆林铝厂空压站进度表.xlsx 20 V2
"""

import sys
import os
import re
from datetime import datetime

# 添加路径
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl
from sqlalchemy import create_engine, text
from dotenv import load_dotenv

load_dotenv()


def excel_date_to_date(excel_date):
    """Excel 日期序列号转换为日期"""
    from datetime import timedelta
    if isinstance(excel_date, int):
        return datetime(1899, 12, 30) + timedelta(days=excel_date)
    return None


def parse_gantt_excel(excel_path: str) -> list:
    """解析甘特图格式 Excel"""
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    
    # 构建日期映射
    col_to_date = {}
    current_date = None
    
    for col_idx in range(4, min(ws.max_column + 1, 320)):
        val_row2 = ws.cell(2, col_idx).value
        val_row3 = ws.cell(3, col_idx).value
        
        if isinstance(val_row2, int) and val_row2 > 40000:
            current_date = datetime(1899, 12, 30) + timedelta(days=val_row2)
        
        if current_date and isinstance(val_row3, int):
            try:
                date = current_date.replace(day=val_row3)
                col_to_date[col_idx] = date
            except:
                pass
    
    # 解析任务
    tasks = []
    for row_idx in range(4, min(ws.max_row + 1, 130)):
        task_name = ws.cell(row_idx, 4).value
        
        if task_name and re.match(r'^\d+\.\d+', str(task_name)):
            # 收集该行甘特图区域的所有非空单元格
            work_cols = []
            for col_idx in sorted(col_to_date.keys()):
                val = ws.cell(row_idx, col_idx).value
                if val is not None and str(val).strip() and str(val).strip() not in ['休', 'NaN']:
                    work_cols.append(col_idx)
            
            if work_cols:
                start_col = min(work_cols)
                end_col = max(work_cols)
                start_date = col_to_date.get(start_col)
                end_date = col_to_date.get(end_col)
                
                if start_date and end_date:
                    # 提取任务编号作为 task_id 的一部分
                    task_num = re.match(r'^(\d+\.\d+)', str(task_name)).group(1)
                    tasks.append({
                        'task_name': str(task_name).strip(),
                        'task_num': task_num,
                        'start_date': start_date.strftime('%Y-%m-%d'),
                        'end_date': end_date.strftime('%Y-%m-%d'),
                        'duration': (end_date - start_date).days + 1
                    })
    
    return tasks


def import_tasks(project_id: int, tasks: list, version: str = "2"):
    """导入任务到数据库"""
    db_url = os.getenv("DATABASE_URL")
    if not db_url:
        raise RuntimeError("DATABASE_URL not set")
    
    engine = create_engine(db_url)
    
    with engine.connect() as conn:
        # 1. 将旧版本的 is_latest 设为 false
        conn.execute(text("""
            UPDATE project_tasks 
            SET is_latest = false 
            WHERE project_id::integer = :project_id AND is_latest = true
        """), {"project_id": project_id})
        
        # 2. 将旧版本记录设为非当前
        conn.execute(text("""
            UPDATE project_plan_versions 
            SET is_current = false 
            WHERE project_id::integer = :project_id
        """), {"project_id": project_id})
        
        # 3. 插入新任务（修复日期格式）
        imported_count = 0
        for task in tasks:
            # 使用任务编号作为task_id（如 1.1 -> 20_1_1_V2）
            task_num = task['task_num'].replace('.', '_')
            task_id = f"{project_id}_{task_num}_V{version}"
            
            # 检查是否已存在
            exists = conn.execute(text("""
                SELECT 1 FROM project_tasks WHERE task_id = :task_id
            """), {"task_id": task_id}).fetchone()
            
            if not exists:
                # 打印调试信息
                if imported_count < 3:
                    print(f"插入任务: {task_id}, 名称: {task['task_name']}, 开始: {task['start_date']}, 结束: {task['end_date']}")
                
                conn.execute(text("""
                    INSERT INTO project_tasks 
                    (task_id, project_id, task_name, start_date, end_date, 
                     planned_hours, status, is_latest, progress, is_deleted)
                    VALUES 
                    (:task_id, :project_id, :task_name, :start_date, :end_date,
                     :planned_hours, '未开始', true, 0, false)
                """), {
                    "task_id": task_id,
                    "project_id": str(project_id),
                    "task_name": task['task_name'],
                    "start_date": task['start_date'],
                    "end_date": task['end_date'],
                    "planned_hours": task['duration'] * 8  # 假设每天8小时
                })
                imported_count += 1
        
        # 4. 创建版本记录
        version_number = f"{version}.0"
        version_name = f"V{version}"
        
        conn.execute(text("""
            INSERT INTO project_plan_versions 
            (project_id, version_number, version_name, description, upload_by, file_name, is_current, task_count)
            VALUES 
            (:project_id, :version_number, :version_name, :description, '0001', :file_name, true, :task_count)
        """), {
            "project_id": str(project_id),
            "version_number": version_number,
            "version_name": version_name,
            "description": "从甘特图格式 Excel 导入",
            "file_name": "隆林铝厂空压站进度表.xlsx",
            "task_count": imported_count
        })
        
        conn.commit()
        
        print(f"\n导入完成：")
        print(f"  - 项目 ID: {project_id}")
        print(f"  - 版本: V{version}")
        print(f"  - 新增任务: {imported_count} 个")
        print(f"  - 解析任务: {len(tasks)} 个")
        
        # 显示版本分布
        result = conn.execute(text("""
            SELECT 
                SUBSTRING(task_id FROM 'V([0-9]+)') as version,
                COUNT(*) as count
            FROM project_tasks 
            WHERE project_id::integer = :project_id AND is_deleted = false
            GROUP BY SUBSTRING(task_id FROM 'V([0-9]+)')
            ORDER BY version
        """), {"project_id": project_id}).fetchall()
        
        print(f"\n版本分布：")
        for row in result:
            print(f"  - 版本 {row[0]}: {row[1]} 个任务")


def main():
    if len(sys.argv) < 3:
        print("用法: python import_gantt_plan.py <excel_path> <project_id> [version]")
        print("示例: python import_gantt_plan.py /home/ubuntu/info/隆林铝厂空压站进度表.xlsx 20 2")
        sys.exit(1)
    
    excel_path = sys.argv[1]
    project_id = int(sys.argv[2])
    version = sys.argv[3] if len(sys.argv) > 3 else "2"
    
    if not os.path.exists(excel_path):
        print(f"错误: 文件不存在 {excel_path}")
        sys.exit(1)
    
    print(f"解析 Excel: {excel_path}")
    tasks = parse_gantt_excel(excel_path)
    print(f"找到 {len(tasks)} 个任务")
    
    print(f"\n开始导入...")
    import_tasks(project_id, tasks, version)


if __name__ == "__main__":
    from datetime import timedelta
    main()
