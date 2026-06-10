#!/usr/bin/env python3
"""
田阳铝厂阳极组装提质增效项目 - 任务导入脚本
从Excel导入最新计划到项目ID 26
"""

import sys
sys.path.insert(0, '/home/ubuntu/.openclaw/workspace/project-agent-v2/backend')

from openpyxl import load_workbook
from datetime import datetime
import psycopg2
from psycopg2.extras import execute_values

# 数据库连接
DB_URL = "postgresql://yjydb:qv52A03xcxAQCoDglUJelm4Sb@localhost:5432/project_cost_tracking"

# 项目配置
PROJECT_ID = 26
PROJECT_NAME = "田阳铝厂阳极组装提质增效项目的技术研究"
NEW_VERSION = "V2"

def parse_excel():
    """解析Excel文件，返回任务列表"""
    file_path = '/home/ubuntu/.openclaw/media/inbound/田阳铝厂阳极组装提质增效技术服务项目进度表---864df333-1ac9-44b4-b8b7-39aacec6fa34.xlsx'
    wb = load_workbook(file_path, data_only=True)
    ws = wb['Sheet1']
    
    tasks = []
    current_result = None  # 当前"结果"节点
    result_idx = 0
    process_idx = 0
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        name = row[0]
        start_date = row[1]
        duration = row[2]
        end_date = row[3]
        
        if not name:
            continue
        
        # 判断是"结果"还是"过程"
        if name.startswith('结果'):
            # 这是Level 1节点（父节点）
            result_idx += 1
            process_idx = 0
            current_result = {
                'name': name,
                'start_date': start_date,
                'end_date': end_date,
                'level': 1,
                'result_idx': result_idx,
                'children': []
            }
            tasks.append(current_result)
        elif name.startswith('过程'):
            # 这是Level 2节点（子任务）
            process_idx += 1
            if current_result:
                child = {
                    'name': name,
                    'start_date': start_date,
                    'end_date': end_date,
                    'level': 2,
                    'parent_idx': result_idx,
                    'process_idx': process_idx
                }
                current_result['children'].append(child)
    
    return tasks

def generate_task_id(project_id, level, idx1, idx2=None, version="V2"):
    """生成任务ID"""
    if level == 1:
        # 父节点: {project_id}_{idx}_V{version}
        return f"{project_id}_{idx1}_{version}"
    else:
        # 子节点: {project_id}_{idx1}_{idx2}_V{version}
        return f"{project_id}_{idx1}_{idx2}_{version}"

def import_tasks():
    """导入任务到数据库"""
    tasks = parse_excel()
    
    conn = psycopg2.connect(DB_URL)
    cur = conn.cursor()
    
    try:
        # 1. 将现有V1任务标记为 is_latest = false
        cur.execute("""
            UPDATE project_tasks 
            SET is_latest = false 
            WHERE project_id = %s AND is_deleted = false
        """, (str(PROJECT_ID),))
        print(f"已将 {cur.rowcount} 个V1任务标记为非最新版本")
        
        # 2. 插入新任务
        task_count = 0
        task_rows = []
        
        for result in tasks:
            # 插入Level 1节点
            task_id = generate_task_id(PROJECT_ID, 1, result['result_idx'])
            task_name = f"{result['result_idx']}. {result['name'].replace('结果', '月度服务')}"
            
            task_rows.append((
                task_id,
                str(PROJECT_ID),
                task_name,
                None,  # parent_task_id
                1,     # task_level
                result['start_date'].date() if hasattr(result['start_date'], 'date') else result['start_date'],
                result['end_date'].date() if hasattr(result['end_date'], 'date') else result['end_date'],
                '未开始',
                True,  # is_latest
                False  # is_deleted
            ))
            task_count += 1
            
            # 插入Level 2子任务
            for proc in result['children']:
                child_task_id = generate_task_id(PROJECT_ID, 2, result['result_idx'], proc['process_idx'])
                parent_task_id = generate_task_id(PROJECT_ID, 1, result['result_idx'])
                # 清理任务名称：去掉"过程N："前缀
                proc_name = proc['name']
                if '：' in proc_name:
                    proc_name = proc_name.split('：', 1)[1]
                task_name = f"{result['result_idx']}.{proc['process_idx']} {proc_name}"
                
                task_rows.append((
                    child_task_id,
                    str(PROJECT_ID),
                    task_name,
                    parent_task_id,
                    2,     # task_level
                    proc['start_date'].date() if hasattr(proc['start_date'], 'date') else proc['start_date'],
                    proc['end_date'].date() if hasattr(proc['end_date'], 'date') else proc['end_date'],
                    '未开始',
                    True,  # is_latest
                    False  # is_deleted
                ))
                task_count += 1
        
        # 批量插入
        execute_values(cur, """
            INSERT INTO project_tasks 
            (task_id, project_id, task_name, parent_task_id, task_level, start_date, end_date, status, is_latest, is_deleted)
            VALUES %s
        """, task_rows)
        
        conn.commit()
        print(f"\n✅ 成功导入 {task_count} 个任务")
        
        # 3. 验证导入结果
        cur.execute("""
            SELECT COUNT(*) FROM project_tasks 
            WHERE project_id = %s AND is_latest = true AND is_deleted = false
        """, (str(PROJECT_ID),))
        count = cur.fetchone()[0]
        print(f"项目{PROJECT_ID}当前最新任务数: {count}")
        
        # 显示部分任务
        cur.execute("""
            SELECT task_id, task_name, task_level, parent_task_id, start_date, end_date
            FROM project_tasks 
            WHERE project_id = %s AND is_latest = true AND is_deleted = false
            ORDER BY task_id
            LIMIT 15
        """, (str(PROJECT_ID),))
        
        print("\n前15个任务:")
        print("-" * 100)
        for row in cur.fetchall():
            print(f"{row[0]:25} | {row[1]:30} | Level {row[2]} | Parent: {row[3] or 'None':15} | {row[4]} ~ {row[5]}")
        
    except Exception as e:
        conn.rollback()
        print(f"❌ 导入失败: {e}")
        import traceback
        traceback.print_exc()
    finally:
        cur.close()
        conn.close()

if __name__ == "__main__":
    print(f"开始导入项目 {PROJECT_ID}: {PROJECT_NAME}")
    print("=" * 80)
    import_tasks()
